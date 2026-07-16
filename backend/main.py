from __future__ import annotations

import base64
import binascii
import copy
import gc
import hashlib
import html
import json
import os
import re
import secrets
import shutil
import smtplib
import sqlite3
import ssl
import time
import urllib.parse
import uuid
import zipfile
import threading
import tempfile
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime
from functools import wraps
from io import BytesIO
import unicodedata
from contextlib import asynccontextmanager
from email.message import EmailMessage
from pathlib import Path
from typing import Any

from argon2 import PasswordHasher
from argon2.exceptions import InvalidHashError, VerificationError, VerifyMismatchError
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request, Response
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.trustedhost import TrustedHostMiddleware
try:
    from opencc import OpenCC
except ImportError:  # Optional fallback for maintenance environments before dependencies are installed.
    class OpenCC:  # type: ignore[override]
        def __init__(self, _: str): pass
        def convert(self, value: str) -> str: return value
from pydantic import BaseModel, EmailStr, Field
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:  # Dependency is required in production; this keeps maintenance imports readable.
    Workbook = None
    load_workbook = None
    Comment = Alignment = Font = PatternFill = None
    class InvalidFileException(Exception):
        pass
from backend.core.game_registry import enabled_projects, get_game_config, get_source_policy, load_registry, resolve_game_index
from backend.services.catalog_health import repair_plan as build_catalog_repair_plan, scan_catalog
from backend.services.achievement_governance import make_issue as make_governance_issue, scan_governance, summarize_plan, stable_hash as governance_hash
from backend.services.governance_contract import (
    ACTION_SPECS as GOVERNANCE_ACTION_SPECS,
    DECISION_ONLY_ACTIONS as GOVERNANCE_DECISION_ONLY_ACTIONS,
    SUPPORTED_ACTIONS as GOVERNANCE_SUPPORTED_ACTIONS,
    action_public_payload as governance_action_public_payload,
    canonical_action as canonical_governance_action,
    allowed_actions_for_issue as governance_allowed_actions_for_issue,
)
from backend.services.sync_engine import apply_decisions as apply_sync_decisions, build_diff as build_sync_diff, default_selection_decisions as build_default_sync_selections, row_fingerprint as shared_catalog_fingerprint
from backend.services.catalog_repository import normalize_catalog_rows, replace_catalog_rows
from backend.services.catalog_sorting import catalog_sort_key, sort_catalog_rows, sync_change_sort_key
from backend.services.official_id_model import (
    migrate_wuwa_to_official_ids,
    normalize_all_game_official_order,
    official_id_number,
    sanitize_legacy_id_display,
    verify_official_id_model,
)
from backend.services.source_pipeline import PIPELINE_VERSION as SOURCE_PIPELINE_VERSION, adapter_id as source_adapter_id, common_metadata as source_common_metadata, source_error_code
from backend.services.game_data_sources import RepositorySourceError, prepare_repository_candidate, test_repository_source, source_cache_path
from backend.services.relation_governance import DERIVED_RELATION_FIELDS, expected_relation_state, validate_relation_state
from backend.core.paths import (
    ACCOUNT_INDEX,
    ADMIN_INDEX,
    CACHE_FILE,
    DATA_DIR,
    GAME_ICON_FILES,
    GENSHIN_CATALOG_FILE,
    HSR_ACHIEVEMENTS_METADATA_CACHE_FILE,
    HSR_CATALOG_FILE,
    HSR_OFFICIAL_REWARD_FILE,
    HUB_INDEX,
    LOG_DIR,
    META_FILE,
    OFFICIAL_ZH_TW_FILE,
    OUTBOX_DIR,
    PROJECTS_DIR,
    ROOT,
    SITE_DIR,
    WUWA_CATALOG_FILE,
    WUWA_CHOICE_GROUPS_FILE,
    ZZZ_CATALOG_FILE,
    ensure_runtime_directories,
    game_catalog_file,
    game_relation_file,
)
from backend.wuwa_catalog import build_wuwa_catalog_file
from backend.wuwa_categories import canonicalize_wuwa_category, sort_wuwa_achievement_rows

load_dotenv(ROOT / ".env")
ensure_runtime_directories()

def resolve_runtime_database_path(value: str | None) -> Path:
    raw=str(value or "").strip()
    default=(DATA_DIR / "app.db").resolve()
    if not raw:
        return default
    # A legacy package stored an installation-specific Windows path in .env.
    # Keep explicit existing paths, but fall back to the current project root when moved.
    windows_absolute=bool(re.match(r"^[A-Za-z]:[\\/]",raw))
    if windows_absolute:
        if os.name!="nt":
            return default
        configured=Path(raw)
        return configured if configured.exists() or configured.parent.exists() else default
    configured=Path(raw)
    if not configured.is_absolute():
        configured=ROOT/configured
    return configured.resolve()

DB_FILE = resolve_runtime_database_path(os.getenv("DATABASE_PATH"))
SYNC_PREVIEW_RETENTION_SECONDS = 30 * 24 * 60 * 60
DB_FILE.parent.mkdir(parents=True, exist_ok=True)
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "http://127.0.0.1:817").rstrip("/")
APP_ENV = os.getenv("APP_ENV", "development").lower()
OPEN_SOURCE_EMPTY_DATA = os.getenv("OPEN_SOURCE_EMPTY_DATA", "1").lower() in {"1","true","yes","on"}
COOKIE_NAME = os.getenv("SESSION_COOKIE_NAME", "game_achievement_session")
COOKIE_SECURE = os.getenv("SESSION_COOKIE_SECURE", "false").lower() in {"1","true","yes","on"}
SESSION_SECONDS = int(os.getenv("SESSION_SECONDS", str(30 * 24 * 60 * 60)))
VERIFY_SECONDS = int(os.getenv("VERIFY_TOKEN_SECONDS", str(24 * 60 * 60)))
RESET_SECONDS = int(os.getenv("RESET_TOKEN_SECONDS", str(30 * 60)))
TRUSTED_HOSTS = [x.strip() for x in os.getenv("TRUSTED_HOSTS", "localhost,127.0.0.1").split(",") if x.strip()]
TRUSTED_ORIGINS = {x.strip().rstrip("/") for x in os.getenv("TRUSTED_ORIGINS", PUBLIC_BASE_URL + ",http://localhost:817,http://127.0.0.1:817").split(",") if x.strip()}
MAIL_DELIVERY = os.getenv("MAIL_DELIVERY", "console").lower()
SMTP_HOST = os.getenv("SMTP_HOST", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USERNAME = os.getenv("SMTP_USERNAME", "")
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "")
SMTP_FROM = os.getenv("SMTP_FROM", SMTP_USERNAME or "noreply@example.com")
SMTP_STARTTLS = os.getenv("SMTP_STARTTLS", "true").lower() in {"1","true","yes","on"}
SMTP_SSL = os.getenv("SMTP_SSL", "false").lower() in {"1","true","yes","on"}
SMTP_VALIDATE_CERT = os.getenv("SMTP_VALIDATE_CERT", "true").lower() in {"1","true","yes","on"}
SITE_OWNER_EMAIL = os.getenv("SITE_OWNER_EMAIL", "").strip().casefold()
ADMIN_EMAILS = {normalize.strip().casefold() for normalize in os.getenv("ADMIN_EMAILS", "").split(",") if normalize.strip()}
if SITE_OWNER_EMAIL:
    ADMIN_EMAILS.add(SITE_OWNER_EMAIL)

ENTRY_ID = "1220879855033786368"
OFFICIAL_PAGE = f"https://wiki.kurobbs.com/mc/item/{ENTRY_ID}"
WUWA_SHARED_MODEL_MIGRATION = "20260624-wuwa-shared-model-v1"
GAME_ICON_LABELS = {
    "wuwa": "Wuthering Waves",
    "hsr": "Honkai: Star Rail",
    "genshin": "Genshin Impact",
    "zzz": "Zenless Zone Zero",
    "hna": "Honkai: Nexus Anima",
}

# 崩壞：星穹鐵道成就資料以 HoYoLAB 官方百科為主要來源。
# 程式會使用該頁面所使用的官方 HoYoWiki API 取得繁體中文清單與官方排列順序。
# StarRailRes 僅在官方清單未提供穩定成就 ID、條件、隱藏狀態或版本時補齊欄位，
# 不會取代 HoYoLAB 的名稱與排列順序。
TRADITIONAL_CONVERTER = OpenCC("s2twp")
PASSWORD_HASHER = PasswordHasher(time_cost=3, memory_cost=65536, parallelism=4, hash_len=32, salt_len=16)
AUTH_RATE_WINDOW_SECONDS = 5 * 60

class EmailPassword(BaseModel):
    email: EmailStr
    password: str = Field(min_length=8, max_length=200)

class LoginPayload(BaseModel):
    identifier: str = Field(default="", max_length=254)
    email: str = Field(default="", max_length=254)
    password: str = Field(min_length=8, max_length=200)

class EmailOnly(BaseModel):
    email: EmailStr

class TokenOnly(BaseModel):
    token: str = Field(min_length=20, max_length=500)

class ResetPassword(BaseModel):
    token: str = Field(min_length=20, max_length=500)
    password: str = Field(min_length=8, max_length=200)

class ChangePasswordPayload(BaseModel):
    current_password: str = Field(min_length=8, max_length=200)
    new_password: str = Field(min_length=8, max_length=200)

class UsernamePayload(BaseModel):
    username: str = Field(min_length=3, max_length=30)

class ProgressSet(BaseModel):
    achievement_id: str = Field(min_length=1, max_length=200)
    completed: bool

class ProgressBatch(BaseModel):
    achievement_ids: list[str]
    completed: bool

class ProgressReplace(BaseModel):
    achievement_ids: list[str]

class AdminRoleUpdate(BaseModel):
    role: str = Field(min_length=4, max_length=10)

class AdminUsernamePayload(BaseModel):
    username: str = Field(min_length=3, max_length=30)

class AdminStatusUpdate(BaseModel):
    active: bool

class AdminVerificationUpdate(BaseModel):
    verified: bool

class AdminUserBlockPayload(BaseModel):
    reason: str = Field(default="", max_length=1000)

class AdminSyncApplyPayload(BaseModel):
    preview_id: str = Field(min_length=8, max_length=100)
    selected_change_ids: list[str] | None = None
    decisions: dict[str, Any] = Field(default_factory=dict)
    reason: str = Field(default="", max_length=1000)

class SyncHistoryRollbackPayload(BaseModel):
    reason: str = Field(min_length=3, max_length=2000)

class CatalogRepairPayload(BaseModel):
    scan_id: str = Field(min_length=8, max_length=100)
    actions: dict[str, Any] = Field(default_factory=dict)
    reason: str = Field(default="", max_length=1000)

class AuditArchivePayload(BaseModel):
    reason: str = Field(min_length=3, max_length=1000)

class AdminEmailUpdatePayload(BaseModel):
    email: EmailStr
    verified: bool = False

class AdminPasswordResetPayload(BaseModel):
    password: str = Field(min_length=8, max_length=200)

class RelationGroupPayload(BaseModel):
    relation_type: str = Field(min_length=5, max_length=20)
    group_id: str = Field(min_length=2, max_length=200)
    achievement_ids: list[str]
    name: str = Field(default="", max_length=300)
    basis: str = Field(default="", max_length=2000)

class RelationValidationPreviewPayload(BaseModel):
    validation_id: str = Field(min_length=8, max_length=100)
    actions: list[dict[str, Any]] = Field(default_factory=list)
    reason: str = Field(default="", max_length=2000)

class RelationValidationExecutePayload(BaseModel):
    confirmation_text: str = Field(default="", max_length=100)

class RelationValidationRollbackPayload(BaseModel):
    reason: str = Field(min_length=3, max_length=2000)

class BackupRestorePayload(BaseModel):
    filename: str = Field(min_length=5, max_length=255)

class AdminTestEmailPayload(BaseModel):
    recipient: EmailStr

class AchievementReportCreate(BaseModel):
    achievement_id: str = Field(min_length=1, max_length=200)
    achievement_name: str = Field(min_length=1, max_length=300)
    report_type: str = Field(min_length=1, max_length=60)
    message: str = Field(min_length=3, max_length=3000)

class AchievementReportUpdate(BaseModel):
    status: str = Field(min_length=2, max_length=30)
    admin_note: str = Field(default="", max_length=3000)

class AnnouncementPayload(BaseModel):
    title: str = Field(min_length=1, max_length=200)
    body: str = Field(min_length=1, max_length=5000)
    level: str = Field(default="info", max_length=20)
    is_active: bool = True
    pinned: bool = False
    starts_at: int | None = None
    ends_at: int | None = None

class NotificationPayload(BaseModel):
    title: str = Field(min_length=1, max_length=200)
    body: str = Field(min_length=1, max_length=3000)
    kind: str = Field(default="info", max_length=30)
    link: str = Field(default="", max_length=500)
    target_email: str = Field(default="", max_length=254)
    target_scope: str = Field(default="all", max_length=30)

class RedeemGamePayload(BaseModel):
    game_id: str = Field(min_length=2, max_length=40)
    name: str = Field(min_length=1, max_length=120)
    display_order: int = Field(default=0, ge=-100000, le=100000)
    enabled: bool = True
    note: str = Field(default="", max_length=1000)

class RedeemGameUpdatePayload(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    display_order: int = Field(default=0, ge=-100000, le=100000)
    enabled: bool = True
    note: str = Field(default="", max_length=1000)

class RedeemServerPayload(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    display_order: int = Field(default=0, ge=-100000, le=100000)
    enabled: bool = True

class RedeemReorderPayload(BaseModel):
    item_ids: list[str] = Field(min_length=1, max_length=500)

class RedeemCodePayload(BaseModel):
    game_id: str = Field(min_length=2, max_length=40)
    code: str = Field(min_length=1, max_length=200)
    source: str = Field(default="", max_length=500)
    description: str = Field(default="", max_length=1000)
    reward: str = Field(default="", max_length=500)
    start_at: int | None = None
    end_at: int | None = None
    server_ids: list[str] = Field(default_factory=list)
    redeem_url: str = Field(default="", max_length=500)
    enabled: bool = True

class RedeemImportPreviewPayload(BaseModel):
    default_game_id: str = Field(default="", max_length=40)
    rows: list[dict[str, Any]] = Field(min_length=1, max_length=500)
    reason: str = Field(default="", max_length=1000)
    input_type: str = Field(default="editor", max_length=20)
    source_filename: str = Field(default="", max_length=255)
    source_sheet: str = Field(default="", max_length=255)

class RedeemExcelImportPayload(BaseModel):
    filename: str = Field(min_length=1, max_length=255)
    content_base64: str = Field(min_length=4, max_length=14_500_000)

class RedeemImportExecutePayload(BaseModel):
    reason: str = Field(default="", max_length=1000)

class RedeemImportRollbackPayload(BaseModel):
    reason: str = Field(min_length=3, max_length=1000)

class RedeemNotificationPreferencesPayload(BaseModel):
    game_ids: list[str] = Field(default_factory=list, max_length=100)

class BlockEntryPayload(BaseModel):
    kind: str = Field(min_length=2, max_length=20)
    value: str = Field(min_length=1, max_length=300)
    reason: str = Field(default="", max_length=1000)
    active: bool = True

class AchievementEditPayload(BaseModel):
    name: str = Field(min_length=1, max_length=300)
    condition: str = Field(default="", max_length=5000)
    version: str = Field(default="未標示", max_length=100)
    category: str = Field(default="未辨識分類", max_length=200)
    reward: int = Field(default=0, ge=0, le=100000)
    hidden: bool = False
    tags: list[str] = Field(default_factory=list)
    is_deleted: bool = False
    source: str = Field(default="override", max_length=30)

class AchievementPermanentDeletePayload(BaseModel):
    confirmation_text: str = Field(min_length=1, max_length=300)
    reason: str = Field(min_length=3, max_length=2000)

class AchievementPermanentRestorePayload(BaseModel):
    reason: str = Field(min_length=3, max_length=2000)


class AchievementCategoryCreatePayload(BaseModel):
    name: str = Field(min_length=1, max_length=200)


class AchievementCategoryUpdatePayload(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    reason: str = Field(min_length=3, max_length=1000)


class AchievementCategoryMergePayload(BaseModel):
    target_category_id: str = Field(min_length=8, max_length=100)
    reason: str = Field(min_length=3, max_length=1000)


class AchievementCategoryDeletePayload(BaseModel):
    reason: str = Field(min_length=3, max_length=1000)


class AchievementCategoryReorderPayload(BaseModel):
    category_ids: list[str] = Field(min_length=1, max_length=500)


class TicketCreatePayload(BaseModel):
    subject: str = Field(min_length=2, max_length=200)
    message: str = Field(min_length=3, max_length=5000)

class TicketReplyPayload(BaseModel):
    message: str = Field(min_length=1, max_length=5000)
    status: str | None = Field(default=None, max_length=30)

class TicketStatusPayload(BaseModel):
    status: str = Field(min_length=2, max_length=30)
    priority: str = Field(default="normal", max_length=20)

class MergeAccountsPayload(BaseModel):
    source_user_id: str = Field(min_length=10, max_length=100)
    target_user_id: str = Field(min_length=10, max_length=100)

class CatalogValidationPayload(BaseModel):
    items: list[dict[str, Any]] = Field(default_factory=list)

class GovernanceScanPayload(BaseModel):
    include_catalog: bool = True
    include_progress: bool = True
    include_relations: bool = True
    include_aliases: bool = True
    include_sources: bool = True
    similarity_threshold: float = Field(default=0.94, ge=0.80, le=1.0)
    background: bool = True

class GovernanceDraftPayload(BaseModel):
    scan_id: str = Field(min_length=8, max_length=100)
    name: str = Field(default="", max_length=200)
    actions: list[dict[str, Any]] = Field(default_factory=list)
    reason: str = Field(default="", max_length=2000)

class GovernanceExecutePayload(BaseModel):
    confirmation_text: str = Field(default="", max_length=100)

class GovernanceIssueStatePayload(BaseModel):
    state: str = Field(min_length=2, max_length=40)
    reason: str = Field(default="", max_length=2000)
    permanent: bool = False
    recheck_on_change: bool = True

class GovernanceRollbackPayload(BaseModel):
    reason: str = Field(min_length=3, max_length=2000)


def now() -> int:
    return int(time.time())


def digest_token(token: str) -> str:
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def normalize_email(value: str) -> str:
    return str(value).strip().casefold()

def normalize_username(value: str) -> str:
    return str(value).strip().casefold()


def validate_username(value: str) -> tuple[str, str]:
    username = str(value or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9]{3,30}", username):
        raise HTTPException(status_code=400, detail="使用者名稱至少 3 個字，且只能使用英文字母與數字。")
    return username, normalize_username(username)


def is_site_owner_email(value: str) -> bool:
    return bool(SITE_OWNER_EMAIL) and normalize_email(value) == SITE_OWNER_EMAIL


class ManagedSQLiteConnection(sqlite3.Connection):
    """SQLite connection that commits or rolls back and then always closes.

    The standard sqlite3 context manager does not close the file handle.  This
    subclass keeps direct-call compatibility while making every
    ``with connect_db()`` block release Windows file locks deterministically.
    """

    def __exit__(self, exc_type, exc_value, traceback):
        try:
            return super().__exit__(exc_type, exc_value, traceback)
        finally:
            self.close()


def connect_db() -> sqlite3.Connection:
    db = sqlite3.connect(DB_FILE, timeout=20, factory=ManagedSQLiteConnection)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA foreign_keys = ON")
    db.execute("PRAGMA busy_timeout = 20000")
    return db




def _load_wuwa_choice_group_payload() -> dict[str,Any]:
    """讀取鳴潮關聯／二選一成就設定。"""
    if not WUWA_CHOICE_GROUPS_FILE.exists():
        return {"groups":[]}
    payload=json.loads(WUWA_CHOICE_GROUPS_FILE.read_text(encoding="utf-8-sig"))
    groups=payload.get("groups") if isinstance(payload,dict) else None
    if not isinstance(groups,list):
        raise ValueError("鳴潮關聯成就資料格式錯誤：缺少 groups。")
    return payload





def _load_relation_group_rows(game_id: str, path: Path, default_type: str="stage") -> list[tuple[str,str,str,str,int]]:
    """讀取通用關聯群組，回傳 game/group/achievement/type/order。"""
    if not path.exists():
        return []
    payload=json.loads(path.read_text(encoding="utf-8-sig"))
    groups=payload.get("groups") if isinstance(payload,dict) else None
    if not isinstance(groups,list):
        raise ValueError(f"{game_id} 關聯成就資料格式錯誤：缺少 groups。")
    rows=[]; seen=set()
    for index,group in enumerate(groups,1):
        if not isinstance(group,dict):
            raise ValueError(f"{game_id} 關聯成就第 {index} 組格式錯誤。")
        group_id=str(group.get("id") or f"{game_id}-{default_type}-{index:03d}").strip()
        relation_type=str(group.get("type") or default_type).strip().lower()
        if relation_type not in {"exclusive","stage"}:
            raise ValueError(f"{game_id} 關聯群組 {group_id} 類型錯誤。")
        if relation_type!=default_type:
            raise ValueError(f"{game_id} 關聯群組 {group_id} 的類型 {relation_type} 與檔案 {default_type} 不一致。")
        raw_ids=[str(value or "").strip() for value in (group.get("achievement_ids") or []) if str(value or "").strip()]
        if len(raw_ids)!=len(set(raw_ids)):
            raise ValueError(f"{game_id} 關聯群組 {group_id or index} 含有重複成員。")
        ids=raw_ids
        if not group_id or len(ids)<2:
            raise ValueError(f"{game_id} 關聯群組 {group_id or index} 至少需要 2 項成就。")
        for order,achievement_id in enumerate(ids,1):
            if achievement_id in seen:
                raise ValueError(f"{game_id} 成就重複出現在不同關聯群組：{achievement_id}")
            seen.add(achievement_id)
            rows.append((game_id,group_id,achievement_id,relation_type,order if relation_type=="stage" else 0))
    return rows


def _sync_relation_groups(db: sqlite3.Connection, game_id: str, path: Path, relation_type: str="stage") -> tuple[int,int]:
    rows=_load_relation_group_rows(game_id,path,relation_type)
    db.execute("delete from game_achievement_choice_groups where game_id=? and relation_type=?",(game_id,relation_type))
    if rows:
        db.executemany(
            "insert into game_achievement_choice_groups(game_id,group_id,achievement_id,relation_type,stage_order,updated_at) values(?,?,?,?,?,?)",
            [(*row,now()) for row in rows],
        )
    return len({row[1] for row in rows}),len(rows)



def _refresh_wuwa_catalog_file() -> dict[str,Any]:
    if not CACHE_FILE.exists():
        raise ValueError("尚未建立鳴潮官方成就快取。")
    return build_wuwa_catalog_file(
        CACHE_FILE,
        WUWA_CATALOG_FILE,
        OFFICIAL_ZH_TW_FILE if OFFICIAL_ZH_TW_FILE.exists() else None,
        OFFICIAL_PAGE,
    )


def _load_wuwa_catalog_rows(catalog_path: Path | None = None) -> list[dict[str,Any]]:
    path=catalog_path or WUWA_CATALOG_FILE
    if not path.exists():
        if catalog_path is not None:
            raise ValueError(f"鳴潮成就目錄不存在：{path}")
        _refresh_wuwa_catalog_file()
    payload=json.loads(path.read_text(encoding="utf-8-sig"))
    raw=payload.get("items") if isinstance(payload,dict) else payload
    if not isinstance(raw,list):
        raise ValueError("鳴潮成就資料格式錯誤：缺少 items。")
    rows=[]
    for index,item in enumerate(raw):
        if not isinstance(item,dict):
            continue
        achievement_id=str(item.get("id") or item.get("achievement_id") or "").strip()
        name=str(item.get("name") or "").strip()
        condition=str(item.get("condition") or "").strip()
        if not achievement_id or not name or not condition:
            continue
        rows.append({
            "achievement_id":achievement_id,
            "name":name,
            "condition":condition,
            "version":str(item.get("version") or "未標示").strip(),
            "category":canonicalize_wuwa_category(item.get("category") or "未辨識分類"),
            "reward":int(item.get("reward") or 0),
            "hidden":1 if item.get("hidden") else 0,
            "tags_json":json.dumps(item.get("tags") if isinstance(item.get("tags"),list) else [],ensure_ascii=False),
            "source":str(item.get("source") or "kuro-official-wiki").strip(),
            "source_order":int(item.get("sourceOrder") if item.get("sourceOrder") is not None else index),
        })
    rows=sort_wuwa_achievement_rows(rows)
    if len(rows)<1000:
        raise ValueError(f"鳴潮成就資料筆數異常：{len(rows)}")
    return rows


def _sync_wuwa_catalog(db: sqlite3.Connection) -> tuple[int,int]:
    rows=_load_wuwa_catalog_rows()
    stamp=now()
    db.execute("delete from game_catalog_items where game_id='wuwa' and lower(source) not in ('manual','admin')")
    db.executemany(
        """insert into game_catalog_items(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order,updated_at)
        values('wuwa',?,?,?,?,?,?,?,?,?,?,?)
        on conflict(game_id,achievement_id) do nothing""",
        [(r["achievement_id"],r["name"],r["condition"],r["version"],r["category"],r["reward"],r["hidden"],r["tags_json"],r["source"],r["source_order"],stamp) for r in rows],
    )
    return len(rows),len({r["achievement_id"] for r in rows})


def _migrate_wuwa_shared_model(db: sqlite3.Connection) -> dict[str,int]:
    marker=db.execute("select name from schema_migrations where name=?",(WUWA_SHARED_MODEL_MIGRATION,)).fetchone()
    if marker:
        return {"already_applied":1}

    # 先修正舊版合併式互斥成就，再把最終結果一次搬入共用表。
    _migrate_wuwa_legacy_choice_progress(db)
    _repair_wuwa_choice_group_progress(db)

    legacy_counts={
        "progress":int(db.execute("select count(*) c from progress").fetchone()["c"] or 0),
        "reports":int(db.execute("select count(*) c from achievement_reports").fetchone()["c"] or 0),
        "overrides":int(db.execute("select count(*) c from achievement_overrides").fetchone()["c"] or 0),
        "deleted":int(db.execute("select count(*) c from deleted_achievements").fetchone()["c"] or 0),
        "revisions":int(db.execute("select count(*) c from achievement_revisions").fetchone()["c"] or 0),
        "featured":int(db.execute("select count(*) c from featured_achievements").fetchone()["c"] or 0),
    }

    db.execute(
        """insert into game_progress(game_id,user_id,achievement_id,completed_at)
        select 'wuwa',user_id,achievement_id,completed_at from progress where true
        on conflict(game_id,user_id,achievement_id) do update set completed_at=min(game_progress.completed_at,excluded.completed_at)"""
    )
    db.execute(
        """insert into game_achievement_reports(id,game_id,user_id,achievement_id,achievement_name,report_type,message,status,admin_note,created_at,updated_at)
        select id,'wuwa',user_id,achievement_id,achievement_name,report_type,message,status,admin_note,created_at,updated_at
        from achievement_reports where true
        on conflict(id) do nothing"""
    )
    db.execute(
        """insert into game_achievement_overrides(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,is_deleted,source,updated_by,updated_at)
        select 'wuwa',achievement_id,name,condition,version,category,reward,hidden,tags_json,is_deleted,source,updated_by,updated_at
        from achievement_overrides where true
        on conflict(game_id,achievement_id) do update set
          name=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.name else game_achievement_overrides.name end,
          condition=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.condition else game_achievement_overrides.condition end,
          version=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.version else game_achievement_overrides.version end,
          category=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.category else game_achievement_overrides.category end,
          reward=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.reward else game_achievement_overrides.reward end,
          hidden=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.hidden else game_achievement_overrides.hidden end,
          tags_json=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.tags_json else game_achievement_overrides.tags_json end,
          is_deleted=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.is_deleted else game_achievement_overrides.is_deleted end,
          source=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.source else game_achievement_overrides.source end,
          updated_by=case when excluded.updated_at>=game_achievement_overrides.updated_at then excluded.updated_by else game_achievement_overrides.updated_by end,
          updated_at=max(game_achievement_overrides.updated_at,excluded.updated_at)"""
    )
    db.execute(
        """insert into game_deleted_achievements(game_id,achievement_id,achievement_name,deleted_by,deleted_at)
        select 'wuwa',achievement_id,achievement_name,deleted_by,deleted_at from deleted_achievements where true
        on conflict(game_id,achievement_id) do update set
          achievement_name=case when excluded.deleted_at>=game_deleted_achievements.deleted_at then excluded.achievement_name else game_deleted_achievements.achievement_name end,
          deleted_by=case when excluded.deleted_at>=game_deleted_achievements.deleted_at then excluded.deleted_by else game_deleted_achievements.deleted_by end,
          deleted_at=max(game_deleted_achievements.deleted_at,excluded.deleted_at)"""
    )
    db.execute(
        """insert into game_featured_achievements(game_id,achievement_id,note,sort_order,is_active,updated_by,updated_at)
        select 'wuwa',achievement_id,note,sort_order,is_active,updated_by,updated_at from featured_achievements where true
        on conflict(game_id,achievement_id) do update set
          note=case when excluded.updated_at>=game_featured_achievements.updated_at then excluded.note else game_featured_achievements.note end,
          sort_order=case when excluded.updated_at>=game_featured_achievements.updated_at then excluded.sort_order else game_featured_achievements.sort_order end,
          is_active=case when excluded.updated_at>=game_featured_achievements.updated_at then excluded.is_active else game_featured_achievements.is_active end,
          updated_by=case when excluded.updated_at>=game_featured_achievements.updated_at then excluded.updated_by else game_featured_achievements.updated_by end,
          updated_at=max(game_featured_achievements.updated_at,excluded.updated_at)"""
    )
    db.execute(
        """insert into game_achievement_revisions(game_id,achievement_id,action,snapshot_json,actor_user_id,created_at)
        select 'wuwa',achievement_id,action,snapshot_json,actor_user_id,created_at from achievement_revisions"""
    )

    # 延續舊版即時版本號，避免更新後已開啟頁面看不到資料變更。
    for scope in ("catalog","stats","reports"):
        old=db.execute("select revision,updated_at from live_revisions where scope=?",(scope,)).fetchone()
        if old:
            db.execute(
                """insert into game_live_revisions(game_id,scope,revision,updated_at) values('wuwa',?,?,?)
                on conflict(game_id,scope) do update set
                  revision=max(game_live_revisions.revision,excluded.revision),
                  updated_at=max(game_live_revisions.updated_at,excluded.updated_at)""",
                (scope,int(old["revision"] or 0),int(old["updated_at"] or 0)),
            )

    _repair_choice_group_progress(db,"wuwa")
    shared_counts={
        "progress":int(db.execute("select count(*) c from game_progress where game_id='wuwa'").fetchone()["c"] or 0),
        "reports":int(db.execute("select count(*) c from game_achievement_reports where game_id='wuwa'").fetchone()["c"] or 0),
        "overrides":int(db.execute("select count(*) c from game_achievement_overrides where game_id='wuwa'").fetchone()["c"] or 0),
        "deleted":int(db.execute("select count(*) c from game_deleted_achievements where game_id='wuwa'").fetchone()["c"] or 0),
        "revisions":int(db.execute("select count(*) c from game_achievement_revisions where game_id='wuwa'").fetchone()["c"] or 0),
        "featured":int(db.execute("select count(*) c from game_featured_achievements where game_id='wuwa'").fetchone()["c"] or 0),
    }
    for key,legacy_count in legacy_counts.items():
        if shared_counts[key] < legacy_count:
            raise RuntimeError(f"鳴潮共用資料模型遷移不完整：{key} 舊資料 {legacy_count}，新資料 {shared_counts[key]}")

    details={"legacy":legacy_counts,"shared":shared_counts}
    db.execute(
        "insert into schema_migrations(name,applied_at,details_json) values(?,?,?)",
        (WUWA_SHARED_MODEL_MIGRATION,now(),json.dumps(details,ensure_ascii=False)),
    )
    return shared_counts


def _verify_wuwa_shared_model(db: sqlite3.Connection) -> dict[str,int]:
    marker=db.execute("select name from schema_migrations where name=?",(WUWA_SHARED_MODEL_MIGRATION,)).fetchone()
    if not marker:
        raise RuntimeError("鳴潮共用資料模型遷移尚未完成。")
    catalog=int(db.execute("select count(*) c from game_catalog_items where game_id='wuwa'").fetchone()["c"] or 0)
    if catalog<1000:
        raise RuntimeError(f"鳴潮共用成就目錄筆數異常：{catalog}")
    relation_missing=int(db.execute(
        """select count(*) c from game_achievement_choice_groups g
        left join game_catalog_items c on c.game_id=g.game_id and c.achievement_id=g.achievement_id
        where g.game_id='wuwa' and c.achievement_id is null"""
    ).fetchone()["c"] or 0)
    if relation_missing:
        raise RuntimeError(f"鳴潮有 {relation_missing} 筆關聯成就不在共用目錄中。")
    duplicate_choice=int(db.execute(
        """select count(*) c from (
          select p.user_id,g.group_id,count(*) n
          from game_progress p join game_achievement_choice_groups g
            on g.game_id=p.game_id and g.achievement_id=p.achievement_id
          where p.game_id='wuwa' and g.relation_type='exclusive'
          group by p.user_id,g.group_id having count(*)>1
        )"""
    ).fetchone()["c"] or 0)
    if duplicate_choice:
        raise RuntimeError(f"鳴潮共用進度仍有 {duplicate_choice} 組互斥成就重複。")
    return {
        "catalog":catalog,
        "progress":int(db.execute("select count(*) c from game_progress where game_id='wuwa'").fetchone()["c"] or 0),
        "users_with_progress":int(db.execute("select count(distinct user_id) c from game_progress where game_id='wuwa'").fetchone()["c"] or 0),
        "relations":int(db.execute("select count(*) c from game_achievement_choice_groups where game_id='wuwa'").fetchone()["c"] or 0),
    }


def _load_hsr_catalog_version_map() -> dict[str,str]:
    """Load the persisted HSR release-version map used by the normalized catalog.

    The HSR catalog stores official Traditional Chinese text and order, while release
    versions are derived from the preserved achievement history metadata.  Keeping
    this lookup here prevents startup catalog synchronization from replacing every
    version with ``未標示``.
    """
    try:
        if not HSR_ACHIEVEMENTS_METADATA_CACHE_FILE.exists():
            return {}
        payload=json.loads(HSR_ACHIEVEMENTS_METADATA_CACHE_FILE.read_text(encoding="utf-8-sig"))
        raw=payload.get("version_by_id") if isinstance(payload,dict) else None
        if not isinstance(raw,dict):
            return {}
        result={}
        for key,value in raw.items():
            achievement_id=str(key or "").strip()
            version=str(value or "").strip()
            if achievement_id and version:
                result[achievement_id]=version
        return result
    except Exception:
        return {}


def _load_hsr_catalog_rows() -> list[dict[str,Any]]:
    if not HSR_CATALOG_FILE.exists():
        return []
    payload=json.loads(HSR_CATALOG_FILE.read_text(encoding="utf-8-sig"))
    raw=payload.get("items") if isinstance(payload,dict) else payload
    version_by_id=_load_hsr_catalog_version_map()
    catalog_source=str(payload.get("source") or "").strip() if isinstance(payload,dict) else ""
    if not isinstance(raw,list):
        raise ValueError("崩鐵成就資料格式錯誤：缺少 items。")
    rows=[]
    for index,item in enumerate(raw):
        if not isinstance(item,dict):
            continue
        achievement_id=str(item.get("id") or item.get("achievement_id") or "").strip()
        name=str(item.get("title") or item.get("name") or "").strip()
        condition=str(item.get("desc") or item.get("condition") or item.get("description") or item.get("hide_desc") or "").strip()
        if not achievement_id or not name:
            continue
        rows.append({
            "achievement_id":achievement_id,
            "name":name,
            "condition":condition,
            "version":str(item.get("version") or version_by_id.get(achievement_id) or "未標示").strip(),
            "category":str(item.get("category") or "未辨識分類").strip(),
            "reward":int(item.get("reward") or 0),
            "hidden":1 if item.get("hide") or item.get("hidden") else 0,
            "tags_json":json.dumps(item.get("tags") if isinstance(item.get("tags"),list) else [],ensure_ascii=False),
            "source":str(item.get("source") or ("hoyolab-official-zh-tw" if version_by_id else catalog_source) or "hsr-local-catalog").strip(),
            "source_order":int(item.get("sourceOrder") if item.get("sourceOrder") is not None else index),
        })
    if len(rows)<1000:
        raise ValueError(f"崩鐵成就資料筆數異常：{len(rows)}")
    return rows


def _sync_hsr_catalog(db: sqlite3.Connection) -> tuple[int,int]:
    rows=_load_hsr_catalog_rows()
    if not rows:
        return 0,0
    stamp=now()
    db.execute("delete from game_catalog_items where game_id='hsr' and lower(source) not in ('manual','admin')")
    db.executemany(
        """insert into game_catalog_items(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order,updated_at)
        values('hsr',?,?,?,?,?,?,?,?,?,?,?)
        on conflict(game_id,achievement_id) do nothing""",
        [(r["achievement_id"],r["name"],r["condition"],r["version"],r["category"],r["reward"],r["hidden"],r["tags_json"],r["source"],r["source_order"],stamp) for r in rows],
    )
    return len(rows),len({r["achievement_id"] for r in rows})


def _load_genshin_catalog_rows() -> list[dict[str,Any]]:
    if not GENSHIN_CATALOG_FILE.exists():
        return []
    payload=json.loads(GENSHIN_CATALOG_FILE.read_text(encoding="utf-8-sig"))
    raw=payload.get("items") if isinstance(payload,dict) else payload
    if not isinstance(raw,list):
        raise ValueError("原神成就資料格式錯誤：缺少 items。")
    rows=[]
    for index,item in enumerate(raw):
        if not isinstance(item,dict):
            continue
        achievement_id=str(item.get("id") or item.get("achievement_id") or "").strip()
        name=str(item.get("name") or "").strip()
        condition=str(item.get("condition") or "").strip()
        if not achievement_id or not name:
            continue
        rows.append({
            "achievement_id":achievement_id,"name":name,"condition":condition,
            "version":str(item.get("version") or "未標示").strip(),
            "category":str(item.get("category") or "未辨識分類").strip(),
            "reward":int(item.get("reward") or 0),"hidden":1 if item.get("hidden") else 0,
            "tags_json":json.dumps(item.get("tags") if isinstance(item.get("tags"),list) else [],ensure_ascii=False),
            "source":str(item.get("source") or "genshin-official").strip(),
            "source_order":int(item.get("sourceOrder") if item.get("sourceOrder") is not None else index),
        })
    if len(rows)<1500:
        raise ValueError(f"原神成就資料筆數異常：{len(rows)}")
    return rows


def _sync_genshin_catalog(db: sqlite3.Connection) -> tuple[int,int]:
    rows=_load_genshin_catalog_rows()
    if not rows:
        return 0,0
    stamp=now()
    source_ids={row["achievement_id"] for row in rows}
    # 保留管理員手動新增的資料；官方來源資料則跟隨內建資料更新。
    db.execute("delete from game_catalog_items where game_id='genshin' and lower(source) not in ('manual','admin')")
    db.executemany(
        """insert into game_catalog_items(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order,updated_at)
        values('genshin',?,?,?,?,?,?,?,?,?,?,?)
        on conflict(game_id,achievement_id) do nothing""",
        [(row["achievement_id"],row["name"],row["condition"],row["version"],row["category"],row["reward"],row["hidden"],row["tags_json"],row["source"],row["source_order"],stamp) for row in rows],
    )
    return len(rows),len(source_ids)



def _load_zzz_catalog_rows() -> list[dict[str,Any]]:
    if not ZZZ_CATALOG_FILE.exists():
        return []
    payload=json.loads(ZZZ_CATALOG_FILE.read_text(encoding="utf-8-sig"))
    raw=payload.get("items") if isinstance(payload,dict) else payload
    if not isinstance(raw,list):
        raise ValueError("絕區零成就資料格式錯誤：缺少 items。")
    rows=[]
    for index,item in enumerate(raw):
        if not isinstance(item,dict):
            continue
        achievement_id=str(item.get("id") or item.get("achievement_id") or "").strip()
        name=str(item.get("name") or "").strip()
        if not achievement_id or not name:
            continue
        category=str(item.get("category") or "未辨識分類").strip()
        if item.get("arcade") and not category.startswith("【街機】"):
            category=f"【街機】{category}"
        rows.append({
            "achievement_id":achievement_id,"name":name,"condition":str(item.get("condition") or "").strip(),
            "version":str(item.get("version") or "未標示").strip(),"category":category,
            "reward":int(item.get("reward") or 0),"hidden":1 if item.get("hidden") else 0,
            "tags_json":json.dumps(item.get("tags") if isinstance(item.get("tags"),list) else [],ensure_ascii=False),
            "source":str(item.get("source") or "stardb-zzz-zh-tw").strip(),
            "source_order":int(item.get("sourceOrder") if item.get("sourceOrder") is not None else index),
        })
    if len(rows)<400:
        raise ValueError(f"絕區零成就資料筆數異常：{len(rows)}")
    normal=sum(1 for row in rows if not str(row["category"]).startswith("【街機】"))
    arcade=len(rows)-normal
    if normal<250 or arcade<30:
        raise ValueError(f"絕區零成就分類數量異常：一般 {normal}、街機 {arcade}")
    return rows


def _sync_zzz_catalog(db: sqlite3.Connection) -> tuple[int,int]:
    rows=_load_zzz_catalog_rows()
    if not rows:
        return 0,0
    stamp=now()
    db.execute("delete from game_catalog_items where game_id='zzz' and lower(source) not in ('manual','admin')")
    db.executemany(
        """insert into game_catalog_items(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order,updated_at)
        values('zzz',?,?,?,?,?,?,?,?,?,?,?)
        on conflict(game_id,achievement_id) do nothing""",
        [(r["achievement_id"],r["name"],r["condition"],r["version"],r["category"],r["reward"],r["hidden"],r["tags_json"],r["source"],r["source_order"],stamp) for r in rows],
    )
    return len(rows),sum(1 for r in rows if str(r["category"]).startswith("【街機】"))

def _load_registered_catalog_rows(game_id: str) -> list[dict[str,Any]]:
    config=get_game_config(game_id) or {}
    return normalize_catalog_rows(
        game_catalog_file(game_id),
        game_id=game_id,
        minimum_count=max(1,int(config.get("minimumCatalogCount") or 1)),
        default_source=f"{game_id}-catalog",
    )


def _sync_registered_catalog(db: sqlite3.Connection, game_id: str) -> tuple[int,int]:
    if game_id=="wuwa": return _sync_wuwa_catalog(db)
    if game_id=="hsr": return _sync_hsr_catalog(db)
    if game_id=="genshin": return _sync_genshin_catalog(db)
    if game_id=="zzz": return _sync_zzz_catalog(db)
    rows=_load_registered_catalog_rows(game_id)
    return replace_catalog_rows(db,game_id=game_id,rows=rows,updated_at=now())


def _bootstrap_achievement_identities(db: sqlite3.Connection) -> None:
    """Register current catalog IDs as permanent internal identities.

    Existing IDs are preserved exactly so user progress, relations, overrides and
    historical audit records remain valid. Source-specific IDs are stored in a
    separate mapping table and may change without changing the internal identity.
    """
    stamp=now()
    rows=db.execute("select game_id,achievement_id,source from game_catalog_items where game_id<>\'wuwa\'").fetchall()
    db.executemany(
        """insert into achievement_identities(game_id,internal_id,display_source_name,display_source_id,created_at,updated_at)
        values(?,?,?,?,?,?) on conflict(game_id,internal_id) do update set updated_at=excluded.updated_at""",
        [(str(row["game_id"]),str(row["achievement_id"]),"legacy_catalog",str(row["achievement_id"]),stamp,stamp) for row in rows],
    )
    db.executemany(
        """insert into achievement_source_ids(game_id,source_name,source_id,internal_id,is_primary,match_status,match_confidence,match_basis,first_seen_at,last_seen_at)
        values(?,?,?,?,0,'confirmed',1.0,'existing_catalog_identity',?,?)
        on conflict(game_id,source_name,source_id) do update set internal_id=excluded.internal_id,last_seen_at=excluded.last_seen_at""",
        [(str(row["game_id"]),"legacy_catalog",str(row["achievement_id"]),str(row["achievement_id"]),stamp,stamp) for row in rows],
    )


def _migrate_orphan_identity_cleanup(db: sqlite3.Connection) -> int:
    migration="2026-07-15-remove-safe-orphan-achievement-identities-v1"
    if db.execute("select 1 from schema_migrations where name=?",(migration,)).fetchone():
        return 0
    rows=db.execute(
        """select i.game_id,i.internal_id
        from achievement_identities i
        left join game_catalog_items c on c.game_id=i.game_id and c.achievement_id=i.internal_id
        where c.achievement_id is null
          and not exists(select 1 from game_progress p where p.game_id=i.game_id and p.achievement_id=i.internal_id)
          and not exists(select 1 from game_achievement_choice_groups g where g.game_id=i.game_id and g.achievement_id=i.internal_id)
        order by i.game_id,i.internal_id"""
    ).fetchall()
    removed=[]
    for row in rows:
        game_id=str(row["game_id"]);achievement_id=str(row["internal_id"])
        source_count=int(db.execute("select count(*) from achievement_source_ids where game_id=? and internal_id=?",(game_id,achievement_id)).fetchone()[0])
        db.execute("delete from achievement_source_ids where game_id=? and internal_id=?",(game_id,achievement_id))
        db.execute("delete from achievement_identities where game_id=? and internal_id=?",(game_id,achievement_id))
        removed.append({"game_id":game_id,"achievement_id":achievement_id,"source_id_count":source_count})
    db.execute(
        "insert into schema_migrations(name,applied_at,details_json) values(?,?,?)",
        (migration,now(),json.dumps({"removed":removed},ensure_ascii=False)),
    )
    return len(removed)


def _sync_registered_relations(db: sqlite3.Connection, game_id: str) -> None:
    for relation_type in ("exclusive","stage"):
        _sync_relation_groups(db,game_id,game_relation_file(game_id,relation_type),relation_type)
    _repair_choice_group_progress(db,game_id)


def _migrate_wuwa_legacy_choice_progress(db: sqlite3.Connection) -> int:
    """舊版把二選一合併成一筆；保留群組完成狀態並移到第一個選項。"""
    payload=_load_wuwa_choice_group_payload()
    migrated=0
    for group in payload.get("groups") or []:
        if not isinstance(group,dict):
            continue
        legacy_id=str(group.get("legacy_combined_id") or "").strip()
        members=validate_ids(group.get("achievement_ids") or [])
        if not legacy_id or not members:
            continue
        placeholders=','.join('?' for _ in members)
        rows=db.execute(
            "select user_id,completed_at from progress where achievement_id=?",
            (legacy_id,),
        ).fetchall()
        for row in rows:
            already=db.execute(
                f"select 1 from progress where user_id=? and achievement_id in ({placeholders}) limit 1",
                [row["user_id"],*members],
            ).fetchone()
            if not already:
                db.execute(
                    "insert into progress(user_id,achievement_id,completed_at) values(?,?,?) on conflict(user_id,achievement_id) do nothing",
                    (row["user_id"],members[0],row["completed_at"]),
                )
                migrated+=1
        db.execute("delete from progress where achievement_id=?",(legacy_id,))
    return migrated


def _load_hsr_official_reward_metadata() -> dict[str, int]:
    """讀取遊戲資料中的官方星瓊獎勵；與排序檔分開保存，避免重建排序時遺失。"""
    if not HSR_OFFICIAL_REWARD_FILE.exists():
        return {}
    try:
        payload=json.loads(HSR_OFFICIAL_REWARD_FILE.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}
    raw_items=payload.get("items") if isinstance(payload,dict) else None
    if not isinstance(raw_items,list):
        return {}
    result: dict[str,int]={}
    for item in raw_items:
        if not isinstance(item,dict):
            continue
        achievement_id=str(item.get("achievement_id") or item.get("id") or "").strip()
        try:
            reward=int(item.get("reward") or 0)
        except (TypeError,ValueError):
            reward=0
        if achievement_id and reward in {5,10,20}:
            result[achievement_id]=reward
    return result




def _repair_hsr_catalog_rewards(db: sqlite3.Connection) -> int:
    """依獨立官方獎勵檔回填崩鐵成就，避免同步或舊快取把獎勵寫回 0。"""
    rewards=_load_hsr_official_reward_metadata()
    if len(rewards)<1500:
        return 0
    changed=0
    stamp=now()
    for achievement_id,reward in rewards.items():
        cursor=db.execute(
            """update game_catalog_items
            set reward=?,updated_at=?
            where game_id='hsr' and achievement_id=? and reward<>?""",
            (reward,stamp,achievement_id,reward),
        )
        changed+=max(0,int(cursor.rowcount or 0))
    return changed


def _repair_choice_group_progress(db: sqlite3.Connection, game_id: str="hsr") -> int:
    """Compatibility no-op.

    Older releases silently deleted conflicting progress during startup and
    relation synchronization. Conflicts are now reported by relation governance
    and require an administrator-reviewed, backed-up resolution.
    """
    return 0

def _repair_wuwa_choice_group_progress(db: sqlite3.Connection) -> int:
    """Compatibility no-op; Wuwa conflicts are handled by relation governance."""
    return 0

def _effective_achievement_count_for_ids(db: sqlite3.Connection, game_id: str, achievement_ids: list[str]) -> int:
    """一般與階段成就逐項計數；互斥成就同組只計一次。"""
    ids=list(dict.fromkeys(str(value or "").strip() for value in achievement_ids if str(value or "").strip()))
    if not ids:
        return 0
    placeholders=','.join('?' for _ in ids)
    rows=db.execute(
        f"""select achievement_id,group_id,relation_type
        from game_achievement_choice_groups
        where game_id=? and achievement_id in ({placeholders})""",
        [game_id,*ids],
    ).fetchall()
    relation_by_id={str(row["achievement_id"]):row for row in rows}
    keys=set()
    for achievement_id in ids:
        relation=relation_by_id.get(achievement_id)
        if relation and str(relation["relation_type"] or "")=="exclusive" and str(relation["group_id"] or ""):
            keys.add(f"exclusive:{relation['group_id']}")
        else:
            keys.add(f"achievement:{achievement_id}")
    return len(keys)


def _effective_catalog_count(db: sqlite3.Connection, game_id: str) -> int:
    rows=db.execute(
        """select achievement_id from (
            select achievement_id from game_catalog_items where game_id=?
            union select achievement_id from game_achievement_overrides where game_id=? and is_deleted=0
        )""",
        (game_id,game_id),
    ).fetchall()
    return _effective_achievement_count_for_ids(db,game_id,[str(row["achievement_id"]) for row in rows])



def _choice_groups_for_ids(db: sqlite3.Connection, game_id: str, ids: list[str]) -> dict[str,str]:
    if not ids:
        return {}
    placeholders=','.join('?' for _ in ids)
    rows=db.execute(
        f"select achievement_id,group_id from game_achievement_choice_groups where game_id=? and relation_type='exclusive' and achievement_id in ({placeholders})",
        [game_id,*ids],
    ).fetchall()
    return {row["achievement_id"]:row["group_id"] for row in rows}


def _normalize_choice_progress_ids(db: sqlite3.Connection, game_id: str, ids: list[str]) -> list[str]:
    """保留輸入順序；同一互斥群組最後出現的成就為有效選擇。"""
    group_by_id=_choice_groups_for_ids(db,game_id,ids)
    last_by_group={}
    for index,achievement_id in enumerate(ids):
        group_id=group_by_id.get(achievement_id)
        if group_id:
            last_by_group[group_id]=index
    result=[]
    for index,achievement_id in enumerate(ids):
        group_id=group_by_id.get(achievement_id)
        if group_id and last_by_group.get(group_id)!=index:
            continue
        result.append(achievement_id)
    return result


def _delete_choice_siblings(db: sqlite3.Connection, game_id: str, user_id: str, ids: list[str]) -> None:
    groups=sorted(set(_choice_groups_for_ids(db,game_id,ids).values()))
    if not groups:
        return
    placeholders=','.join('?' for _ in groups)
    db.execute(
        f"""delete from game_progress
        where game_id=? and user_id=? and achievement_id in (
          select achievement_id from game_achievement_choice_groups
          where game_id=? and relation_type='exclusive' and group_id in ({placeholders})
        )""",
        [game_id,user_id,game_id,*groups],
    )



def _stage_group_for_id(db: sqlite3.Connection, game_id: str, achievement_id: str) -> sqlite3.Row | None:
    return db.execute(
        "select group_id,stage_order from game_achievement_choice_groups where game_id=? and achievement_id=? and relation_type='stage'",
        (game_id,achievement_id),
    ).fetchone()



def _normalize_stage_ids(db: sqlite3.Connection, game_id: str, ids: list[str]) -> list[str]:
    """只保留每個階段群組從第一階段開始的連續完成前綴。"""
    selected=set(ids); result=[]
    stage_rows=db.execute(
        "select group_id,achievement_id,stage_order from game_achievement_choice_groups where game_id=? and relation_type='stage' order by group_id,stage_order",
        (game_id,),
    ).fetchall()
    stage_ids={str(row["achievement_id"]) for row in stage_rows}
    valid_stage=set(); current_group=None; prefix_ok=True
    for row in stage_rows:
        group=str(row["group_id"])
        if group!=current_group:
            current_group=group; prefix_ok=True
        aid=str(row["achievement_id"])
        if prefix_ok and aid in selected:
            valid_stage.add(aid)
        else:
            prefix_ok=False
    for aid in ids:
        if aid not in stage_ids or aid in valid_stage:
            result.append(aid)
    return result


def _stage_prerequisites_completed(db: sqlite3.Connection, game_id: str, user_id: str, achievement_id: str, wuwa: bool=False) -> bool:
    row=_stage_group_for_id(db,game_id,achievement_id)
    if not row or int(row["stage_order"] or 0)<=1:
        return True
    earlier=db.execute(
        "select achievement_id from game_achievement_choice_groups where game_id=? and group_id=? and relation_type='stage' and stage_order<? order by stage_order",
        (game_id,row["group_id"],int(row["stage_order"])),
    ).fetchall()
    if not earlier:
        return True
    ids=[str(x["achievement_id"]) for x in earlier]
    placeholders=','.join('?' for _ in ids)
    if wuwa:
        count=int(db.execute(f"select count(*) c from progress where user_id=? and achievement_id in ({placeholders})",[user_id,*ids]).fetchone()["c"] or 0)
    else:
        count=int(db.execute(f"select count(*) c from game_progress where game_id=? and user_id=? and achievement_id in ({placeholders})",[game_id,user_id,*ids]).fetchone()["c"] or 0)
    return count==len(ids)


def _delete_stage_from(db: sqlite3.Connection, game_id: str, user_id: str, achievement_id: str, wuwa: bool=False) -> None:
    row=_stage_group_for_id(db,game_id,achievement_id)
    if not row:
        if wuwa:
            db.execute("delete from progress where user_id=? and achievement_id=?",(user_id,achievement_id))
        else:
            db.execute("delete from game_progress where game_id=? and user_id=? and achievement_id=?",(game_id,user_id,achievement_id))
        return
    later=db.execute(
        "select achievement_id from game_achievement_choice_groups where game_id=? and group_id=? and relation_type='stage' and stage_order>=?",
        (game_id,row["group_id"],int(row["stage_order"])),
    ).fetchall()
    ids=[str(x["achievement_id"]) for x in later]
    placeholders=','.join('?' for _ in ids)
    if wuwa:
        db.execute(f"delete from progress where user_id=? and achievement_id in ({placeholders})",[user_id,*ids])
    else:
        db.execute(f"delete from game_progress where game_id=? and user_id=? and achievement_id in ({placeholders})",[game_id,user_id,*ids])


def _normalize_achievement_category_name(game_id: str, value: Any) -> str:
    name=re.sub(r"\s+"," ",str(value or "").replace("\u00a0"," ")).strip()
    if game_id=="wuwa":
        name=canonicalize_wuwa_category(name)
    if not name:
        raise HTTPException(status_code=400,detail="分類名稱不可空白。")
    if len(name)>200:
        raise HTTPException(status_code=400,detail="分類名稱不可超過 200 個字。")
    return name


def _effective_category_values(db: sqlite3.Connection, game_id: str) -> list[dict[str,Any]]:
    deleted={str(row["achievement_id"]) for row in db.execute(
        "select achievement_id from game_deleted_achievements where game_id=?",(game_id,)
    ).fetchall()}
    values: dict[str,dict[str,Any]]={}
    for row in db.execute(
        "select achievement_id,category,source_order from game_catalog_items where game_id=? order by source_order,achievement_id",
        (game_id,),
    ).fetchall():
        achievement_id=str(row["achievement_id"])
        if achievement_id in deleted:
            continue
        values[achievement_id]={
            "achievement_id":achievement_id,
            "category":_normalize_achievement_category_name(game_id,row["category"] or "未辨識分類"),
            "source_order":int(row["source_order"] or 0),
        }
    for row in db.execute(
        "select achievement_id,category,is_deleted from game_achievement_overrides where game_id=? order by updated_at,achievement_id",
        (game_id,),
    ).fetchall():
        achievement_id=str(row["achievement_id"])
        if achievement_id in deleted:
            continue
        current=values.get(achievement_id)
        raw_category=str(row["category"] or "").strip()
        if current and (not raw_category or raw_category=="未辨識分類"):
            category=current["category"]
        else:
            category=_normalize_achievement_category_name(game_id,raw_category or "未辨識分類")
        values[achievement_id]={
            "achievement_id":achievement_id,
            "category":category,
            "source_order":int(current["source_order"] if current else 10**9+len(values)),
        }
    return sorted(values.values(),key=lambda row:(int(row["source_order"]),str(row["achievement_id"])))


def _achievement_category_alias_map(db: sqlite3.Connection, game_id: str) -> dict[str,str]:
    rows=db.execute(
        """select a.source_name,c.name from game_achievement_category_aliases a
        join game_achievement_categories c on c.id=a.category_id and c.game_id=a.game_id
        where a.game_id=?""",(game_id,),
    ).fetchall()
    return {str(row["source_name"]):str(row["name"]) for row in rows}


def _sync_achievement_categories(db: sqlite3.Connection, game_id: str) -> None:
    existing=db.execute(
        "select id,name,display_order from game_achievement_categories where game_id=? order by display_order,name",
        (game_id,),
    ).fetchall()
    existing_names={str(row["name"]):str(row["id"]) for row in existing}
    aliases=_achievement_category_alias_map(db,game_id)
    discovered=[]
    seen=set()
    for row in _effective_category_values(db,game_id):
        name=aliases.get(str(row["category"]),str(row["category"]))
        if name not in seen:
            seen.add(name);discovered.append(name)
    next_order=(max((int(row["display_order"]) for row in existing),default=-1)+1)
    stamp=now()
    for name in discovered:
        if name in existing_names:
            continue
        category_id=str(uuid.uuid4())
        db.execute(
            """insert into game_achievement_categories(id,game_id,name,display_order,is_custom,created_by,updated_by,created_at,updated_at)
            values(?,?,?,?,0,null,null,?,?)""",
            (category_id,game_id,name,next_order,stamp,stamp),
        )
        existing_names[name]=category_id
        next_order+=1


def _achievement_category_rows(db: sqlite3.Connection, game_id: str) -> list[dict[str,Any]]:
    _sync_achievement_categories(db,game_id)
    aliases=_achievement_category_alias_map(db,game_id)
    counts: dict[str,int]={}
    for row in _effective_category_values(db,game_id):
        name=aliases.get(str(row["category"]),str(row["category"]))
        counts[name]=counts.get(name,0)+1
    rows=db.execute(
        """select id,name,display_order,is_custom,created_at,updated_at
        from game_achievement_categories where game_id=? order by display_order,name""",
        (game_id,),
    ).fetchall()
    return [{
        "id":str(row["id"]),"name":str(row["name"]),"display_order":index,
        "achievement_count":int(counts.get(str(row["name"]),0)),"is_custom":bool(row["is_custom"]),
        "created_at":int(row["created_at"] or 0),"updated_at":int(row["updated_at"] or 0),
    } for index,row in enumerate(rows)]


def _apply_managed_category_aliases(db: sqlite3.Connection, game_id: str, rows: list[dict[str,Any]]) -> list[dict[str,Any]]:
    aliases=_achievement_category_alias_map(db,game_id)
    if not aliases:
        return rows
    normalized=[]
    for raw in rows:
        row=dict(raw)
        source_name=str(row.get("category") or "未辨識分類").strip()
        target_name=aliases.get(source_name)
        if target_name:
            row["category"]=target_name
            if str(row.get("group_name") or "").strip()==source_name:
                row["group_name"]=target_name
        normalized.append(row)
    return normalized


def _achievement_display_official_id(row: dict[str,Any]) -> int:
    raw=next((row.get(key) for key in ("officialId","official_source_id","displayId","display_id","achievement_id","id") if row.get(key) not in (None,"")),"")
    value=str(raw).strip()
    if not re.fullmatch(r"\d+",value):
        achievement_id=str(row.get("id") or row.get("achievement_id") or "").strip() or "（空白）"
        raise RuntimeError(f"成就顯示排序資料錯誤：成就 {achievement_id} 缺少有效的數字官方 ID。請先至成就資料治理中心處理。")
    return int(value)


def _sort_achievement_display_rows(
    db: sqlite3.Connection,
    game_id: str,
    rows: list[dict[str,Any]],
    category_rows: list[dict[str,Any]] | None = None,
) -> list[dict[str,Any]]:
    categories=category_rows if category_rows is not None else _achievement_category_rows(db,game_id)
    category_rank={str(row["name"]):index for index,row in enumerate(categories)}
    normalized=_apply_managed_category_aliases(db,game_id,rows)
    stage_rows=db.execute(
        """select group_id,achievement_id,stage_order
        from game_achievement_choice_groups
        where game_id=? and relation_type='stage'
        order by group_id,stage_order,achievement_id""",
        (game_id,),
    ).fetchall()
    stage_membership: dict[str,tuple[str,int]]={}
    stage_orders: dict[str,list[int]]={}
    for relation in stage_rows:
        achievement_id=str(relation["achievement_id"])
        group_id=str(relation["group_id"])
        stage_order=int(relation["stage_order"] or 0)
        previous=stage_membership.get(achievement_id)
        if previous and previous[0]!=group_id:
            raise RuntimeError(f"成就顯示排序資料錯誤：成就 {achievement_id} 同時屬於多個階段型關聯群組。請先至關聯資料驗證處理。")
        stage_membership[achievement_id]=(group_id,stage_order)
        stage_orders.setdefault(group_id,[]).append(stage_order)
    for group_id,orders in stage_orders.items():
        if sorted(orders)!=list(range(1,len(orders)+1)):
            raise RuntimeError(f"成就顯示排序資料錯誤：階段型關聯群組 {group_id} 的可完成順序不連續或重複。請先至關聯資料驗證處理。")

    category_singles: dict[str,list[dict[str,Any]]]={}
    category_stage_groups: dict[tuple[str,str],list[dict[str,Any]]]={}
    for row in normalized:
        category=str(row.get("category") or "").strip()
        achievement_id=str(row.get("id") or row.get("achievement_id") or "").strip() or "（空白）"
        if category not in category_rank:
            raise RuntimeError(f"成就顯示排序資料錯誤：成就 {achievement_id} 的分類「{category or '（空白）'}」尚未納入成就類別管理。請先至成就資料治理中心處理。")
        _achievement_display_official_id(row)
        membership=stage_membership.get(achievement_id)
        if membership:
            group_id,_=membership
            category_stage_groups.setdefault((category,group_id),[]).append(row)
        else:
            category_singles.setdefault(category,[]).append(row)

    ordered: list[dict[str,Any]]=[]
    for category in sorted(category_rank,key=category_rank.get):
        blocks: list[tuple[int,int,str,list[dict[str,Any]]]]=[]
        for row in category_singles.get(category,[]):
            official_id=_achievement_display_official_id(row)
            achievement_id=str(row.get("id") or row.get("achievement_id") or "")
            blocks.append((official_id,1,achievement_id,[row]))
        for (group_category,group_id),members in category_stage_groups.items():
            if group_category!=category:
                continue
            members.sort(key=lambda row:(stage_membership[str(row.get("id") or row.get("achievement_id") or "")][1],_achievement_display_official_id(row)))
            anchor=min(_achievement_display_official_id(row) for row in members)
            blocks.append((anchor,0,group_id,members))
        blocks.sort(key=lambda block:(block[0],block[1],block[2]))
        for _,_,_,members in blocks:
            ordered.extend(members)
    return ordered


def _rewrite_catalog_category_names(game_id: str, replacements: dict[str,str]) -> tuple[bytes,int]:
    path=game_catalog_file(game_id)
    if not path.exists():
        raise RuntimeError(f"找不到 {game_display_name(game_id)} 正式成就目錄。")
    original=path.read_bytes()
    payload=json.loads(original.decode("utf-8-sig"))
    items=payload.get("items") if isinstance(payload,dict) else None
    if not isinstance(items,list):
        raise RuntimeError("正式成就目錄格式錯誤，缺少 items 清單。")
    changed=0
    for item in items:
        if not isinstance(item,dict):
            continue
        category=str(item.get("category") or "").strip()
        target=replacements.get(category)
        if not target or target==category:
            continue
        item["category"]=target
        for key in ("groupName","subcategory","detailCategory"):
            if str(item.get(key) or "").strip()==category:
                item[key]=target
        changed+=1
    payload["category_management_updated_at"]=time.strftime("%Y-%m-%dT%H:%M:%S%z")
    temp=path.with_suffix(path.suffix+".category.tmp")
    temp.write_text(json.dumps(payload,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    temp.replace(path)
    return original,changed


def _restore_catalog_bytes(game_id: str, content: bytes) -> None:
    path=game_catalog_file(game_id)
    path.parent.mkdir(parents=True,exist_ok=True)
    temp=path.with_suffix(path.suffix+".category.rollback.tmp")
    temp.write_bytes(content)
    temp.replace(path)


def init_db() -> None:
    with connect_db() as db:
        db.execute("PRAGMA journal_mode = WAL")
        db.executescript("""
        create table if not exists users (
            id text primary key,
            email text not null,
            email_key text not null unique,
            password_hash text not null,
            email_verified integer not null default 0,
            role text not null default 'user',
            created_at integer not null,
            updated_at integer not null
        );
        create table if not exists sessions (
            token_hash text primary key,
            user_id text not null references users(id) on delete cascade,
            expires_at integer not null,
            created_at integer not null,
            ip_address text not null default ''
        );
        create table if not exists progress (
            user_id text not null references users(id) on delete cascade,
            achievement_id text not null,
            completed_at integer not null,
            primary key (user_id, achievement_id)
        );
        create table if not exists email_verification_tokens (
            token_hash text primary key,
            user_id text not null references users(id) on delete cascade,
            expires_at integer not null,
            used_at integer,
            created_at integer not null
        );
        create table if not exists password_reset_tokens (
            token_hash text primary key,
            user_id text not null references users(id) on delete cascade,
            expires_at integer not null,
            used_at integer,
            created_at integer not null
        );
        create table if not exists rate_limits (
            action text not null,
            rate_key text not null,
            window_start integer not null,
            count integer not null,
            primary key(action, rate_key, window_start)
        );
        create table if not exists admin_audit_logs (
            id integer primary key autoincrement,
            event_id text,
            actor_user_id text references users(id) on delete set null,
            actor_email_snapshot text not null default '',
            actor_ip text not null default '',
            action text not null,
            category text not null default 'administration',
            status text not null default 'success',
            game_id text not null default '',
            target_user_id text references users(id) on delete set null,
            target_type text not null default '',
            target_id text not null default '',
            summary text not null default '',
            details text,
            before_json text not null default '',
            after_json text not null default '',
            metadata_json text not null default '{}',
            request_id text not null default '',
            backup_name text not null default '',
            error_message text not null default '',
            archived integer not null default 0,
            locked integer not null default 0,
            created_at integer not null
        );
        create table if not exists achievement_reports (
            id text primary key, user_id text references users(id) on delete set null,
            achievement_id text not null, achievement_name text not null, report_type text not null,
            message text not null, status text not null default 'open', admin_note text not null default '',
            created_at integer not null, updated_at integer not null
        );
        create table if not exists announcements (
            id text primary key, title text not null, body text not null, level text not null default 'info',
            is_active integer not null default 1, pinned integer not null default 0, starts_at integer, ends_at integer,
            created_by text references users(id) on delete set null, created_at integer not null, updated_at integer not null
        );
        create table if not exists notifications (
            id text primary key, target_user_id text references users(id) on delete cascade,
            title text not null, body text not null, kind text not null default 'info', link text not null default '',
            created_by text references users(id) on delete set null, created_at integer not null
        );
        create table if not exists notification_reads (
            user_id text not null references users(id) on delete cascade,
            notification_id text not null references notifications(id) on delete cascade,
            read_at integer not null, primary key(user_id, notification_id)
        );
        create table if not exists notification_deletions (
            user_id text not null references users(id) on delete cascade,
            notification_id text not null references notifications(id) on delete cascade,
            deleted_at integer not null, primary key(user_id, notification_id)
        );
        create table if not exists redeem_games (
            game_id text primary key,
            name text not null,
            display_order integer not null default 0,
            enabled integer not null default 1,
            note text not null default '',
            created_by text references users(id) on delete set null,
            updated_by text references users(id) on delete set null,
            created_at integer not null,
            updated_at integer not null
        );
        create table if not exists redeem_servers (
            id text primary key,
            game_id text not null references redeem_games(game_id) on delete cascade,
            name text not null,
            display_order integer not null default 0,
            enabled integer not null default 1,
            created_by text references users(id) on delete set null,
            updated_by text references users(id) on delete set null,
            created_at integer not null,
            updated_at integer not null,
            unique(game_id,name)
        );
        create table if not exists redeem_codes (
            id text primary key,
            game_id text not null references redeem_games(game_id) on delete restrict,
            code text not null,
            source text not null default '',
            description text not null default '',
            reward text not null default '',
            start_at integer,
            end_at integer,
            server_ids_json text not null default '[]',
            redeem_url text not null default '',
            enabled integer not null default 1,
            created_by text references users(id) on delete set null,
            updated_by text references users(id) on delete set null,
            created_at integer not null,
            updated_at integer not null
        );
        create table if not exists redeem_import_batches (
            id text primary key,
            admin_user_id text references users(id) on delete set null,
            status text not null default 'preview_ready',
            default_game_id text not null default '',
            reason text not null default '',
            input_type text not null default '',
            source_filename text not null default '',
            source_sheet text not null default '',
            summary_json text not null default '{}',
            plan_json text not null default '{}',
            snapshot_json text not null default '{}',
            pre_state_hash text not null default '',
            post_state_hash text not null default '',
            created_at integer not null,
            completed_at integer,
            completed_by text references users(id) on delete set null,
            backup_name text not null default '',
            rolled_back_at integer,
            rolled_back_by text references users(id) on delete set null,
            rollback_reason text not null default '',
            rollback_backup_name text not null default ''
        );
        create table if not exists redeem_import_items (
            id text primary key,
            batch_id text not null references redeem_import_batches(id) on delete cascade,
            row_number integer not null,
            game_id text not null default '',
            code text not null default '',
            action text not null,
            target_id text not null default '',
            candidate_json text not null default '{}',
            before_json text not null default '{}',
            error_text text not null default ''
        );
        create table if not exists redeem_notification_preferences (
            user_id text not null references users(id) on delete cascade,
            game_id text not null references redeem_games(game_id) on delete cascade,
            updated_at integer not null,
            primary key(user_id,game_id)
        );
        create table if not exists redeem_notification_events (
            id text primary key,
            user_id text not null references users(id) on delete cascade,
            game_id text not null,
            code_id text not null,
            event_key text not null,
            message_item_id text references message_center_items(id) on delete set null,
            created_at integer not null,
            unique(user_id,event_key)
        );
        create table if not exists email_logs (
            id integer primary key autoincrement, recipient text not null, subject text not null,
            mail_type text not null default 'generic', status text not null, error text not null default '', created_at integer not null
        );
        create table if not exists blocked_entries (
            id text primary key, kind text not null, value_key text not null, reason text not null default '',
            active integer not null default 1, created_by text references users(id) on delete set null, created_at integer not null
        );
        create unique index if not exists blocked_entries_unique on blocked_entries(kind,value_key);
        create table if not exists achievement_overrides (
            achievement_id text primary key, name text not null, condition text not null default '',
            version text not null default '未標示', category text not null default '未辨識分類', reward integer not null default 0,
            hidden integer not null default 0, tags_json text not null default '[]', is_deleted integer not null default 0,
            source text not null default 'override', updated_by text references users(id) on delete set null, updated_at integer not null
        );
        create table if not exists deleted_achievements (
            achievement_id text primary key, achievement_name text not null default '',
            deleted_by text references users(id) on delete set null, deleted_at integer not null
        );
        create table if not exists achievement_revisions (
            id integer primary key autoincrement, achievement_id text not null, action text not null,
            snapshot_json text not null, actor_user_id text references users(id) on delete set null, created_at integer not null
        );
        create table if not exists featured_achievements (
            achievement_id text primary key, note text not null default '', sort_order integer not null default 0,
            is_active integer not null default 1, updated_by text references users(id) on delete set null, updated_at integer not null
        );
        create table if not exists support_tickets (
            id text primary key, user_id text references users(id) on delete set null, subject text not null,
            status text not null default 'open', priority text not null default 'normal', created_at integer not null, updated_at integer not null
        );
        create table if not exists support_ticket_messages (
            id text primary key, ticket_id text not null references support_tickets(id) on delete cascade,
            sender_user_id text references users(id) on delete set null, message text not null, created_at integer not null
        );
        create index if not exists sessions_user_idx on sessions(user_id);
        create index if not exists progress_user_idx on progress(user_id);
        create index if not exists verification_user_idx on email_verification_tokens(user_id);
        create index if not exists reset_user_idx on password_reset_tokens(user_id);
        create index if not exists reports_status_idx on achievement_reports(status,created_at);
        create index if not exists notifications_target_idx on notifications(target_user_id,created_at);
        create index if not exists redeem_games_order_idx on redeem_games(enabled,display_order,name);
        create index if not exists redeem_servers_game_idx on redeem_servers(game_id,enabled,display_order,name);
        create index if not exists redeem_codes_game_idx on redeem_codes(game_id,enabled,start_at,end_at);
        create unique index if not exists redeem_codes_game_code_unique on redeem_codes(game_id,lower(trim(code)));
        create index if not exists redeem_import_batches_created_idx on redeem_import_batches(created_at,status);
        create index if not exists redeem_import_items_batch_idx on redeem_import_items(batch_id,row_number);
        create index if not exists redeem_notification_events_lookup_idx on redeem_notification_events(game_id,code_id,created_at);
        create index if not exists email_logs_created_idx on email_logs(created_at);
        create index if not exists tickets_user_idx on support_tickets(user_id,updated_at);

        create table if not exists live_revisions (
            scope text primary key, revision integer not null default 0, updated_at integer not null default 0
        );
        insert or ignore into live_revisions(scope,revision,updated_at) values
            ('catalog',0,0),('stats',0,0),('announcements',0,0),('notifications',0,0),
            ('support',0,0),('reports',0,0),('users',0,0),('redeem_codes',0,0);

        create trigger if not exists live_progress_insert after insert on progress begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='stats'; end;
        create trigger if not exists live_progress_update after update on progress begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='stats'; end;
        create trigger if not exists live_progress_delete after delete on progress begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='stats'; end;
        create trigger if not exists live_users_insert after insert on users begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope in ('users','stats'); end;
        create trigger if not exists live_users_update after update on users begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope in ('users','stats'); end;
        create trigger if not exists live_users_delete after delete on users begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope in ('users','stats'); end;
        create trigger if not exists live_overrides_insert after insert on achievement_overrides begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_overrides_update after update on achievement_overrides begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_overrides_delete after delete on achievement_overrides begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_deleted_achievements_insert after insert on deleted_achievements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_deleted_achievements_delete after delete on deleted_achievements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_featured_insert after insert on featured_achievements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_featured_update after update on featured_achievements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_featured_delete after delete on featured_achievements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='catalog'; end;
        create trigger if not exists live_announcements_insert after insert on announcements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='announcements'; end;
        create trigger if not exists live_announcements_update after update on announcements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='announcements'; end;
        create trigger if not exists live_announcements_delete after delete on announcements begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='announcements'; end;
        create trigger if not exists live_notifications_insert after insert on notifications begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='notifications'; end;
        create trigger if not exists live_notifications_update after update on notifications begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='notifications'; end;
        create trigger if not exists live_notifications_delete after delete on notifications begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='notifications'; end;
        create trigger if not exists live_notification_reads_insert after insert on notification_reads begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='notifications'; end;
        create trigger if not exists live_notification_reads_update after update on notification_reads begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='notifications'; end;
        create trigger if not exists live_notification_deletions_insert after insert on notification_deletions begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='notifications'; end;
        create trigger if not exists live_redeem_games_insert after insert on redeem_games begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_games_update after update on redeem_games begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_games_delete after delete on redeem_games begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_servers_insert after insert on redeem_servers begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_servers_update after update on redeem_servers begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_servers_delete after delete on redeem_servers begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_codes_insert after insert on redeem_codes begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_codes_update after update on redeem_codes begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_redeem_codes_delete after delete on redeem_codes begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='redeem_codes'; end;
        create trigger if not exists live_support_tickets_insert after insert on support_tickets begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='support'; end;
        create trigger if not exists live_support_tickets_update after update on support_tickets begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='support'; end;
        create trigger if not exists live_support_tickets_delete after delete on support_tickets begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='support'; end;
        create trigger if not exists live_support_messages_insert after insert on support_ticket_messages begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='support'; end;
        create trigger if not exists live_reports_insert after insert on achievement_reports begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='reports'; end;
        create trigger if not exists live_reports_update after update on achievement_reports begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='reports'; end;
        create trigger if not exists live_reports_delete after delete on achievement_reports begin update live_revisions set revision=revision+1,updated_at=cast(strftime('%s','now') as integer) where scope='reports'; end;
        """)

        # 多遊戲專案共用帳號，但各遊戲的成就資料與進度完全分離。
        db.executescript("""
        create table if not exists game_catalog_items (
            game_id text not null,
            achievement_id text not null,
            name text not null,
            condition text not null default '',
            version text not null default '未標示',
            category text not null default '未辨識分類',
            reward integer not null default 0,
            hidden integer not null default 0,
            tags_json text not null default '[]',
            source text not null default 'catalog',
            source_order integer not null default 0,
            updated_at integer not null,
            primary key(game_id, achievement_id)
        );
        create table if not exists game_progress (
            game_id text not null,
            user_id text not null references users(id) on delete cascade,
            achievement_id text not null,
            completed_at integer not null,
            primary key(game_id, user_id, achievement_id)
        );
        create table if not exists game_achievement_choice_groups (
            game_id text not null,
            group_id text not null,
            achievement_id text not null,
            relation_type text not null default 'exclusive',
            stage_order integer not null default 0,
            updated_at integer not null,
            primary key(game_id, achievement_id)
        );
        create table if not exists game_achievement_reports (
            id text primary key,
            game_id text not null,
            user_id text references users(id) on delete set null,
            achievement_id text not null,
            achievement_name text not null,
            report_type text not null,
            message text not null,
            status text not null default 'open',
            admin_note text not null default '',
            created_at integer not null,
            updated_at integer not null
        );
        create table if not exists game_achievement_overrides (
            game_id text not null,
            achievement_id text not null,
            name text not null,
            condition text not null default '',
            version text not null default '未標示',
            category text not null default '未辨識分類',
            reward integer not null default 0,
            hidden integer not null default 0,
            tags_json text not null default '[]',
            is_deleted integer not null default 0,
            source text not null default 'override',
            updated_by text references users(id) on delete set null,
            updated_at integer not null,
            primary key(game_id, achievement_id)
        );
        create table if not exists game_deleted_achievements (
            game_id text not null,
            achievement_id text not null,
            achievement_name text not null default '',
            deleted_by text references users(id) on delete set null,
            deleted_at integer not null,
            primary key(game_id, achievement_id)
        );
        create table if not exists game_achievement_revisions (
            id integer primary key autoincrement,
            game_id text not null,
            achievement_id text not null,
            action text not null,
            snapshot_json text not null,
            actor_user_id text references users(id) on delete set null,
            created_at integer not null
        );
        create table if not exists achievement_delete_backups (
            id text primary key,
            game_id text not null,
            achievement_id text not null,
            achievement_name text not null default '',
            admin_user_id text references users(id) on delete set null,
            reason text not null default '',
            confirmation_text text not null default '',
            snapshot_json text not null default '{}',
            backup_name text not null default '',
            status text not null default 'deleted',
            created_at integer not null,
            restored_at integer,
            restored_by text references users(id) on delete set null,
            restore_reason text not null default ''
        );
        create index if not exists achievement_delete_backups_lookup_idx
            on achievement_delete_backups(game_id,achievement_id,created_at);
        create table if not exists game_featured_achievements (
            game_id text not null,
            achievement_id text not null,
            note text not null default '',
            sort_order integer not null default 0,
            is_active integer not null default 1,
            updated_by text references users(id) on delete set null,
            updated_at integer not null,
            primary key(game_id, achievement_id)
        );
        create table if not exists game_live_revisions (
            game_id text not null,
            scope text not null,
            revision integer not null default 0,
            updated_at integer not null default 0,
            primary key(game_id, scope)
        );
        create table if not exists schema_migrations (
            name text primary key,
            applied_at integer not null,
            details_json text not null default '{}'
        );
        create table if not exists game_achievement_categories (
            id text primary key,
            game_id text not null,
            name text not null,
            display_order integer not null default 0,
            is_custom integer not null default 0,
            created_by text references users(id) on delete set null,
            updated_by text references users(id) on delete set null,
            created_at integer not null,
            updated_at integer not null,
            unique(game_id,name)
        );
        create table if not exists game_achievement_category_aliases (
            game_id text not null,
            source_name text not null,
            category_id text not null references game_achievement_categories(id) on delete cascade,
            created_at integer not null,
            primary key(game_id,source_name)
        );
        create index if not exists game_achievement_categories_order_idx
            on game_achievement_categories(game_id,display_order,name);
        create index if not exists game_achievement_category_aliases_category_idx
            on game_achievement_category_aliases(category_id);
        create table if not exists game_sync_previews (
            id text primary key,
            game_id text not null,
            admin_user_id text references users(id) on delete cascade,
            candidate_json text not null,
            source_payload_json text not null default '',
            metadata_json text not null default '{}',
            diff_json text not null default '{}',
            created_at integer not null,
            expires_at integer not null
        );
        create table if not exists achievement_identities (
            game_id text not null,
            internal_id text not null,
            display_source_name text not null default '',
            display_source_id text not null default '',
            created_at integer not null,
            updated_at integer not null,
            primary key(game_id,internal_id)
        );
        create table if not exists achievement_source_ids (
            game_id text not null,
            source_name text not null,
            source_id text not null,
            internal_id text not null,
            is_primary integer not null default 0,
            match_status text not null default 'confirmed',
            match_confidence real not null default 1.0,
            match_basis text not null default '',
            first_seen_at integer not null,
            last_seen_at integer not null,
            primary key(game_id,source_name,source_id),
            foreign key(game_id,internal_id) references achievement_identities(game_id,internal_id) on delete cascade
        );
        create unique index if not exists achievement_source_internal_idx on achievement_source_ids(game_id,source_name,internal_id);
        create index if not exists achievement_source_lookup_idx on achievement_source_ids(game_id,internal_id);
        create table if not exists game_catalog_source_records (
            game_id text not null,
            achievement_id text not null,
            official_source_id text not null default '',
            category_id text not null default '',
            group_id text not null default '',
            group_name text not null default '',
            progress_value integer not null default 0,
            level integer not null default 0,
            next_link text not null default '',
            reward_id text not null default '',
            primary_source_id text not null default '',
            secondary_source_id text not null default '',
            source_ref text not null default '',
            raw_json text not null default '{}',
            provenance_json text not null default '{}',
            updated_at integer not null,
            primary key(game_id,achievement_id),
            foreign key(game_id,achievement_id) references game_catalog_items(game_id,achievement_id) on delete cascade
        );
        create index if not exists game_catalog_source_group_idx on game_catalog_source_records(game_id,group_id);
        create table if not exists catalog_scan_previews (
            id text primary key,
            game_id text not null,
            admin_user_id text references users(id) on delete cascade,
            items_json text not null,
            result_json text not null,
            created_at integer not null,
            expires_at integer not null
        );
        create table if not exists catalog_issue_decisions (
            game_id text not null,
            issue_key text not null,
            decision text not null,
            details_json text not null default '{}',
            updated_by text references users(id) on delete set null,
            updated_at integer not null,
            primary key(game_id,issue_key)
        );
        create table if not exists achievement_id_aliases (
            game_id text not null,
            alias_id text not null,
            canonical_id text not null,
            reason text not null default '',
            created_by text references users(id) on delete set null,
            created_at integer not null,
            primary key(game_id,alias_id)
        );
        create table if not exists source_sync_history (
            id text primary key,
            game_id text not null,
            preview_id text not null default '',
            source_id text not null default '',
            source_mode text not null default '',
            source_hash text not null default '',
            summary_json text not null default '{}',
            backup_name text not null default '',
            snapshot_dir text not null default '',
            pre_state_hash text not null default '',
            post_state_hash text not null default '',
            status text not null,
            actor_user_id text references users(id) on delete set null,
            created_at integer not null,
            rolled_back_at integer,
            rolled_back_by text references users(id) on delete set null,
            rollback_reason text not null default ''
        );
        create table if not exists source_sync_decisions (
            id text primary key,
            history_id text not null default '',
            preview_id text not null default '',
            game_id text not null,
            change_id text not null,
            achievement_id text not null default '',
            action text not null,
            fields_json text not null default '[]',
            reason text not null default '',
            data_changed integer not null default 0,
            actor_user_id text references users(id) on delete set null,
            created_at integer not null
        );
        create index if not exists source_sync_decisions_game_idx on source_sync_decisions(game_id,created_at);
        create index if not exists game_progress_user_idx on game_progress(game_id,user_id);
        create index if not exists game_choice_groups_idx on game_achievement_choice_groups(game_id,group_id);
        create index if not exists game_reports_status_idx on game_achievement_reports(game_id,status,created_at);
        create index if not exists game_overrides_updated_idx on game_achievement_overrides(game_id,updated_at);
        create index if not exists game_sync_previews_expiry_idx on game_sync_previews(expires_at);
        create table if not exists achievement_scan_runs (
            id text primary key, game_id text not null, status text not null, admin_user_id text references users(id) on delete set null,
            rules_version text not null, catalog_hash text not null, database_hash text not null, options_json text not null default '{}',
            summary_json text not null default '{}', started_at integer not null, completed_at integer, expires_at integer
        );
        create table if not exists achievement_issues (
            id text primary key, game_id text not null, fingerprint text not null, kind text not null, severity text not null, risk text not null,
            title text not null, message text not null, state text not null default 'new', first_seen_at integer not null, last_seen_at integer not null,
            occurrence_count integer not null default 1, last_scan_id text references achievement_scan_runs(id) on delete set null,
            progress_count integer not null default 0, relation_count integer not null default 0, auto_fixable integer not null default 0,
            evidence_json text not null default '{}', actions_json text not null default '[]', resolution_json text not null default '{}',
            resolved_by text references users(id) on delete set null, resolved_at integer, unique(game_id,fingerprint)
        );
        create table if not exists achievement_issue_entities (
            issue_id text not null references achievement_issues(id) on delete cascade, entity_type text not null default 'achievement',
            entity_id text not null, snapshot_json text not null default '{}', primary key(issue_id,entity_type,entity_id)
        );
        create table if not exists achievement_resolution_drafts (
            id text primary key, game_id text not null, scan_id text not null references achievement_scan_runs(id) on delete cascade,
            admin_user_id text references users(id) on delete set null, name text not null default '', reason text not null default '',
            actions_json text not null default '[]', plan_json text not null default '{}', status text not null default 'draft',
            created_at integer not null, updated_at integer not null
        );
        create table if not exists achievement_resolution_batches (
            id text primary key, game_id text not null, draft_id text references achievement_resolution_drafts(id) on delete set null,
            admin_user_id text references users(id) on delete set null, status text not null, plan_json text not null default '{}',
            result_json text not null default '{}', backup_name text not null default '', snapshot_dir text not null default '',
            started_at integer not null, completed_at integer, rolled_back_at integer, rollback_reason text not null default ''
        );
        create table if not exists achievement_resolution_actions (
            id integer primary key autoincrement, batch_id text not null references achievement_resolution_batches(id) on delete cascade,
            issue_id text, action text not null, status text not null, before_json text not null default '{}', after_json text not null default '{}',
            error_message text not null default '', created_at integer not null
        );
        create table if not exists achievement_resolution_snapshots (
            id text primary key, batch_id text not null references achievement_resolution_batches(id) on delete cascade,
            snapshot_type text not null, file_path text not null default '', payload_json text not null default '', checksum text not null default '', created_at integer not null
        );
        create table if not exists relation_validation_runs (
            id text primary key, game_id text not null, admin_user_id text references users(id) on delete set null,
            state_hash text not null, result_json text not null default '{}', created_at integer not null, expires_at integer not null
        );
        create table if not exists relation_resolution_batches (
            id text primary key, game_id text not null, validation_id text references relation_validation_runs(id) on delete set null,
            admin_user_id text references users(id) on delete set null, status text not null, reason text not null default '',
            plan_json text not null default '{}', result_json text not null default '{}', backup_name text not null default '',
            snapshot_dir text not null default '', created_at integer not null, completed_at integer, rolled_back_at integer,
            rollback_reason text not null default ''
        );
        create table if not exists relation_validation_exceptions (
            game_id text not null, fingerprint text not null, reason text not null default '', active integer not null default 1,
            created_by text references users(id) on delete set null, created_at integer not null, updated_at integer not null,
            primary key(game_id,fingerprint)
        );
        create index if not exists relation_validation_runs_game_idx on relation_validation_runs(game_id,created_at);
        create index if not exists relation_resolution_batches_game_idx on relation_resolution_batches(game_id,created_at);
        create table if not exists achievement_field_registry (
            game_id text not null, field_name text not null, classification text not null default 'retained_metadata',
            mapped_field text not null default '', active integer not null default 1,
            created_by text references users(id) on delete set null, created_at integer not null, updated_at integer not null,
            primary key(game_id,field_name)
        );
        create table if not exists achievement_exception_rules (
            id text primary key, game_id text not null, fingerprint text not null, reason text not null, source_basis text not null default '',
            snapshot_hash text not null, permanent integer not null default 0, recheck_on_change integer not null default 1, active integer not null default 1,
            created_by text references users(id) on delete set null, created_at integer not null, updated_at integer not null, unique(game_id,fingerprint)
        );
        create table if not exists achievement_governance_decisions (
            id text primary key, game_id text not null, fingerprint text not null, decision_type text not null, reason text not null default '',
            snapshot_hash text not null, evidence_json text not null default '{}', active integer not null default 1,
            created_by text references users(id) on delete set null, created_at integer not null, updated_at integer not null,
            invalidated_at integer, invalidation_reason text not null default '', unique(game_id,fingerprint,decision_type)
        );
        create table if not exists message_center_items (
            id text primary key, item_type text not null, target_user_id text references users(id) on delete cascade, title text not null, body text not null,
            level text not null default 'info', kind text not null default 'info', link text not null default '', is_active integer not null default 1,
            pinned integer not null default 0, starts_at integer, ends_at integer, created_by text references users(id) on delete set null,
            created_at integer not null, updated_at integer not null
        );
        create table if not exists message_center_reads (
            user_id text not null references users(id) on delete cascade, item_id text not null references message_center_items(id) on delete cascade,
            read_at integer not null, primary key(user_id,item_id)
        );
        create table if not exists message_center_deletions (
            user_id text not null references users(id) on delete cascade, item_id text not null references message_center_items(id) on delete cascade,
            deleted_at integer not null, primary key(user_id,item_id)
        );
        create index if not exists achievement_scan_runs_game_idx on achievement_scan_runs(game_id,started_at);
        create index if not exists achievement_issues_filter_idx on achievement_issues(game_id,state,severity,kind,last_seen_at);
        create index if not exists achievement_batches_game_idx on achievement_resolution_batches(game_id,started_at);
        create index if not exists achievement_governance_decisions_active_idx on achievement_governance_decisions(game_id,decision_type,active,updated_at);
        create index if not exists message_center_active_idx on message_center_items(item_type,is_active,created_at);
        create trigger if not exists message_center_announcement_insert after insert on announcements begin
            insert or replace into message_center_items(id,item_type,target_user_id,title,body,level,kind,link,is_active,pinned,starts_at,ends_at,created_by,created_at,updated_at)
            values(new.id,'announcement',null,new.title,new.body,new.level,'announcement','',new.is_active,new.pinned,new.starts_at,new.ends_at,new.created_by,new.created_at,new.updated_at);
        end;
        create trigger if not exists message_center_announcement_update after update on announcements begin
            update message_center_items set title=new.title,body=new.body,level=new.level,is_active=new.is_active,pinned=new.pinned,starts_at=new.starts_at,ends_at=new.ends_at,updated_at=new.updated_at where id=new.id;
        end;
        create trigger if not exists message_center_announcement_delete after delete on announcements begin
            delete from message_center_items where id=old.id;
        end;
        create trigger if not exists message_center_notification_insert after insert on notifications begin
            insert or replace into message_center_items(id,item_type,target_user_id,title,body,level,kind,link,is_active,pinned,starts_at,ends_at,created_by,created_at,updated_at)
            values(new.id,'notification',new.target_user_id,new.title,new.body,'info',new.kind,new.link,1,0,null,null,new.created_by,new.created_at,new.created_at);
        end;
        create trigger if not exists message_center_notification_delete after delete on notifications begin
            delete from message_center_items where id=old.id;
        end;
        create trigger if not exists message_center_read_insert after insert on notification_reads begin
            insert or replace into message_center_reads(user_id,item_id,read_at) values(new.user_id,new.notification_id,new.read_at);
        end;
        create trigger if not exists message_center_read_update after update on notification_reads begin
            insert or replace into message_center_reads(user_id,item_id,read_at) values(new.user_id,new.notification_id,new.read_at);
        end;
        create trigger if not exists message_center_delete_insert after insert on notification_deletions begin
            insert or replace into message_center_deletions(user_id,item_id,deleted_at) values(new.user_id,new.notification_id,new.deleted_at);
        end;
        create trigger if not exists message_center_delete_update after update on notification_deletions begin
            insert or replace into message_center_deletions(user_id,item_id,deleted_at) values(new.user_id,new.notification_id,new.deleted_at);
        end;
        """)
        # Migrate legacy announcements and notifications into the shared message center.
        db.execute("""insert or ignore into message_center_items(id,item_type,target_user_id,title,body,level,kind,link,is_active,pinned,starts_at,ends_at,created_by,created_at,updated_at)
            select id,'announcement',null,title,body,level,'announcement','',is_active,pinned,starts_at,ends_at,created_by,created_at,updated_at from announcements""")
        db.execute("""insert or ignore into message_center_items(id,item_type,target_user_id,title,body,level,kind,link,is_active,pinned,starts_at,ends_at,created_by,created_at,updated_at)
            select id,'notification',target_user_id,title,body,'info',kind,link,1,0,null,null,created_by,created_at,created_at from notifications""")
        db.execute("""insert or ignore into message_center_reads(user_id,item_id,read_at) select user_id,notification_id,read_at from notification_reads""")
        db.execute("""insert or ignore into message_center_deletions(user_id,item_id,deleted_at) select user_id,notification_id,deleted_at from notification_deletions""")
        redeem_code_columns={row["name"] for row in db.execute("pragma table_info(redeem_codes)").fetchall()}
        if "source" not in redeem_code_columns:
            db.execute("alter table redeem_codes add column source text not null default ''")
        if "reward" not in redeem_code_columns:
            db.execute("alter table redeem_codes add column reward text not null default ''")
        db.execute("update redeem_codes set source=description where (source is null or source='') and coalesce(description,'')<>''")
        redeem_preference_columns={row["name"] for row in db.execute("pragma table_info(redeem_notification_preferences)").fetchall()}
        if "notify_new" in redeem_preference_columns:
            db.execute("drop table if exists redeem_notification_preferences_legacy")
            db.execute("alter table redeem_notification_preferences rename to redeem_notification_preferences_legacy")
            db.execute("""create table redeem_notification_preferences (
                user_id text not null references users(id) on delete cascade,
                game_id text not null references redeem_games(game_id) on delete cascade,
                updated_at integer not null,
                primary key(user_id,game_id)
            )""")
            db.execute("""insert or ignore into redeem_notification_preferences(user_id,game_id,updated_at)
                select user_id,game_id,updated_at from redeem_notification_preferences_legacy where notify_new=1""")
            db.execute("drop table redeem_notification_preferences_legacy")
        redeem_event_columns={row["name"] for row in db.execute("pragma table_info(redeem_notification_events)").fetchall()}
        if "event_type" in redeem_event_columns:
            db.execute("drop index if exists redeem_notification_events_lookup_idx")
            db.execute("drop table if exists redeem_notification_events_legacy")
            db.execute("alter table redeem_notification_events rename to redeem_notification_events_legacy")
            db.execute("""create table redeem_notification_events (
                id text primary key,
                user_id text not null references users(id) on delete cascade,
                game_id text not null,
                code_id text not null,
                event_key text not null,
                message_item_id text references message_center_items(id) on delete set null,
                created_at integer not null,
                unique(user_id,event_key)
            )""")
            db.execute("""insert or ignore into redeem_notification_events(id,user_id,game_id,code_id,event_key,message_item_id,created_at)
                select id,user_id,game_id,code_id,event_key,message_item_id,created_at
                from redeem_notification_events_legacy where event_type='new'""")
            db.execute("drop table redeem_notification_events_legacy")
        db.execute("create index if not exists redeem_notification_events_lookup_idx on redeem_notification_events(game_id,code_id,created_at)")
        db.execute(
            "insert or ignore into schema_migrations(name,applied_at,details_json) values(?,?,?)",
            ("2026-07-12-redeem-new-code-only-notifications-v1", now(), json.dumps({"preference": "selected_games", "event": "new_code_only"}, ensure_ascii=False)),
        )
        audit_columns={row["name"] for row in db.execute("pragma table_info(admin_audit_logs)").fetchall()}
        audit_additions={
            "event_id":"text","actor_email_snapshot":"text not null default ''","actor_ip":"text not null default ''",
            "category":"text not null default 'administration'","status":"text not null default 'success'",
            "game_id":"text not null default ''","target_type":"text not null default ''","target_id":"text not null default ''",
            "summary":"text not null default ''","before_json":"text not null default ''","after_json":"text not null default ''",
            "metadata_json":"text not null default '{}'","request_id":"text not null default ''","backup_name":"text not null default ''",
            "error_message":"text not null default ''","archived":"integer not null default 0","locked":"integer not null default 0",
        }
        for column_name,column_type in audit_additions.items():
            if column_name not in audit_columns:
                db.execute(f"alter table admin_audit_logs add column {column_name} {column_type}")
        db.execute("update admin_audit_logs set event_id='legacy-'||id where event_id is null or event_id=''")
        db.execute("create unique index if not exists admin_audit_logs_event_idx on admin_audit_logs(event_id)")
        db.execute("create index if not exists admin_audit_logs_filter_idx on admin_audit_logs(created_at,category,status,game_id)")
        db.execute("create index if not exists catalog_scan_previews_expiry_idx on catalog_scan_previews(expires_at)")
        db.execute("create index if not exists source_sync_history_game_idx on source_sync_history(game_id,created_at)")
        source_sync_columns={row["name"] for row in db.execute("pragma table_info(source_sync_history)").fetchall()}
        source_sync_additions={
            "snapshot_dir":"text not null default ''","pre_state_hash":"text not null default ''","post_state_hash":"text not null default ''",
            "rolled_back_at":"integer","rolled_back_by":"text","rollback_reason":"text not null default ''",
        }
        for column_name,column_type in source_sync_additions.items():
            if column_name not in source_sync_columns:
                db.execute(f"alter table source_sync_history add column {column_name} {column_type}")
        redeem_import_columns={row["name"] for row in db.execute("pragma table_info(redeem_import_batches)").fetchall()}
        redeem_import_additions={
            "input_type":"text not null default ''","source_filename":"text not null default ''","source_sheet":"text not null default ''",
            "completed_by":"text","backup_name":"text not null default ''","rollback_backup_name":"text not null default ''",
        }
        for column_name,column_type in redeem_import_additions.items():
            if column_name not in redeem_import_columns:
                db.execute(f"alter table redeem_import_batches add column {column_name} {column_type}")
        source_record_columns={row["name"] for row in db.execute("pragma table_info(game_catalog_source_records)").fetchall()}
        if "official_source_id" not in source_record_columns:
            db.execute("alter table game_catalog_source_records add column official_source_id text not null default ''")

        relation_columns={row["name"] for row in db.execute("pragma table_info(game_achievement_choice_groups)").fetchall()}
        if "relation_type" not in relation_columns:
            db.execute("alter table game_achievement_choice_groups add column relation_type text not null default 'exclusive'")
        if "stage_order" not in relation_columns:
            db.execute("alter table game_achievement_choice_groups add column stage_order integer not null default 0")
        db.execute("update game_achievement_choice_groups set relation_type='exclusive' where relation_type is null or relation_type not in ('exclusive','stage')")

        for project in load_registry().get("projects", []):
            game_id=str(project.get("id") or "").strip()
            if project.get("enabled") and game_id:
                for scope in ("catalog","stats","reports"):
                    db.execute(
                        "insert or ignore into game_live_revisions(game_id,scope,revision,updated_at) values(?,?,0,0)",
                        (game_id,scope),
                    )

        # Wuwa must be re-keyed before the catalog and relation files are loaded.
        # The migration consumes the verified WW_Data evidence once, rewrites every
        # active reference, and removes the generated legacy key space completely.
        migration_stamp=now()
        if not OPEN_SOURCE_EMPTY_DATA:
            existing_wuwa_catalog_count = int(
                db.execute("select count(*) from game_catalog_items where game_id='wuwa'").fetchone()[0]
            )
            if existing_wuwa_catalog_count:
                migrate_wuwa_to_official_ids(db,migration_stamp)
            for project in enabled_projects():
                game_id=str(project.get("id") or "").strip()
                if game_id:
                    _sync_registered_catalog(db,game_id)
                    _sync_achievement_categories(db,game_id)
            migrate_wuwa_to_official_ids(db,migration_stamp)
            _bootstrap_achievement_identities(db)

        for project in enabled_projects():
            game_id=str(project.get("id") or "").strip()
            if game_id:
                if not OPEN_SOURCE_EMPTY_DATA:
                    _sync_registered_relations(db,game_id)
                for field_name in sorted(DERIVED_RELATION_FIELDS):
                    db.execute(
                        """insert into achievement_field_registry(
                        game_id,field_name,classification,mapped_field,active,created_by,created_at,updated_at)
                        values(?,?, 'system_derived_relation','',1,null,?,?)
                        on conflict(game_id,field_name) do update set
                        classification='system_derived_relation',mapped_field='',active=1,updated_at=excluded.updated_at""",
                        (game_id,field_name,migration_stamp,migration_stamp),
                    )
        _migrate_orphan_identity_cleanup(db)
        _migrate_wuwa_shared_model(db)
        if not OPEN_SOURCE_EMPTY_DATA:
            _verify_wuwa_shared_model(db)
            _repair_hsr_catalog_rewards(db)
            normalize_all_game_official_order(db,migration_stamp)
            _verify_wuwa_shared_model(db)
            verify_official_id_model(db)
        db.execute(
            "insert or ignore into schema_migrations(name,applied_at,details_json) values(?,?,?)",
            (
                "2026-06-26-governance-control-v1",
                now(),
                json.dumps(
                    {
                        "action_contract": "backend/services/governance_contract.py",
                        "real_dry_run": True,
                        "differential_catalog_persistence": True,
                        "safe_permanent_delete_restore": True,
                        "official_numeric_id_guard": True,
                    },
                    ensure_ascii=False,
                ),
            ),
        )
        _migrate_completed_keep_decisions(db)
        _repair_final_governance_lifecycle(db)
        _migrate_governance_semantic_decision_hashes(db)
        _repair_invalid_audit_json(db)

        recommendation_migration="2026-07-14-remove-achievement-recommendations-v1"
        if not db.execute("select 1 from schema_migrations where name=?",(recommendation_migration,)).fetchone():
            legacy_count=int(db.execute("select count(*) from featured_achievements").fetchone()[0])
            shared_count=int(db.execute("select count(*) from game_featured_achievements").fetchone()[0])
            db.execute("delete from featured_achievements")
            db.execute("delete from game_featured_achievements")
            db.execute(
                "insert into schema_migrations(name,applied_at,details_json) values(?,?,?)",
                (recommendation_migration,now(),json.dumps({"legacy_removed":legacy_count,"shared_removed":shared_count},ensure_ascii=False)),
            )


        user_columns={row["name"] for row in db.execute("pragma table_info(users)").fetchall()}
        if "username" not in user_columns:
            db.execute("alter table users add column username text")
        if "username_key" not in user_columns:
            db.execute("alter table users add column username_key text")
        db.execute("create unique index if not exists users_username_key_unique on users(username_key) where username_key is not null and username_key <> ''")
        if "role" not in user_columns:
            db.execute("alter table users add column role text not null default 'user'")
        if "is_active" not in user_columns:
            db.execute("alter table users add column is_active integer not null default 1")
        if "last_login_at" not in user_columns:
            db.execute("alter table users add column last_login_at integer")
        if "last_login_ip" not in user_columns:
            db.execute("alter table users add column last_login_ip text")
        session_columns={row["name"] for row in db.execute("pragma table_info(sessions)").fetchall()}
        if "ip_address" not in session_columns:
            db.execute("alter table sessions add column ip_address text not null default ''")
        db.execute("update users set role='user' where role not in ('admin','user') or role is null")
        db.execute("update users set is_active=1 where is_active is null")
        for admin_email in ADMIN_EMAILS:
            db.execute("update users set role='admin', updated_at=? where email_key=?", (now(), admin_email))
        if SITE_OWNER_EMAIL:
            admin_email = SITE_OWNER_EMAIL
            db.execute("update users set role='admin', is_active=1, updated_at=? where email_key=?", (now(), admin_email))
        # 站內通知功能已移除，清除舊通知資料但保留資料表以相容既有資料庫。
        db.execute("delete from notification_reads")
        db.execute("delete from notification_deletions")
        db.execute("delete from notifications")
        cleanup(db)


def cleanup(db: sqlite3.Connection) -> None:
    t = now()
    db.execute("delete from sessions where expires_at <= ?", (t,))
    db.execute("delete from email_verification_tokens where expires_at <= ? or used_at is not null", (t,))
    db.execute("delete from password_reset_tokens where expires_at <= ? or used_at is not null", (t,))
    db.execute("delete from rate_limits where window_start < ?", (t - 172800,))
    try:
        db.execute("delete from game_sync_previews where expires_at <= ?", (t,))
        db.execute("delete from catalog_scan_previews where expires_at <= ?", (t,))
    except sqlite3.OperationalError:
        pass


def _audit_json_text(value: Any, *, limit: int = 200000) -> str:
    """Serialize audit data without ever storing malformed JSON.

    Oversized payloads are represented by a valid metadata envelope with a
    digest and a bounded preview instead of cutting JSON in the middle.
    """
    text = json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
    if len(text) <= limit:
        return text
    preview_limit = max(1000, limit - 1200)
    envelope = {
        "_truncated": True,
        "_original_characters": len(text),
        "_sha256": hashlib.sha256(text.encode("utf-8")).hexdigest(),
        "_preview": text[:preview_limit],
    }
    encoded = json.dumps(envelope, ensure_ascii=False, separators=(",", ":"))
    return encoded[:limit] if len(encoded) <= limit else json.dumps({
        "_truncated": True,
        "_original_characters": len(text),
        "_sha256": envelope["_sha256"],
    }, ensure_ascii=False, separators=(",", ":"))


def _repair_invalid_audit_json(db: sqlite3.Connection) -> dict[str, int]:
    migration = "2026-06-26-final-audit-json-v1"
    if db.execute("select 1 from schema_migrations where name=?", (migration,)).fetchone():
        return {"repaired": 0}
    repaired = 0
    rows = db.execute("select id,before_json,after_json,metadata_json from admin_audit_logs").fetchall()
    for row in rows:
        updates = {}
        for field, default in (("before_json", None), ("after_json", None), ("metadata_json", {})):
            raw = row[field]
            if raw in (None, ""):
                continue
            try:
                json.loads(str(raw))
            except Exception:
                updates[field] = _audit_json_text({
                    "_legacy_invalid_json": True,
                    "_field": field,
                    "_raw": str(raw),
                })
        if updates:
            assignments = ",".join(f"{key}=?" for key in updates)
            db.execute(f"update admin_audit_logs set {assignments} where id=?", (*updates.values(), row["id"]))
            repaired += len(updates)
    db.execute("insert into schema_migrations(name,applied_at,details_json) values(?,?,?)",
               (migration, now(), json.dumps({"repaired_fields": repaired}, ensure_ascii=False)))
    return {"repaired": repaired}


def log_admin_action(
    actor_user_id: str | None,
    action: str,
    target_user_id: str | None = None,
    details: str | None = None,
    *,
    category: str = "administration",
    status: str = "success",
    game_id: str = "",
    target_type: str = "",
    target_id: str = "",
    summary: str = "",
    before: Any = None,
    after: Any = None,
    metadata: Any = None,
    request_id: str = "",
    backup_name: str = "",
    error_message: str = "",
    actor_ip: str = "",
    locked: bool = False,
) -> str:
    event_id = str(uuid.uuid4())
    with connect_db() as db:
        actor_email = ""
        if actor_user_id:
            row = db.execute("select email from users where id=?", (actor_user_id,)).fetchone()
            actor_email = str(row["email"] or "") if row else ""
        db.execute(
            """insert into admin_audit_logs(
            event_id,actor_user_id,actor_email_snapshot,actor_ip,action,category,status,game_id,target_user_id,
            target_type,target_id,summary,details,before_json,after_json,metadata_json,request_id,backup_name,
            error_message,archived,locked,created_at)
            values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                event_id, actor_user_id, actor_email[:254], actor_ip[:100], action[:100], category[:50], status[:30],
                game_id[:50], target_user_id, target_type[:80], target_id[:300], summary[:1000], (details or "")[:5000],
                _audit_json_text(before) if before is not None else "",
                _audit_json_text(after) if after is not None else "",
                _audit_json_text(metadata if metadata is not None else {}),
                request_id[:100], backup_name[:255], error_message[:5000], 0, 1 if locked else 0, now(),
            ),
        )
    return event_id


# 20260621-sync-log-game-label-v1
def game_display_name(game_id: str) -> str:
    value=str(game_id or "").strip()
    fallback={"wuwa":"鳴潮","hsr":"崩壞：星穹鐵道"}.get(value,value or "未指定遊戲")
    try:
        project=next((row for row in load_registry().get("projects", []) if str(row.get("id") or "").strip()==value),None)
        name=str((project or {}).get("name") or "").strip()
        return name or fallback
    except Exception:
        return fallback


def enabled_game_projects() -> list[dict[str,Any]]:
    projects=[]
    for row in load_registry().get("projects", []):
        game_id=str(row.get("id") or "").strip()
        if row.get("enabled") and game_id:
            projects.append({"id":game_id,"name":str(row.get("name") or game_display_name(game_id)).strip()})
    return projects


def _progress_counts_by_user_game(db: sqlite3.Connection, user_id: str | None = None) -> dict[str,dict[str,int]]:
    where="where p.user_id=?" if user_id else ""
    params=(user_id,) if user_id else ()
    rows=db.execute(f"""select p.user_id,p.game_id,
        count(distinct case
          when rg.relation_type='exclusive' and rg.group_id is not null
            then 'exclusive:'||rg.group_id
          else 'achievement:'||p.achievement_id
        end) as completed_count
      from game_progress p
      left join game_achievement_choice_groups rg
        on rg.game_id=p.game_id and rg.achievement_id=p.achievement_id
      {where}
      group by p.user_id,p.game_id""",params).fetchall()
    result: dict[str,dict[str,int]]={}
    for row in rows:
        result.setdefault(str(row["user_id"]),{})[str(row["game_id"])]=int(row["completed_count"] or 0)
    return result


def _game_catalog_summaries(db: sqlite3.Connection) -> list[dict[str,Any]]:
    summaries=[]
    for project in enabled_game_projects():
        game_id=project["id"]
        summaries.append({
            "id":game_id,
            "name":project["name"],
            "catalog_count":_effective_catalog_count(db,game_id),
        })
    return summaries


def _user_progress_summary(db: sqlite3.Connection, user_id: str) -> dict[str,Any]:
    counts=_progress_counts_by_user_game(db,user_id).get(user_id,{})
    games=[]
    for project in _game_catalog_summaries(db):
        completed=int(counts.get(project["id"],0))
        games.append({**project,"completed_count":completed})
    return {
        "games":games,
        "completed_count":sum(int(row["completed_count"]) for row in games),
        "catalog_count":sum(int(row["catalog_count"]) for row in games),
    }



HIGH_RISK_OPERATION_GUARD = threading.RLock()
BACKUP_OPERATION_GUARD = HIGH_RISK_OPERATION_GUARD


def high_risk_operation(func):
    """Serialize all catalog, relation, backup and governance write workflows."""
    @wraps(func)
    def wrapped(*args, **kwargs):
        if not HIGH_RISK_OPERATION_GUARD.acquire(blocking=False):
            raise HTTPException(status_code=409, detail="另一項高風險資料操作正在執行，請完成後再試。")
        try:
            return func(*args, **kwargs)
        finally:
            HIGH_RISK_OPERATION_GUARD.release()
    return wrapped


def _copy_sqlite_database(source: Path, target: Path) -> None:
    """Copy a SQLite database while deterministically releasing Windows file handles."""
    src: sqlite3.Connection | None = None
    dst: sqlite3.Connection | None = None
    try:
        src = sqlite3.connect(source)
        dst = sqlite3.connect(target)
        src.backup(dst)
        dst.commit()
    finally:
        if dst is not None:
            try:
                dst.close()
            except Exception:
                pass
        if src is not None:
            try:
                src.close()
            except Exception:
                pass


def _remove_temporary_tree(path: Path, *, attempts: int = 8, delay_seconds: float = 0.08) -> str:
    """Best-effort Windows-safe cleanup. Cleanup errors must not hide the real operation error."""
    last_error = ""
    for attempt in range(max(1, attempts)):
        try:
            shutil.rmtree(path)
            return ""
        except FileNotFoundError:
            return ""
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
            gc.collect()
            if attempt + 1 < attempts:
                time.sleep(delay_seconds * (attempt + 1))
    return last_error


def create_database_backup() -> Path:
    with BACKUP_OPERATION_GUARD:
        backup_dir = ROOT / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp=time.strftime('%Y%m%d-%H%M%S')
        target=backup_dir/f"app-{stamp}.db"
        suffix=1
        while target.exists():
            target=backup_dir/f"app-{stamp}-{suffix:02d}.db"
            suffix+=1
        _copy_sqlite_database(DB_FILE, target)
        return target


def json_list(value: str | None) -> list[str]:
    try:
        parsed=json.loads(value or "[]")
        return [str(x)[:80] for x in parsed if str(x).strip()] if isinstance(parsed,list) else []
    except Exception:
        return []

def enforce_blocklist(email_address: str, ip_address: str) -> None:
    email_key=normalize_email(email_address)
    domain=email_key.rsplit("@",1)[-1] if "@" in email_key else ""
    with connect_db() as db:
        rows=db.execute("select kind,value_key,reason from blocked_entries where active=1").fetchall()
    for row in rows:
        kind=row["kind"]; value=row["value_key"]
        matched=(kind=="ip" and ip_address==value) or (kind=="email" and email_key==value) or (kind=="domain" and domain==value)
        if matched:
            raise HTTPException(status_code=403,detail="此來源目前無法使用帳號功能。" + ((" 原因："+row["reason"]) if row["reason"] else ""))

def create_notification(title: str, body: str, kind: str="info", link: str="", target_user_id: str | None=None, created_by: str | None=None) -> str:
    """Create a broadcast or account-specific in-site notification."""
    notification_id=str(uuid.uuid4())
    with connect_db() as db:
        if target_user_id:
            exists=db.execute("select 1 from users where id=?",(target_user_id,)).fetchone()
            if not exists:
                raise HTTPException(status_code=404,detail="找不到通知目標帳號。")
        db.execute(
            "insert into notifications(id,target_user_id,title,body,kind,link,created_by,created_at) values(?,?,?,?,?,?,?,?)",
            (notification_id,target_user_id,title.strip()[:200],body.strip()[:3000],kind.strip()[:30] or "info",link.strip()[:500],created_by,now()),
        )
    return notification_id


def client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for", "")
    if forwarded:
        return forwarded.split(",", 1)[0].strip()[:100]
    return (request.client.host if request.client else "unknown")[:100]


def consume_rate_limit(action: str, key: str, limit: int, seconds: int) -> None:
    bucket = now() // seconds * seconds
    with connect_db() as db:
        db.execute("begin immediate")
        row = db.execute("select count from rate_limits where action=? and rate_key=? and window_start=?", (action,key,bucket)).fetchone()
        count = int(row["count"]) + 1 if row else 1
        if row:
            db.execute("update rate_limits set count=? where action=? and rate_key=? and window_start=?", (count,action,key,bucket))
        else:
            db.execute("insert into rate_limits(action,rate_key,window_start,count) values(?,?,?,?)", (action,key,bucket,count))
        if count > limit:
            raise HTTPException(status_code=429, detail="操作太頻繁，請等待最多 5 分鐘後再試。")


def hash_password(password: str) -> str:
    return PASSWORD_HASHER.hash(password)


def verify_password(password: str, encoded: str) -> bool:
    try:
        return PASSWORD_HASHER.verify(encoded, password)
    except (VerifyMismatchError, VerificationError, InvalidHashError):
        return False


def create_session(user_id: str, ip_address: str = "") -> str:
    token = secrets.token_urlsafe(40)
    t = now()
    with connect_db() as db:
        cleanup(db)
        db.execute(
            "insert into sessions(token_hash,user_id,expires_at,created_at,ip_address) values(?,?,?,?,?)",
            (digest_token(token),user_id,t+SESSION_SECONDS,t,(ip_address or "")[:100])
        )
    return token


def session_user(token: str | None) -> dict[str, Any] | None:
    if not token:
        return None
    with connect_db() as db:
        row = db.execute("""
        select users.id, users.email, users.username, users.role
        from sessions join users on users.id=sessions.user_id
        where sessions.token_hash=? and sessions.expires_at>? and users.email_verified=1 and users.is_active=1
        """, (digest_token(token), now())).fetchone()
    return {"id":row["id"],"email":row["email"],"username":row["username"],"role":row["role"],"is_site_owner":is_site_owner_email(row["email"])} if row else None


def set_session_cookie(response: Response, token: str) -> None:
    response.set_cookie(COOKIE_NAME, token, max_age=SESSION_SECONDS, httponly=True, secure=COOKIE_SECURE, samesite="strict", path="/")


def clear_session_cookie(response: Response) -> None:
    response.delete_cookie(COOKIE_NAME, path="/", secure=COOKIE_SECURE, httponly=True, samesite="strict")


def current_user(request: Request) -> dict[str, Any] | None:
    return session_user(request.cookies.get(COOKIE_NAME))


def require_user(request: Request) -> dict[str, Any]:
    user = current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="請先登入。")
    return user


def require_admin(request: Request) -> dict[str, Any]:
    user = require_user(request)
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="此功能僅限管理員使用。")
    return user


def require_site_owner(request: Request) -> dict[str, Any]:
    user = require_admin(request)
    if not is_site_owner_email(str(user.get("email") or "")):
        raise HTTPException(status_code=403, detail="帳號管理僅限站長使用。")
    return user


def validate_ids(values: list[str]) -> list[str]:
    if len(values) > 10000:
        raise HTTPException(status_code=400, detail="一次處理的成就數量過多。")
    result=[]; seen=set()
    for value in values:
        item=str(value or "").strip()
        if item and len(item)<=200 and item not in seen:
            seen.add(item); result.append(item)
    return result


def build_email(to: str, subject: str, text: str, html_body: str) -> EmailMessage:
    message=EmailMessage()
    message["From"]=SMTP_FROM
    message["To"]=to
    message["Subject"]=subject
    message.set_content(text)
    message.add_alternative(html_body, subtype="html")
    return message


def deliver_email(message: EmailMessage, mail_type: str="generic") -> None:
    recipient=str(message.get("To", ""))[:300]
    subject=str(message.get("Subject", ""))[:300]
    status="sent"; error=""
    try:
        if MAIL_DELIVERY == "console":
            stamp=time.strftime("%Y%m%d-%H%M%S")
            path=OUTBOX_DIR/f"{stamp}-{secrets.token_hex(4)}.eml"
            path.write_bytes(message.as_bytes())
            print(f"[郵件測試模式] 已寫入 {path}")
            status="saved"
            return
        context=ssl.create_default_context()
        if not SMTP_VALIDATE_CERT:
            context.check_hostname=False
            context.verify_mode=ssl.CERT_NONE
        if SMTP_SSL:
            client=smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT, timeout=25, context=context)
        else:
            client=smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=25)
        try:
            client.ehlo()
            if SMTP_STARTTLS and not SMTP_SSL:
                client.starttls(context=context); client.ehlo()
            if SMTP_USERNAME:
                client.login(SMTP_USERNAME, SMTP_PASSWORD)
            client.send_message(message)
        finally:
            try: client.quit()
            except Exception: client.close()
    except Exception as exc:
        status="failed"; error=f"{type(exc).__name__}: {exc}"[:2000]
        raise
    finally:
        try:
            with connect_db() as db:
                db.execute("insert into email_logs(recipient,subject,mail_type,status,error,created_at) values(?,?,?,?,?,?)",
                           (recipient,subject,mail_type[:40],status,error,now()))
        except Exception as log_exc:
            print("[郵件紀錄寫入失敗]",repr(log_exc))


def send_verification(email_address: str, token: str) -> None:
    link=f"{PUBLIC_BASE_URL}/account/?verify_token={urllib.parse.quote(token)}"
    safe=html.escape(link)
    message=build_email(email_address, "驗證你的遊戲成就紀錄器帳號", f"請開啟以下連結完成信箱驗證：\n{link}\n\n連結將在 24 小時後失效。", f"<h2>驗證信箱</h2><p>請點擊下方連結完成驗證：</p><p><a href=\"{safe}\">完成信箱驗證</a></p><p>連結將在 24 小時後失效。</p>")
    deliver_email(message, "verification")


def send_reset(email_address: str, token: str) -> None:
    link=f"{PUBLIC_BASE_URL}/account/?reset_token={urllib.parse.quote(token)}"
    safe=html.escape(link)
    message=build_email(email_address, "重設遊戲成就紀錄器密碼", f"請開啟以下連結設定新密碼：\n{link}\n\n連結將在 30 分鐘後失效。若不是你提出申請，可忽略此信。", f"<h2>重設密碼</h2><p>請點擊下方連結設定新密碼：</p><p><a href=\"{safe}\">設定新密碼</a></p><p>連結將在 30 分鐘後失效。若不是你提出申請，可忽略此信。</p>")
    deliver_email(message, "password_reset")


def issue_verification(user_id: str, email_address: str) -> None:
    token=secrets.token_urlsafe(40); t=now()
    with connect_db() as db:
        db.execute("delete from email_verification_tokens where user_id=?", (user_id,))
        db.execute("insert into email_verification_tokens(token_hash,user_id,expires_at,created_at) values(?,?,?,?)", (digest_token(token),user_id,t+VERIFY_SECONDS,t))
    send_verification(email_address, token)


def issue_reset(user_id: str, email_address: str) -> None:
    token=secrets.token_urlsafe(40); t=now()
    with connect_db() as db:
        db.execute("delete from password_reset_tokens where user_id=?", (user_id,))
        db.execute("insert into password_reset_tokens(token_hash,user_id,expires_at,created_at) values(?,?,?,?)", (digest_token(token),user_id,t+RESET_SECONDS,t))
    send_reset(email_address, token)


def convert_to_traditional(value: Any) -> Any:
    if isinstance(value,str): return TRADITIONAL_CONVERTER.convert(value)
    if isinstance(value,list): return [convert_to_traditional(x) for x in value]
    if isinstance(value,dict): return {k:convert_to_traditional(v) for k,v in value.items()}
    return value


def _static_game_icon_response(path: Path, label: str):
    if not path.is_file():
        raise HTTPException(status_code=404, detail=f"{label} fixed icon is missing.")
    return FileResponse(
        path,
        media_type="image/png",
        headers={"Cache-Control": "no-cache", "X-Content-Type-Options": "nosniff"},
    )



@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()
    yield

app=FastAPI(title="遊戲成就紀錄器",docs_url=None,redoc_url=None,openapi_url=None,lifespan=lifespan)
app.add_middleware(TrustedHostMiddleware,allowed_hosts=TRUSTED_HOSTS)

@app.middleware("http")
async def security_middleware(request: Request, call_next):
    if request.method in {"POST","PUT","PATCH","DELETE"} and request.url.path.startswith("/api/"):
        origin=request.headers.get("origin")
        if APP_ENV=="production" and origin and origin.rstrip("/") not in TRUSTED_ORIGINS:
            return JSONResponse(status_code=403,content={"ok":False,"message":"來源驗證失敗。"})
    response=await call_next(request)
    response.headers["X-Content-Type-Options"]="nosniff"
    response.headers["Referrer-Policy"]="strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"]="camera=(), microphone=(), geolocation=()"
    icon_paths = {
        "/api/wuwa-official-logo",
        "/api/hsr-official-logo",
        "/api/genshin-logo",
        "/api/zzz-logo",
    }
    is_game_icon = request.url.path in icon_paths or request.url.path.startswith("/api/games/") and request.url.path.endswith("/icon")
    if is_game_icon and response.status_code == 200:
        response.headers["Cache-Control"] = "public, max-age=86400"
    else:
        response.headers["Cache-Control"] = "no-store" if request.url.path.startswith("/api/") else "no-cache"
    return response

@app.exception_handler(HTTPException)
async def http_error(_: Request, exc: HTTPException):
    return JSONResponse(status_code=exc.status_code,content={"ok":False,"message":str(exc.detail)})

@app.get("/api/health")
def health():
    return {"ok":True,"service":"game-achievement-hub","mail_delivery":MAIL_DELIVERY}

@app.get("/api/games/{game_id}/icon")
def game_icon(game_id: str):
    icon = GAME_ICON_FILES.get(game_id)
    if icon is None:
        raise HTTPException(status_code=404, detail="找不到此遊戲圖像。")
    return _static_game_icon_response(icon, GAME_ICON_LABELS.get(game_id, game_id))


# 舊圖像 API 保留相容性；所有端點都回傳同一套統一資產。
@app.get("/api/wuwa-official-logo")
def wuwa_official_logo():
    return game_icon("wuwa")


@app.get("/api/hsr-official-logo")
def hsr_official_logo():
    return game_icon("hsr")


@app.get("/api/genshin-logo")
def genshin_logo():
    return game_icon("genshin")


@app.get("/api/zzz-logo")
def zzz_logo():
    return game_icon("zzz")


@app.post("/api/auth/register")
def register(body: EmailPassword, request: Request):
    email_key=normalize_email(body.email)
    enforce_blocklist(str(body.email),client_ip(request))
    consume_rate_limit("register-ip-5m",client_ip(request),10,AUTH_RATE_WINDOW_SECONDS)
    consume_rate_limit("register-email-5m",email_key,3,AUTH_RATE_WINDOW_SECONDS)
    user_id=str(uuid.uuid4()); t=now()
    try:
        with connect_db() as db:
            role="admin" if email_key in ADMIN_EMAILS else "user"
            db.execute("insert into users(id,email,email_key,password_hash,email_verified,role,created_at,updated_at) values(?,?,?,?,0,?,?,?)",(user_id,str(body.email),email_key,hash_password(body.password),role,t,t))
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409,detail="此電子信箱已註冊。")
    try:
        issue_verification(user_id,str(body.email))
    except Exception as exc:
        print("[寄送驗證信失敗]",repr(exc))
        raise HTTPException(status_code=503,detail="帳號已建立，但驗證信寄送失敗。請檢查郵件伺服器後使用『重新寄送驗證信』。")
    return {"ok":True,"requires_verification":True}

@app.post("/api/auth/resend-verification")
def resend_verification(body: EmailOnly, request: Request):
    email_key=normalize_email(body.email)
    consume_rate_limit("resend-ip-5m",client_ip(request),10,AUTH_RATE_WINDOW_SECONDS)
    consume_rate_limit("resend-email-5m",email_key,3,AUTH_RATE_WINDOW_SECONDS)
    with connect_db() as db:
        row=db.execute("select id,email,email_verified from users where email_key=?",(email_key,)).fetchone()
    if row and not row["email_verified"]:
        try: issue_verification(row["id"],row["email"])
        except Exception as exc:
            print("[重寄驗證信失敗]",repr(exc))
            raise HTTPException(status_code=503,detail="驗證信寄送失敗，請檢查郵件伺服器。")
    return {"ok":True}

@app.post("/api/auth/verify-email")
def verify_email(body: TokenOnly, request: Request, response: Response):
    token_hash=digest_token(body.token); t=now(); login_ip=client_ip(request)
    with connect_db() as db:
        db.execute("begin immediate")
        row=db.execute("select v.user_id,u.email,u.username,u.role from email_verification_tokens v join users u on u.id=v.user_id where v.token_hash=? and v.expires_at>? and v.used_at is null",(token_hash,t)).fetchone()
        if not row: raise HTTPException(status_code=400,detail="驗證連結無效或已過期。")
        enforce_blocklist(row["email"],login_ip)
        db.execute("update users set email_verified=1,last_login_at=?,last_login_ip=?,updated_at=? where id=?",(t,login_ip,t,row["user_id"]))
        db.execute("update email_verification_tokens set used_at=? where token_hash=?",(t,token_hash))
    token=create_session(row["user_id"],login_ip); set_session_cookie(response,token)
    return {"ok":True,"user":{"id":row["user_id"],"email":row["email"],"username":row["username"],"role":row["role"],"is_site_owner":is_site_owner_email(row["email"])}}

@app.post("/api/auth/login")
def login(body: LoginPayload, request: Request, response: Response):
    identifier = str(body.identifier or body.email or "").strip()
    if len(identifier) < 3:
        raise HTTPException(status_code=400, detail="請輸入使用者名稱或電子信箱。")
    identifier_key = identifier.casefold()
    login_ip = client_ip(request)
    enforce_blocklist(identifier, login_ip)
    consume_rate_limit("login-ip-5m", login_ip, 30, AUTH_RATE_WINDOW_SECONDS)
    consume_rate_limit("login-identifier-5m", identifier_key, 10, AUTH_RATE_WINDOW_SECONDS)
    with connect_db() as db:
        row = db.execute(
            "select id,email,username,password_hash,email_verified,role,is_active from users where email_key=? or username_key=?",
            (identifier_key, identifier_key),
        ).fetchone()
    if not row or not verify_password(body.password, row["password_hash"]):
        raise HTTPException(status_code=401, detail="使用者名稱、電子信箱或密碼錯誤。")
    enforce_blocklist(row["email"], login_ip)
    if not row["is_active"]:
        raise HTTPException(status_code=403, detail="此帳號已被管理員停用。")
    if not row["email_verified"]:
        raise HTTPException(status_code=403, detail="請先到信箱完成帳號驗證。")
    with connect_db() as db:
        db.execute("update users set last_login_at=?,last_login_ip=?,updated_at=? where id=?", (now(),login_ip,now(),row["id"]))
    token = create_session(row["id"], login_ip)
    set_session_cookie(response, token)
    return {"ok":True,"user":{"id":row["id"],"email":row["email"],"username":row["username"],"role":row["role"],"is_site_owner":is_site_owner_email(row["email"])}}

@app.post("/api/auth/logout")
def logout(request: Request, response: Response):
    token=request.cookies.get(COOKIE_NAME)
    if token:
        with connect_db() as db: db.execute("delete from sessions where token_hash=?",(digest_token(token),))
    clear_session_cookie(response)
    return {"ok":True}

@app.get("/api/auth/me")
def me(request: Request):
    user=current_user(request)
    return {"ok":True,"authenticated":bool(user),"user":user}

@app.put("/api/auth/username")
@app.post("/api/auth/username")
def update_username(body: UsernamePayload, request: Request):
    user = require_user(request)
    username, username_key = validate_username(body.username)
    consume_rate_limit("username-user-5m", user["id"], 5, AUTH_RATE_WINDOW_SECONDS)
    with connect_db() as db:
        current = db.execute("select username from users where id=? and is_active=1", (user["id"],)).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="找不到帳號。")
        duplicate = db.execute("select id from users where username_key=? and id<>?", (username_key, user["id"])).fetchone()
        if duplicate:
            raise HTTPException(status_code=409, detail="此使用者名稱已被使用。")
        try:
            db.execute("update users set username=?,username_key=?,updated_at=? where id=?", (username,username_key,now(),user["id"]))
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="此使用者名稱已被使用。")
    action = "設定使用者名稱" if not current["username"] else "更改使用者名稱"
    log_admin_action(user["id"], action, user["id"], f"使用者名稱：{username}")
    return {"ok":True,"user":{"id":user["id"],"email":user["email"],"username":username,"role":user["role"],"is_site_owner":is_site_owner_email(user["email"])}}

@app.post("/api/auth/change-password")
def change_password(body: ChangePasswordPayload, request: Request, response: Response):
    user = require_user(request)
    ip_address = client_ip(request)
    consume_rate_limit("change-password-ip-5m", ip_address, 10, AUTH_RATE_WINDOW_SECONDS)
    consume_rate_limit("change-password-user-5m", user["id"], 5, AUTH_RATE_WINDOW_SECONDS)
    if body.current_password == body.new_password:
        raise HTTPException(status_code=400, detail="新密碼不可與目前密碼相同。")

    with connect_db() as db:
        row = db.execute(
            "select password_hash from users where id=? and is_active=1",
            (user["id"],),
        ).fetchone()
        if not row or not verify_password(body.current_password, row["password_hash"]):
            raise HTTPException(status_code=401, detail="目前密碼錯誤。")
        db.execute(
            "update users set password_hash=?, updated_at=? where id=?",
            (hash_password(body.new_password), now(), user["id"]),
        )
        db.execute("delete from sessions where user_id=?", (user["id"],))

    token = create_session(user["id"], ip_address)
    set_session_cookie(response, token)
    log_admin_action(user["id"], "更換密碼", user["id"], "使用者自行更換密碼")
    return {"ok": True}
@app.post("/api/auth/forgot-password")
def forgot_password(body: EmailOnly, request: Request):
    email_key=normalize_email(body.email)
    consume_rate_limit("forgot-ip-5m",client_ip(request),10,AUTH_RATE_WINDOW_SECONDS)
    consume_rate_limit("forgot-email-5m",email_key,3,AUTH_RATE_WINDOW_SECONDS)
    with connect_db() as db:
        row=db.execute("select id,email,email_verified from users where email_key=?",(email_key,)).fetchone()
    if row and row["email_verified"]:
        try: issue_reset(row["id"],row["email"])
        except Exception as exc: print("[寄送重設信失敗]",repr(exc))
    return {"ok":True}

@app.post("/api/auth/reset-password/validate")
def validate_reset_password_token(body: TokenOnly, request: Request):
    consume_rate_limit("reset-validate-ip-5m", client_ip(request), 30, AUTH_RATE_WINDOW_SECONDS)
    token_hash = digest_token(body.token)
    t = now()
    with connect_db() as db:
        row = db.execute(
            "select 1 from password_reset_tokens where token_hash=? and expires_at>? and used_at is null",
            (token_hash, t),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=400, detail="密碼重設連結無效或已過期。")
    return {"ok": True, "valid": True}

@app.post("/api/auth/reset-password")
def reset_password(body: ResetPassword, request: Request):
    consume_rate_limit("reset-ip-5m",client_ip(request),10,AUTH_RATE_WINDOW_SECONDS)
    token_hash=digest_token(body.token); t=now()
    with connect_db() as db:
        db.execute("begin immediate")
        row=db.execute("select user_id from password_reset_tokens where token_hash=? and expires_at>? and used_at is null",(token_hash,t)).fetchone()
        if not row: raise HTTPException(status_code=400,detail="密碼重設連結無效或已過期。")
        db.execute("update users set password_hash=?,updated_at=? where id=?",(hash_password(body.password),t,row["user_id"]))
        db.execute("update password_reset_tokens set used_at=? where token_hash=?",(t,token_hash))
        db.execute("delete from sessions where user_id=?",(row["user_id"],))
    return {"ok":True}

@app.get("/api/admin/overview")
def admin_overview(request: Request):
    return extra_game_admin_overview("wuwa",request)



@app.get("/api/live-state")
def live_state():
    return extra_game_live_state("wuwa")


@app.get("/api/admin/users")
def admin_users(request: Request):
    """Return shared account information with effective per-game completion totals."""
    admin=require_site_owner(request)
    with connect_db() as db:
        rows=db.execute("""select u.id,u.email,u.username,u.email_verified,u.role,u.is_active,u.created_at,u.updated_at,
               u.last_login_at,u.last_login_ip,
               exists(select 1 from blocked_entries b where b.active=1 and b.kind='email' and b.value_key=u.email_key) is_blocked,
               (select b.id from blocked_entries b where b.active=1 and b.kind='email' and b.value_key=u.email_key order by b.created_at desc limit 1) block_id,
               (select count(*) from sessions s where s.user_id=u.id and s.expires_at>?) session_count
        from users u
        order by case when u.role='admin' then 0 else 1 end,u.created_at,u.email_key""",(now(),)).fetchall()
        progress_map=_progress_counts_by_user_game(db)
        games=_game_catalog_summaries(db)
    users=[]
    for r in rows:
        by_game={game["id"]:int(progress_map.get(str(r["id"]),{}).get(game["id"],0)) for game in games}
        users.append({
            "id":r["id"],"email":r["email"],"username":r["username"] or "","email_verified":bool(r["email_verified"]),"role":r["role"],
            "is_active":bool(r["is_active"]),"created_at":r["created_at"],"updated_at":r["updated_at"],
            "last_login_at":r["last_login_at"],"last_login_ip":r["last_login_ip"] or "",
            "is_blocked":bool(r["is_blocked"]),"block_id":r["block_id"] or "",
            "progress_count":sum(by_game.values()),"progress_by_game":by_game,"session_count":int(r["session_count"] or 0),
            "is_self":r["id"]==admin["id"],"is_site_owner":is_site_owner_email(r["email"])
        })
    return {"ok":True,"users":users,"games":games}


def _sanitize_admin_history_value(value: Any) -> Any:
    if isinstance(value, dict):
        return {sanitize_legacy_id_display(str(key)): _sanitize_admin_history_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_sanitize_admin_history_value(item) for item in value]
    if isinstance(value, str):
        return sanitize_legacy_id_display(value)
    return value


def _safe_audit_json(value: Any, default: Any) -> Any:
    """Parse historical audit JSON without allowing one damaged row to break the list API."""
    if value in (None, ""):
        return default
    if isinstance(value, (dict, list, int, float, bool)):
        return value
    text = str(value)
    try:
        return _sanitize_admin_history_value(json.loads(text))
    except Exception as exc:
        return {
            "_parse_error": "legacy_json_invalid",
            "_message": str(exc),
            "_raw": text[:4000],
        }


@app.get("/api/admin/audit-logs")
def admin_audit_logs(
    request: Request,
    category: str = "",
    status: str = "",
    game_id: str = "",
    actor: str = "",
    search: str = "",
    include_archived: bool = False,
    limit: int = 200,
    offset: int = 0,
):
    require_admin(request)
    limit=max(1,min(500,int(limit))); offset=max(0,int(offset))
    clauses=[]; params=[]
    if not include_archived: clauses.append("coalesce(l.archived,0)=0")
    if category: clauses.append("l.category=?"); params.append(category[:50])
    if status: clauses.append("l.status=?"); params.append(status[:30])
    if game_id: clauses.append("l.game_id=?"); params.append(game_id[:50])
    if actor:
        clauses.append("(l.actor_email_snapshot like ? or actor.email like ?)")
        pattern=f"%{actor[:200]}%"; params.extend([pattern,pattern])
    if search:
        clauses.append("(l.action like ? or l.summary like ? or l.details like ? or l.target_id like ? or l.event_id like ?)")
        pattern=f"%{search[:300]}%"; params.extend([pattern]*5)
    where=("where "+" and ".join(clauses)) if clauses else ""
    with connect_db() as db:
        total=int(db.execute(f"select count(*) c from admin_audit_logs l left join users actor on actor.id=l.actor_user_id {where}",params).fetchone()["c"] or 0)
        rows=db.execute(f"""
        select l.*,coalesce(nullif(l.actor_email_snapshot,''),actor.email,'已刪除帳號') actor_email,
               target.email target_email
        from admin_audit_logs l
        left join users actor on actor.id=l.actor_user_id
        left join users target on target.id=l.target_user_id
        {where}
        order by l.id desc limit ? offset ?
        """,(*params,limit,offset)).fetchall()
    return {"ok":True,"total":total,"limit":limit,"offset":offset,"logs":[{
        "id":r["id"],"event_id":r["event_id"] or f"legacy-{r['id']}","action":r["action"],
        "category":r["category"] or "administration","status":r["status"] or "success",
        "game_id":r["game_id"] or "","details":sanitize_legacy_id_display(r["details"] or ""),"summary":sanitize_legacy_id_display(r["summary"] or ""),
        "created_at":r["created_at"],"actor_email":r["actor_email"] or "已刪除帳號","actor_ip":r["actor_ip"] or "",
        "target_email":r["target_email"] or "","target_type":r["target_type"] or "","target_id":sanitize_legacy_id_display(r["target_id"] or ""),
        "before":_safe_audit_json(r["before_json"], None),
        "after":_safe_audit_json(r["after_json"], None),
        "metadata":_safe_audit_json(r["metadata_json"], {}),"request_id":r["request_id"] or "",
        "backup_name":r["backup_name"] or "","error_message":sanitize_legacy_id_display(r["error_message"] or ""),
        "archived":bool(r["archived"]),"locked":bool(r["locked"]),
    } for r in rows]}


@app.post("/api/admin/audit-logs/{log_id}/archive")
def admin_archive_audit_log(log_id: int, body: AuditArchivePayload, request: Request):
    admin=require_admin(request)
    with connect_db() as db:
        row=db.execute("select * from admin_audit_logs where id=?",(log_id,)).fetchone()
        if not row: raise HTTPException(status_code=404,detail="找不到操作紀錄。")
        if bool(row["locked"]): raise HTTPException(status_code=409,detail="此紀錄已鎖定，不能封存。")
        db.execute("update admin_audit_logs set archived=1 where id=?",(log_id,))
    log_admin_action(admin["id"],"archive_audit_log",details=f"archived_log={log_id}; reason={body.reason}",category="audit",target_type="audit_log",target_id=str(log_id),summary="封存操作紀錄",metadata={"reason":body.reason},locked=True,actor_ip=client_ip(request))
    return {"ok":True}


@app.delete("/api/admin/audit-logs/{log_id}")
def admin_delete_audit_log(log_id: int, request: Request):
    require_admin(request)
    raise HTTPException(status_code=405,detail="操作紀錄不可永久刪除，請改用封存功能。")


# 20260621-admin-username-management-v1
@app.patch("/api/admin/users/{user_id}/username")
def admin_update_username(user_id: str, body: AdminUsernamePayload, request: Request):
    admin = require_site_owner(request)
    username, username_key = validate_username(body.username)
    with connect_db() as db:
        target = db.execute("select id,email,username from users where id=?", (user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="找不到帳號。")
        duplicate = db.execute("select id from users where username_key=? and id<>?", (username_key,user_id)).fetchone()
        if duplicate:
            raise HTTPException(status_code=409, detail="此使用者名稱已被使用。")
        old_username = target["username"] or ""
        try:
            db.execute("update users set username=?,username_key=?,updated_at=? where id=?", (username,username_key,now(),user_id))
        except sqlite3.IntegrityError:
            raise HTTPException(status_code=409, detail="此使用者名稱已被使用。")
    action = "新增使用者名稱" if not old_username else "更改使用者名稱"
    log_admin_action(admin["id"], action, user_id, f"舊使用者名稱：{old_username or '未設定'}｜新使用者名稱：{username}")
    return {"ok":True,"username":username}


@app.delete("/api/admin/users/{user_id}/username")
def admin_delete_username(user_id: str, request: Request):
    admin = require_site_owner(request)
    with connect_db() as db:
        target = db.execute("select id,email,username from users where id=?", (user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404, detail="找不到帳號。")
        old_username = target["username"] or ""
        if old_username:
            db.execute("update users set username=null,username_key=null,updated_at=? where id=?", (now(),user_id))
    if old_username:
        log_admin_action(admin["id"], "刪除使用者名稱", user_id, f"使用者名稱：{old_username}")
    return {"ok":True,"deleted":bool(old_username)}


@app.patch("/api/admin/users/{user_id}/email")
def admin_update_email(user_id: str, body: AdminEmailUpdatePayload, request: Request):
    admin=require_site_owner(request)
    email=str(body.email).strip()
    email_key=normalize_email(email)
    with connect_db() as db:
        db.execute("begin immediate")
        target=db.execute("select id,email,email_key,role from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到帳號。")
        if is_site_owner_email(target["email"]):
            raise HTTPException(status_code=400,detail="站長信箱不可由後台修改。")
        duplicate=db.execute("select id from users where email_key=? and id<>?",(email_key,user_id)).fetchone()
        if duplicate:
            raise HTTPException(status_code=409,detail="此電子信箱已被使用。")
        old_email=target["email"]
        db.execute("update users set email=?,email_key=?,email_verified=?,updated_at=? where id=?",(email,email_key,1 if body.verified else 0,now(),user_id))
        db.execute("delete from email_verification_tokens where user_id=?",(user_id,))
        db.execute("delete from password_reset_tokens where user_id=?",(user_id,))
        db.execute("delete from sessions where user_id=?",(user_id,))
    log_admin_action(admin["id"],"update_email",user_id,f"old={old_email}; new={email}; verified={body.verified}")
    return {"ok":True,"email":email,"verified":bool(body.verified),"logged_out":True}


@app.post("/api/admin/users/{user_id}/reset-password")
def admin_reset_user_password(user_id: str, body: AdminPasswordResetPayload, request: Request):
    admin=require_site_owner(request)
    if user_id==admin["id"]:
        raise HTTPException(status_code=400,detail="請從帳號頁使用『更換密碼』修改自己的密碼。")
    with connect_db() as db:
        target=db.execute("select id,email from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到帳號。")
        db.execute("update users set password_hash=?,updated_at=? where id=?",(hash_password(body.password),now(),user_id))
        db.execute("delete from sessions where user_id=?",(user_id,))
        db.execute("delete from password_reset_tokens where user_id=?",(user_id,))
    log_admin_action(admin["id"],"reset_password",user_id,"管理員已設定新密碼並登出全部裝置")
    create_notification("密碼已由管理員重設","你的登入密碼已由管理員更新，請使用新密碼重新登入。","account","/_projects/account/index.html",user_id,admin["id"])
    return {"ok":True,"logged_out":True}


@app.get("/api/admin/users/{user_id}/sessions")
def admin_user_sessions(user_id: str, request: Request):
    admin=require_site_owner(request)
    with connect_db() as db:
        target=db.execute("select id,email from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到帳號。")
        rows=db.execute("select token_hash,created_at,expires_at,ip_address from sessions where user_id=? order by created_at desc",(user_id,)).fetchall()
    current_hash=digest_token(request.cookies.get(COOKIE_NAME,"")) if request.cookies.get(COOKIE_NAME) else ""
    return {"ok":True,"email":target["email"],"sessions":[{
        "session_id":row["token_hash"],"created_at":row["created_at"],"expires_at":row["expires_at"],
        "ip_address":row["ip_address"] or "","is_current":row["token_hash"]==current_hash,
    } for row in rows]}


@app.delete("/api/admin/users/{user_id}/sessions/{session_id}")
def admin_delete_user_session(user_id: str, session_id: str, request: Request):
    admin=require_site_owner(request)
    if not re.fullmatch(r"[0-9a-f]{64}",session_id):
        raise HTTPException(status_code=400,detail="登入階段識別碼無效。")
    current_hash=digest_token(request.cookies.get(COOKIE_NAME,"")) if request.cookies.get(COOKIE_NAME) else ""
    if user_id==admin["id"] and session_id==current_hash:
        raise HTTPException(status_code=400,detail="不能從後台刪除目前正在使用的登入階段。")
    with connect_db() as db:
        deleted=db.execute("delete from sessions where user_id=? and token_hash=?",(user_id,session_id)).rowcount
    if not deleted:
        raise HTTPException(status_code=404,detail="找不到登入階段。")
    log_admin_action(admin["id"],"delete_session",user_id,f"session={session_id[:12]}")
    return {"ok":True}


@app.patch("/api/admin/users/{user_id}/role")
def admin_update_role(user_id: str, body: AdminRoleUpdate, request: Request):
    admin=require_site_owner(request)
    role=body.role.strip().lower()
    if role not in {"admin","user"}:
        raise HTTPException(status_code=400,detail="角色只能是 admin 或 user。")
    if user_id==admin["id"] and role!="admin":
        raise HTTPException(status_code=400,detail="不能取消自己的管理員權限。")
    with connect_db() as db:
        db.execute("begin immediate")
        target=db.execute("select id,email,role from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        if is_site_owner_email(target["email"]) and role!="admin":
            raise HTTPException(status_code=400,detail="此帳號是站長，無法移除管理員權限。")
        if target["role"]=="admin" and role!="admin":
            count=db.execute("select count(*) as c from users where role='admin' and is_active=1").fetchone()["c"]
            if int(count)<=1:
                raise HTTPException(status_code=400,detail="系統至少必須保留一位啟用中的管理員。")
        db.execute("update users set role=?,updated_at=? where id=?",(role,now(),user_id))
        db.execute("delete from sessions where user_id=?",(user_id,))
    log_admin_action(admin["id"],"update_role",user_id,f"role={role}")
    return {"ok":True}


@app.patch("/api/admin/users/{user_id}/status")
def admin_update_status(user_id: str, body: AdminStatusUpdate, request: Request):
    admin=require_site_owner(request)
    if user_id==admin["id"] and not body.active:
        raise HTTPException(status_code=400,detail="不能停用目前登入中的管理員帳號。")
    with connect_db() as db:
        db.execute("begin immediate")
        target=db.execute("select id,email,role,is_active from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        if is_site_owner_email(target["email"]) and not body.active:
            raise HTTPException(status_code=400,detail="站長帳號無法停用。")
        if target["role"]=="admin" and not body.active and target["is_active"]:
            count=db.execute("select count(*) as c from users where role='admin' and is_active=1").fetchone()["c"]
            if int(count)<=1:
                raise HTTPException(status_code=400,detail="系統至少必須保留一位啟用中的管理員。")
        db.execute("update users set is_active=?,updated_at=? where id=?",(1 if body.active else 0,now(),user_id))
        if not body.active:
            db.execute("delete from sessions where user_id=?",(user_id,))
    log_admin_action(admin["id"],"update_status",user_id,f"active={body.active}")
    return {"ok":True}


@app.patch("/api/admin/users/{user_id}/verification")
def admin_update_verification(user_id: str, body: AdminVerificationUpdate, request: Request):
    admin=require_site_owner(request)
    if user_id==admin["id"] and not body.verified:
        raise HTTPException(status_code=400,detail="不能取消自己目前帳號的信箱驗證。")
    with connect_db() as db:
        target=db.execute("select id,email from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        if is_site_owner_email(target["email"]) and not body.verified:
            raise HTTPException(status_code=400,detail="站長的信箱驗證無法取消。")
        db.execute("update users set email_verified=?,updated_at=? where id=?",(1 if body.verified else 0,now(),user_id))
        if not body.verified:
            db.execute("delete from sessions where user_id=?",(user_id,))
    log_admin_action(admin["id"],"update_verification",user_id,f"verified={body.verified}")
    return {"ok":True}


@app.post("/api/admin/users/{user_id}/logout")
def admin_force_logout(user_id: str, request: Request):
    admin=require_site_owner(request)
    if user_id==admin["id"]:
        raise HTTPException(status_code=400,detail="請使用右上角的登出按鈕登出自己。")
    with connect_db() as db:
        target=db.execute("select id from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        db.execute("delete from sessions where user_id=?",(user_id,))
    log_admin_action(admin["id"],"force_logout",user_id)
    return {"ok":True}


@app.post("/api/admin/users/{user_id}/block")
def admin_block_user(user_id: str, body: AdminUserBlockPayload, request: Request):
    admin=require_site_owner(request)
    if user_id==admin["id"]:
        raise HTTPException(status_code=400,detail="不能封鎖目前登入中的管理員帳號。")
    block_id=str(uuid.uuid4())
    reason=body.reason.strip() or "由帳號管理封鎖"
    try:
        with connect_db() as db:
            db.execute("begin immediate")
            target=db.execute("select id,email,email_key from users where id=?",(user_id,)).fetchone()
            if not target:
                raise HTTPException(status_code=404,detail="找不到此帳號。")
            if is_site_owner_email(target["email"]):
                raise HTTPException(status_code=400,detail="站長帳號無法封鎖。")
            session_count=int(db.execute("select count(*) as c from sessions where user_id=?",(user_id,)).fetchone()["c"] or 0)
            db.execute(
                "insert into blocked_entries(id,kind,value_key,reason,active,created_by,created_at) values(?,?,?,?,?,?,?)",
                (block_id,"email",target["email_key"],reason,1,admin["id"],now())
            )
            db.execute("delete from sessions where user_id=?",(user_id,))
            target_email=target["email"]
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409,detail="此帳號已在封鎖名單中。")
    log_admin_action(admin["id"],"create_block",user_id,f"email:{normalize_email(target_email)}; sessions={session_count}; source=account_management")
    return {"ok":True,"id":block_id,"email":target_email,"logged_out_users":1,"logged_out_sessions":session_count}


@app.delete("/api/admin/users/{user_id}/block")
def admin_unblock_user(user_id: str, request: Request):
    admin=require_site_owner(request)
    with connect_db() as db:
        db.execute("begin immediate")
        target=db.execute("select id,email,email_key from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        deleted=db.execute(
            "delete from blocked_entries where kind='email' and value_key=?",
            (target["email_key"],)
        ).rowcount
        target_email=target["email"]
    if not deleted:
        raise HTTPException(status_code=404,detail="此帳號目前不在電子信箱封鎖名單中。")
    log_admin_action(admin["id"],"delete_block",user_id,f"email:{normalize_email(target_email)}; source=account_management")
    return {"ok":True,"email":target_email,"removed_blocks":int(deleted)}


@app.get("/api/admin/users/{user_id}/progress-summary")
def admin_user_progress_summary(user_id: str, request: Request):
    require_site_owner(request)
    with connect_db() as db:
        target=db.execute("select id,email,username from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        summary=_user_progress_summary(db,user_id)
    return {"ok":True,"user":{"id":target["id"],"email":target["email"],"username":target["username"] or ""},**summary}


@app.delete("/api/admin/users/{user_id}/progress")
@high_risk_operation
def admin_reset_user_progress(user_id: str, request: Request, game_id: str = "all"):
    """Clear one game or every game after creating an automatic SQLite backup."""
    admin=require_site_owner(request)
    scope=str(game_id or "all").strip().lower()
    projects=enabled_game_projects()
    valid_ids={row["id"] for row in projects}
    if scope!="all" and scope not in valid_ids:
        raise HTTPException(status_code=400,detail="不支援的遊戲範圍。")
    with connect_db() as db:
        target=db.execute("select id,email from users where id=?",(user_id,)).fetchone()
        if not target:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        before=_user_progress_summary(db,user_id)
    selected_ids=list(valid_ids) if scope=="all" else [scope]
    selected_counts={row["id"]:int(row["completed_count"]) for row in before["games"] if row["id"] in selected_ids}
    if not any(selected_counts.values()):
        return {"ok":True,"removed":0,"removed_by_game":selected_counts,"backup":"","scope":scope}
    backup=create_database_backup()
    removed_by_game: dict[str,int]={}
    legacy_removed=0
    with connect_db() as db:
        db.execute("begin immediate")
        for selected in selected_ids:
            removed_by_game[selected]=int(db.execute(
                "delete from game_progress where game_id=? and user_id=?",(selected,user_id)
            ).rowcount or 0)
        if "wuwa" in selected_ids:
            legacy_removed=int(db.execute("delete from progress where user_id=?",(user_id,)).rowcount or 0)
    for selected in selected_ids:
        bump_game_live_scope(selected,"stats")
    removed_total=sum(removed_by_game.values())
    log_admin_action(
        admin["id"],"reset_progress",user_id,
        f"scope={scope}; removed={removed_total}; legacy_wuwa={legacy_removed}; backup={backup.name}; by_game={json.dumps(removed_by_game,ensure_ascii=False,sort_keys=True)}",
    )
    return {"ok":True,"removed":removed_total,"removed_by_game":removed_by_game,"backup":backup.name,"scope":scope}


@app.post("/api/admin/users/{user_id}/resend-verification")
def admin_resend_user_verification(user_id: str, request: Request):
    admin=require_site_owner(request)
    with connect_db() as db:
        target=db.execute("select id,email,email_verified,is_active from users where id=?",(user_id,)).fetchone()
    if not target:
        raise HTTPException(status_code=404,detail="找不到此帳號。")
    if target["email_verified"]:
        raise HTTPException(status_code=400,detail="此帳號已完成信箱驗證。")
    if not target["is_active"]:
        raise HTTPException(status_code=400,detail="請先啟用此帳號。")
    issue_verification(target["id"],target["email"])
    log_admin_action(admin["id"],"resend_verification",user_id)
    return {"ok":True}


@app.delete("/api/admin/users/{user_id}")
def admin_delete_user(user_id: str, request: Request):
    admin=require_site_owner(request)
    if user_id==admin["id"]:
        raise HTTPException(status_code=400,detail="不能刪除目前登入中的管理員帳號。")
    with connect_db() as db:
        row=db.execute("select id,email,role,is_active from users where id=?",(user_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        if is_site_owner_email(row["email"]):
            raise HTTPException(status_code=400,detail="站長帳號無法刪除。")
        if row["role"]=="admin" and row["is_active"]:
            count=db.execute("select count(*) as c from users where role='admin' and is_active=1").fetchone()["c"]
            if int(count)<=1:
                raise HTTPException(status_code=400,detail="不能刪除最後一位啟用中的管理員。")
        target_email=row["email"]
        db.execute("delete from users where id=?",(user_id,))
    log_admin_action(admin["id"],"delete_user",None,f"deleted={target_email}")
    return {"ok":True}


@app.post("/api/admin/system/clear-rate-limits")
def admin_clear_rate_limits(request: Request):
    admin=require_admin(request)
    with connect_db() as db:
        deleted=db.execute("delete from rate_limits").rowcount
    log_admin_action(admin["id"],"clear_rate_limits",details=f"deleted={deleted}")
    return {"ok":True,"deleted":deleted}


@app.post("/api/admin/system/backup")
@high_risk_operation
def admin_backup_database(request: Request):
    admin=require_admin(request)
    target=create_database_backup()
    log_admin_action(admin["id"],"backup_database",details=target.name)
    return {"ok":True,"filename":target.name,"path":str(target)}


def _safe_backup_path(filename: str) -> Path:
    name=Path(filename).name
    if name!=filename or not re.fullmatch(r"app-[0-9]{8}-[0-9]{6}(?:-[0-9]{2})?\.db",name):
        raise HTTPException(status_code=400,detail="備份檔名無效。")
    path=(ROOT/"backups"/name).resolve()
    backup_root=(ROOT/"backups").resolve()
    if path.parent!=backup_root:
        raise HTTPException(status_code=400,detail="備份路徑無效。")
    return path


def _sqlite_integrity(path: Path) -> str:
    db: sqlite3.Connection | None = None
    try:
        db=sqlite3.connect(path)
        row=db.execute("pragma integrity_check").fetchone()
        return str(row[0] if row else "unknown")
    except Exception as exc:
        return f"error: {type(exc).__name__}: {exc}"
    finally:
        if db is not None:
            try:
                db.close()
            except Exception:
                pass


@app.get("/api/admin/system/backups")
def admin_list_backups(request: Request):
    require_admin(request)
    backup_dir=ROOT/"backups"
    backup_dir.mkdir(parents=True,exist_ok=True)
    values=[]
    for path in sorted(backup_dir.glob("app-*.db"),key=lambda item:item.stat().st_mtime,reverse=True):
        stat=path.stat()
        values.append({"filename":path.name,"size_bytes":stat.st_size,"modified_at":int(stat.st_mtime),"integrity":_sqlite_integrity(path)})
    return {"ok":True,"backups":values,"database":{"filename":DB_FILE.name,"size_bytes":DB_FILE.stat().st_size if DB_FILE.exists() else 0,"integrity":_sqlite_integrity(DB_FILE)}}


@app.post("/api/admin/system/backups/restore")
def admin_restore_backup(body: BackupRestorePayload, request: Request):
    admin=require_admin(request)
    if not BACKUP_OPERATION_GUARD.acquire(blocking=False):
        raise HTTPException(status_code=409,detail="另一項備份建立、刪除或還原作業正在執行，請稍後再試。")
    try:
        source=_safe_backup_path(body.filename)
        if not source.exists():
            raise HTTPException(status_code=404,detail="找不到備份檔。")
        if _sqlite_integrity(source)!="ok":
            raise HTTPException(status_code=400,detail="備份資料庫完整性檢查未通過。")
        safety=create_database_backup()
        try:
            _copy_sqlite_database(source,DB_FILE)
            if _sqlite_integrity(DB_FILE)!="ok":
                raise RuntimeError("還原後資料庫完整性檢查未通過")
        except Exception as exc:
            try:
                _copy_sqlite_database(safety,DB_FILE)
            except Exception as rollback_exc:
                raise HTTPException(status_code=500,detail=f"備份還原失敗，且安全備份回復也失敗：{exc}；{rollback_exc}") from rollback_exc
            raise HTTPException(status_code=500,detail=f"備份還原失敗，已回到還原前狀態：{exc}") from exc
        log_admin_action(admin["id"],"restore_database_backup",details=f"source={source.name}; safety={safety.name}")
        return {"ok":True,"restored":source.name,"safety_backup":safety.name,"requires_relogin":True}
    finally:
        BACKUP_OPERATION_GUARD.release()


@app.post("/api/admin/system/backups/delete-all")
def admin_delete_all_backups(request: Request):
    admin=require_admin(request)
    if not BACKUP_OPERATION_GUARD.acquire(blocking=False):
        raise HTTPException(status_code=409,detail="另一項備份建立、刪除或還原作業正在執行，請稍後再試。")
    try:
        backup_dir=ROOT/"backups"
        backup_dir.mkdir(parents=True,exist_ok=True)
        paths=sorted(backup_dir.glob("app-*.db"),key=lambda item:item.stat().st_mtime,reverse=True)
        before={"count":len(paths),"size_bytes":sum(path.stat().st_size for path in paths),"files":[path.name for path in paths]}
        deleted=[]
        failed=[]
        for path in paths:
            try:
                path.unlink()
                deleted.append(path.name)
            except Exception as exc:
                failed.append({"filename":path.name,"error":f"{type(exc).__name__}: {exc}"})
        status="success" if not failed else "failed"
        log_admin_action(
            admin["id"],"delete_all_database_backups",
            details=f"deleted={len(deleted)}; failed={len(failed)}",
            category="administration",status=status,target_type="backup_collection",target_id="all",
            summary="一鍵刪除全部資料庫備份",before=before,
            after={"deleted_count":len(deleted),"remaining_count":len(failed)},
            metadata={"deleted":deleted,"failed":failed},actor_ip=client_ip(request),locked=True,
        )
        if failed:
            raise HTTPException(status_code=500,detail=f"已刪除 {len(deleted)} 份，但有 {len(failed)} 份刪除失敗。")
        return {"ok":True,"deleted_count":len(deleted),"deleted":deleted}
    finally:
        BACKUP_OPERATION_GUARD.release()


@app.delete("/api/admin/system/backups/{filename}")
def admin_delete_backup(filename: str, request: Request):
    admin=require_admin(request)
    if not BACKUP_OPERATION_GUARD.acquire(blocking=False):
        raise HTTPException(status_code=409,detail="另一項備份建立、刪除或還原作業正在執行，請稍後再試。")
    try:
        path=_safe_backup_path(filename)
        if not path.exists():
            raise HTTPException(status_code=404,detail="找不到備份檔。")
        path.unlink()
        log_admin_action(admin["id"],"delete_database_backup",details=path.name)
        return {"ok":True}
    finally:
        BACKUP_OPERATION_GUARD.release()


def _relation_health_for_game(db: sqlite3.Connection, game_id: str) -> dict[str,Any]:
    catalog_ids={str(row["achievement_id"]) for row in db.execute("select achievement_id from game_catalog_items where game_id=?",(game_id,)).fetchall()}
    rows=db.execute("select group_id,achievement_id,relation_type,stage_order from game_achievement_choice_groups where game_id=? order by relation_type,group_id,stage_order",(game_id,)).fetchall()
    missing=[str(row["achievement_id"]) for row in rows if str(row["achievement_id"]) not in catalog_ids]
    groups={}
    for row in rows:
        key=f"{row['relation_type']}:{row['group_id']}"
        groups.setdefault(key,[]).append(str(row["achievement_id"]))
    invalid=[key for key,members in groups.items() if len(members)<2]
    return {"members":len(rows),"groups":len(groups),"missing_catalog_ids":missing[:100],"invalid_groups":invalid[:100],"ok":not missing and not invalid}


@app.get("/api/admin/system/health")
def admin_system_health(request: Request):
    require_admin(request)
    disk=shutil.disk_usage(ROOT)
    directory_checks={}
    for name,path in {"data":DATA_DIR,"logs":LOG_DIR,"backups":ROOT/"backups","outbox":OUTBOX_DIR}.items():
        try:
            path.mkdir(parents=True,exist_ok=True)
            probe=path/f".write-test-{secrets.token_hex(4)}"
            probe.write_text("ok",encoding="utf-8"); probe.unlink()
            directory_checks[name]={"path":str(path),"writable":True}
        except Exception as exc:
            directory_checks[name]={"path":str(path),"writable":False,"error":str(exc)}
    with connect_db() as db:
        integrity=str(db.execute("pragma integrity_check").fetchone()[0])
        tables={str(row["name"]) for row in db.execute("select name from sqlite_master where type='table'").fetchall()}
        required={"users","sessions","game_catalog_items","game_progress","game_achievement_choice_groups","notifications","email_logs","game_sync_previews","redeem_games","redeem_servers","redeem_codes"}
        game_rows=[]
        for project in enabled_game_projects():
            gid=project["id"]
            count=int(db.execute("select count(*) c from game_catalog_items where game_id=?",(gid,)).fetchone()["c"] or 0)
            progress=int(db.execute("select count(*) c from game_progress where game_id=?",(gid,)).fetchone()["c"] or 0)
            game_rows.append({"id":gid,"name":project["name"],"catalog_count":count,"progress_count":progress,"relations":_relation_health_for_game(db,gid),"icon_exists":bool(GAME_ICON_FILES.get(gid) and GAME_ICON_FILES[gid].exists())})
        pending={
            "achievement_reports":int(db.execute("select count(*) c from game_achievement_reports where status in ('open','reviewing')").fetchone()["c"] or 0),
            "support_tickets":int(db.execute("select count(*) c from support_tickets where status not in ('resolved','closed')").fetchone()["c"] or 0),
            "failed_emails":int(db.execute("select count(*) c from email_logs where status='failed'").fetchone()["c"] or 0),
            "sync_previews":int(db.execute("select count(*) c from game_sync_previews where expires_at> ?",(now(),)).fetchone()["c"] or 0),
        }
    backup_files=list((ROOT/"backups").glob("app-*.db"))
    latest=max(backup_files,key=lambda item:item.stat().st_mtime) if backup_files else None
    checks={
        "database_ok":integrity=="ok",
        "required_tables_ok":required.issubset(tables),
        "directories_writable":all(value.get("writable") for value in directory_checks.values()),
        "game_catalogs_ok":all(row["catalog_count"]>0 for row in game_rows),
        "relations_ok":all(row["relations"]["ok"] for row in game_rows),
        "icons_ok":all(row["icon_exists"] for row in game_rows),
    }
    return {"ok":True,"healthy":all(checks.values()),"checks":checks,"database":{"path":str(DB_FILE),"size_bytes":DB_FILE.stat().st_size if DB_FILE.exists() else 0,"integrity":integrity,"missing_tables":sorted(required-tables)},"directories":directory_checks,"disk":{"total":disk.total,"used":disk.used,"free":disk.free},"games":game_rows,"mail":{"delivery":MAIL_DELIVERY,"host":SMTP_HOST,"port":SMTP_PORT,"from":SMTP_FROM,"configured":MAIL_DELIVERY=="console" or bool(SMTP_HOST and SMTP_FROM)},"backups":{"count":len(backup_files),"latest":latest.name if latest else "","latest_at":int(latest.stat().st_mtime) if latest else None},"pending":pending,"time":now()}


# ----- 成就回報 -----
@app.post("/api/achievement-reports")
@high_risk_operation
def create_achievement_report(body: AchievementReportCreate, request: Request):
    return extra_game_create_report("wuwa",body,request)


@app.get("/api/my/achievement-reports")
def my_achievement_reports(request: Request):
    return extra_game_my_reports("wuwa",request)


@app.get("/api/admin/achievement-reports")
def admin_achievement_reports(request: Request):
    return extra_game_admin_reports("wuwa",request)


@app.patch("/api/admin/achievement-reports/{report_id}")
@high_risk_operation
def admin_update_achievement_report(report_id: str, body: AchievementReportUpdate, request: Request):
    return extra_game_admin_update_report("wuwa",report_id,body,request)


@app.delete("/api/admin/achievement-reports/{report_id}")
@high_risk_operation
def admin_delete_achievement_report(report_id: str, request: Request):
    return extra_game_admin_delete_report("wuwa",report_id,request)


# ----- 兌換碼中心 -----
REDEEM_GAME_ID_RE = re.compile(r"^[a-z0-9][a-z0-9_-]{1,39}$")


def normalize_redeem_game_id(value: str) -> str:
    game_id = str(value or "").strip().casefold()
    if not REDEEM_GAME_ID_RE.fullmatch(game_id):
        raise HTTPException(status_code=422, detail="兌換碼遊戲 ID 只能使用小寫英文、數字、底線或連字號，長度 2-40。")
    return game_id


def normalize_redeem_text(value: str, *, field: str, limit: int, required: bool = False) -> str:
    text = str(value or "").strip()
    if required and not text:
        raise HTTPException(status_code=422, detail=f"{field}不可空白。")
    if len(text) > limit:
        raise HTTPException(status_code=422, detail=f"{field}長度不可超過 {limit} 字。")
    return text


def normalize_redeem_url(value: str) -> str:
    text = normalize_redeem_text(value, field="兌換連結", limit=500)
    if not text:
        return ""
    parsed = urllib.parse.urlparse(text)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise HTTPException(status_code=422, detail="兌換連結必須是 http 或 https 網址。")
    return text


def normalize_redeem_time_range(start_at: int | None, end_at: int | None) -> tuple[int | None, int | None]:
    start = int(start_at) if start_at is not None else None
    end = int(end_at) if end_at is not None else None
    if start is not None and start < 0:
        raise HTTPException(status_code=422, detail="開始時間無效。")
    if end is not None and end < 0:
        raise HTTPException(status_code=422, detail="結束時間無效。")
    if start is not None and end is not None and end <= start:
        raise HTTPException(status_code=422, detail="結束時間必須晚於開始時間。")
    return start, end


def redeem_json_list(value: Any) -> list[str]:
    if isinstance(value, list):
        raw = value
    else:
        try:
            raw = json.loads(str(value or "[]"))
        except Exception:
            raw = []
    result: list[str] = []
    seen: set[str] = set()
    for item in raw if isinstance(raw, list) else []:
        text = str(item or "").strip()
        if text and text not in seen:
            seen.add(text)
            result.append(text)
    return result


def require_redeem_game(db: sqlite3.Connection, game_id: str) -> sqlite3.Row:
    row = db.execute("select * from redeem_games where game_id=?", (game_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="找不到兌換碼遊戲；請先新增遊戲。")
    return row


def validate_redeem_servers(db: sqlite3.Connection, game_id: str, server_ids: list[str]) -> list[str]:
    ids: list[str] = []
    seen: set[str] = set()
    for raw in server_ids:
        server_id = str(raw or "").strip()
        if server_id and server_id not in seen:
            seen.add(server_id)
            ids.append(server_id)
    if not ids:
        raise HTTPException(status_code=422, detail="請先選擇至少一個服務器；不得手動輸入服務器文字。")
    if len(ids) > 50:
        raise HTTPException(status_code=422, detail="單筆兌換碼最多選擇 50 個服務器。")
    rows = db.execute(
        f"select id,enabled from redeem_servers where game_id=? and id in ({','.join('?' for _ in ids)})",
        (game_id, *ids),
    ).fetchall()
    by_id = {str(row["id"]): row for row in rows}
    missing = [server_id for server_id in ids if server_id not in by_id]
    if missing:
        raise HTTPException(status_code=422, detail="兌換碼服務器選項不存在，請先在該遊戲底下建立服務器。")
    disabled = [server_id for server_id, row in by_id.items() if not bool(row["enabled"])]
    if disabled:
        raise HTTPException(status_code=422, detail="兌換碼不可選用已停用的服務器。")
    return ids


def redeem_status(row: sqlite3.Row | dict[str, Any], stamp: int | None = None) -> str:
    current = stamp if stamp is not None else now()
    end_at = int(row["end_at"] or 0)
    return "expired" if end_at and current > end_at else "open"


def redeem_code_sort_key(item: dict[str, Any]) -> tuple[int, int, int, str]:
    status_rank = 1 if item.get("status") == "expired" else 0
    start_at = int(item.get("start_at") or 0)
    unknown_start = 1 if not start_at else 0
    return (status_rank, unknown_start, -start_at, str(item.get("code") or ""))


def redeem_game_payload(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "game_id": row["game_id"],
        "name": row["name"],
        "display_order": int(row["display_order"] or 0),
        "enabled": bool(row["enabled"]),
        "note": row["note"] or "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def redeem_server_payload(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "game_id": row["game_id"],
        "name": row["name"],
        "display_order": int(row["display_order"] or 0),
        "enabled": bool(row["enabled"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def redeem_code_payload(row: sqlite3.Row, games: dict[str, dict[str, Any]], servers: dict[str, dict[str, Any]], *, stamp: int | None = None) -> dict[str, Any]:
    server_ids = redeem_json_list(row["server_ids_json"])
    server_rows = [servers[server_id] for server_id in server_ids if server_id in servers]
    item = {
        "id": row["id"],
        "game_id": row["game_id"],
        "game_name": games.get(str(row["game_id"]), {}).get("name", row["game_id"]),
        "code": row["code"],
        "source": row["source"] or row["description"] or "",
        "description": row["description"] or "",
        "reward": row["reward"] or "",
        "start_at": row["start_at"],
        "end_at": row["end_at"],
        "server_ids": server_ids,
        "servers": server_rows,
        "server_names": "、".join(server["name"] for server in server_rows),
        "redeem_url": row["redeem_url"] or "",
        "enabled": bool(row["enabled"]),
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }
    item["status"] = redeem_status(row, stamp)
    return item


def redeem_dataset(db: sqlite3.Connection, *, public_only: bool = False) -> dict[str, Any]:
    game_where = "where enabled=1" if public_only else ""
    games_rows = db.execute(f"select * from redeem_games {game_where} order by display_order,name,game_id").fetchall()
    games = [redeem_game_payload(row) for row in games_rows]
    games_by_id = {row["game_id"]: row for row in games}
    if not games_by_id:
        return {"games": [], "servers": [], "codes": []}
    placeholders = ",".join("?" for _ in games_by_id)
    servers_rows = db.execute(
        f"select * from redeem_servers where game_id in ({placeholders}) order by game_id,display_order,name,id",
        tuple(games_by_id.keys()),
    ).fetchall()
    servers = [redeem_server_payload(row) for row in servers_rows]
    servers_by_id = {row["id"]: row for row in servers}
    code_where = f"where game_id in ({placeholders})"
    if public_only:
        code_where += " and enabled=1"
    code_rows = db.execute(
        f"select * from redeem_codes {code_where} order by updated_at desc",
        tuple(games_by_id.keys()),
    ).fetchall()
    stamp = now()
    codes = [redeem_code_payload(row, games_by_id, servers_by_id, stamp=stamp) for row in code_rows]
    codes.sort(key=redeem_code_sort_key)
    return {"games": games, "servers": servers, "codes": codes}


REDEEM_CODE_DB_COLUMNS = (
    "id", "game_id", "code", "source", "description", "reward", "start_at", "end_at",
    "server_ids_json", "redeem_url", "enabled", "created_by", "updated_by", "created_at", "updated_at",
)


def redeem_code_snapshot(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    return {column: row[column] for column in REDEEM_CODE_DB_COLUMNS}


def redeem_import_key(game_id: str, code: str) -> str:
    return f"{game_id}|{code.strip().casefold()}"


REDEEM_EXCEL_MAX_BYTES = 10 * 1024 * 1024
REDEEM_EXCEL_HEADER_ALIASES = {
    "遊戲": "game_id", "game": "game_id", "gameid": "game_id",
    "兌換碼": "code", "代碼": "code", "code": "code", "redeemcode": "code",
    "來源": "source", "說明": "source", "source": "source", "description": "source",
    "獎勵": "reward", "reward": "reward",
    "開始日期": "start_at", "開始時間": "start_at", "startdate": "start_at", "startat": "start_at",
    "結束日期": "end_at", "結束時間": "end_at", "enddate": "end_at", "endat": "end_at",
    "服務器": "server_names", "伺服器": "server_names", "server": "server_names",
    "servers": "server_names", "serverid": "server_names", "serverids": "server_names",
    "servername": "server_names", "servernames": "server_names",
    "兌換連結": "redeem_url", "兌換網址": "redeem_url", "連結": "redeem_url",
    "redeemurl": "redeem_url", "url": "redeem_url",
}


def normalize_redeem_excel_header(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).strip().casefold()
    return re.sub(r"[\s_-]+", "", text)


def redeem_excel_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_redeem_excel_workbook(filename: str, content: bytes) -> dict[str, Any]:
    suffix = Path(filename).suffix.casefold()
    if suffix == ".xls":
        raise HTTPException(status_code=422, detail="不支援舊版 .xls，請在 Excel 另存為 .xlsx 後再匯入。")
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=422, detail="只支援 .xlsx 或 .xlsm Excel 檔案。")
    if not content:
        raise HTTPException(status_code=422, detail="Excel 檔案是空白的。")
    if len(content) > REDEEM_EXCEL_MAX_BYTES:
        raise HTTPException(status_code=413, detail="Excel 檔案不可超過 10 MB。")
    if load_workbook is None:
        raise HTTPException(status_code=503, detail="後端尚未安裝 Excel 解析套件，請重新安裝 requirements.txt。")
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True, keep_links=False)
    except (InvalidFileException, OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        raise HTTPException(status_code=422, detail="Excel 檔案損壞、受密碼保護或格式無法讀取。") from exc

    empty_sheet: tuple[str, int, list[str], list[str]] | None = None
    try:
        worksheets = sorted(workbook.worksheets, key=lambda sheet: sheet.sheet_state != "visible")
        for worksheet in worksheets:
            header_fields: list[str] | None = None
            header_labels: list[str] = []
            unknown_headers: list[str] = []
            header_row = 0
            rows: list[dict[str, str]] = []
            skipped_without_code = 0
            for row_number, values in enumerate(worksheet.iter_rows(values_only=True), 1):
                cells = [redeem_excel_cell_text(value) for value in values]
                if header_fields is None:
                    if row_number > 50:
                        break
                    mapped = [REDEEM_EXCEL_HEADER_ALIASES.get(normalize_redeem_excel_header(value), "") for value in cells]
                    if "code" not in mapped:
                        continue
                    duplicates = sorted({field for field in mapped if field and mapped.count(field) > 1})
                    if duplicates:
                        raise HTTPException(status_code=422, detail=f"Excel 標題欄重複：{'、'.join(duplicates)}。")
                    header_fields = mapped
                    header_labels = cells
                    header_row = row_number
                    unknown_headers = [label for label, field in zip(cells, mapped) if label and not field]
                    continue
                if not any(cells):
                    continue
                row = {field: cells[index] if index < len(cells) else "" for index, field in enumerate(header_fields) if field}
                if not row.get("code", "").strip():
                    skipped_without_code += 1
                    continue
                rows.append(row)
                if len(rows) > 500:
                    raise HTTPException(status_code=422, detail="Excel 內的兌換碼超過每批 500 筆上限。")
            if header_fields is None:
                continue
            if rows:
                return {
                    "filename": Path(filename).name,
                    "sheet_name": worksheet.title,
                    "header_row": header_row,
                    "columns": [field for field in header_fields if field],
                    "unknown_headers": unknown_headers,
                    "skipped_without_code": skipped_without_code,
                    "rows": rows,
                }
            empty_sheet = (worksheet.title, header_row, header_labels, unknown_headers)
    finally:
        workbook.close()
    if empty_sheet:
        raise HTTPException(status_code=422, detail=f"工作表「{empty_sheet[0]}」有標題列，但沒有可匯入的兌換碼。")
    raise HTTPException(status_code=422, detail="找不到含「兌換碼」或「code」欄位的工作表；標題列須位於前 50 列。")


def build_redeem_excel_template(game_id: str, game_name: str, server_names: list[str]) -> bytes:
    if Workbook is None:
        raise HTTPException(status_code=503, detail="後端尚未安裝 Excel 解析套件，請重新安裝 requirements.txt。")
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "兌換碼匯入"
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = f"{game_name}｜兌換碼批次匯入"
    worksheet["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    worksheet["A1"].fill = PatternFill("solid", fgColor="174A5B")
    worksheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    worksheet["A2"] = "填寫兌換碼後，可直接由後台的「匯入 Excel」載入；每列服務器可不同，多個服務器請用「、」分隔。"
    worksheet.merge_cells("A2:H2")
    worksheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    headers = ["遊戲", "兌換碼", "來源", "獎勵", "開始日期", "結束日期", "服務器", "兌換連結"]
    worksheet.append([])
    worksheet.append(headers)
    header_notes = {
        "遊戲": f"固定使用 {game_id}；同一檔案只能有一個遊戲。",
        "兌換碼": "必填。系統會在同一遊戲內檢查重複。",
        "來源": "例如：官方直播、官方社群、活動頁。",
        "獎勵": "可在儲存格內換行。",
        "開始日期": "格式：YYYY-MM-DD，可留空。",
        "結束日期": "格式：YYYY-MM-DD，可留空。",
        "服務器": "必填。每列可不同；多個服務器以「、」分隔。",
        "兌換連結": "必須是 http 或 https 網址，可留空。",
    }
    for column, title in enumerate(headers, 1):
        cell = worksheet.cell(row=4, column=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="276477")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.comment = Comment(header_notes[title], "Milora_tool")
    for row in range(5, 25):
        worksheet.cell(row=row, column=1, value=game_id)
        worksheet.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        for column in range(1, 9):
            if row % 2 == 0:
                worksheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor="EAF3F6")
    worksheet.freeze_panes = "A5"
    worksheet.auto_filter.ref = "A4:H24"
    worksheet.row_dimensions[1].height = 26
    worksheet.row_dimensions[2].height = 34
    widths = {"A": 16, "B": 24, "C": 22, "D": 34, "E": 15, "F": 15, "G": 28, "H": 36}
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    servers = workbook.create_sheet("可用服務器")
    servers["A1"] = f"{game_name} 可用服務器"
    servers["A1"].font = Font(bold=True, color="FFFFFF")
    servers["A1"].fill = PatternFill("solid", fgColor="174A5B")
    servers["A2"] = "每一列兌換碼可選不同服務器；需要多選時，以「、」連接名稱。"
    servers["A2"].alignment = Alignment(wrap_text=True)
    for index, name in enumerate(server_names, 4):
        servers.cell(row=index, column=1, value=name)
    servers.column_dimensions["A"].width = 42
    servers.freeze_panes = "A4"

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


def find_redeem_code_duplicate(
    db: sqlite3.Connection,
    game_id: str,
    code: str,
    *,
    exclude_id: str = "",
) -> sqlite3.Row | None:
    query = "select * from redeem_codes where game_id=? and lower(trim(code))=?"
    params: list[Any] = [game_id, code.strip().casefold()]
    if exclude_id:
        query += " and id<>?"
        params.append(exclude_id)
    return db.execute(query, params).fetchone()


def redeem_import_state(db: sqlite3.Connection, keys: list[str]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key in sorted(set(keys)):
        game_id, _, code_key = key.partition("|")
        row = db.execute(
            "select * from redeem_codes where game_id=? and lower(trim(code))=?",
            (game_id, code_key),
        ).fetchone()
        result[key] = redeem_code_snapshot(row) if row else None
    return result


def redeem_import_state_hash(db: sqlite3.Connection, keys: list[str]) -> str:
    payload = redeem_import_state(db, keys)
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")).hexdigest()


def redeem_import_server_ids(db: sqlite3.Connection, game_id: str, raw: dict[str, Any]) -> list[str]:
    raw_ids = raw.get("server_ids")
    if isinstance(raw_ids, list) and raw_ids:
        return validate_redeem_servers(db, game_id, [str(value) for value in raw_ids])
    raw_names = raw.get("server_names")
    if isinstance(raw_names, str):
        names = [part.strip() for part in re.split(r"[,，、;；|]", raw_names) if part.strip()]
    elif isinstance(raw_names, list):
        names = [str(part or "").strip() for part in raw_names if str(part or "").strip()]
    else:
        names = []
    rows = db.execute(
        "select id,name from redeem_servers where game_id=? and enabled=1 order by display_order,name,id",
        (game_id,),
    ).fetchall()
    if len(names) == 1 and names[0].casefold() in {"all", "全部", "全服務器", "全部服務器"}:
        return validate_redeem_servers(db, game_id, [str(row["id"]) for row in rows])
    by_name: dict[str, list[str]] = {}
    for row in rows:
        by_name.setdefault(str(row["name"]).strip().casefold(), []).append(str(row["id"]))
    ids: list[str] = []
    missing: list[str] = []
    ambiguous: list[str] = []
    for name in names:
        matches = by_name.get(name.casefold(), [])
        if not matches:
            missing.append(name)
        elif len(matches) > 1:
            ambiguous.append(name)
        else:
            ids.append(matches[0])
    if missing:
        raise HTTPException(status_code=422, detail=f"找不到服務器：{'、'.join(missing)}。請先在該遊戲建立並啟用服務器。")
    if ambiguous:
        raise HTTPException(status_code=422, detail=f"服務器名稱不唯一：{'、'.join(ambiguous)}。請改用服務器 ID。")
    return validate_redeem_servers(db, game_id, ids)


def redeem_import_server_names(db: sqlite3.Connection, server_ids: list[str]) -> list[str]:
    if not server_ids:
        return []
    rows = db.execute(
        f"select id,name from redeem_servers where id in ({','.join('?' for _ in server_ids)})",
        server_ids,
    ).fetchall()
    by_id = {str(row["id"]): str(row["name"]) for row in rows}
    return [by_id.get(server_id, server_id) for server_id in server_ids]


def normalize_redeem_import_row(db: sqlite3.Connection, raw: dict[str, Any], default_game_id: str) -> dict[str, Any]:
    game_id = normalize_redeem_game_id(str(raw.get("game_id") or default_game_id or ""))
    require_redeem_game(db, game_id)
    code = normalize_redeem_text(str(raw.get("code") or ""), field="兌換碼", limit=200, required=True)
    source = normalize_redeem_text(str(raw.get("source") or raw.get("description") or ""), field="來源", limit=500)
    reward = normalize_redeem_text(str(raw.get("reward") or ""), field="獎勵", limit=500)
    start_raw = raw.get("start_at")
    end_raw = raw.get("end_at")
    start_at = int(start_raw) if start_raw not in (None, "") else None
    end_at = int(end_raw) if end_raw not in (None, "") else None
    start_at, end_at = normalize_redeem_time_range(start_at, end_at)
    server_ids = redeem_import_server_ids(db, game_id, raw)
    redeem_url = normalize_redeem_url(str(raw.get("redeem_url") or ""))
    return {
        "game_id": game_id,
        "code": code,
        "source": source,
        "description": "",
        "reward": reward,
        "start_at": start_at,
        "end_at": end_at,
        "server_ids": server_ids,
        "server_names": redeem_import_server_names(db, server_ids),
        "redeem_url": redeem_url,
        "enabled": bool(raw.get("enabled", True)),
    }


def redeem_import_candidate_matches(before: dict[str, Any], candidate: dict[str, Any]) -> bool:
    return (
        str(before.get("code") or "") == candidate["code"]
        and str(before.get("source") or before.get("description") or "") == candidate["source"]
        and str(before.get("reward") or "") == candidate["reward"]
        and before.get("start_at") == candidate["start_at"]
        and before.get("end_at") == candidate["end_at"]
        and redeem_json_list(before.get("server_ids_json")) == candidate["server_ids"]
        and str(before.get("redeem_url") or "") == candidate["redeem_url"]
        and bool(before.get("enabled")) == candidate["enabled"]
    )


def redeem_import_batch_payload(row: sqlite3.Row | dict[str, Any]) -> dict[str, Any]:
    data = dict(row)
    summary = _json_object(data.pop("summary_json", "{}"), {})
    data.pop("plan_json", None)
    data.pop("snapshot_json", None)
    data.pop("pre_state_hash", None)
    data.pop("post_state_hash", None)
    actors: dict[str, dict[str, str] | None] = {}
    for actor_key, prefix, id_key in (
        ("creator", "creator", "admin_user_id"),
        ("executor", "executor", "completed_by"),
        ("rollback_actor", "rollback_actor", "rolled_back_by"),
    ):
        username = str(data.pop(f"{prefix}_username", "") or "").strip()
        email = str(data.pop(f"{prefix}_email", "") or "").strip()
        actor_id = str(data.get(id_key) or "").strip()
        actors[actor_key] = {"id": actor_id, "username": username, "email": email, "label": username or email or "已刪除帳號"} if actor_id else None
    status = str(data.get("status") or "")
    can_execute = status == "preview_ready" and int(summary.get("errors") or 0) == 0 and int(summary.get("create") or 0) + int(summary.get("update") or 0) > 0
    return {
        **data,
        "summary": summary,
        **actors,
        "can_execute": can_execute,
        "can_delete": status not in {"completed", "rolled_back"},
        "can_rollback": status == "completed" and not data.get("rolled_back_at"),
    }


REDEEM_IMPORT_BATCH_SELECT = """
    select b.*,g.name default_game_name,
           creator.username creator_username,creator.email creator_email,
           executor.username executor_username,executor.email executor_email,
           rollback_actor.username rollback_actor_username,rollback_actor.email rollback_actor_email
    from redeem_import_batches b
    left join redeem_games g on g.game_id=b.default_game_id
    left join users creator on creator.id=b.admin_user_id
    left join users executor on executor.id=b.completed_by
    left join users rollback_actor on rollback_actor.id=b.rolled_back_by
"""


def redeem_import_batch_where(
    *,
    game_id: str = "",
    status: str = "",
    date_from: int | None = None,
    date_to: int | None = None,
    search: str = "",
) -> tuple[str, list[Any]]:
    conditions: list[str] = []
    params: list[Any] = []
    if game_id:
        conditions.append("b.default_game_id=?")
        params.append(game_id)
    if status:
        conditions.append("b.status=?")
        params.append(status)
    if date_from is not None:
        conditions.append("b.created_at>=?")
        params.append(int(date_from))
    if date_to is not None:
        conditions.append("b.created_at<=?")
        params.append(int(date_to))
    needle = search.strip().casefold()
    if needle:
        conditions.append(
            "(instr(lower(b.id),?)>0 or instr(lower(coalesce(b.reason,'')),?)>0 "
            "or exists(select 1 from redeem_import_items search_item where search_item.batch_id=b.id and instr(lower(search_item.code),?)>0))"
        )
        params.extend((needle, needle, needle))
    return (" where " + " and ".join(conditions) if conditions else ""), params


def redeem_import_batches(
    db: sqlite3.Connection,
    limit: int = 20,
    offset: int = 0,
    *,
    game_id: str = "",
    status: str = "",
    date_from: int | None = None,
    date_to: int | None = None,
    search: str = "",
) -> list[dict[str, Any]]:
    where_sql, params = redeem_import_batch_where(game_id=game_id, status=status, date_from=date_from, date_to=date_to, search=search)
    rows = db.execute(
        REDEEM_IMPORT_BATCH_SELECT + where_sql + " order by b.created_at desc,b.id desc limit ? offset ?",
        (*params, max(1, min(100, int(limit))), max(0, int(offset))),
    ).fetchall()
    return [redeem_import_batch_payload(row) for row in rows]


def redeem_import_batch_detail(db: sqlite3.Connection, batch_id: str) -> dict[str, Any]:
    row = db.execute(REDEEM_IMPORT_BATCH_SELECT + " where b.id=?", (batch_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="找不到兌換碼批次匯入紀錄。")
    payload = redeem_import_batch_payload(row)
    item_rows = db.execute(
        "select * from redeem_import_items where batch_id=? order by row_number,id",
        (batch_id,),
    ).fetchall()
    items: list[dict[str, Any]] = []
    for item_row in item_rows:
        item = dict(item_row)
        candidate = _json_object(item.pop("candidate_json", "{}"), {})
        before = _json_object(item.pop("before_json", "{}"), {})
        if before and "server_ids" not in before:
            before["server_ids"] = redeem_json_list(before.get("server_ids_json"))
        item["candidate"] = candidate
        item["before"] = before
        item["error"] = item.pop("error_text", "")
        items.append(item)
    payload["items"] = items
    return payload


def redeem_notification_event_key(row: dict[str, Any]) -> str:
    return f"redeem:{row.get('id')}:new:{int(row.get('created_at') or 0)}"


def create_message_center_notification(db: sqlite3.Connection, *, user_id: str, title: str, body: str, link: str, kind: str = "redeem", level: str = "info") -> str:
    item_id = str(uuid.uuid4()); stamp = now()
    db.execute(
        """insert into message_center_items(id,item_type,target_user_id,title,body,level,kind,link,is_active,pinned,created_by,created_at,updated_at)
           values(?,'notification',?,?,?,?,?, ?,1,0,null,?,?)""",
        (item_id, user_id, title[:200], body[:3000], level, kind[:30], link[:500], stamp, stamp),
    )
    return item_id


def dispatch_redeem_code_notifications(code_ids: list[str]) -> int:
    ids = list(dict.fromkeys(str(value) for value in code_ids if str(value)))
    if not ids:
        return 0
    with connect_db() as db:
        rows = [dict(row) for row in db.execute(
            f"""select c.*,g.name game_name from redeem_codes c join redeem_games g on g.game_id=c.game_id
                where c.id in ({','.join('?' for _ in ids)}) and c.enabled=1 and g.enabled=1""",
            ids,
        ).fetchall()]
        by_game: dict[str, list[dict[str, Any]]] = {}
        for row in rows:
            by_game.setdefault(str(row["game_id"]), []).append(row)
        sent = 0
        for game_id, game_rows in by_game.items():
            subscribers = db.execute(
                """select p.user_id from redeem_notification_preferences p join users u on u.id=p.user_id
                    where p.game_id=? and u.is_active=1""",
                (game_id,),
            ).fetchall()
            for subscriber in subscribers:
                user_id = str(subscriber["user_id"])
                pending = []
                for row in game_rows:
                    event_key = redeem_notification_event_key(row)
                    exists = db.execute("select 1 from redeem_notification_events where user_id=? and event_key=?", (user_id, event_key)).fetchone()
                    if not exists:
                        pending.append((row, event_key))
                if not pending:
                    continue
                game_name = str(pending[0][0].get("game_name") or game_id)
                lines = []
                for row, _ in pending[:20]:
                    reward = str(row.get("reward") or "").strip().replace("\n", "、")
                    lines.append(f"{row.get('code')}" + (f"｜{reward}" if reward else ""))
                if len(pending) > 20:
                    lines.append(f"另有 {len(pending)-20} 筆")
                link_params = {"game": game_id}
                if len(pending) == 1:
                    link_params["code"] = str(pending[0][0].get("code") or "")
                link = f"/redeem?{urllib.parse.urlencode(link_params)}"
                item_id = create_message_center_notification(db, user_id=user_id, title=f"{game_name}新增兌換碼", body="\n".join(lines), link=link)
                stamp = now()
                db.executemany(
                    "insert into redeem_notification_events(id,user_id,game_id,code_id,event_key,message_item_id,created_at) values(?,?,?,?,?,?,?)",
                    [(str(uuid.uuid4()), user_id, game_id, str(row["id"]), event_key, item_id, stamp) for row, event_key in pending],
                )
                sent += 1
    return sent


def safe_dispatch_redeem_code_notifications(code_ids: list[str]) -> int:
    try:
        return dispatch_redeem_code_notifications(code_ids)
    except Exception:
        return 0


@app.get("/api/redeem-codes")
def public_redeem_codes():
    with connect_db() as db:
        data = redeem_dataset(db, public_only=True)
    return {"ok": True, **data}


@app.get("/api/redeem-notification-preferences")
def user_redeem_notification_preferences(request: Request):
    user = require_user(request)
    with connect_db() as db:
        games = [redeem_game_payload(row) for row in db.execute("select * from redeem_games where enabled=1 order by display_order,name,game_id").fetchall()]
        rows = db.execute("select * from redeem_notification_preferences where user_id=?", (user["id"],)).fetchall()
    selected = {str(row["game_id"]) for row in rows}
    preferences = [{"game_id": game["game_id"], "name": game["name"], "selected": game["game_id"] in selected} for game in games]
    return {"ok": True, "game_ids": sorted(selected), "preferences": preferences}


@app.put("/api/redeem-notification-preferences")
def update_user_redeem_notification_preferences(body: RedeemNotificationPreferencesPayload, request: Request):
    user = require_user(request); stamp = now()
    selected = list(dict.fromkeys(str(game_id or "").strip() for game_id in body.game_ids if str(game_id or "").strip()))
    with connect_db() as db:
        valid = {str(row["game_id"]) for row in db.execute("select game_id from redeem_games where enabled=1").fetchall()}
        unknown = sorted(set(selected) - valid)
        if unknown:
            raise HTTPException(status_code=422, detail=f"通知設定包含不存在或已停用的遊戲：{'、'.join(unknown)}")
        db.execute("delete from redeem_notification_preferences where user_id=?", (user["id"],))
        db.executemany(
            "insert into redeem_notification_preferences(user_id,game_id,updated_at) values(?,?,?)",
            [(user["id"], game_id, stamp) for game_id in selected],
        )
    return {"ok": True, "game_ids": selected}


@app.get("/api/admin/redeem")
def admin_redeem_dataset(request: Request):
    require_admin(request)
    with connect_db() as db:
        data = redeem_dataset(db, public_only=False)
        import_total = int(db.execute("select count(*) from redeem_import_batches").fetchone()[0])
        data["import_batches"] = redeem_import_batches(db, 20, 0)
        data["import_batch_pagination"] = {"page": 1, "page_size": 20, "total": import_total, "pages": max(1, (import_total + 19) // 20)}
    return {"ok": True, **data}


@app.get("/api/admin/redeem/import/batches")
def admin_redeem_import_batches(
    request: Request,
    page: int = 1,
    page_size: int = 20,
    game_id: str = "",
    status: str = "",
    date_from: int | None = None,
    date_to: int | None = None,
    search: str = "",
):
    require_admin(request)
    size = max(1, min(100, int(page_size)))
    gid = normalize_redeem_game_id(game_id) if game_id.strip() else ""
    normalized_status = status.strip().casefold()
    if normalized_status not in {"", "preview_ready", "completed", "rolled_back"}:
        raise HTTPException(status_code=422, detail="批次匯入紀錄狀態篩選無效。")
    if (date_from is not None and date_from < 0) or (date_to is not None and date_to < 0):
        raise HTTPException(status_code=422, detail="批次匯入紀錄日期篩選無效。")
    if date_from is not None and date_to is not None and date_to < date_from:
        raise HTTPException(status_code=422, detail="批次匯入紀錄結束日期不可早於開始日期。")
    normalized_search = search.strip()
    if len(normalized_search) > 200:
        raise HTTPException(status_code=422, detail="批次匯入紀錄搜尋文字不可超過 200 個字。")
    where_sql, params = redeem_import_batch_where(
        game_id=gid,
        status=normalized_status,
        date_from=date_from,
        date_to=date_to,
        search=normalized_search,
    )
    with connect_db() as db:
        total = int(db.execute("select count(*) from redeem_import_batches b" + where_sql, params).fetchone()[0])
        pages = max(1, (total + size - 1) // size)
        current = max(1, min(int(page), pages))
        items = redeem_import_batches(
            db,
            size,
            (current - 1) * size,
            game_id=gid,
            status=normalized_status,
            date_from=date_from,
            date_to=date_to,
            search=normalized_search,
        )
    return {"ok": True, "items": items, "page": current, "page_size": size, "total": total, "pages": pages}


@app.get("/api/admin/redeem/import/batches/{batch_id}")
def admin_redeem_import_batch_detail(batch_id: str, request: Request):
    require_admin(request)
    with connect_db() as db:
        batch = redeem_import_batch_detail(db, batch_id)
    return {"ok": True, "batch": batch}


@app.delete("/api/admin/redeem/import/batches/{batch_id}")
@high_risk_operation
def admin_delete_redeem_import_batch(batch_id: str, request: Request):
    admin = require_admin(request)
    with connect_db() as db:
        batch = db.execute("select * from redeem_import_batches where id=?", (batch_id,)).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="找不到兌換碼批次匯入紀錄。")
        status = str(batch["status"] or "")
        if status in {"completed", "rolled_back"}:
            raise HTTPException(status_code=409, detail="已完成或已復原的批次必須保留稽核紀錄，不可刪除。")
        summary = _json_object(batch["summary_json"], {})
        game_id = str(batch["default_game_id"] or "")
        db.execute("delete from redeem_import_batches where id=?", (batch_id,))
    log_admin_action(
        admin["id"],
        "delete_incomplete_redeem_import",
        category="redeem",
        game_id=game_id,
        target_type="redeem_import_batch",
        target_id=batch_id,
        summary="刪除未完成的兌換碼批次匯入紀錄",
        before={"status": status, "summary": summary},
        actor_ip=client_ip(request),
        locked=True,
    )
    return {"ok": True, "batch_id": batch_id, "deleted": True}


@app.post("/api/admin/redeem/import/excel")
def admin_parse_redeem_excel(body: RedeemExcelImportPayload, request: Request):
    admin = require_admin(request)
    try:
        content = base64.b64decode(body.content_base64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise HTTPException(status_code=422, detail="Excel 檔案內容編碼錯誤，請重新選擇檔案。") from exc
    result = parse_redeem_excel_workbook(body.filename, content)
    log_admin_action(
        admin["id"],
        "parse_redeem_excel",
        category="redeem",
        target_type="redeem_import_file",
        target_id="",
        summary=f"讀取兌換碼 Excel：{result['filename']}",
        after={"sheet_name": result["sheet_name"], "rows": len(result["rows"]), "header_row": result["header_row"]},
        actor_ip=client_ip(request),
    )
    return {"ok": True, **result}


@app.get("/api/admin/redeem/import/excel-template")
def admin_redeem_excel_template(request: Request, game_id: str):
    admin = require_admin(request)
    gid = normalize_redeem_game_id(game_id)
    with connect_db() as db:
        game = require_redeem_game(db, gid)
        server_names = [
            str(row["name"])
            for row in db.execute(
                "select name from redeem_servers where game_id=? and enabled=1 order by display_order,name",
                (gid,),
            ).fetchall()
        ]
    content = build_redeem_excel_template(gid, str(game["name"]), server_names)
    log_admin_action(
        admin["id"],
        "generate_redeem_excel_template",
        category="redeem",
        game_id=gid,
        target_type="redeem_import_template",
        target_id=gid,
        summary=f"生成兌換碼 Excel 範例：{game['name']}",
        after={"server_count": len(server_names)},
        actor_ip=client_ip(request),
    )
    filename = f"redeem-code-template-{gid}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/admin/redeem/import/preview")
def admin_preview_redeem_import(body: RedeemImportPreviewPayload, request: Request):
    admin = require_admin(request)
    reason = body.reason.strip()
    input_type = body.input_type.strip().casefold() or "editor"
    if input_type not in {"editor", "clipboard", "excel", "advanced"}:
        raise HTTPException(status_code=422, detail="批次匯入輸入方式無效。")
    source_filename = body.source_filename.strip() if input_type == "excel" else ""
    source_sheet = body.source_sheet.strip() if input_type == "excel" else ""
    input_count = sum(bool(str(row.get("code") or "").strip()) for row in body.rows)
    if input_count > 1 and len(reason) < 3:
        raise HTTPException(status_code=422, detail="新增 2 筆以上兌換碼時，請填寫至少 3 個字的批次原因。")
    default_game_id = normalize_redeem_game_id(body.default_game_id) if body.default_game_id.strip() else ""
    batch_id = str(uuid.uuid4())
    items: list[dict[str, Any]] = []
    keys: list[str] = []
    seen_keys: set[str] = set()
    with connect_db() as db:
        for row_number, raw in enumerate(body.rows, 1):
            item_id = str(uuid.uuid4())
            try:
                existing_action = str(raw.get("existing_action") or "").strip().casefold()
                if existing_action not in {"", "skip", "update"}:
                    raise HTTPException(status_code=422, detail="既有兌換碼只能選擇略過或更新。")
                candidate = normalize_redeem_import_row(db, dict(raw), default_game_id)
                key = redeem_import_key(candidate["game_id"], candidate["code"])
                if key in seen_keys:
                    raise HTTPException(status_code=422, detail="同一批匯入資料中出現重複兌換碼。")
                seen_keys.add(key); keys.append(key)
                existing = db.execute(
                    "select * from redeem_codes where game_id=? and lower(trim(code))=?",
                    (candidate["game_id"], candidate["code"].strip().casefold()),
                ).fetchone()
                before = redeem_code_snapshot(existing) if existing else {}
                if before:
                    before["server_ids"] = redeem_json_list(before.get("server_ids_json"))
                    before["server_names"] = redeem_import_server_names(db, before["server_ids"])
                if existing and existing_action == "skip":
                    action = "skip"
                else:
                    action = "unchanged" if existing and redeem_import_candidate_matches(before, candidate) else "update" if existing else "create"
                target_id = str(existing["id"]) if existing else str(uuid.uuid4())
                item = {"id": item_id, "row_number": row_number, "game_id": candidate["game_id"], "code": candidate["code"], "action": action, "target_id": target_id, "candidate": candidate, "before": before, "error": ""}
            except HTTPException as exc:
                item = {"id": item_id, "row_number": row_number, "game_id": str(raw.get("game_id") or default_game_id or ""), "code": str(raw.get("code") or ""), "action": "error", "target_id": "", "candidate": {}, "before": {}, "error": str(exc.detail)}
            except Exception as exc:
                item = {"id": item_id, "row_number": row_number, "game_id": str(raw.get("game_id") or default_game_id or ""), "code": str(raw.get("code") or ""), "action": "error", "target_id": "", "candidate": {}, "before": {}, "error": str(exc)}
            items.append(item)
        summary = {
            "total": len(items),
            "create": sum(item["action"] == "create" for item in items),
            "update": sum(item["action"] == "update" for item in items),
            "skip": sum(item["action"] == "skip" for item in items),
            "unchanged": sum(item["action"] == "unchanged" for item in items),
            "errors": sum(item["action"] == "error" for item in items),
        }
        pre_state_hash = redeem_import_state_hash(db, keys)
        plan = {"keys": keys, "items": items}
        snapshot = {"before": [item["before"] for item in items if item["before"]], "created_ids": [item["target_id"] for item in items if item["action"] == "create"]}
        db.execute(
            """insert into redeem_import_batches(id,admin_user_id,status,default_game_id,reason,input_type,source_filename,source_sheet,summary_json,plan_json,snapshot_json,pre_state_hash,created_at)
               values(?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (batch_id, admin["id"], "preview_ready", default_game_id, reason, input_type, source_filename, source_sheet, json.dumps(summary, ensure_ascii=False), json.dumps(plan, ensure_ascii=False), json.dumps(snapshot, ensure_ascii=False), pre_state_hash, now()),
        )
        db.executemany(
            """insert into redeem_import_items(id,batch_id,row_number,game_id,code,action,target_id,candidate_json,before_json,error_text)
               values(?,?,?,?,?,?,?,?,?,?)""",
            [(item["id"], batch_id, item["row_number"], item["game_id"], item["code"], item["action"], item["target_id"], json.dumps(item["candidate"], ensure_ascii=False), json.dumps(item["before"], ensure_ascii=False), item["error"]) for item in items],
        )
    log_admin_action(admin["id"], "preview_redeem_import", category="redeem", target_type="redeem_import_batch", target_id=batch_id, summary="建立兌換碼批次匯入預覽", after=summary, metadata={"input_type": input_type, "source_filename": source_filename, "source_sheet": source_sheet}, actor_ip=client_ip(request), locked=True)
    return {"ok": True, "batch_id": batch_id, "status": "preview_ready", "summary": summary, "items": items, "can_execute": summary["errors"] == 0 and summary["create"] + summary["update"] > 0}


@app.post("/api/admin/redeem/import/batches/{batch_id}/execute")
@high_risk_operation
def admin_execute_redeem_import(batch_id: str, body: RedeemImportExecutePayload, request: Request):
    admin = require_admin(request)
    with connect_db() as db:
        batch = db.execute("select * from redeem_import_batches where id=?", (batch_id,)).fetchone()
        if not batch or batch["status"] != "preview_ready":
            raise HTTPException(status_code=409, detail="找不到可套用的兌換碼匯入預覽。")
        summary = _json_object(batch["summary_json"], {})
        plan = _json_object(batch["plan_json"], {})
        reason = body.reason.strip()
        if int(summary.get("total") or 0) > 1 and len(reason) < 3:
            raise HTTPException(status_code=422, detail="新增 2 筆以上兌換碼時，請填寫至少 3 個字的批次原因。")
        if int(summary.get("errors") or 0):
            raise HTTPException(status_code=409, detail="匯入預覽仍有錯誤，請修正後重新產生預覽。")
        keys = [str(key) for key in plan.get("keys") or []]
        if redeem_import_state_hash(db, keys) != str(batch["pre_state_hash"] or ""):
            raise HTTPException(status_code=409, detail="兌換碼資料已在預覽後變更，請重新產生預覽。")
        backup = create_database_backup()
        stamp = now()
        try:
            db.execute("begin immediate")
            for item in plan.get("items") or []:
                action = str(item.get("action") or "")
                if action not in {"create", "update"}:
                    continue
                candidate = dict(item.get("candidate") or {})
                target_id = str(item.get("target_id") or "")
                server_json = json.dumps(candidate.get("server_ids") or [], ensure_ascii=False)
                if action == "create":
                    db.execute(
                        """insert into redeem_codes(id,game_id,code,source,description,reward,start_at,end_at,server_ids_json,redeem_url,enabled,created_by,updated_by,created_at,updated_at)
                           values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (target_id, candidate["game_id"], candidate["code"], candidate["source"], "", candidate["reward"], candidate.get("start_at"), candidate.get("end_at"), server_json, candidate["redeem_url"], 1 if candidate.get("enabled", True) else 0, admin["id"], admin["id"], stamp, stamp),
                    )
                else:
                    db.execute(
                        """update redeem_codes set game_id=?,code=?,source=?,description='',reward=?,start_at=?,end_at=?,server_ids_json=?,redeem_url=?,enabled=?,updated_by=?,updated_at=? where id=?""",
                        (candidate["game_id"], candidate["code"], candidate["source"], candidate["reward"], candidate.get("start_at"), candidate.get("end_at"), server_json, candidate["redeem_url"], 1 if candidate.get("enabled", True) else 0, admin["id"], stamp, target_id),
                    )
            post_hash = redeem_import_state_hash(db, keys)
            completed_summary = {**summary, "applied": int(summary.get("create") or 0) + int(summary.get("update") or 0)}
            db.execute(
                "update redeem_import_batches set status='completed',reason=?,summary_json=?,post_state_hash=?,completed_at=?,completed_by=?,backup_name=? where id=?",
                (reason, json.dumps(completed_summary, ensure_ascii=False), post_hash, stamp, admin["id"], backup.name, batch_id),
            )
        except Exception as exc:
            raise HTTPException(status_code=409, detail=f"兌換碼批次匯入失敗，交易已回復：{exc}") from exc
    log_admin_action(admin["id"], "execute_redeem_import", category="redeem", game_id=str(batch["default_game_id"] or ""), target_type="redeem_import_batch", target_id=batch_id, summary="套用兌換碼批次匯入", before=summary, after=completed_summary, metadata={"reason": reason}, backup_name=backup.name, actor_ip=client_ip(request), locked=True)
    created_ids = [str(item.get("target_id") or "") for item in plan.get("items") or [] if item.get("action") == "create"]
    safe_dispatch_redeem_code_notifications(created_ids)
    return {"ok": True, "batch_id": batch_id, "status": "completed", "summary": completed_summary, "backup": backup.name}


@app.post("/api/admin/redeem/import/batches/{batch_id}/rollback")
@high_risk_operation
def admin_rollback_redeem_import(batch_id: str, body: RedeemImportRollbackPayload, request: Request):
    admin = require_admin(request)
    with connect_db() as db:
        batch = db.execute("select * from redeem_import_batches where id=?", (batch_id,)).fetchone()
        if not batch or batch["status"] != "completed" or batch["rolled_back_at"]:
            raise HTTPException(status_code=409, detail="此兌換碼匯入批次目前不可復原。")
        plan = _json_object(batch["plan_json"], {})
        snapshot = _json_object(batch["snapshot_json"], {})
        keys = [str(key) for key in plan.get("keys") or []]
        if redeem_import_state_hash(db, keys) != str(batch["post_state_hash"] or ""):
            raise HTTPException(status_code=409, detail="此批次完成後已有相關兌換碼被修改，為避免覆蓋新資料已阻擋復原。")
        safety = create_database_backup()
        try:
            db.execute("begin immediate")
            created_ids = [str(value) for value in snapshot.get("created_ids") or [] if str(value)]
            if created_ids:
                db.execute(f"delete from redeem_codes where id in ({','.join('?' for _ in created_ids)})", created_ids)
            for before in snapshot.get("before") or []:
                values = [before.get(column) for column in REDEEM_CODE_DB_COLUMNS]
                db.execute(
                    f"insert into redeem_codes({','.join(REDEEM_CODE_DB_COLUMNS)}) values({','.join('?' for _ in REDEEM_CODE_DB_COLUMNS)}) on conflict(id) do update set " + ",".join(f"{column}=excluded.{column}" for column in REDEEM_CODE_DB_COLUMNS if column != "id"),
                    values,
                )
            restored_hash = redeem_import_state_hash(db, keys)
            if restored_hash != str(batch["pre_state_hash"] or ""):
                raise RuntimeError("redeem_import_restore_hash_mismatch")
            db.execute("update redeem_import_batches set status='rolled_back',rolled_back_at=?,rolled_back_by=?,rollback_reason=?,rollback_backup_name=? where id=?", (now(), admin["id"], body.reason.strip(), safety.name, batch_id))
        except Exception as exc:
            raise HTTPException(status_code=409, detail=f"兌換碼批次復原失敗，交易已回復：{exc}") from exc
    log_admin_action(admin["id"], "rollback_redeem_import", category="redeem", game_id=str(batch["default_game_id"] or ""), target_type="redeem_import_batch", target_id=batch_id, summary="復原兌換碼批次匯入", metadata={"reason": body.reason}, backup_name=safety.name, actor_ip=client_ip(request), locked=True)
    return {"ok": True, "batch_id": batch_id, "status": "rolled_back", "backup": safety.name}


@app.post("/api/admin/redeem/games")
def admin_create_redeem_game(body: RedeemGamePayload, request: Request):
    admin = require_admin(request)
    game_id = normalize_redeem_game_id(body.game_id)
    name = normalize_redeem_text(body.name, field="遊戲名稱", limit=120, required=True)
    stamp = now()
    try:
        with connect_db() as db:
            db.execute(
                """insert into redeem_games(game_id,name,display_order,enabled,note,created_by,updated_by,created_at,updated_at)
                   values(?,?,?,?,?,?,?,?,?)""",
                (game_id, name, int(body.display_order), 1 if body.enabled else 0, body.note.strip(), admin["id"], admin["id"], stamp, stamp),
            )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="兌換碼遊戲 ID 已存在。") from exc
    log_admin_action(admin["id"], "create_redeem_game", category="redeem", target_type="redeem_game", target_id=game_id, summary=f"新增兌換碼遊戲：{name}", after=body.model_dump(), actor_ip=client_ip(request))
    return {"ok": True, "game_id": game_id}


@app.patch("/api/admin/redeem/games/{game_id}")
def admin_update_redeem_game(game_id: str, body: RedeemGameUpdatePayload, request: Request):
    admin = require_admin(request)
    gid = normalize_redeem_game_id(game_id)
    name = normalize_redeem_text(body.name, field="遊戲名稱", limit=120, required=True)
    with connect_db() as db:
        before = db.execute("select * from redeem_games where game_id=?", (gid,)).fetchone()
        if not before:
            raise HTTPException(status_code=404, detail="找不到兌換碼遊戲。")
        db.execute(
            "update redeem_games set name=?,display_order=?,enabled=?,note=?,updated_by=?,updated_at=? where game_id=?",
            (name, int(body.display_order), 1 if body.enabled else 0, body.note.strip(), admin["id"], now(), gid),
        )
    log_admin_action(admin["id"], "update_redeem_game", category="redeem", target_type="redeem_game", target_id=gid, summary=f"更新兌換碼遊戲：{name}", before=dict(before), after=body.model_dump(), actor_ip=client_ip(request))
    return {"ok": True}


@app.post("/api/admin/redeem/games/reorder")
def admin_reorder_redeem_games(body: RedeemReorderPayload, request: Request):
    admin = require_admin(request)
    requested = [normalize_redeem_game_id(value) for value in body.item_ids]
    if len(requested) != len(set(requested)):
        raise HTTPException(status_code=400, detail="兌換碼遊戲排序清單包含重複項目。")
    with connect_db() as db:
        db.execute("begin immediate")
        current_rows = db.execute("select game_id from redeem_games order by display_order,name,game_id").fetchall()
        current_ids = [str(row["game_id"]) for row in current_rows]
        if len(requested) != len(current_ids) or set(requested) != set(current_ids):
            raise HTTPException(status_code=409, detail="兌換碼遊戲清單已變更，請重新整理後再調整順序。")
        stamp = now()
        db.executemany(
            "update redeem_games set display_order=?,updated_by=?,updated_at=? where game_id=?",
            [(index, admin["id"], stamp, game_id) for index, game_id in enumerate(requested)],
        )
        rows = db.execute("select * from redeem_games order by display_order,name,game_id").fetchall()
    log_admin_action(admin["id"], "reorder_redeem_games", category="redeem", target_type="redeem_games", target_id="all", summary="調整兌換碼遊戲順序", before={"item_ids": current_ids}, after={"item_ids": requested}, actor_ip=client_ip(request))
    return {"ok": True, "games": [redeem_game_payload(row) for row in rows]}


@app.delete("/api/admin/redeem/games/{game_id}")
def admin_delete_redeem_game(game_id: str, request: Request):
    admin = require_admin(request)
    gid = normalize_redeem_game_id(game_id)
    with connect_db() as db:
        row = db.execute("select * from redeem_games where game_id=?", (gid,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="找不到兌換碼遊戲。")
        code_count = int(db.execute("select count(*) c from redeem_codes where game_id=?", (gid,)).fetchone()["c"] or 0)
        if code_count:
            raise HTTPException(status_code=409, detail="此遊戲已有兌換碼，請先刪除兌換碼或改為停用。")
        db.execute("delete from redeem_games where game_id=?", (gid,))
    log_admin_action(admin["id"], "delete_redeem_game", category="redeem", target_type="redeem_game", target_id=gid, summary=f"刪除兌換碼遊戲：{row['name']}", before=dict(row), actor_ip=client_ip(request))
    return {"ok": True}


@app.post("/api/admin/redeem/games/{game_id}/servers")
def admin_create_redeem_server(game_id: str, body: RedeemServerPayload, request: Request):
    admin = require_admin(request)
    gid = normalize_redeem_game_id(game_id)
    name = normalize_redeem_text(body.name, field="服務器名稱", limit=120, required=True)
    server_id = str(uuid.uuid4())
    stamp = now()
    try:
        with connect_db() as db:
            require_redeem_game(db, gid)
            db.execute(
                """insert into redeem_servers(id,game_id,name,display_order,enabled,created_by,updated_by,created_at,updated_at)
                   values(?,?,?,?,?,?,?,?,?)""",
                (server_id, gid, name, int(body.display_order), 1 if body.enabled else 0, admin["id"], admin["id"], stamp, stamp),
            )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="同一遊戲底下已有相同服務器名稱。") from exc
    log_admin_action(admin["id"], "create_redeem_server", category="redeem", game_id=gid, target_type="redeem_server", target_id=server_id, summary=f"新增兌換碼服務器：{name}", after=body.model_dump(), actor_ip=client_ip(request))
    return {"ok": True, "id": server_id}


@app.patch("/api/admin/redeem/servers/{server_id}")
def admin_update_redeem_server(server_id: str, body: RedeemServerPayload, request: Request):
    admin = require_admin(request)
    name = normalize_redeem_text(body.name, field="服務器名稱", limit=120, required=True)
    try:
        with connect_db() as db:
            before = db.execute("select * from redeem_servers where id=?", (server_id,)).fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="找不到兌換碼服務器。")
            db.execute(
                "update redeem_servers set name=?,display_order=?,enabled=?,updated_by=?,updated_at=? where id=?",
                (name, int(body.display_order), 1 if body.enabled else 0, admin["id"], now(), server_id),
            )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail="同一遊戲底下已有相同服務器名稱。") from exc
    log_admin_action(admin["id"], "update_redeem_server", category="redeem", game_id=before["game_id"], target_type="redeem_server", target_id=server_id, summary=f"更新兌換碼服務器：{name}", before=dict(before), after=body.model_dump(), actor_ip=client_ip(request))
    return {"ok": True}


@app.post("/api/admin/redeem/games/{game_id}/servers/reorder")
def admin_reorder_redeem_servers(game_id: str, body: RedeemReorderPayload, request: Request):
    admin = require_admin(request)
    gid = normalize_redeem_game_id(game_id)
    requested = [str(value or "").strip() for value in body.item_ids if str(value or "").strip()]
    if len(requested) != len(set(requested)):
        raise HTTPException(status_code=400, detail="服務器排序清單包含重複項目。")
    with connect_db() as db:
        db.execute("begin immediate")
        require_redeem_game(db, gid)
        current_rows = db.execute(
            "select id from redeem_servers where game_id=? order by display_order,name,id",
            (gid,),
        ).fetchall()
        current_ids = [str(row["id"]) for row in current_rows]
        if len(requested) != len(current_ids) or set(requested) != set(current_ids):
            raise HTTPException(status_code=409, detail="此遊戲的服務器清單已變更，請重新整理後再調整順序。")
        stamp = now()
        db.executemany(
            "update redeem_servers set display_order=?,updated_by=?,updated_at=? where game_id=? and id=?",
            [(index, admin["id"], stamp, gid, server_id) for index, server_id in enumerate(requested)],
        )
        rows = db.execute(
            "select * from redeem_servers where game_id=? order by display_order,name,id",
            (gid,),
        ).fetchall()
    log_admin_action(admin["id"], "reorder_redeem_servers", category="redeem", game_id=gid, target_type="redeem_servers", target_id=gid, summary="調整兌換碼服務器順序", before={"item_ids": current_ids}, after={"item_ids": requested}, actor_ip=client_ip(request))
    return {"ok": True, "game_id": gid, "servers": [redeem_server_payload(row) for row in rows]}


@app.delete("/api/admin/redeem/servers/{server_id}")
def admin_delete_redeem_server(server_id: str, request: Request):
    admin = require_admin(request)
    with connect_db() as db:
        row = db.execute("select * from redeem_servers where id=?", (server_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="找不到兌換碼服務器。")
        code_rows = db.execute("select id,server_ids_json from redeem_codes where game_id=?", (row["game_id"],)).fetchall()
        used = [code["id"] for code in code_rows if server_id in redeem_json_list(code["server_ids_json"])]
        if used:
            raise HTTPException(status_code=409, detail="此服務器仍被兌換碼使用，請先修改或刪除相關兌換碼。")
        db.execute("delete from redeem_servers where id=?", (server_id,))
    log_admin_action(admin["id"], "delete_redeem_server", category="redeem", game_id=row["game_id"], target_type="redeem_server", target_id=server_id, summary=f"刪除兌換碼服務器：{row['name']}", before=dict(row), actor_ip=client_ip(request))
    return {"ok": True}


@app.post("/api/admin/redeem/codes")
def admin_create_redeem_code(body: RedeemCodePayload, request: Request):
    admin = require_admin(request)
    gid = normalize_redeem_game_id(body.game_id)
    code = normalize_redeem_text(body.code, field="兌換碼", limit=200, required=True)
    source = normalize_redeem_text(body.source or body.description, field="來源", limit=500)
    reward = normalize_redeem_text(body.reward, field="獎勵", limit=500)
    start_at, end_at = normalize_redeem_time_range(body.start_at, body.end_at)
    redeem_url = normalize_redeem_url(body.redeem_url)
    code_id = str(uuid.uuid4())
    stamp = now()
    try:
        with connect_db() as db:
            require_redeem_game(db, gid)
            duplicate = find_redeem_code_duplicate(db, gid, code)
            if duplicate:
                raise HTTPException(status_code=409, detail=f"此遊戲已有兌換碼「{duplicate['code']}」，請改為編輯現有資料。")
            server_ids = validate_redeem_servers(db, gid, body.server_ids)
            db.execute(
                """insert into redeem_codes(id,game_id,code,source,description,reward,start_at,end_at,server_ids_json,redeem_url,enabled,created_by,updated_by,created_at,updated_at)
                   values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (code_id, gid, code, source, "", reward, start_at, end_at, json.dumps(server_ids, ensure_ascii=False), redeem_url, 1 if body.enabled else 0, admin["id"], admin["id"], stamp, stamp),
            )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail=f"此遊戲已有兌換碼「{code}」，請改為編輯現有資料。") from exc
    log_admin_action(admin["id"], "create_redeem_code", category="redeem", game_id=gid, target_type="redeem_code", target_id=code_id, summary=f"新增兌換碼：{code}", after=body.model_dump(), actor_ip=client_ip(request))
    safe_dispatch_redeem_code_notifications([code_id])
    return {"ok": True, "id": code_id}


@app.patch("/api/admin/redeem/codes/{code_id}")
def admin_update_redeem_code(code_id: str, body: RedeemCodePayload, request: Request):
    admin = require_admin(request)
    gid = normalize_redeem_game_id(body.game_id)
    code = normalize_redeem_text(body.code, field="兌換碼", limit=200, required=True)
    source = normalize_redeem_text(body.source or body.description, field="來源", limit=500)
    reward = normalize_redeem_text(body.reward, field="獎勵", limit=500)
    start_at, end_at = normalize_redeem_time_range(body.start_at, body.end_at)
    redeem_url = normalize_redeem_url(body.redeem_url)
    try:
        with connect_db() as db:
            before = db.execute("select * from redeem_codes where id=?", (code_id,)).fetchone()
            if not before:
                raise HTTPException(status_code=404, detail="找不到兌換碼。")
            require_redeem_game(db, gid)
            duplicate = find_redeem_code_duplicate(db, gid, code, exclude_id=code_id)
            if duplicate:
                raise HTTPException(status_code=409, detail=f"此遊戲已有兌換碼「{duplicate['code']}」，請改為編輯該筆資料。")
            server_ids = validate_redeem_servers(db, gid, body.server_ids)
            candidate = {
                "code": code,
                "source": source,
                "reward": reward,
                "start_at": start_at,
                "end_at": end_at,
                "server_ids": server_ids,
                "redeem_url": redeem_url,
                "enabled": bool(body.enabled),
            }
            materially_changed = str(before["game_id"] or "") != gid or not redeem_import_candidate_matches(dict(before), candidate)
            db.execute(
                """update redeem_codes set game_id=?,code=?,source=?,description=?,reward=?,start_at=?,end_at=?,server_ids_json=?,redeem_url=?,enabled=?,updated_by=?,updated_at=? where id=?""",
                (gid, code, source, "", reward, start_at, end_at, json.dumps(server_ids, ensure_ascii=False), redeem_url, 1 if body.enabled else 0, admin["id"], now(), code_id),
            )
    except sqlite3.IntegrityError as exc:
        raise HTTPException(status_code=409, detail=f"此遊戲已有兌換碼「{code}」，請改為編輯現有資料。") from exc
    log_admin_action(admin["id"], "update_redeem_code", category="redeem", game_id=gid, target_type="redeem_code", target_id=code_id, summary=f"更新兌換碼：{code}", before=dict(before), after=body.model_dump(), actor_ip=client_ip(request))
    return {"ok": True, "changed": materially_changed}


@app.delete("/api/admin/redeem/codes/{code_id}")
def admin_delete_redeem_code(code_id: str, request: Request):
    admin = require_admin(request)
    with connect_db() as db:
        row = db.execute("select * from redeem_codes where id=?", (code_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="找不到兌換碼。")
        db.execute("delete from redeem_codes where id=?", (code_id,))
    log_admin_action(admin["id"], "delete_redeem_code", category="redeem", game_id=row["game_id"], target_type="redeem_code", target_id=code_id, summary=f"刪除兌換碼：{row['code']}", before=dict(row), actor_ip=client_ip(request))
    return {"ok": True}


# ----- 公告與站內通知 -----
ANNOUNCEMENT_LEVELS={"info","success","warning","danger","update"}


def normalize_announcement_level(value: str) -> str:
    level=str(value or "info").strip().casefold()
    if level not in ANNOUNCEMENT_LEVELS:
        raise HTTPException(status_code=422,detail="公告類別無效，可使用：一般、成功、提醒、重要、更新。")
    return level


@app.get("/api/announcements")
def public_announcements():
    t=now()
    with connect_db() as db:
        rows=db.execute("select id,title,body,level,pinned,starts_at,ends_at,created_at,updated_at from announcements where is_active=1 and (starts_at is null or starts_at<=?) and (ends_at is null or ends_at>?) order by pinned desc,created_at desc",(t,t)).fetchall()
    return {"ok":True,"announcements":[dict(r) for r in rows]}

@app.get("/api/admin/announcements")
def admin_announcements(request: Request):
    require_admin(request)
    with connect_db() as db: rows=db.execute("select * from announcements order by pinned desc,created_at desc").fetchall()
    return {"ok":True,"announcements":[dict(r) for r in rows]}

@app.post("/api/admin/announcements")
def admin_create_announcement(body: AnnouncementPayload, request: Request):
    admin=require_admin(request); aid=str(uuid.uuid4()); t=now(); level=normalize_announcement_level(body.level)
    with connect_db() as db:
        db.execute("insert into announcements(id,title,body,level,is_active,pinned,starts_at,ends_at,created_by,created_at,updated_at) values(?,?,?,?,?,?,?,?,?,?,?)",
                   (aid,body.title.strip(),body.body.strip(),level,1 if body.is_active else 0,1 if body.pinned else 0,body.starts_at,body.ends_at,admin["id"],t,t))
    log_admin_action(admin["id"],"create_announcement",details=body.title)
    return {"ok":True,"id":aid}

@app.put("/api/admin/announcements/{announcement_id}")
def admin_update_announcement(announcement_id: str, body: AnnouncementPayload, request: Request):
    admin=require_admin(request); level=normalize_announcement_level(body.level)
    with connect_db() as db:
        if not (current:=db.execute("select id,updated_at from announcements where id=?",(announcement_id,)).fetchone()): raise HTTPException(status_code=404,detail="找不到公告。")
        db.execute("update announcements set title=?,body=?,level=?,is_active=?,pinned=?,starts_at=?,ends_at=?,updated_at=? where id=?",
                   (body.title.strip(),body.body.strip(),level,1 if body.is_active else 0,1 if body.pinned else 0,body.starts_at,body.ends_at,max(now(),int(current["updated_at"] or 0)+1),announcement_id))
    log_admin_action(admin["id"],"update_announcement",details=announcement_id)
    return {"ok":True}

@app.delete("/api/admin/announcements/{announcement_id}")
def admin_delete_announcement(announcement_id: str, request: Request):
    admin=require_admin(request)
    with connect_db() as db: deleted=db.execute("delete from announcements where id=?",(announcement_id,)).rowcount
    if not deleted: raise HTTPException(status_code=404,detail="找不到公告。")
    log_admin_action(admin["id"],"delete_announcement",details=announcement_id)
    return {"ok":True}

@app.get("/api/notifications")
def user_notifications(request: Request):
    user=require_user(request)
    with connect_db() as db:
        rows=db.execute("""select n.*,case when r.read_at is null then 0 else 1 end as is_read
        from notifications n
        left join notification_reads r on r.notification_id=n.id and r.user_id=?
        left join notification_deletions d on d.notification_id=n.id and d.user_id=?
        where (n.target_user_id is null or n.target_user_id=?) and d.user_id is null
        order by n.created_at desc limit 100""",(user["id"],user["id"],user["id"])).fetchall()
    values=[dict(r) for r in rows]
    return {"ok":True,"notifications":values,"unread":sum(1 for r in values if not r["is_read"])}

@app.post("/api/notifications/{notification_id}/read")
def mark_notification_read(notification_id: str, request: Request):
    user=require_user(request)
    with connect_db() as db:
        n=db.execute("select id,target_user_id from notifications where id=?",(notification_id,)).fetchone()
        if not n or (n["target_user_id"] and n["target_user_id"]!=user["id"]): raise HTTPException(status_code=404,detail="找不到通知。")
        db.execute("insert into notification_reads(user_id,notification_id,read_at) values(?,?,?) on conflict(user_id,notification_id) do update set read_at=excluded.read_at",(user["id"],notification_id,now()))
    return {"ok":True}

@app.post("/api/notifications/read-all")
def mark_all_notifications_read(request: Request):
    user=require_user(request); stamp=now()
    with connect_db() as db:
        rows=db.execute("""select n.id from notifications n
          left join notification_deletions d on d.notification_id=n.id and d.user_id=?
          where (n.target_user_id is null or n.target_user_id=?) and d.notification_id is null""",(user["id"],user["id"])).fetchall()
        db.executemany("insert into notification_reads(user_id,notification_id,read_at) values(?,?,?) on conflict(user_id,notification_id) do update set read_at=excluded.read_at",[(user["id"],row["id"],stamp) for row in rows])
    return {"ok":True,"updated":len(rows)}


@app.get("/api/admin/notifications")
def admin_notifications(request: Request):
    require_admin(request)
    with connect_db() as db:
        rows=db.execute("""select n.*,u.email target_email,creator.email creator_email
        from notifications n
        left join users u on u.id=n.target_user_id
        left join users creator on creator.id=n.created_by
        order by n.created_at desc limit 200""").fetchall()
        recipients=db.execute("""select email,username,role,is_active
        from users
        order by case when role='admin' then 0 else 1 end,coalesce(nullif(username,''),email_key)""").fetchall()
    return {
        "ok":True,
        "notifications":[dict(r) for r in rows],
        "recipients":[{
            "email":r["email"],"username":r["username"] or "","role":r["role"],"is_active":bool(r["is_active"]),
        } for r in recipients],
    }


@app.delete("/api/admin/notifications/{notification_id}")
def admin_delete_notification(notification_id: str, request: Request):
    admin=require_admin(request)
    with connect_db() as db:
        deleted=db.execute("delete from notifications where id=?",(notification_id,)).rowcount
    if not deleted: raise HTTPException(status_code=404,detail="找不到通知。")
    log_admin_action(admin["id"],"delete_notification",details=notification_id)
    return {"ok":True}


@app.post("/api/admin/notifications")
def admin_create_notification(body: NotificationPayload, request: Request):
    admin=require_admin(request)
    scope=(body.target_scope or "all").strip().casefold()
    if scope not in {"all","admins","user"}:
        raise HTTPException(status_code=400,detail="通知對象類型無效。")
    target_ids: list[str|None]=[]
    target_label="全體使用者"
    recipient_count=0
    with connect_db() as db:
        if scope=="admins":
            rows=db.execute("select id,email from users where role='admin' and is_active=1 order by email_key").fetchall()
            target_ids=[str(row["id"]) for row in rows]
            target_label="全部管理員"
            recipient_count=len(target_ids)
            if not target_ids:
                raise HTTPException(status_code=400,detail="目前沒有可接收通知的啟用中管理員。")
        elif scope=="user":
            email=body.target_email.strip()
            if not email:
                raise HTTPException(status_code=400,detail="請選擇通知目標帳號。")
            row=db.execute("select id,email from users where email_key=?",(normalize_email(email),)).fetchone()
            if not row:
                raise HTTPException(status_code=404,detail="找不到目標帳號。")
            target_ids=[str(row["id"])]
            target_label=str(row["email"])
            recipient_count=1
        else:
            # A broadcast is stored once, but the response and audit record show
            # the number of currently active accounts that can receive it.
            target_ids=[None]
            recipient_count=int(db.execute("select count(*) from users where is_active=1").fetchone()[0] or 0)
    ids=[create_notification(body.title,body.body,body.kind,body.link,target_id,admin["id"]) for target_id in target_ids]
    log_admin_action(
        admin["id"],"create_notification",None,
        f"target_scope={scope}; target={target_label}; recipients={recipient_count}; title={body.title.strip()}"
    )
    return {"ok":True,"id":ids[0],"ids":ids,"target_scope":scope,"target_label":target_label,"recipient_count":recipient_count}

# ----- 郵件紀錄 -----
@app.get("/api/admin/email-logs")
def admin_email_logs(request: Request, status: str = "", mail_type: str = "", search: str = "", limit: int = 300):
    require_admin(request)
    limit=max(1,min(int(limit or 300),1000))
    clauses=[]; params=[]
    if status.strip(): clauses.append("status=?"); params.append(status.strip())
    if mail_type.strip(): clauses.append("mail_type=?"); params.append(mail_type.strip())
    if search.strip(): clauses.append("(recipient like ? or subject like ? or error like ?)"); term=f"%{search.strip()}%"; params.extend([term,term,term])
    where=(" where "+" and ".join(clauses)) if clauses else ""
    with connect_db() as db:
        rows=db.execute(f"select * from email_logs{where} order by id desc limit ?",(*params,limit)).fetchall()
        types=[str(r["mail_type"]) for r in db.execute("select distinct mail_type from email_logs order by mail_type").fetchall()]
    return {"ok":True,"logs":[dict(r) for r in rows],"mail_types":types}

@app.post("/api/admin/email-logs/{log_id}/retry")
def admin_retry_email(log_id: int, request: Request):
    admin=require_admin(request)
    with connect_db() as db: row=db.execute("select recipient,mail_type from email_logs where id=?",(log_id,)).fetchone()
    if not row: raise HTTPException(status_code=404,detail="找不到郵件紀錄。")
    with connect_db() as db: user=db.execute("select id,email,email_verified from users where email_key=?",(normalize_email(row["recipient"]),)).fetchone()
    if not user: raise HTTPException(status_code=404,detail="此收件者不是系統帳號，無法自動重寄。")
    if row["mail_type"]=="verification": issue_verification(user["id"],user["email"])
    elif row["mail_type"]=="password_reset": issue_reset(user["id"],user["email"])
    else: raise HTTPException(status_code=400,detail="此類郵件不支援自動重寄。")
    log_admin_action(admin["id"],"retry_email",user["id"],f"log={log_id} type={row['mail_type']}")
    return {"ok":True}

@app.delete("/api/admin/email-logs/{log_id}")
def admin_delete_email_log(log_id: int, request: Request):
    admin=require_admin(request)
    with connect_db() as db:
        deleted=db.execute("delete from email_logs where id=?",(log_id,)).rowcount
    if not deleted:
        raise HTTPException(status_code=404,detail="找不到郵件紀錄。")
    log_admin_action(admin["id"],"delete_email_log",details=f"log={log_id}")
    return {"ok":True}


@app.post("/api/admin/email/test")
def admin_test_email(body: AdminTestEmailPayload, request: Request):
    admin=require_admin(request)
    recipient=str(body.recipient)
    message=build_email(recipient,"遊戲成就紀錄器測試郵件","這是一封由後台寄出的測試郵件。","<p>這是一封由後台寄出的測試郵件。</p>")
    try:
        deliver_email(message,"admin_test")
    except Exception as exc:
        raise HTTPException(status_code=502,detail=f"測試郵件寄送失敗：{exc}")
    log_admin_action(admin["id"],"send_test_email",details=f"recipient={recipient}")
    return {"ok":True}


# ----- 封鎖名單 -----
@app.get("/api/admin/blocks")
def admin_blocks(request: Request):
    require_admin(request)
    with connect_db() as db: rows=db.execute("select * from blocked_entries order by created_at desc").fetchall()
    return {"ok":True,"blocks":[dict(r) for r in rows]}

@app.post("/api/admin/blocks")
def admin_create_block(body: BlockEntryPayload, request: Request):
    admin=require_admin(request); kind=body.kind.strip().lower(); value=body.value.strip().casefold()
    if kind not in {"ip","email","domain"}: raise HTTPException(status_code=400,detail="封鎖類型只能是 IP、信箱或網域。")
    if kind=="domain": value=value.lstrip("@")
    if not value: raise HTTPException(status_code=400,detail="封鎖值不能為空白。")
    bid=str(uuid.uuid4()); logged_out_users=0; logged_out_sessions=0
    try:
        with connect_db() as db:
            db.execute("begin immediate")
            db.execute("insert into blocked_entries(id,kind,value_key,reason,active,created_by,created_at) values(?,?,?,?,?,?,?)",(bid,kind,value,body.reason.strip(),1 if body.active else 0,admin["id"],now()))
            if body.active:
                if kind=="email":
                    rows=db.execute("select id from users where email_key=?",(value,)).fetchall()
                elif kind=="domain":
                    rows=db.execute("select id from users where instr(email_key,'@')>0 and substr(email_key,instr(email_key,'@')+1)=?",(value,)).fetchall()
                else:
                    rows=db.execute(
                        "select distinct u.id from users u left join sessions s on s.user_id=u.id where u.last_login_ip=? or s.ip_address=?",
                        (value,value)
                    ).fetchall()
                user_ids=[row["id"] for row in rows]
                logged_out_users=len(user_ids)
                if user_ids:
                    placeholders=",".join("?" for _ in user_ids)
                    logged_out_sessions=int(db.execute(f"select count(*) as c from sessions where user_id in ({placeholders})",user_ids).fetchone()["c"] or 0)
                    db.execute(f"delete from sessions where user_id in ({placeholders})",user_ids)
    except sqlite3.IntegrityError: raise HTTPException(status_code=409,detail="此封鎖項目已存在。")
    log_admin_action(admin["id"],"create_block",details=f"{kind}:{value}; users={logged_out_users}; sessions={logged_out_sessions}")
    return {"ok":True,"id":bid,"logged_out_users":logged_out_users,"logged_out_sessions":logged_out_sessions}

@app.delete("/api/admin/blocks/{block_id}")
def admin_delete_block(block_id: str, request: Request):
    admin=require_admin(request)
    with connect_db() as db: deleted=db.execute("delete from blocked_entries where id=?",(block_id,)).rowcount
    if not deleted: raise HTTPException(status_code=404,detail="找不到封鎖項目。")
    log_admin_action(admin["id"],"delete_block",details=block_id)
    return {"ok":True}

# ----- 成就編輯、自訂標籤、精選與版本紀錄 -----
@app.get("/api/achievement-customizations")
def achievement_customizations():
    return extra_game_customizations("wuwa")


@app.get("/api/admin/achievement-overrides")
def admin_achievement_overrides(request: Request):
    return extra_game_admin_overrides("wuwa",request)


@app.put("/api/admin/achievements/{achievement_id}")
def admin_save_achievement(achievement_id: str, body: AchievementEditPayload, request: Request):
    return extra_game_save_achievement("wuwa",achievement_id,body,request)


@app.delete("/api/admin/achievements/{achievement_id}")
def admin_hide_achievement(achievement_id: str, request: Request):
    return extra_game_hide_achievement("wuwa",achievement_id,request)


@app.get("/api/admin/achievements/{achievement_id}/permanent-preview")
def admin_permanent_delete_preview(achievement_id: str, request: Request):
    return extra_game_permanent_delete_preview("wuwa",achievement_id,request)


@app.delete("/api/admin/achievements/{achievement_id}/permanent")
def admin_legacy_permanent_delete_disabled(achievement_id: str, request: Request):
    return extra_game_legacy_permanent_delete_disabled("wuwa",achievement_id,request)


@app.post("/api/admin/achievements/{achievement_id}/permanent")
def admin_permanently_delete_achievement(achievement_id: str, body: AchievementPermanentDeletePayload, request: Request):
    return extra_game_permanently_delete_achievement("wuwa",achievement_id,body,request)


@app.get("/api/admin/achievement-deletions")
def admin_list_achievement_deletions(request: Request):
    return extra_game_list_achievement_deletions("wuwa",request)


@app.post("/api/admin/achievement-deletions/{delete_id}/restore")
def admin_restore_permanent_deletion(delete_id: str, body: AchievementPermanentRestorePayload, request: Request):
    return extra_game_restore_permanent_deletion("wuwa",delete_id,body,request)


@app.post("/api/admin/achievements/{achievement_id}/restore")
def admin_restore_achievement(achievement_id: str, request: Request):
    return extra_game_restore_achievement("wuwa",achievement_id,request)


@app.get("/api/admin/achievements/{achievement_id}/revisions")
def admin_achievement_revisions(achievement_id: str, request: Request):
    return extra_game_achievement_revisions("wuwa",achievement_id,request)


@app.post("/api/admin/catalog/rebuild")
def admin_rebuild_achievement_catalog(request: Request):
    return extra_game_rebuild_catalog("wuwa",request)


def _catalog_validation_context(game_id: str) -> tuple[dict[str,int],dict[str,int]]:
    with connect_db() as db:
        progress={str(row["achievement_id"]):int(row["c"] or 0) for row in db.execute(
            "select achievement_id,count(*) c from game_progress where game_id=? group by achievement_id",(game_id,)
        ).fetchall()}
        relations={str(row["achievement_id"]):int(row["c"] or 0) for row in db.execute(
            "select achievement_id,count(*) c from game_achievement_choice_groups where game_id=? group by achievement_id",(game_id,)
        ).fetchall()}
    return progress,relations


def _database_catalog_fingerprint(game_id: str) -> str:
    with connect_db() as db:
        rows=[dict(row) for row in db.execute(
            "select achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order from game_catalog_items where game_id=? order by source_order,achievement_id",
            (game_id,),
        ).fetchall()]
    return hashlib.sha256(json.dumps(rows,ensure_ascii=False,sort_keys=True,separators=(",",":"),default=str).encode("utf-8")).hexdigest()


def _load_catalog_items_for_health(game_id: str) -> list[dict[str,Any]]:
    path=game_catalog_file(game_id)
    if not path.exists():
        raise HTTPException(status_code=404,detail="找不到成就目錄檔。")
    try:
        payload=json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        raise HTTPException(status_code=400,detail=f"成就目錄檔無法讀取：{exc}")
    rows=payload.get("items") if isinstance(payload,dict) else payload
    if not isinstance(rows,list):
        raise HTTPException(status_code=400,detail="成就目錄格式錯誤。")
    return [dict(row) for row in rows if isinstance(row,dict)]


CATALOG_DECISIONS_THAT_RESOLVE_ISSUES={
    "mark_legal_duplicate","create_alias","create_stage_group","create_exclusive_group","ignore_permanent"
}


def _apply_catalog_issue_decisions(game_id: str, result: dict[str,Any]) -> dict[str,Any]:
    with connect_db() as db:
        decisions={str(row["issue_key"]):str(row["decision"] or "") for row in db.execute(
            "select issue_key,decision from catalog_issue_decisions where game_id=?",(game_id,)
        ).fetchall()}
    resolved=[]; active=[]
    for issue in result.get("issues") or []:
        decision=decisions.get(str(issue.get("issue_id") or ""),"")
        if decision in CATALOG_DECISIONS_THAT_RESOLVE_ISSUES:
            resolved.append({**issue,"resolution":decision})
        else:
            active.append(issue)
    counts: dict[str,int]={}; risks={"confirmed":0,"needs_review":0,"blocked":0}
    for issue in active:
        kind=str(issue.get("kind") or "unknown"); counts[kind]=counts.get(kind,0)+1
        risk=str(issue.get("risk") or "needs_review"); risks[risk]=risks.get(risk,0)+1
    result={**result,"issues":active,"resolved_issues":resolved,"resolved_count":len(resolved),"by_kind":counts,"by_risk":risks}
    result["errors"]=sum(1 for issue in active if issue.get("level")=="error")
    result["warnings"]=sum(1 for issue in active if issue.get("level")=="warning")
    result["info"]=sum(1 for issue in active if issue.get("level")=="info")
    return result


def _scan_catalog_for_admin(game_id: str, items: list[dict[str,Any]], admin_id: str) -> dict[str,Any]:
    authoritative=_load_catalog_items_for_health(game_id)
    rows=items or authoritative
    if len(rows)>20000: raise HTTPException(status_code=400,detail="資料筆數過多。")
    config=get_game_config(game_id) or {}
    minimum=max(1,int(config.get("minimumCatalogCount") or 1))
    if len(rows)<minimum:
        raise HTTPException(status_code=400,detail=f"待檢查資料只有 {len(rows)} 項，低於安全門檻 {minimum} 項。")
    authoritative_ids={str(row.get("id") or row.get("achievement_id") or "").strip() for row in authoritative}
    supplied_ids={str(row.get("id") or row.get("achievement_id") or "").strip() for row in rows}
    if items and (len(rows)!=len(authoritative) or supplied_ids!=authoritative_ids):
        raise HTTPException(status_code=409,detail="管理畫面中的成就資料不是目前完整目錄，請重新整理後再檢查。")
    progress,relations=_catalog_validation_context(game_id)
    result=_apply_catalog_issue_decisions(game_id,scan_catalog(rows,progress_counts=progress,relation_counts=relations))
    result["database_fingerprint"]=_database_catalog_fingerprint(game_id)
    result["catalog_fingerprint"]=hashlib.sha256(json.dumps(authoritative,ensure_ascii=False,sort_keys=True,separators=(",",":"),default=str).encode("utf-8")).hexdigest()
    scan_id=secrets.token_urlsafe(24); created=now(); expires=created+2*60*60
    result={**result,"scan_id":scan_id,"game_id":game_id,"expires_at":expires}
    with connect_db() as db:
        db.execute("delete from catalog_scan_previews where admin_user_id=? and game_id=?",(admin_id,game_id))
        db.execute("insert into catalog_scan_previews(id,game_id,admin_user_id,items_json,result_json,created_at,expires_at) values(?,?,?,?,?,?,?)",(
            scan_id,game_id,admin_id,json.dumps(rows,ensure_ascii=False,separators=(",",":")),json.dumps(result,ensure_ascii=False,separators=(",",":")),created,expires,
        ))
    return result


@app.post("/api/admin/catalog/validate")
def admin_validate_catalog(body: CatalogValidationPayload, request: Request):
    admin=require_admin(request)
    return {"ok":True,**_scan_catalog_for_admin("wuwa",body.items,admin["id"])}


def _write_catalog_items_payload(game_id: str, items: list[dict[str,Any]]) -> None:
    path=game_catalog_file(game_id)
    payload=json.loads(path.read_text(encoding="utf-8-sig")) if path.exists() else {"schema_version":1,"game_id":game_id}
    normalized=[]
    for source_item in items:
        item=dict(source_item)
        achievement_id=str(item.get("id") or item.get("achievement_id") or "").strip()
        official_order=official_id_number(achievement_id)
        item["sourceOrder"]=official_order
        item.pop("source_order",None)
        normalized.append(item)
    normalized.sort(key=lambda item:(int(str(item.get("id") or item.get("achievement_id"))),str(item.get("id") or item.get("achievement_id"))))
    payload["items"]=normalized; payload["count"]=len(normalized); payload["generated_at"]=time.strftime("%Y-%m-%dT%H:%M:%S%z")
    temp=path.with_suffix(path.suffix+".tmp")
    temp.write_text(json.dumps(payload,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    temp.replace(path)


def _replace_relation_members(document: dict[str,Any], remove_ids: list[str], keep_id: str) -> bool:
    changed=False; next_groups=[]; remove_set=set(remove_ids)
    for group in document.get("groups") or []:
        if not isinstance(group,dict):
            continue
        original=[str(value) for value in group.get("achievement_ids") or [] if str(value)]
        replaced=[]
        for value in original:
            candidate=keep_id if value in remove_set else value
            if candidate not in replaced:
                replaced.append(candidate)
        if replaced!=original:
            changed=True
        if len(replaced)>=2:
            next_groups.append({**group,"achievement_ids":replaced})
        else:
            changed=True
    if changed:
        document["groups"]=next_groups
    return changed


def _upsert_relation_group(document: dict[str,Any], relation_type: str, group_id: str, member_ids: list[str], reason: str) -> None:
    groups=document.get("groups") or []
    payload={"id":group_id,"type":relation_type,"achievement_ids":member_ids}
    if reason:
        payload["basis"]=reason
    existing=next((row for row in groups if isinstance(row,dict) and str(row.get("id") or "")==group_id),None)
    if existing is None:
        groups.append(payload)
    else:
        groups[groups.index(existing)]={**existing,**payload}
    document["groups"]=groups


def _apply_catalog_repair(game_id: str, record: sqlite3.Row, body: CatalogRepairPayload, admin: dict[str,Any], request: Request) -> dict[str,Any]:
    if int(record["expires_at"] or 0)<=now(): raise HTTPException(status_code=410,detail="檢查結果已過期，請重新檢查。")
    original_items=json.loads(record["items_json"] or "[]")
    scan_result=json.loads(record["result_json"] or "{}")
    if not isinstance(original_items,list) or not original_items:
        raise HTTPException(status_code=409,detail="檢查結果未保存完整目錄，請重新執行檢查。")
    if scan_result.get("database_fingerprint") and scan_result["database_fingerprint"]!=_database_catalog_fingerprint(game_id):
        raise HTTPException(status_code=409,detail="正式成就資料已在檢查後變更，請重新執行檢查。")
    authoritative=_load_catalog_items_for_health(game_id)
    current_catalog_fingerprint=hashlib.sha256(json.dumps(authoritative,ensure_ascii=False,sort_keys=True,separators=(",",":"),default=str).encode("utf-8")).hexdigest()
    if scan_result.get("catalog_fingerprint") and scan_result["catalog_fingerprint"]!=current_catalog_fingerprint:
        raise HTTPException(status_code=409,detail="成就目錄檔已在檢查後變更，請重新執行檢查。")
    plan=build_catalog_repair_plan(original_items,scan_result.get("issues") or [],body.actions)
    unsupported=[op for op in plan["operations"] if op.get("status") in {"blocked","unsupported"}]
    if unsupported: raise HTTPException(status_code=400,detail={"message":"部分處置無法安全執行。","operations":unsupported})
    if not plan["operations"]:
        raise HTTPException(status_code=400,detail="沒有選擇任何可執行處置。")

    backup=create_database_backup(); catalog_path=game_catalog_file(game_id); old_catalog=catalog_path.read_bytes() if catalog_path.exists() else None
    relation_documents={kind:_read_relation_document(game_id,kind) for kind in ("stage","exclusive")}
    relation_snapshots={kind:(game_relation_file(game_id,kind).read_bytes() if game_relation_file(game_id,kind).exists() else None) for kind in ("stage","exclusive")}
    relation_changed=False; operation_summary=[]; changed_count=0; no_change_count=0
    try:
        for op in plan["operations"]:
            action=str(op.get("action") or "")
            if action in {"merge_keep_first","merge_keep_selected"}:
                for document in relation_documents.values():
                    relation_changed=_replace_relation_members(document,[str(v) for v in op.get("remove_ids") or []],str(op.get("keep_id") or "")) or relation_changed
            elif action in {"create_stage_group","create_exclusive_group"}:
                relation_type=str(op.get("relation_type") or "stage")
                _upsert_relation_group(relation_documents[relation_type],relation_type,str(op.get("group_id") or ""),[str(v) for v in op.get("member_ids") or []],body.reason.strip())
                relation_changed=True
        if relation_changed:
            final_ids={str(row.get("id") or row.get("achievement_id") or "") for row in plan["items"]}
            relation_issues=_validate_relation_documents(game_id,relation_documents,final_ids)
            if relation_issues:
                raise HTTPException(status_code=400,detail={"message":"處置後關聯資料驗證失敗。","issues":relation_issues[:30]})
            for kind,document in relation_documents.items():
                _write_relation_document(game_id,kind,document)

        with connect_db() as db:
            db.execute("begin immediate")
            for op in plan["operations"]:
                action=str(op.get("action") or ""); issue_id=str(op.get("issue_id") or "")
                if op.get("status")=="no_change":
                    no_change_count+=1
                elif op.get("status")=="resolved_decision":
                    db.execute("insert into catalog_issue_decisions(game_id,issue_key,decision,details_json,updated_by,updated_at) values(?,?,?,?,?,?) on conflict(game_id,issue_key) do update set decision=excluded.decision,details_json=excluded.details_json,updated_by=excluded.updated_by,updated_at=excluded.updated_at",(game_id,issue_id,action,json.dumps(op,ensure_ascii=False),admin["id"],now()))
                    changed_count+=1
                elif action in {"merge_keep_first","merge_keep_selected"}:
                    keep_id=str(op.get("keep_id") or ""); remove_ids=[str(v) for v in op.get("remove_ids") or []]
                    for remove_id in remove_ids:
                        db.execute("insert or ignore into game_progress(game_id,user_id,achievement_id,completed_at) select game_id,user_id,?,completed_at from game_progress where game_id=? and achievement_id=?",(keep_id,game_id,remove_id))
                        db.execute("delete from game_progress where game_id=? and achievement_id=?",(game_id,remove_id))
                        db.execute("update game_achievement_reports set achievement_id=? where game_id=? and achievement_id=?",(keep_id,game_id,remove_id))
                        db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,remove_id))
                        db.execute("delete from game_deleted_achievements where game_id=? and achievement_id=?",(game_id,remove_id))
                        db.execute("delete from game_featured_achievements where game_id=? and achievement_id=?",(game_id,remove_id))
                        db.execute("delete from game_catalog_items where game_id=? and achievement_id=?",(game_id,remove_id))
                        db.execute("insert into achievement_id_aliases(game_id,alias_id,canonical_id,reason,created_by,created_at) values(?,?,?,?,?,?) on conflict(game_id,alias_id) do update set canonical_id=excluded.canonical_id,reason=excluded.reason,created_by=excluded.created_by,created_at=excluded.created_at",(game_id,remove_id,keep_id,body.reason or "duplicate_merge",admin["id"],now()))
                    changed_count+=max(1,len(remove_ids))
                elif action=="recalculate_order":
                    for item in plan["items"]:
                        achievement_id=str(item.get("id") or item.get("achievement_id") or "").strip()
                        if achievement_id:
                            db.execute("update game_catalog_items set source_order=?,updated_at=? where game_id=? and achievement_id=?",(official_id_number(achievement_id),now(),game_id,achievement_id))
                    plan["items"].sort(key=lambda item:(official_id_number(item.get("id") or item.get("achievement_id")),str(item.get("id") or item.get("achievement_id"))))
                    changed_count+=len(plan["items"])
                elif action=="create_alias":
                    db.execute("insert into achievement_id_aliases(game_id,alias_id,canonical_id,reason,created_by,created_at) values(?,?,?,?,?,?) on conflict(game_id,alias_id) do update set canonical_id=excluded.canonical_id,reason=excluded.reason,created_by=excluded.created_by,created_at=excluded.created_at",(game_id,str(op.get("alias_id") or ""),str(op.get("canonical_id") or ""),body.reason or "catalog_health_alias",admin["id"],now()))
                    db.execute("insert into catalog_issue_decisions(game_id,issue_key,decision,details_json,updated_by,updated_at) values(?,?,?,?,?,?) on conflict(game_id,issue_key) do update set decision=excluded.decision,details_json=excluded.details_json,updated_by=excluded.updated_by,updated_at=excluded.updated_at",(game_id,issue_id,action,json.dumps(op,ensure_ascii=False),admin["id"],now()))
                    changed_count+=1
                elif action in {"create_stage_group","create_exclusive_group"}:
                    db.execute("insert into catalog_issue_decisions(game_id,issue_key,decision,details_json,updated_by,updated_at) values(?,?,?,?,?,?) on conflict(game_id,issue_key) do update set decision=excluded.decision,details_json=excluded.details_json,updated_by=excluded.updated_by,updated_at=excluded.updated_at",(game_id,issue_id,action,json.dumps(op,ensure_ascii=False),admin["id"],now()))
                    changed_count+=1
                elif action=="manual_edit":
                    achievement_id=str(op.get("achievement_id") or ""); changes=op.get("changes") or {}
                    allowed={k:v for k,v in changes.items() if k in {"name","condition","version","category","reward","hidden","source_order"}}
                    if allowed:
                        assignments=",".join(f"{key}=?" for key in allowed)
                        cursor=db.execute(f"update game_catalog_items set {assignments},updated_at=? where game_id=? and achievement_id=?",(*allowed.values(),now(),game_id,achievement_id))
                        changed_count+=int(cursor.rowcount or 0)
                operation_summary.append(op)
            if relation_changed:
                _sync_relation_groups(db,game_id,game_relation_file(game_id,"stage"),"stage")
                _sync_relation_groups(db,game_id,game_relation_file(game_id,"exclusive"),"exclusive")
                _repair_choice_group_progress(db,game_id)
        _write_catalog_items_payload(game_id,plan["items"])
        with connect_db() as db:
            integrity=str(db.execute("pragma integrity_check").fetchone()[0])
            if integrity!="ok": raise RuntimeError(f"database_integrity:{integrity}")
            db_count=int(db.execute("select count(*) from game_catalog_items where game_id=?",(game_id,)).fetchone()[0])
            if db_count!=len(plan["items"]):
                raise RuntimeError(f"catalog_count_mismatch:{db_count}:{len(plan['items'])}")
            db.execute("delete from catalog_scan_previews where id=?",(body.scan_id,))
        progress,relations=_catalog_validation_context(game_id)
        remaining=_apply_catalog_issue_decisions(game_id,scan_catalog(plan["items"],progress_counts=progress,relation_counts=relations))
    except Exception as exc:
        rollback_errors=[]
        try:
            if old_catalog is None: catalog_path.unlink(missing_ok=True)
            else: catalog_path.write_bytes(old_catalog)
        except Exception as rollback_exc:
            rollback_errors.append(f"catalog_file:{rollback_exc}")
        for kind,data in relation_snapshots.items():
            path=game_relation_file(game_id,kind)
            try:
                if data is None: path.unlink(missing_ok=True)
                else: path.write_bytes(data)
            except Exception as rollback_exc:
                rollback_errors.append(f"relation_{kind}:{rollback_exc}")
        try:
            _restore_governance_database_scope(backup,game_id)
        except Exception as rollback_exc:
            rollback_errors.append(f"database_scope:{rollback_exc}")
        if rollback_errors:
            raise HTTPException(status_code=500,detail=f"成就目錄處置失敗，且安全回復未完整完成：{exc}；{'；'.join(rollback_errors)}") from exc
        raise HTTPException(status_code=409,detail=f"成就目錄處置失敗，已回到操作前狀態：{exc}") from exc
    bump_game_live_scope(game_id,"catalog"); bump_game_live_scope(game_id,"stats")
    log_admin_action(admin["id"],"repair_catalog_issues",details=f"game={game_id}; operations={len(operation_summary)}; changed={changed_count}; reason={body.reason}",category="catalog",game_id=game_id,target_type="catalog",target_id=game_id,summary="處置成就重複與異常",before={"count":len(original_items)},after={"count":len(plan["items"]),"remaining_issues":len(remaining.get("issues") or [])},metadata={"operations":operation_summary,"changed_count":changed_count,"no_change_count":no_change_count},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
    return {"ok":True,"game_id":game_id,"backup":backup.name,"operations":operation_summary,"count":len(plan["items"]),"changed_count":changed_count,"no_change_count":no_change_count,"remaining_issues":len(remaining.get("issues") or []),"remaining_summary":{"errors":remaining.get("errors",0),"warnings":remaining.get("warnings",0),"info":remaining.get("info",0),"resolved_count":remaining.get("resolved_count",0)}}


@app.post("/api/admin/catalog/repair")
@high_risk_operation
def admin_repair_catalog(body: CatalogRepairPayload, request: Request):
    admin=require_admin(request)
    with connect_db() as db: record=db.execute("select * from catalog_scan_previews where id=? and game_id='wuwa' and admin_user_id=?",(body.scan_id,admin["id"])).fetchone()
    if not record: raise HTTPException(status_code=404,detail="找不到檢查結果，請重新檢查。")
    return _apply_catalog_repair("wuwa",record,body,admin,request)


# ----- 全站完成率 -----
@app.get("/api/completion-stats")
def public_completion_stats(request: Request):
    return extra_game_public_completion_stats("wuwa",request)


@app.get("/api/admin/completion-stats")
def admin_completion_stats(request: Request):
    return extra_game_admin_completion_stats("wuwa",request)


# ----- 帳號合併 -----
@app.post("/api/admin/users/merge")
def admin_merge_accounts(body: MergeAccountsPayload, request: Request):
    admin=require_site_owner(request)
    if body.source_user_id==body.target_user_id: raise HTTPException(status_code=400,detail="來源與目標帳號不能相同。")
    if body.source_user_id==admin["id"]: raise HTTPException(status_code=400,detail="不能合併目前登入中的管理員帳號。")
    with connect_db() as db:
        db.execute("begin immediate")
        source=db.execute("select id,email,role from users where id=?",(body.source_user_id,)).fetchone(); target=db.execute("select id,email from users where id=?",(body.target_user_id,)).fetchone()
        if not source or not target: raise HTTPException(status_code=404,detail="找不到來源或目標帳號。")
        if is_site_owner_email(source["email"]): raise HTTPException(status_code=400,detail="站長帳號無法作為合併來源。")
        if source["role"]=="admin": raise HTTPException(status_code=400,detail="管理員帳號不能直接合併，請先降為一般用戶。")
        db.execute("""insert into game_progress(game_id,user_id,achievement_id,completed_at)
        select game_id,?,achievement_id,completed_at from game_progress where user_id=?
        on conflict(game_id,user_id,achievement_id) do update set completed_at=min(game_progress.completed_at,excluded.completed_at)""",
        (body.target_user_id,body.source_user_id))
        db.execute("update achievement_reports set user_id=? where user_id=?",(body.target_user_id,body.source_user_id))
        db.execute("update game_achievement_reports set user_id=? where user_id=?",(body.target_user_id,body.source_user_id))
        db.execute("update support_tickets set user_id=? where user_id=?",(body.target_user_id,body.source_user_id))
        db.execute("delete from sessions where user_id in (?,?)",(body.source_user_id,body.target_user_id))
        source_email=source["email"]; target_email=target["email"]
        db.execute("delete from users where id=?",(body.source_user_id,))
    create_notification("帳號進度已合併",f"已將 {source_email} 的成就進度合併至此帳號。","account",target_user_id=body.target_user_id,created_by=admin["id"])
    log_admin_action(admin["id"],"merge_accounts",body.target_user_id,f"source={source_email} target={target_email}")
    return {"ok":True}

# ----- 問題回報／客服單 -----
def ticket_payload(db: sqlite3.Connection, row: sqlite3.Row) -> dict[str,Any]:
    messages=db.execute("select m.*,u.email sender_email,u.role sender_role from support_ticket_messages m left join users u on u.id=m.sender_user_id where ticket_id=? order by created_at",(row["id"],)).fetchall()
    d=dict(row); d["messages"]=[dict(m) for m in messages]; return d

@app.post("/api/support/tickets")
def create_support_ticket(body: TicketCreatePayload, request: Request):
    user=require_user(request); tid=str(uuid.uuid4()); mid=str(uuid.uuid4()); t=now()
    with connect_db() as db:
        db.execute("insert into support_tickets(id,user_id,subject,status,priority,created_at,updated_at) values(?,?,?,'open','normal',?,?)",(tid,user["id"],body.subject.strip(),t,t))
        db.execute("insert into support_ticket_messages(id,ticket_id,sender_user_id,message,created_at) values(?,?,?,?,?)",(mid,tid,user["id"],body.message.strip(),t))
    return {"ok":True,"id":tid}

@app.get("/api/support/tickets")
def my_support_tickets(request: Request):
    user=require_user(request)
    with connect_db() as db:
        rows=db.execute("select * from support_tickets where user_id=? order by updated_at desc",(user["id"],)).fetchall(); tickets=[ticket_payload(db,r) for r in rows]
    return {"ok":True,"tickets":tickets}

@app.post("/api/support/tickets/{ticket_id}/reply")
def user_reply_ticket(ticket_id: str, body: TicketReplyPayload, request: Request):
    user=require_user(request); t=now()
    with connect_db() as db:
        row=db.execute("select id,status from support_tickets where id=? and user_id=?",(ticket_id,user["id"])).fetchone()
        if not row: raise HTTPException(status_code=404,detail="找不到客服單。")
        db.execute("insert into support_ticket_messages(id,ticket_id,sender_user_id,message,created_at) values(?,?,?,?,?)",(str(uuid.uuid4()),ticket_id,user["id"],body.message.strip(),t))
        db.execute("update support_tickets set status='open',updated_at=? where id=?",(t,ticket_id))
    return {"ok":True}

@app.delete("/api/support/tickets/{ticket_id}")
def user_delete_support_ticket(ticket_id: str, request: Request):
    user=require_user(request)
    with connect_db() as db:
        deleted=db.execute("delete from support_tickets where id=? and user_id=?",(ticket_id,user["id"])).rowcount
    if not deleted: raise HTTPException(status_code=404,detail="找不到客服單。")
    return {"ok":True}


@app.get("/api/admin/support/tickets")
def admin_support_tickets(request: Request):
    require_admin(request)
    with connect_db() as db:
        rows=db.execute("select t.*,u.email from support_tickets t left join users u on u.id=t.user_id order by case t.status when 'open' then 0 when 'pending' then 1 else 2 end,t.updated_at desc").fetchall(); tickets=[ticket_payload(db,r) for r in rows]
    return {"ok":True,"tickets":tickets}

@app.post("/api/admin/support/tickets/{ticket_id}/reply")
def admin_reply_ticket(ticket_id: str, body: TicketReplyPayload, request: Request):
    admin=require_admin(request); t=now(); status=(body.status or "pending").strip().lower()
    if status not in {"open","pending","resolved","closed"}: raise HTTPException(status_code=400,detail="客服單狀態錯誤。")
    with connect_db() as db:
        row=db.execute("select user_id,subject from support_tickets where id=?",(ticket_id,)).fetchone()
        if not row: raise HTTPException(status_code=404,detail="找不到客服單。")
        db.execute("insert into support_ticket_messages(id,ticket_id,sender_user_id,message,created_at) values(?,?,?,?,?)",(str(uuid.uuid4()),ticket_id,admin["id"],body.message.strip(),t))
        db.execute("update support_tickets set status=?,updated_at=? where id=?",(status,t,ticket_id))
    if row["user_id"]: create_notification("客服單有新回覆",f"「{row['subject']}」收到管理員回覆。","support",f"#ticket-{ticket_id}",row["user_id"],admin["id"])
    return {"ok":True}

@app.delete("/api/admin/support/tickets/{ticket_id}")
def admin_delete_support_ticket(ticket_id: str, request: Request):
    admin=require_admin(request)
    with connect_db() as db:
        deleted=db.execute("delete from support_tickets where id=?",(ticket_id,)).rowcount
    if not deleted: raise HTTPException(status_code=404,detail="找不到客服單。")
    # 客服單操作不寫入一般操作紀錄，避免紀錄被大量客服資料淹沒。
    return {"ok":True}


@app.patch("/api/admin/support/tickets/{ticket_id}")
def admin_update_ticket(ticket_id: str, body: TicketStatusPayload, request: Request):
    admin=require_admin(request); status=body.status.strip().lower(); priority=body.priority.strip().lower()
    if status not in {"open","pending","resolved","closed"}: raise HTTPException(status_code=400,detail="客服單狀態錯誤。")
    if priority not in {"low","normal","high","urgent"}: raise HTTPException(status_code=400,detail="優先級錯誤。")
    with connect_db() as db:
        if not db.execute("select id from support_tickets where id=?",(ticket_id,)).fetchone(): raise HTTPException(status_code=404,detail="找不到客服單。")
        db.execute("update support_tickets set status=?,priority=?,updated_at=? where id=?",(status,priority,now(),ticket_id))
    return {"ok":True}

@app.get("/api/progress")
def get_progress(request: Request):
    return extra_game_get_progress("wuwa",request)


@app.post("/api/progress/set")
@high_risk_operation
def set_progress(body: ProgressSet, request: Request):
    return extra_game_set_progress("wuwa",body,request)


@app.post("/api/progress/batch")
@high_risk_operation
def batch_progress(body: ProgressBatch, request: Request):
    return extra_game_batch_progress("wuwa",body,request)


@app.post("/api/progress/replace")
@high_risk_operation
def replace_progress(body: ProgressReplace, request: Request):
    return extra_game_replace_progress("wuwa",body,request)


def load_official_cache() -> dict[str,Any]:
    if not CACHE_FILE.exists():
        return {"ok":False,"message":"尚未建立官方成就快取，請由管理員執行同步。"}
    try:
        payload=json.loads(CACHE_FILE.read_text(encoding="utf-8"))
        meta={}
        if META_FILE.exists():
            try: meta=json.loads(META_FILE.read_text(encoding="utf-8"))
            except Exception: meta={}
        return {"ok":True,"cached":True,"payload":payload,"meta":meta}
    except Exception as exc:
        return {"ok":False,"message":"官方成就快取損壞。","errors":[str(exc)]}


@app.get("/api/official-achievements")
def official_achievements():
    return load_official_cache()


@app.get("/api/official-zh-tw-achievements")
def official_zh_tw_achievements():
    if not OFFICIAL_ZH_TW_FILE.exists():
        return {"ok":False,"records":[],"message":"尚未同步遊戲正式繁中成就文字。"}
    try:
        payload=json.loads(OFFICIAL_ZH_TW_FILE.read_text(encoding="utf-8-sig"))
        if not isinstance(payload,dict) or not isinstance(payload.get("records"),list):
            raise ValueError("資料格式不正確")
        return payload
    except Exception as exc:
        return {"ok":False,"records":[],"message":"遊戲正式繁中成就文字資料損壞。","errors":[str(exc)]}


@app.post("/api/admin/official-achievements/sync")
def admin_sync_official_achievements(request: Request):
    admin=require_admin(request)
    log_admin_action(
        admin["id"],
        "blocked_direct_sync",
        details="game=wuwa; reason=preview_required",
        category="sync",
        status="blocked",
        game_id="wuwa",
        target_type="official_catalog",
        summary="阻擋舊版直接同步；必須先建立差異預覽",
        actor_ip=client_ip(request),
    )
    raise HTTPException(status_code=409,detail="直接覆蓋同步已停用，請使用管理後台的「抓取並預覽差異」流程。")



# ===== 四遊戲共用成就專案 API =====
def require_extra_game(game_id: str) -> str:
    value=str(game_id or "").strip()
    if not value or any(ch not in "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-_" for ch in value):
        raise HTTPException(status_code=404, detail="找不到此遊戲專案。")
    project=next((row for row in load_registry().get("projects", []) if row.get("id")==value and row.get("enabled")), None)
    if not project:
        raise HTTPException(status_code=404, detail="找不到或尚未啟用此遊戲專案。")
    return value


def bump_game_live_scope(game_id: str, scope: str) -> None:
    if scope not in {"catalog","stats","reports"}:
        return
    with connect_db() as db:
        db.execute(
            """insert into game_live_revisions(game_id,scope,revision,updated_at) values(?,?,1,?)
            on conflict(game_id,scope) do update set revision=revision+1,updated_at=excluded.updated_at""",
            (game_id,scope,now()),
        )


def serialize_game_override(row: sqlite3.Row) -> dict[str,Any]:
    value=dict(row)
    value["hidden"]=bool(value.get("hidden"))
    value["is_deleted"]=bool(value.get("is_deleted"))
    value["tags"]=json_list(value.pop("tags_json", "[]"))
    return value


OFFICIAL_ACHIEVEMENT_ID_PATTERN = re.compile(r"^[0-9]+$")


def _validate_official_achievement_id(game_id: str, achievement_id: Any) -> str:
    aid = str(achievement_id or "").strip()
    if not aid or len(aid) > 64 or not OFFICIAL_ACHIEVEMENT_ID_PATTERN.fullmatch(aid):
        raise HTTPException(status_code=422, detail=f"{game_id} 成就 ID 必須是官方純數字 ID，不可使用中文、名稱、分類或自訂代碼。")
    return aid


def _effective_achievement_row(db: sqlite3.Connection, game_id: str, achievement_id: str) -> dict[str, Any] | None:
    aid = str(achievement_id or "").strip()
    if not aid:
        return None
    deleted = db.execute(
        "select 1 from game_deleted_achievements where game_id=? and achievement_id=?",
        (game_id, aid),
    ).fetchone()
    if deleted:
        return None
    base = db.execute(
        "select * from game_catalog_items where game_id=? and achievement_id=?",
        (game_id, aid),
    ).fetchone()
    override = db.execute(
        "select * from game_achievement_overrides where game_id=? and achievement_id=?",
        (game_id, aid),
    ).fetchone()
    if override and bool(override["is_deleted"]):
        return None
    if not base and not override:
        return None
    result = dict(base or {})
    if not result:
        result = {
            "game_id": game_id,
            "achievement_id": aid,
            "name": aid,
            "condition": "",
            "version": "未標示",
            "category": "未辨識分類",
            "reward": 0,
            "hidden": 0,
            "tags_json": "[]",
            "source": "admin_manual",
            "source_order": official_id_number(aid),
            "updated_at": int(override["updated_at"] or 0) if override else 0,
        }
    if override:
        ov = dict(override)
        for key in ("name", "condition", "version", "category", "reward", "hidden", "tags_json", "source"):
            if key in ov and ov[key] is not None:
                # A hide-only override intentionally stores placeholder text; do not replace
                # the official presentation fields with that placeholder.
                if ov.get("source") == "hide-only" and key not in {"hidden"}:
                    continue
                result[key] = ov[key]
        result["is_override"] = True
        result["override_updated_at"] = ov.get("updated_at")
    else:
        result["is_override"] = False
    result["achievement_id"] = aid
    result["id"] = aid
    result["hidden"] = bool(result.get("hidden"))
    result["tags"] = json_list(result.get("tags_json") or "[]")
    return result


def _resolve_effective_achievement_id(db: sqlite3.Connection, game_id: str, achievement_id: Any) -> str:
    aid = str(achievement_id or "").strip()
    if not aid:
        raise HTTPException(status_code=422, detail="成就 ID 不可為空。")
    if _effective_achievement_row(db, game_id, aid):
        return aid
    visited: set[str] = set()
    current = aid
    while current and current not in visited:
        visited.add(current)
        alias = db.execute(
            "select canonical_id from achievement_id_aliases where game_id=? and alias_id=?",
            (game_id, current),
        ).fetchone()
        if not alias:
            break
        current = str(alias["canonical_id"] or "").strip()
        if _effective_achievement_row(db, game_id, current):
            return current
    raise HTTPException(status_code=404, detail=f"找不到有效成就 ID：{aid}。請重新整理正式成就列表。")


def _effective_catalog_items(db: sqlite3.Connection, game_id: str) -> list[dict[str, Any]]:
    ids = {
        str(row["achievement_id"])
        for row in db.execute(
            """select achievement_id from game_catalog_items where game_id=?
               union select achievement_id from game_achievement_overrides where game_id=?""",
            (game_id, game_id),
        ).fetchall()
    }
    rows = [row for aid in ids if (row := _effective_achievement_row(db, game_id, aid))]
    rows.sort(key=lambda row: (official_id_number(row.get("achievement_id")), str(row.get("achievement_id"))))
    return rows


def record_game_achievement_revision(game_id: str, achievement_id: str, action: str, snapshot: dict[str,Any], actor_user_id: str | None) -> None:
    with connect_db() as db:
        db.execute(
            "insert into game_achievement_revisions(game_id,achievement_id,action,snapshot_json,actor_user_id,created_at) values(?,?,?,?,?,?)",
            (game_id,achievement_id,action,json.dumps(snapshot,ensure_ascii=False),actor_user_id,now()),
        )














@app.get("/api/games/{game_id}/catalog")
def extra_game_catalog(game_id: str):
    game_id=require_extra_game(game_id)
    with connect_db() as db:
        rows=db.execute(
            """select c.*,coalesce(s.official_source_id,c.achievement_id) as official_source_id,
            g.group_id as relation_group,g.relation_type,g.stage_order,
            (select count(*) from game_achievement_choice_groups x
             where x.game_id=c.game_id and x.group_id=g.group_id) as relation_group_size
            from game_catalog_items c
            left join game_catalog_source_records s
              on s.game_id=c.game_id and s.achievement_id=c.achievement_id
            left join game_achievement_choice_groups g
              on g.game_id=c.game_id and g.achievement_id=c.achievement_id
            where c.game_id=?
            order by c.source_order,c.achievement_id""",
            (game_id,),
        ).fetchall()
    items=[]
    for index,row in enumerate(rows):
        value=dict(row)
        items.append({
            "id":value["achievement_id"],
            "officialId":value.get("official_source_id") or value["achievement_id"],
            "displayId":value.get("official_source_id") or value["achievement_id"],
            "name":value["name"],
            "condition":value["condition"],
            "version":value["version"],
            "category":value["category"],
            "reward":int(value["reward"] or 0),
            "hidden":bool(value["hidden"]),
            "tags":json_list(value.get("tags_json") or "[]"),
            "source":value.get("source") or "catalog",
            "sourceOrder":int(value.get("source_order") or index),
            "relationGroup":value.get("relation_group") or "",
            "relationType":value.get("relation_type") or "",
            "stageOrder":int(value.get("stage_order") or 0),
            "relationGroupSize":int(value.get("relation_group_size") or 0),
            "choiceGroup":value.get("relation_group") if value.get("relation_type")=="exclusive" else "",
            "choiceGroupSize":int(value.get("relation_group_size") or 0) if value.get("relation_type")=="exclusive" else 0,
            "isChoiceGroup":bool(value.get("relation_group") and value.get("relation_type")=="exclusive"),
        })
    with connect_db() as db:
        categories=_achievement_category_rows(db,game_id)
        items=_sort_achievement_display_rows(db,game_id,items,categories)
        effective_count=_effective_achievement_count_for_ids(db,game_id,[str(item["id"]) for item in items])
    return {"ok":True,"items":items,"categories":categories,"count":effective_count,"raw_count":len(items),"game_id":game_id}


@app.get("/api/games/{game_id}/relation-groups")
def game_relation_groups(game_id: str):
    value=str(game_id or "").strip()
    if value!="wuwa":
        value=require_extra_game(value)
    with connect_db() as db:
        rows=db.execute(
            """select group_id,achievement_id,relation_type,stage_order
            from game_achievement_choice_groups
            where game_id=?
            order by relation_type,group_id,stage_order,achievement_id""",
            (value,),
        ).fetchall()
    return {
        "ok":True,
        "game_id":value,
        "relations":[{
            "group_id":str(row["group_id"]),
            "achievement_id":str(row["achievement_id"]),
            "relation_type":str(row["relation_type"] or ""),
            "stage_order":int(row["stage_order"] or 0),
        } for row in rows],
    }




def _candidate_row_for_json(row: dict[str,Any]) -> dict[str,Any]:
    achievement_id=str(row.get("achievement_id") or row.get("id") or "").strip()
    numeric_official_id=official_id_number(achievement_id)
    value={
        "achievement_id":achievement_id,
        "name":str(row.get("name") or "").strip(),
        "condition":str(row.get("condition") or "").strip(),
        "version":str(row.get("version") or "未標示").strip(),
        "category":str(row.get("category") or "未辨識分類").strip(),
        "reward":int(row.get("reward") or 0),
        "hidden":1 if row.get("hidden") else 0,
        "tags_json":row.get("tags_json") if isinstance(row.get("tags_json"),str) else json.dumps(row.get("tags") if isinstance(row.get("tags"),list) else [],ensure_ascii=False),
        "source":str(row.get("source") or "official").strip(),
        "source_order":numeric_official_id,
    }
    for key,default in (
        ("category_id",""),("group_id",""),("group_name",""),("progress_value",0),("level",0),
        ("next_link",""),("reward_id",""),("primary_source_id",""),("secondary_source_id",""),
        ("official_source_id",""),("internal_id",""),("identity_match_status",""),
        ("identity_match_confidence",0),("identity_match_basis",""),
        ("source_ref",""),("raw_json","{}"),("provenance_json","{}"),
    ):
        raw=row.get(key,default)
        if key in {"progress_value","level"}:
            try: value[key]=int(raw or 0)
            except (TypeError,ValueError): value[key]=0
        elif key=="identity_match_confidence":
            try: value[key]=float(raw or 0)
            except (TypeError,ValueError): value[key]=0.0
        else:
            value[key]=str(raw if raw is not None else default)
    return value









def _prepare_game_sync_candidate(game_id: str) -> tuple[list[dict[str,Any]],dict[str,Any],Any]:
    with connect_db() as db:
        verified_snapshot_rows=_current_official_catalog_rows(db,game_id)
    rows,metadata,source_payload=prepare_repository_candidate(
        game_id,data_dir=DATA_DIR,verified_snapshot_rows=verified_snapshot_rows
    )
    policy=get_source_policy(game_id)
    identity_metadata={"mode":"official_id_primary_key","mapped":len(rows)} if game_id=="wuwa" else {}
    if game_id=="wuwa":
        official_rows=[]
        for source_row in rows:
            value=dict(source_row)
            official_id=str(value.get("achievement_id") or value.get("id") or "").strip()
            official_id_number(official_id)
            value["achievement_id"]=official_id
            value["official_source_id"]=official_id
            value["source_order"]=int(official_id)
            for legacy_key in ("internal_id","identity_match_status","identity_match_confidence","identity_match_basis","_auxiliary_achievement_id","_auxiliary_match_method","_auxiliary_match_confidence"):
                value.pop(legacy_key,None)
            official_rows.append(value)
        rows=official_rows
    metadata={
        **metadata,
        "identity_model":identity_metadata,
        "source_policy":policy,
        "primary_source":policy.get("primary"),
        "secondary_source":policy.get("secondary"),
        "fallback_source":policy.get("fallback"),
        "requires_admin_confirmation":True,
    }
    return [_candidate_row_for_json(row) for row in rows],metadata,source_payload


def _current_official_catalog_rows(db: sqlite3.Connection,game_id: str) -> list[dict[str,Any]]:
    rows=db.execute("""select c.achievement_id,c.name,c.condition,c.version,c.category,c.reward,c.hidden,c.tags_json,c.source,c.source_order,
      coalesce(s.category_id,'') category_id,coalesce(s.group_id,'') group_id,coalesce(s.group_name,'') group_name,
      coalesce(s.progress_value,0) progress_value,coalesce(s.level,0) level,coalesce(s.next_link,'') next_link,
      coalesce(s.reward_id,'') reward_id,coalesce(s.primary_source_id,'') primary_source_id,
      coalesce(s.secondary_source_id,'') secondary_source_id,coalesce(s.official_source_id,'') official_source_id,
      c.achievement_id internal_id,coalesce(s.source_ref,'') source_ref,
      coalesce(s.raw_json,'{}') raw_json,coalesce(s.provenance_json,'{}') provenance_json
      from game_catalog_items c left join game_catalog_source_records s
        on s.game_id=c.game_id and s.achievement_id=c.achievement_id
      where c.game_id=? and lower(c.source) not in ('manual','admin')
      order by c.source_order,c.achievement_id""",(game_id,)).fetchall()
    return [_candidate_row_for_json(dict(row)) for row in rows]


def _safe_degraded_sync_candidate(game_id: str, exc: BaseException) -> tuple[list[dict[str, Any]], dict[str, Any], dict[str, Any]]:
    """Build a non-destructive preview when a remote source is unavailable.

    The candidate intentionally mirrors the current official catalog so the
    resulting preview contains no automatic changes. Source diagnostics are
    attached to every row as review-only conflicts. This guarantees that an
    upstream 502/503/504 or parser failure never becomes an IIS 502 response.
    """
    policy = get_source_policy(game_id)
    primary = policy.get("primary") if isinstance(policy.get("primary"), dict) else {}
    with connect_db() as db:
        rows = _current_official_catalog_rows(db, game_id)
    error = str(exc) if isinstance(exc,RepositorySourceError) else f"{type(exc).__name__}: {exc}"
    error_code = exc.code if isinstance(exc,RepositorySourceError) else source_error_code(exc)
    source_diagnostics = dict(exc.diagnostics) if isinstance(exc,RepositorySourceError) else {}
    # Keep diagnostics global instead of duplicating the same error for every
    # achievement. This keeps the preview metadata small even for large games.
    source_conflicts: dict[str, list[dict[str, Any]]] = {}
    metadata = {
        **source_common_metadata(
            game_id,
            fetch_status="degraded",
            source_mode="remote_degraded",
            purpose=str(primary.get("purpose") or "official_catalog"),
            error=error,
            error_code=error_code,
        ),
        "source": str(primary.get("name") or game_display_name(game_id)),
        "source_page": str(primary.get("url") or ""),
        "cached": False,
        "count": len(rows),
        "minimum_count": int((get_game_config(game_id) or {}).get("minimumCatalogCount") or 1),
        "source_policy": policy,
        "primary_source": primary,
        "secondary_source": policy.get("secondary"),
        "fallback_source": policy.get("fallback"),
        "cross_validation": {
            "pairing_status": "source_unavailable",
            "official_fetch_error": error,
            "error_code": error_code,
            "matched": 0,
            "official_observations": 0,
            "source_diagnostics": source_diagnostics,
            "cache_used": False,
        },
        "source_conflicts": source_conflicts,
        "source_notice": {
            "kind": "primary_fetch_unavailable",
            "message": "主要官方來源暫時無法取得。本次只建立診斷預覽，不會自動修改正式資料。",
            "error_code": error_code,
            "error": error,
        },
        "requires_admin_confirmation": True,
        "diagnostic_preview": True,
    }
    payload = {
        "pipeline_version": SOURCE_PIPELINE_VERSION,
        "adapter_id": source_adapter_id(game_id),
        "diagnostic_only": True,
        "error_code": error_code,
        "error": error,
    }
    return rows, metadata, payload


def _catalog_rows_fingerprint(rows: list[dict[str,Any]]) -> str:
    return shared_catalog_fingerprint(rows)


def _catalog_sync_context(db: sqlite3.Connection, game_id: str) -> tuple[dict[str,int],dict[str,int],set[str]]:
    progress={str(row["achievement_id"]):int(row["c"] or 0) for row in db.execute(
        "select achievement_id,count(*) c from game_progress where game_id=? group by achievement_id",(game_id,)
    ).fetchall()}
    relations={str(row["achievement_id"]):int(row["c"] or 0) for row in db.execute(
        "select achievement_id,count(*) c from game_achievement_choice_groups where game_id=? group by achievement_id",(game_id,)
    ).fetchall()}
    overrides={str(row["achievement_id"]) for row in db.execute(
        "select achievement_id from game_achievement_overrides where game_id=?",(game_id,)
    ).fetchall()}
    return progress,relations,overrides


def _catalog_sync_diff(current_rows: list[dict[str,Any]],candidate_rows: list[dict[str,Any]], *, game_id: str="", metadata: dict[str,Any] | None=None) -> dict[str,Any]:
    progress: dict[str,int]={}; relations: dict[str,int]={}; overrides: set[str]=set()
    if game_id:
        with connect_db() as db:
            progress,relations,overrides=_catalog_sync_context(db,game_id)
    metadata=metadata or {}
    source_conflicts: dict[str,list[dict[str,Any]]]={
        str(key): [dict(row) for row in value if isinstance(row,dict)]
        for key,value in (metadata.get("source_conflicts") or {}).items()
        if isinstance(value,list)
    }
    suspected_removed_ids={
        str(value).strip() for value in (metadata.get("suspected_removed_ids") or [])
        if str(value).strip()
    }
    global_confirmation_modes = {"fallback", "bundled_snapshot", "official_cache", "remote_degraded"}
    global_confirmation_required = bool(metadata.get("requires_admin_confirmation")) and (
        bool(metadata.get("diagnostic_preview"))
        or bool(metadata.get("apply_blocked"))
        or str(metadata.get("source_mode") or "") in global_confirmation_modes
    )
    if global_confirmation_required:
        reason={
            "kind":"source_confirmation_required",
            "message":"候選目錄來自輔助快照、備援或快取；主要官方來源只作參考，本次所有差異必須由管理員確認。",
            "source_mode":metadata.get("source_mode"),
            "primary_source":metadata.get("primary_source"),
            "candidate_source":metadata.get("source"),
        }
        for row in candidate_rows:
            achievement_id=str(row.get("achievement_id") or "")
            if achievement_id:
                source_conflicts.setdefault(achievement_id,[]).append(reason)
    return build_sync_diff(
        current_rows,candidate_rows,
        progress_counts=progress,relation_counts=relations,override_ids=overrides,
        source_conflicts=source_conflicts,suspected_removed_ids=suspected_removed_ids,
        game_id=game_id,
    )


def _write_game_catalog_candidate(game_id: str,rows: list[dict[str,Any]],metadata: dict[str,Any]) -> None:
    path=game_catalog_file(game_id)
    path.parent.mkdir(parents=True,exist_ok=True)
    items=[]
    for row in sort_catalog_rows(game_id, rows):
        items.append({
            "id":row["achievement_id"],"officialId":str(row.get("official_source_id") or row["achievement_id"]),
            "name":row["name"],"condition":row["condition"],
            "version":row["version"],"category":row["category"],"reward":int(row["reward"] or 0),
            "hidden":bool(row["hidden"]),"tags":json_list(row.get("tags_json")),"source":row["source"],
            "sourceOrder":int(row["source_order"] or 0),
            "categoryId":str(row.get("category_id") or ""),"groupId":str(row.get("group_id") or ""),
            "groupName":str(row.get("group_name") or ""),"progress":int(row.get("progress_value") or 0),
            "level":int(row.get("level") or 0),"nextLink":str(row.get("next_link") or ""),
            "rewardId":str(row.get("reward_id") or ""),
            "sourceDetails":{
                "primary":str(row.get("primary_source_id") or ""),
                "secondary":str(row.get("secondary_source_id") or ""),
                "officialId":str(row.get("official_source_id") or row.get("achievement_id") or ""),
                "internalId":str(row.get("achievement_id") or ""),
                "identityMatchStatus":str(row.get("identity_match_status") or ""),
                "identityMatchBasis":str(row.get("identity_match_basis") or ""),
                "ref":str(row.get("source_ref") or ""),
                "provenance":_json_object(row.get("provenance_json"), {}),
                "raw":_json_object(row.get("raw_json"), {}),
            },
        })
    payload={
        "schema_version":4,"game_id":game_id,"source":metadata.get("source") or (items[0].get("source") if items else "official-sync-preview"),
        "source_page":metadata.get("source_page") or "","source_mode":metadata.get("source_mode") or "",
        "source_architecture_version":SOURCE_PIPELINE_VERSION,
        "primary_source":metadata.get("primary_source") or {},"secondary_source":metadata.get("secondary_source") or {},
        "generated_at":time.strftime("%Y-%m-%dT%H:%M:%S%z"),"count":len(items),"items":items,
    }
    temp=path.with_suffix(path.suffix+".tmp")
    temp.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding="utf-8")
    temp.replace(path)


def _merge_applied_source_metadata(rows: list[dict[str,Any]],candidate_rows: list[dict[str,Any]],current_rows: list[dict[str,Any]],metadata: dict[str,Any]) -> list[dict[str,Any]]:
    candidate_by_id={str(row.get("achievement_id") or ""):row for row in candidate_rows}
    current_by_id={str(row.get("achievement_id") or ""):row for row in current_rows}
    primary_id=str((metadata.get("primary_source") or {}).get("id") or "")
    source_ref=str(metadata.get("source_ref") or "")
    result=[]
    extra_keys=("category_id","group_id","group_name","progress_value","level","next_link","reward_id","primary_source_id","secondary_source_id","official_source_id","internal_id","identity_match_status","identity_match_confidence","identity_match_basis","source_ref","raw_json","provenance_json")
    for core in rows:
        achievement_id=str(core.get("achievement_id") or "")
        value=dict(core)
        source=candidate_by_id.get(achievement_id) or current_by_id.get(achievement_id) or {}
        for key in extra_keys:
            value[key]=source.get(key,0 if key in {"progress_value","level"} else ("{}" if key in {"raw_json","provenance_json"} else ""))
        if not value.get("primary_source_id"): value["primary_source_id"]=primary_id
        if not value.get("source_ref"): value["source_ref"]=source_ref
        result.append(_candidate_row_for_json(value))
    return result


def _upsert_achievement_source_rows(db: sqlite3.Connection, source_rows: list[tuple[Any, ...]]) -> dict[str,int]:
    """Safely reconcile source IDs onto permanent internal identities.

    A source may previously have been bootstrapped with the legacy internal ID as
    its source_id.  When the real official ID becomes available, update that row
    in place instead of inserting a second mapping that violates the unique
    (game_id, source_name, internal_id) identity constraint.
    """
    by_source: dict[tuple[str,str,str],tuple[Any,...]]={}
    by_internal: dict[tuple[str,str,str],tuple[Any,...]]={}
    normalized: list[tuple[Any,...]]=[]
    for raw in source_rows:
        row=tuple(raw)
        game_id,source_name,source_id,internal_id=(str(row[0]),str(row[1]),str(row[2]),str(row[3]))
        if not game_id or not source_name or not source_id or not internal_id:
            raise RuntimeError("來源身分映射缺少 game_id、source_name、source_id 或 internal_id。")
        source_key=(game_id,source_name,source_id)
        internal_key=(game_id,source_name,internal_id)
        prior_source=by_source.get(source_key)
        if prior_source is not None and str(prior_source[3])!=internal_id:
            raise RuntimeError(f"來源身分衝突：{game_id}/{source_name}/{source_id} 同時指向多個內部成就。")
        prior_internal=by_internal.get(internal_key)
        if prior_internal is not None and str(prior_internal[2])!=source_id:
            raise RuntimeError(f"來源身分衝突：{game_id}/{source_name}/{internal_id} 同時對應多個來源 ID。")
        if prior_source is None and prior_internal is None:
            normalized.append(row)
        by_source[source_key]=row
        by_internal[internal_key]=row

    inserted=0; updated=0; rekeyed=0
    for row in normalized:
        game_id,source_name,source_id,internal_id,is_primary,match_status,match_confidence,match_basis,first_seen_at,last_seen_at=row
        game_id=str(game_id); source_name=str(source_name); source_id=str(source_id); internal_id=str(internal_id)
        existing_by_source=db.execute(
            "select rowid,source_id,internal_id,first_seen_at from achievement_source_ids where game_id=? and source_name=? and source_id=?",
            (game_id,source_name,source_id),
        ).fetchone()
        existing_by_internal=db.execute(
            "select rowid,source_id,internal_id,first_seen_at from achievement_source_ids where game_id=? and source_name=? and internal_id=?",
            (game_id,source_name,internal_id),
        ).fetchone()
        if existing_by_source is not None and str(existing_by_source["internal_id"])!=internal_id:
            raise RuntimeError(
                f"來源身分衝突：{game_id}/{source_name}/{source_id} 已屬於內部成就 {existing_by_source['internal_id']}，"
                f"不能改指向 {internal_id}。"
            )
        if existing_by_internal is not None:
            rowid=int(existing_by_internal["rowid"])
            old_source_id=str(existing_by_internal["source_id"])
            if existing_by_source is not None and int(existing_by_source["rowid"])!=rowid:
                raise RuntimeError(
                    f"來源身分衝突：{game_id}/{source_name}/{source_id} 與內部成就 {internal_id} 已分別存在不同映射。"
                )
            db.execute(
                """update achievement_source_ids set source_id=?,is_primary=?,match_status=?,match_confidence=?,match_basis=?,last_seen_at=?
                where rowid=?""",
                (source_id,int(is_primary),str(match_status),float(match_confidence),str(match_basis),int(last_seen_at),rowid),
            )
            if old_source_id!=source_id:
                rekeyed+=1
            else:
                updated+=1
            continue
        if existing_by_source is not None:
            db.execute(
                """update achievement_source_ids set is_primary=?,match_status=?,match_confidence=?,match_basis=?,last_seen_at=?
                where rowid=?""",
                (int(is_primary),str(match_status),float(match_confidence),str(match_basis),int(last_seen_at),int(existing_by_source["rowid"])),
            )
            updated+=1
            continue
        db.execute(
            """insert into achievement_source_ids(game_id,source_name,source_id,internal_id,is_primary,match_status,match_confidence,match_basis,first_seen_at,last_seen_at)
            values(?,?,?,?,?,?,?,?,?,?)""",
            (game_id,source_name,source_id,internal_id,int(is_primary),str(match_status),float(match_confidence),str(match_basis),int(first_seen_at),int(last_seen_at)),
        )
        inserted+=1
    return {"inserted":inserted,"updated":updated,"rekeyed":rekeyed,"total":len(normalized)}


def _replace_official_catalog_rows(db: sqlite3.Connection,game_id: str,rows: list[dict[str,Any]]) -> None:
    stamp=now()
    db.execute("delete from game_catalog_items where game_id=? and lower(source) not in ('manual','admin')",(game_id,))
    db.execute("delete from game_catalog_source_records where game_id=?",(game_id,))
    db.executemany("""insert into game_catalog_items(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order,updated_at)
      values(?,?,?,?,?,?,?,?,?,?,?,?)""",[
        (game_id,row["achievement_id"],row["name"],row["condition"],row["version"],row["category"],
         int(row["reward"] or 0),1 if row["hidden"] else 0,row["tags_json"],row["source"],int(row["source_order"] or 0),stamp)
        for row in rows
    ])
    db.executemany("""insert into game_catalog_source_records(
      game_id,achievement_id,official_source_id,category_id,group_id,group_name,progress_value,level,next_link,reward_id,
      primary_source_id,secondary_source_id,source_ref,raw_json,provenance_json,updated_at)
      values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",[
        (game_id,row["achievement_id"],str(row.get("official_source_id") or row["achievement_id"]),str(row.get("category_id") or ""),str(row.get("group_id") or ""),str(row.get("group_name") or ""),
         int(row.get("progress_value") or 0),int(row.get("level") or 0),str(row.get("next_link") or ""),str(row.get("reward_id") or ""),
         str(row.get("primary_source_id") or ""),str(row.get("secondary_source_id") or ""),str(row.get("source_ref") or ""),
         str(row.get("raw_json") or "{}"),str(row.get("provenance_json") or "{}"),stamp)
        for row in rows
    ])
    if game_id=="wuwa":
        db.execute("delete from achievement_source_ids where game_id='wuwa'")
        db.execute("delete from achievement_identities where game_id='wuwa'")
        db.execute("delete from achievement_id_aliases where game_id='wuwa'")
        return
    identity_rows=[]; source_rows=[]
    for row in rows:
        internal_id=str(row.get("achievement_id") or "")
        source_name=str(row.get("primary_source_id") or row.get("source") or "official")
        official_id=str(row.get("official_source_id") or internal_id)
        if not internal_id:
            continue
        identity_rows.append((game_id,internal_id,source_name,official_id,stamp,stamp))
        source_rows.append((game_id,source_name,official_id,internal_id,1,str(row.get("identity_match_status") or "confirmed"),float(row.get("identity_match_confidence") or 1.0),str(row.get("identity_match_basis") or "catalog_apply"),stamp,stamp))
    db.executemany("""insert into achievement_identities(game_id,internal_id,display_source_name,display_source_id,created_at,updated_at)
      values(?,?,?,?,?,?) on conflict(game_id,internal_id) do update set display_source_name=excluded.display_source_name,display_source_id=excluded.display_source_id,updated_at=excluded.updated_at""",identity_rows)
    _upsert_achievement_source_rows(db,source_rows)


SYNC_ROLLBACK_TABLES = (
    ("game_catalog_items", "achievement_id"),
    ("achievement_identities", "internal_id"),
    ("achievement_source_ids", "source_name,source_id"),
    ("game_catalog_source_records", "achievement_id"),
    ("game_achievement_overrides", "achievement_id"),
)


def sync_rollback_context(game_id: str) -> dict[str, Any]:
    payload: dict[str, Any] = {"game_id": game_id, "tables": {}}
    with connect_db() as db:
        for table, order_by in SYNC_ROLLBACK_TABLES:
            payload["tables"][table] = [dict(row) for row in db.execute(f"select * from {table} where game_id=? order by {order_by}", (game_id,)).fetchall()]
    catalog_path = game_catalog_file(game_id)
    cache_path = source_cache_path(DATA_DIR, game_id)
    payload["catalog_sha256"] = hashlib.sha256(catalog_path.read_bytes()).hexdigest() if catalog_path.exists() else ""
    payload["source_cache_sha256"] = hashlib.sha256(cache_path.read_bytes()).hexdigest() if cache_path.exists() else ""
    payload["state_hash"] = hashlib.sha256(json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"), default=str).encode("utf-8")).hexdigest()
    return payload


def restore_sync_database_scope(backup: Path, game_id: str) -> dict[str, int]:
    source = sqlite3.connect(backup); source.row_factory = sqlite3.Row
    restored: dict[str, int] = {}
    try:
        snapshots: dict[str, tuple[list[str], list[sqlite3.Row]]] = {}
        for table, order_by in SYNC_ROLLBACK_TABLES:
            columns = [str(row[1]) for row in source.execute(f"pragma table_info({table})").fetchall()]
            rows = source.execute(f"select * from {table} where game_id=? order by {order_by}", (game_id,)).fetchall()
            snapshots[table] = (columns, rows)
        with connect_db() as db:
            db.execute("pragma defer_foreign_keys=on"); db.execute("begin immediate")
            for table in ("achievement_source_ids", "game_catalog_source_records", "game_achievement_overrides", "achievement_identities", "game_catalog_items"):
                db.execute(f"delete from {table} where game_id=?", (game_id,))
            for table in ("game_catalog_items", "achievement_identities", "achievement_source_ids", "game_catalog_source_records", "game_achievement_overrides"):
                columns, rows = snapshots[table]
                if rows:
                    placeholders = ",".join("?" for _ in columns)
                    db.executemany(f"insert into {table}({','.join(columns)}) values({placeholders})", [tuple(row[column] for column in columns) for row in rows])
                restored[table] = len(rows)
    finally:
        source.close()
    return restored


@app.get("/api/games/{game_id}/admin/source-status")
def admin_game_source_status(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    policy=get_source_policy(game_id); path=game_catalog_file(game_id)
    with connect_db() as db:
        catalog_count=int(db.execute("select count(*) c from game_catalog_items where game_id=?",(game_id,)).fetchone()["c"] or 0)
        histories=[dict(row) for row in db.execute("select id,source_id,source_mode,source_hash,summary_json,backup_name,snapshot_dir,pre_state_hash,post_state_hash,status,created_at,rolled_back_at,rollback_reason from source_sync_history where game_id=? order by created_at desc limit 20",(game_id,)).fetchall()]
        active_previews=int(db.execute("select count(*) c from game_sync_previews where game_id=? and expires_at>?",(game_id,now())).fetchone()["c"] or 0)
        identity_count=int(db.execute("select count(*) c from achievement_identities where game_id=?",(game_id,)).fetchone()["c"] or 0)
        primary_mapping_count=int(db.execute("select count(*) c from achievement_source_ids where game_id=? and is_primary=1",(game_id,)).fetchone()["c"] or 0)
        legacy_mapping_count=int(db.execute("select count(*) c from achievement_source_ids where game_id=? and source_name='legacy_catalog'",(game_id,)).fetchone()["c"] or 0)
        review_mapping_count=int(db.execute("select count(*) c from achievement_source_ids where game_id=? and match_status in ('needs_review','ambiguous_new_identity')",(game_id,)).fetchone()["c"] or 0)
        orphan_progress=int(db.execute("""select count(*) c from game_progress p left join game_catalog_items c
          on c.game_id=p.game_id and c.achievement_id=p.achievement_id
          where p.game_id=? and c.achievement_id is null""",(game_id,)).fetchone()["c"] or 0)
    for row in histories:
        row["summary"]=json.loads(row.pop("summary_json") or "{}")
    sources=[]
    for role in ("primary","secondary","fallback"):
        source=policy.get(role)
        if not isinstance(source,dict): continue
        sources.append({**source,"role":role,"configured":True,"last_test":None})
    identity_status=(
        {"mode":"official_id_primary_key","official_catalog_ids":catalog_count,"orphan_progress":orphan_progress}
        if game_id=="wuwa" else
        {"mode":"identity_bridge","internal_identities":identity_count,"primary_source_mappings":primary_mapping_count,"legacy_alias_mappings":legacy_mapping_count,"needs_review":review_mapping_count,"orphan_progress":orphan_progress}
    )
    return {"ok":True,"game_id":game_id,"game_name":game_display_name(game_id),"policy":policy,"sources":sources,
            "catalog":{"path":str(path),"exists":path.exists(),"size_bytes":path.stat().st_size if path.exists() else 0,"count":catalog_count},
            "identity_model":identity_status,
            "identity_bridge":identity_status,
            "active_previews":active_previews,"history":histories}


@app.post("/api/games/{game_id}/admin/source-test/{role}")
def admin_test_game_source(game_id: str, role: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    if role not in {"primary","secondary","fallback"}: raise HTTPException(status_code=400,detail="來源角色無效。")
    source=get_source_policy(game_id).get(role)
    if not isinstance(source,dict): raise HTTPException(status_code=404,detail="此遊戲未設定該來源。")
    result=test_repository_source(game_id,role,timeout=15)
    status="ok" if result.get("ok") else "error"
    log_admin_action(admin["id"],"test_official_source",details=f"game={game_id}; role={role}; status={status}; http={result.get('http_status')}; elapsed_ms={result.get('elapsed_ms')}",category="sync",status="success" if status=="ok" else "failed",game_id=game_id,target_type="source",target_id=str(source.get("id") or role),summary="測試遊戲資料來源",metadata={"role":role,"source":source,"diagnostics":result.get("diagnostics") or {},"http_status":result.get("http_status"),"elapsed_ms":result.get("elapsed_ms")},error_message=str(result.get("error") or ""),actor_ip=client_ip(request))
    return {"game_id":game_id,"role":role,"source":source,**result}



def _sync_preview_response(record: sqlite3.Row) -> dict[str, Any]:
    diff=json.loads(record["diff_json"] or "{}")
    metadata=json.loads(record["metadata_json"] or "{}")
    candidate_rows=json.loads(record["candidate_json"] or "[]")
    response_changes=list(diff.get("changes") or [])[:200]
    return {
        "ok":True,
        "preview_id":str(record["id"]),
        "game_id":str(record["game_id"]),
        "game_name":game_display_name(str(record["game_id"])),
        "created_at":int(record["created_at"] or 0),
        "expires_at":int(record["expires_at"] or 0),
        "metadata":metadata,
        "summary":diff.get("summary") or {},
        "current_count":int(diff.get("current_count") or 0),
        "candidate_count":int(diff.get("candidate_count") or len(candidate_rows)),
        "changes":response_changes,
        "changes_total":len(diff.get("changes") or []),
        "default_selections":build_default_sync_selections(diff.get("changes") or []),
        "changes_offset":0,
        "changes_limit":len(response_changes),
        "changes_truncated":len(diff.get("changes") or [])>len(response_changes),
        "restored":True,
    }


def _sync_preview_pipeline_version(record: sqlite3.Row | dict[str,Any]) -> str:
    try:
        metadata=json.loads(record["metadata_json"] or "{}")
    except Exception:
        return ""
    return str(metadata.get("pipeline_version") or metadata.get("source_architecture_version") or "").strip()


def _reject_stale_sync_preview(db: sqlite3.Connection, record: sqlite3.Row | dict[str,Any]) -> None:
    preview_version=_sync_preview_pipeline_version(record)
    if preview_version==SOURCE_PIPELINE_VERSION:
        return
    preview_id=str(record["id"] or "")
    if preview_id:
        db.execute("delete from game_sync_previews where id=?",(preview_id,))
        db.commit()
    shown=preview_version or "未記錄版本"
    raise HTTPException(
        status_code=409,
        detail=f"此差異預覽由舊版來源管線建立（{shown}），已自動作廢。請按「抓取並預覽差異」重新建立最新版預覽。",
    )


def _sync_preview_apply_block_reason(metadata: dict[str, Any]) -> str:
    if not isinstance(metadata, dict):
        return ""
    if metadata.get("apply_blocked"):
        return str(metadata.get("apply_block_reason") or "此預覽只供來源診斷，不能套用正式資料。")
    if metadata.get("diagnostic_preview"):
        return str(metadata.get("apply_block_reason") or "此預覽只供來源診斷，不能套用正式資料。")
    return ""


def _owned_sync_preview(game_id: str, preview_id: str, admin_id: str) -> sqlite3.Row:
    with connect_db() as db:
        record=db.execute(
            "select * from game_sync_previews where id=? and game_id=? and admin_user_id=?",
            (preview_id,game_id,admin_id),
        ).fetchone()
        if not record:
            raise HTTPException(status_code=404,detail="找不到同步預覽。")
        if int(record["expires_at"] or 0)<=now():
            db.execute("delete from game_sync_previews where id=?",(preview_id,))
            raise HTTPException(status_code=410,detail="同步預覽已過期，請重新抓取資料。")
        _reject_stale_sync_preview(db,record)
        renewed=now()+SYNC_PREVIEW_RETENTION_SECONDS
        db.execute("update game_sync_previews set expires_at=? where id=?",(renewed,preview_id))
        values=dict(record)
        values["expires_at"]=renewed
    return values  # type: ignore[return-value]


@app.get("/api/games/{game_id}/admin/official-achievements/preview/active")
def admin_active_sync_preview(game_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    with connect_db() as db:
        db.execute(
            "delete from game_sync_previews where admin_user_id=? and game_id=? and expires_at<=?",
            (admin["id"],game_id,now()),
        )
        record=db.execute(
            "select * from game_sync_previews where game_id=? and admin_user_id=? order by created_at desc limit 1",
            (game_id,admin["id"]),
        ).fetchone()
    if not record:
        raise HTTPException(status_code=404,detail="尚無可恢復的同步預覽。")
    return _sync_preview_response(_owned_sync_preview(game_id,str(record["id"]),str(admin["id"])))


@app.get("/api/games/{game_id}/admin/official-achievements/preview/{preview_id}")
def admin_get_sync_preview(game_id: str, preview_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    return _sync_preview_response(_owned_sync_preview(game_id,preview_id,str(admin["id"])))


@app.get("/api/games/{game_id}/admin/official-achievements/preview/{preview_id}/changes")
def admin_sync_preview_changes(
    game_id: str, preview_id: str, request: Request, offset: int=0, limit: int=200,
    risk: str="", change_type: str="", search: str="", category: str="", group_name: str="",
    version: str="", tag: str="", reward_min: int | None=None, reward_max: int | None=None,
    hidden: str="", source: str="", progress: str="", relation: str="", conflict: str="",
    changed_field: str="", template_format: str="",
):
    game_id=require_extra_game(game_id); admin=require_admin(request); offset=max(0,int(offset)); limit=max(1,min(500,int(limit)))
    record=_owned_sync_preview(game_id,preview_id,str(admin["id"]))
    with connect_db() as db:
        relation_types: dict[str,set[str]]={}
        for relation_row in db.execute(
            "select achievement_id,relation_type from game_achievement_choice_groups where game_id=?",(game_id,)
        ).fetchall():
            achievement_id=str(relation_row["achievement_id"] or "")
            relation_type=str(relation_row["relation_type"] or "")
            if achievement_id and relation_type:
                relation_types.setdefault(achievement_id,set()).add(relation_type)
    diff=json.loads(record["diff_json"] or "{}")
    rows=[dict(row) for row in diff.get("changes") or []]
    for row in rows:
        row["relation_types"]=sorted(value for value in relation_types.get(str(row.get("achievement_id") or ""),set()) if value)
    rows.sort(key=lambda row: sync_change_sort_key(game_id,row))
    if risk: rows=[row for row in rows if row.get("risk")==risk]
    if change_type: rows=[row for row in rows if row.get("type")==change_type]
    if search:
        key=search.casefold(); rows=[row for row in rows if key in str(row.get("achievement_id") or "").casefold() or key in str(row.get("name") or "").casefold()]
    def contains(value: Any, query: str) -> bool:
        return not query or query.casefold() in str(value or "").casefold()
    if category: rows=[row for row in rows if contains(row.get("category"),category)]
    if group_name: rows=[row for row in rows if contains(row.get("group_name"),group_name)]
    if version: rows=[row for row in rows if contains(row.get("version"),version)]
    if tag:
        rows=[row for row in rows if contains(row.get("tags_json"),tag)]
    if reward_min is not None: rows=[row for row in rows if int(row.get("reward") or 0)>=int(reward_min)]
    if reward_max is not None: rows=[row for row in rows if int(row.get("reward") or 0)<=int(reward_max)]
    if hidden in {"true","1","yes"}: rows=[row for row in rows if bool(row.get("hidden"))]
    elif hidden in {"false","0","no"}: rows=[row for row in rows if not bool(row.get("hidden"))]
    if source: rows=[row for row in rows if contains(row.get("source"),source)]
    if progress == "with": rows=[row for row in rows if int(row.get("progress_count") or 0)>0]
    elif progress == "without": rows=[row for row in rows if int(row.get("progress_count") or 0)==0]
    if relation == "with": rows=[row for row in rows if int(row.get("relation_count") or 0)>0]
    elif relation == "without": rows=[row for row in rows if int(row.get("relation_count") or 0)==0]
    elif relation in {"stage","exclusive"}: rows=[row for row in rows if relation in set(row.get("relation_types") or [])]
    if conflict == "with": rows=[row for row in rows if bool(row.get("source_conflicts"))]
    elif conflict == "without": rows=[row for row in rows if not bool(row.get("source_conflicts"))]
    if changed_field: rows=[row for row in rows if changed_field in set(row.get("changed_fields") or [])]
    if template_format == "only": rows=[row for row in rows if bool(row.get("template_format_only"))]
    elif template_format == "exclude": rows=[row for row in rows if not bool(row.get("template_format_only"))]
    return {"ok":True,"preview_id":preview_id,"total":len(rows),"offset":offset,"limit":limit,"changes":rows[offset:offset+limit]}


@app.delete("/api/games/{game_id}/admin/official-achievements/preview/{preview_id}")
def admin_cancel_sync_preview(game_id: str, preview_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    with connect_db() as db: deleted=db.execute("delete from game_sync_previews where id=? and game_id=? and admin_user_id=?",(preview_id,game_id,admin["id"])).rowcount
    if not deleted: raise HTTPException(status_code=404,detail="找不到同步預覽。")
    log_admin_action(admin["id"],"cancel_official_preview",details=f"game={game_id}; preview={preview_id}",category="sync",game_id=game_id,target_type="sync_preview",target_id=preview_id,summary="取消官方成就同步預覽",actor_ip=client_ip(request))
    return {"ok":True}


@app.post("/api/games/{game_id}/admin/official-achievements/preview")
def preview_game_official_achievements(game_id: str, request: Request):
    game_id=require_extra_game(game_id)
    admin=require_admin(request)
    try:
        rows,metadata,source_payload=_prepare_game_sync_candidate(game_id)
    except Exception as exc:
        rows,metadata,source_payload=_safe_degraded_sync_candidate(game_id, exc)
    # The preview row must always be stamped by the same compatibility version
    # that the loader validates.  Individual source adapters may keep their own
    # source_architecture_version, but must never control preview compatibility.
    metadata={
        **metadata,
        "pipeline_version":SOURCE_PIPELINE_VERSION,
        "adapter_id":source_adapter_id(game_id),
    }
    with connect_db() as db:
        current=_current_official_catalog_rows(db,game_id)
    metadata={**metadata,"current_fingerprint":_catalog_rows_fingerprint(current)}
    diff=_catalog_sync_diff(current,rows,game_id=game_id,metadata=metadata)
    preview_id=secrets.token_urlsafe(24)
    created=now(); expires=created+SYNC_PREVIEW_RETENTION_SECONDS
    with connect_db() as db:
        db.execute("delete from game_sync_previews where admin_user_id=? and game_id=?",(admin["id"],game_id))
        db.execute("""insert into game_sync_previews(id,game_id,admin_user_id,candidate_json,source_payload_json,metadata_json,diff_json,created_at,expires_at)
          values(?,?,?,?,?,?,?,?,?)""",(
            preview_id,game_id,admin["id"],json.dumps(rows,ensure_ascii=False,separators=(",",":")),
            json.dumps(source_payload,ensure_ascii=False,separators=(",",":")) if source_payload is not None else "",
            json.dumps(metadata,ensure_ascii=False,separators=(",",":")),json.dumps(diff,ensure_ascii=False,separators=(",",":")),created,expires,
        ))
    response_changes=diff["changes"][:200]
    log_admin_action(admin["id"],"preview_official_achievements",details=f"game={game_id}; changes={diff['summary']['total_changes']}; candidate={len(rows)}",category="sync",game_id=game_id,target_type="sync_preview",target_id=preview_id,summary="建立官方成就同步預覽",after=diff["summary"],metadata=metadata,actor_ip=client_ip(request))
    return {"ok":True,"preview_id":preview_id,"game_id":game_id,"game_name":game_display_name(game_id),
            "expires_at":expires,"metadata":metadata,"summary":diff["summary"],"current_count":diff["current_count"],
            "candidate_count":diff["candidate_count"],"changes":response_changes,"changes_total":len(diff["changes"]),
            "default_selections":build_default_sync_selections(diff["changes"]),
            "changes_offset":0,"changes_limit":len(response_changes),"changes_truncated":len(diff["changes"])>len(response_changes)}


def _verify_sync_apply_result(game_id: str, rows: list[dict[str,Any]], applied_changes: list[dict[str,Any]]) -> dict[str,Any]:
    expected={str(row.get("achievement_id") or ""):row for row in rows}
    with connect_db() as db:
        database={str(row["achievement_id"]):dict(row) for row in db.execute(
            "select achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order from game_catalog_items where game_id=?",(game_id,)
        ).fetchall()}
    payload=json.loads(game_catalog_file(game_id).read_text(encoding="utf-8-sig"))
    json_rows={str(row.get("id") or ""):row for row in payload.get("items") or []}
    failures=[]
    field_map={"tags_json":"tags","source_order":"sourceOrder"}
    for change in applied_changes:
        achievement_id=str(change.get("achievement_id") or "")
        fields=list(change.get("fields") or [])
        if "__remove__" in fields:
            if achievement_id in database or achievement_id in json_rows:
                failures.append({"achievement_id":achievement_id,"field":"__remove__","reason":"removed_row_still_exists"})
            continue
        expected_row=expected.get(achievement_id) or {}
        db_row=database.get(achievement_id)
        json_row=json_rows.get(achievement_id)
        if db_row is None or json_row is None:
            failures.append({"achievement_id":achievement_id,"field":"__row__","reason":"applied_row_missing"})
            continue
        for field in fields:
            expected_value=expected_row.get(field)
            db_value=db_row.get(field)
            json_value=json_row.get(field_map.get(field,field))
            if field=="hidden":
                expected_value=bool(expected_value); db_value=bool(db_value); json_value=bool(json_value)
            elif field=="reward" or field=="source_order":
                expected_value=int(expected_value or 0); db_value=int(db_value or 0); json_value=int(json_value or 0)
            elif field=="tags_json":
                def normalize_tags(value: Any) -> list[str]:
                    if isinstance(value, list):
                        parsed=value
                    elif isinstance(value, (tuple, set)):
                        parsed=list(value)
                    else:
                        try:
                            parsed=json.loads(str(value or "[]"))
                        except Exception:
                            parsed=[]
                    return [str(item)[:80] for item in parsed if str(item).strip()] if isinstance(parsed,list) else []
                expected_value=normalize_tags(expected_value)
                db_value=normalize_tags(db_value)
                json_value=normalize_tags(json_value)
            if db_value!=expected_value or json_value!=expected_value:
                failures.append({"achievement_id":achievement_id,"field":field,"expected":expected_value,"database":db_value,"json":json_value})
    if failures:
        raise RuntimeError("同步套用後逐欄驗證失敗："+json.dumps(failures[:10],ensure_ascii=False))
    return {"ok":True,"checked_changes":len(applied_changes),"checked_fields":sum(len(row.get("fields") or []) for row in applied_changes),"failures":0}


def _write_sync_admin_overrides(db: sqlite3.Connection, game_id: str, rows: list[dict[str,Any]], applied_changes: list[dict[str,Any]], admin_id: str) -> int:
    by_id={str(row.get("achievement_id") or ""):row for row in rows}
    count=0; stamp=now()
    for change in applied_changes:
        if change.get("action")!="admin_override":
            continue
        row=by_id.get(str(change.get("achievement_id") or ""))
        if not row:
            continue
        db.execute("""insert into game_achievement_overrides(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,is_deleted,source,updated_by,updated_at)
          values(?,?,?,?,?,?,?,?,?,0,'sync-admin-override',?,?)
          on conflict(game_id,achievement_id) do update set name=excluded.name,condition=excluded.condition,version=excluded.version,category=excluded.category,reward=excluded.reward,hidden=excluded.hidden,tags_json=excluded.tags_json,is_deleted=0,source=excluded.source,updated_by=excluded.updated_by,updated_at=excluded.updated_at""",(
            game_id,row["achievement_id"],row["name"],row["condition"],row["version"],row["category"],int(row["reward"] or 0),1 if row["hidden"] else 0,row["tags_json"],admin_id,stamp
        ))
        count+=1
    return count


@app.post("/api/games/{game_id}/admin/official-achievements/apply")
@high_risk_operation
def apply_game_official_achievements(game_id: str, body: AdminSyncApplyPayload, request: Request):
    game_id=require_extra_game(game_id)
    admin=require_admin(request)
    with connect_db() as db:
        record=db.execute(
            "select * from game_sync_previews where id=? and game_id=? and admin_user_id=?",
            (body.preview_id,game_id,admin["id"]),
        ).fetchone()
        if record and int(record["expires_at"] or 0)>now():
            _reject_stale_sync_preview(db,record)
    if not record:
        raise HTTPException(status_code=404,detail="找不到同步預覽，請重新抓取資料。")
    if int(record["expires_at"] or 0)<=now():
        with connect_db() as db: db.execute("delete from game_sync_previews where id=?",(body.preview_id,))
        raise HTTPException(status_code=410,detail="同步預覽已過期，請重新抓取資料。")
    candidate_rows=json.loads(record["candidate_json"])
    metadata=json.loads(record["metadata_json"] or "{}")
    apply_block_reason=_sync_preview_apply_block_reason(metadata)
    if apply_block_reason:
        raise HTTPException(status_code=409,detail=apply_block_reason)
    configured_minimum=max(1,int((get_game_config(game_id) or {}).get("minimumCatalogCount") or 1))
    minimum_count=max(configured_minimum,1000 if game_id=="wuwa" else configured_minimum)
    if len(candidate_rows)<minimum_count:
        with connect_db() as db:
            db.execute("delete from game_sync_previews where id=? and admin_user_id=?",(body.preview_id,admin["id"]))
        raise HTTPException(
            status_code=409,
            detail=f"同步預覽候選只有 {len(candidate_rows)} 筆，低於 {game_display_name(game_id)} 安全下限 {minimum_count} 筆；預覽已自動作廢，請重新抓取資料。",
        )
    diff=json.loads(record["diff_json"] or "{}")
    source_payload=json.loads(record["source_payload_json"]) if record["source_payload_json"] else None
    with connect_db() as db:
        current_rows=_current_official_catalog_rows(db,game_id)
    if metadata.get("current_fingerprint") and metadata["current_fingerprint"]!=_catalog_rows_fingerprint(current_rows):
        raise HTTPException(status_code=409,detail="正式成就目錄已在預覽後發生變更，請重新建立同步預覽。")
    all_changes={str(change.get("change_id")):change for change in diff.get("changes") or []}
    if body.selected_change_ids is None:
        selected_ids={change_id for change_id,change in all_changes.items() if change.get("default_selected") and change.get("risk")!="blocked"}
    else:
        selected_ids=set(validate_ids(body.selected_change_ids))
        unknown=selected_ids-set(all_changes)
        if unknown:
            raise HTTPException(status_code=400,detail=f"選取的差異項目不存在：{', '.join(sorted(unknown)[:5])}")
    blocked=[change_id for change_id in selected_ids if all_changes[change_id].get("risk")=="blocked"]
    if blocked: raise HTTPException(status_code=409,detail="選取內容包含被安全規則阻擋的項目，請先處理異常。")
    normalized_decisions={str(key):dict(value) for key,value in (body.decisions or {}).items() if isinstance(value,dict)}
    for decision in normalized_decisions.values():
        if not str(decision.get("reason") or "").strip() and str(body.reason or "").strip():
            decision["reason"]=str(body.reason).strip()
    try:
        rows,applied_summary=apply_sync_decisions(current_rows,candidate_rows,list(all_changes.values()),selected_ids,normalized_decisions,game_id=game_id)
    except ValueError as exc:
        raise HTTPException(status_code=400,detail=str(exc)) from exc
    rows=_merge_applied_source_metadata(rows,candidate_rows,current_rows,metadata)
    with connect_db() as db:
        rows=_apply_managed_category_aliases(db,game_id,rows)
    applied_summary["unchanged"]=int((diff.get("summary") or {}).get("unchanged") or 0)
    backup=create_database_backup()
    pre_context=sync_rollback_context(game_id)
    snapshot_dir=ROOT/"backups"/f"sync-{time.strftime('%Y%m%d-%H%M%S')}-{body.preview_id[-8:]}"
    snapshot_dir.mkdir(parents=True,exist_ok=True)
    affected_files=[game_catalog_file(game_id),source_cache_path(DATA_DIR,game_id)]
    snapshots={path:(path.read_bytes() if path.exists() else None) for path in affected_files}
    if snapshots[affected_files[0]] is not None:
        (snapshot_dir/"catalog.json").write_bytes(snapshots[affected_files[0]])
    if snapshots[affected_files[1]] is not None:
        (snapshot_dir/"source-cache.bin").write_bytes(snapshots[affected_files[1]])
    (snapshot_dir/"snapshot.json").write_text(json.dumps({"catalog_exists":snapshots[affected_files[0]] is not None,"source_cache_exists":snapshots[affected_files[1]] is not None},ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    try:
        _write_game_catalog_candidate(game_id,rows,metadata)
        with connect_db() as db:
            db.execute("begin immediate")
            _replace_official_catalog_rows(db,game_id,rows)
            applied_summary["admin_overrides_written"]=_write_sync_admin_overrides(db,game_id,rows,list(applied_summary.get("applied_changes") or []),admin["id"])
            db.execute("delete from game_sync_previews where id=?",(body.preview_id,))
            if game_id=="wuwa":
                _verify_wuwa_shared_model(db)
        applied_summary["verification"]=_verify_sync_apply_result(game_id,rows,list(applied_summary.get("applied_changes") or []))
        post_context=sync_rollback_context(game_id)
    except Exception as exc:
        rollback_errors=[]
        for path,data in snapshots.items():
            try:
                if data is None:
                    path.unlink(missing_ok=True)
                else:
                    path.parent.mkdir(parents=True,exist_ok=True)
                    path.write_bytes(data)
            except Exception as rollback_exc:
                rollback_errors.append(f"file:{path.name}:{rollback_exc}")
        try:
            _restore_governance_database_scope(backup,game_id)
        except Exception as rollback_exc:
            rollback_errors.append(f"database_scope:{rollback_exc}")
        if rollback_errors:
            log_admin_action(admin["id"],"apply_official_achievements_recovery_failed",category="sync",status="recovery_failed",game_id=game_id,target_type="sync_preview",target_id=body.preview_id,summary="官方同步失敗且自動回復未完整成功",error_message=str(exc),metadata={"rollback_errors":rollback_errors},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
            raise HTTPException(status_code=500,detail=f"同步套用失敗，且自動回復未完整成功：{exc}；回復錯誤：{'；'.join(rollback_errors)}") from exc
        _remove_temporary_tree(snapshot_dir)
        if game_id=="wuwa" and isinstance(exc,RuntimeError):
            raise HTTPException(
                status_code=409,
                detail=f"鳴潮同步安全驗證已阻擋套用：{exc}。資料已完整回復，請重新抓取最新版差異預覽。",
            ) from exc
        raise HTTPException(status_code=409,detail=f"同步套用失敗，資料已完整回復：{exc}") from exc
    bump_game_live_scope(game_id,"catalog")
    summary=applied_summary
    history_id=str(uuid.uuid4())
    with connect_db() as db:
        db.execute("insert into source_sync_history(id,game_id,preview_id,source_id,source_mode,source_hash,summary_json,backup_name,snapshot_dir,pre_state_hash,post_state_hash,status,actor_user_id,created_at) values(?,?,?,?,?,?,?,?,?,?,?,?,?,?)",(
            history_id,game_id,body.preview_id,str((metadata.get("primary_source") or {}).get("id") or metadata.get("source") or ""),str(metadata.get("source_mode") or ""),
            hashlib.sha256(json.dumps(source_payload,ensure_ascii=False,sort_keys=True,default=str).encode("utf-8")).hexdigest() if source_payload is not None else "",
            json.dumps(summary,ensure_ascii=False,separators=(",",":")),backup.name,str(snapshot_dir),pre_context["state_hash"],post_context["state_hash"],"success",admin["id"],now(),
        ))
        for decision in summary.get("recorded_decisions") or []:
            db.execute("""insert into source_sync_decisions(id,history_id,preview_id,game_id,change_id,achievement_id,action,fields_json,reason,data_changed,actor_user_id,created_at)
              values(?,?,?,?,?,?,?,?,?,?,?,?)""",(
                str(uuid.uuid4()),history_id,body.preview_id,game_id,str(decision.get("change_id") or ""),str(decision.get("achievement_id") or ""),
                str(decision.get("action") or ""),json.dumps(decision.get("fields") or [],ensure_ascii=False,separators=(",",":")),
                str(decision.get("reason") or body.reason or ""),1 if decision.get("data_changed") else 0,admin["id"],now(),
            ))
    log_admin_action(admin["id"],"apply_official_achievements",details=f"game={game_id}; added={summary.get('added',0)}; removed={summary.get('removed',0)}; modified={summary.get('modified',0)}; selected={summary.get('selected',0)}; backup={backup.name}",category="sync",game_id=game_id,target_type="sync_preview",target_id=body.preview_id,summary="套用官方成就同步差異",before=diff.get("summary"),after=summary,metadata={"decisions":normalized_decisions,"reason":body.reason,"history_id":history_id,"source":metadata,"verification":summary.get("verification")},backup_name=backup.name,actor_ip=client_ip(request),locked=True)

    return {"ok":True,"game_id":game_id,"game_name":game_display_name(game_id),"count":len(rows),"summary":summary,"backup":backup.name,"history_id":history_id}


@app.post("/api/games/{game_id}/admin/official-achievements/history/{history_id}/rollback")
@high_risk_operation
def rollback_game_official_sync(game_id: str, history_id: str, body: SyncHistoryRollbackPayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request)
    with connect_db() as db:
        history=db.execute("select * from source_sync_history where id=? and game_id=?",(history_id,game_id)).fetchone()
    if not history or history["status"]!="success" or history["rolled_back_at"]:
        raise HTTPException(status_code=409,detail="此同步紀錄目前不可復原。")
    current_context=sync_rollback_context(game_id)
    if not history["post_state_hash"] or current_context["state_hash"]!=str(history["post_state_hash"]):
        raise HTTPException(status_code=409,detail="同步完成後正式目錄或來源對應已再次修改，為避免覆蓋新資料已阻擋復原。")
    backup=ROOT/"backups"/str(history["backup_name"] or "")
    snapshot_dir=Path(str(history["snapshot_dir"] or ""))
    if not backup.exists() or not snapshot_dir.exists() or not (snapshot_dir/"snapshot.json").exists():
        raise HTTPException(status_code=409,detail="同步復原需要的資料庫備份或檔案快照不存在。")
    metadata=_json_object((snapshot_dir/"snapshot.json").read_text(encoding="utf-8"),{})
    catalog_path=game_catalog_file(game_id);cache_path=source_cache_path(DATA_DIR,game_id)
    current_backup=create_database_backup()
    current_catalog=catalog_path.read_bytes() if catalog_path.exists() else None
    current_cache=cache_path.read_bytes() if cache_path.exists() else None
    rollback_errors=[]
    try:
        restored_rows=restore_sync_database_scope(backup,game_id)
        if metadata.get("catalog_exists"):
            catalog_path.parent.mkdir(parents=True,exist_ok=True);shutil.copy2(snapshot_dir/"catalog.json",catalog_path)
        else:
            catalog_path.unlink(missing_ok=True)
        if metadata.get("source_cache_exists"):
            cache_path.parent.mkdir(parents=True,exist_ok=True);shutil.copy2(snapshot_dir/"source-cache.bin",cache_path)
        else:
            cache_path.unlink(missing_ok=True)
        restored_context=sync_rollback_context(game_id)
        if restored_context["state_hash"]!=str(history["pre_state_hash"] or ""):
            raise RuntimeError("sync_history_restore_hash_mismatch")
        with connect_db() as db:
            db.execute("update source_sync_history set status='rolled_back',rolled_back_at=?,rolled_back_by=?,rollback_reason=? where id=?",(now(),admin["id"],body.reason.strip(),history_id))
    except Exception as exc:
        try:
            restore_sync_database_scope(current_backup,game_id)
        except Exception as recovery_exc:
            rollback_errors.append(f"database:{recovery_exc}")
        for path,data,label in ((catalog_path,current_catalog,"catalog"),(cache_path,current_cache,"source_cache")):
            try:
                if data is None:path.unlink(missing_ok=True)
                else:path.parent.mkdir(parents=True,exist_ok=True);path.write_bytes(data)
            except Exception as recovery_exc:
                rollback_errors.append(f"{label}:{recovery_exc}")
        if rollback_errors:
            raise HTTPException(status_code=500,detail=f"同步復原失敗，且安全回復未完整成功：{exc}；{'；'.join(rollback_errors)}") from exc
        raise HTTPException(status_code=409,detail=f"同步復原失敗，資料已回到復原前狀態：{exc}") from exc
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"rollback_official_sync",category="sync",game_id=game_id,target_type="source_sync_history",target_id=history_id,summary="復原官方成就同步批次",metadata={"reason":body.reason,"restored_rows":restored_rows},backup_name=current_backup.name,actor_ip=client_ip(request),locked=True)
    return {"ok":True,"history_id":history_id,"status":"rolled_back","safety_backup":current_backup.name,"restored_rows":restored_rows}


@app.post("/api/games/{game_id}/admin/official-achievements/sync")
def extra_game_sync_official_achievements(game_id: str, request: Request):
    game_id=require_extra_game(game_id)
    admin=require_admin(request)
    log_admin_action(
        admin["id"],
        "blocked_direct_sync",
        details=f"game={game_id}; reason=preview_required",
        category="sync",
        status="blocked",
        game_id=game_id,
        target_type="official_catalog",
        summary="阻擋舊版直接同步；必須先建立差異預覽",
        actor_ip=client_ip(request),
    )
    raise HTTPException(status_code=409,detail="直接覆蓋同步已停用，請使用「抓取並預覽差異」流程。")


@app.get("/api/games/{game_id}/official-zh-tw-achievements")
def extra_game_official_texts(game_id: str):
    game_id=require_extra_game(game_id)
    if game_id=="wuwa":
        return official_zh_tw_achievements()
    return {"ok":True,"records":[]}


@app.get("/api/games/{game_id}/live-state")
def extra_game_live_state(game_id: str):
    game_id=require_extra_game(game_id)
    with connect_db() as db:
        global_rows=db.execute("select scope,revision from live_revisions").fetchall()
        game_rows=db.execute("select scope,revision from game_live_revisions where game_id=?",(game_id,)).fetchall()
    revisions={r["scope"]:int(r["revision"] or 0) for r in global_rows}
    for row in game_rows:
        revisions[row["scope"]]=int(row["revision"] or 0)
    return {"ok":True,"revisions":revisions,"server_time":now()}


@app.get("/api/games/{game_id}/admin/overview")
def extra_game_admin_overview(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request); t=now()
    with connect_db() as db:
        summary=db.execute("""select count(*) total_users,
            sum(case when role='admin' then 1 else 0 end) admins,
            sum(case when email_verified=1 then 1 else 0 end) verified,
            sum(case when email_verified=0 then 1 else 0 end) unverified,
            sum(case when is_active=0 then 1 else 0 end) disabled from users""").fetchone()
        active_sessions=int(db.execute("select count(*) c from sessions where expires_at>?",(t,)).fetchone()["c"] or 0)
        progress_records=int(db.execute("select count(*) c from game_progress where game_id=?",(game_id,)).fetchone()["c"] or 0)
        users_with_progress=int(db.execute("select count(distinct user_id) c from game_progress where game_id=?",(game_id,)).fetchone()["c"] or 0)
        catalog_count=_effective_catalog_count(db,game_id)
    return {"ok":True,"overview":{
        "total_users":int(summary["total_users"] or 0),"admins":int(summary["admins"] or 0),
        "verified":int(summary["verified"] or 0),"unverified":int(summary["unverified"] or 0),
        "disabled":int(summary["disabled"] or 0),"active_sessions":active_sessions,
        "progress_records":progress_records,"users_with_progress":users_with_progress,
        "database_size_bytes":DB_FILE.stat().st_size if DB_FILE.exists() else 0,
        "official_cache_exists":True,"official_cache_size_bytes":0,"official_count":catalog_count,
        "official_synced_at":None,"mail_delivery":MAIL_DELIVERY,"public_base_url":PUBLIC_BASE_URL,"app_env":APP_ENV,
    }}


@app.get("/api/games/{game_id}/admin/users")
def extra_game_admin_users(game_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_site_owner(request)
    with connect_db() as db:
        rows=db.execute("""select u.id,u.email,u.username,u.email_verified,u.role,u.is_active,u.created_at,u.updated_at,
               u.last_login_at,u.last_login_ip,
               exists(select 1 from blocked_entries b where b.active=1 and b.kind='email' and b.value_key=u.email_key) is_blocked,
               (select b.id from blocked_entries b where b.active=1 and b.kind='email' and b.value_key=u.email_key order by b.created_at desc limit 1) block_id,
               count(distinct case
                   when rg.relation_type='exclusive' and rg.group_id is not null then 'exclusive:'||rg.group_id
                   when p.achievement_id is not null then 'achievement:'||p.achievement_id
               end) progress_count,
               count(distinct case when s.expires_at>? then s.token_hash end) session_count
        from users u
        left join game_progress p on p.user_id=u.id and p.game_id=?
        left join game_achievement_choice_groups rg
          on rg.game_id=p.game_id and rg.achievement_id=p.achievement_id
        left join sessions s on s.user_id=u.id
        group by u.id,u.email,u.username,u.email_verified,u.role,u.is_active,u.created_at,u.updated_at,u.last_login_at,u.last_login_ip
        order by case when u.role='admin' then 0 else 1 end,u.created_at,u.email_key""",(now(),game_id)).fetchall()
    return {"ok":True,"users":[{
        "id":r["id"],"email":r["email"],"username":r["username"] or "","email_verified":bool(r["email_verified"]),"role":r["role"],
        "is_active":bool(r["is_active"]),"created_at":r["created_at"],"updated_at":r["updated_at"],
        "last_login_at":r["last_login_at"],"last_login_ip":r["last_login_ip"] or "",
        "is_blocked":bool(r["is_blocked"]),"block_id":r["block_id"] or "",
        "progress_count":int(r["progress_count"] or 0),"session_count":int(r["session_count"] or 0),
        "is_self":r["id"]==admin["id"],"is_site_owner":is_site_owner_email(r["email"])
    } for r in rows]}


@app.delete("/api/games/{game_id}/admin/users/{user_id}/progress")
@high_risk_operation
def extra_game_admin_reset_user_progress(game_id: str, user_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_site_owner(request)
    with connect_db() as db:
        if not db.execute("select id from users where id=?",(user_id,)).fetchone():
            raise HTTPException(status_code=404,detail="找不到此帳號。")
        db.execute("delete from game_progress where game_id=? and user_id=?",(game_id,user_id))
    bump_game_live_scope(game_id,"stats")
    log_admin_action(admin["id"],"reset_progress",user_id,f"game={game_id}")
    return {"ok":True}


@app.post("/api/games/{game_id}/achievement-reports")
@high_risk_operation
def extra_game_create_report(game_id: str, body: AchievementReportCreate, request: Request):
    game_id=require_extra_game(game_id); user=require_user(request); rid=str(uuid.uuid4()); t=now()
    with connect_db() as db:
        aid=_resolve_effective_achievement_id(db,game_id,body.achievement_id)
        achievement=_effective_achievement_row(db,game_id,aid)
        achievement_name=str((achievement or {}).get("name") or body.achievement_name or aid).strip()
        db.execute("""insert into game_achievement_reports(id,game_id,user_id,achievement_id,achievement_name,report_type,message,status,admin_note,created_at,updated_at)
        values(?,?,?,?,?,?,?,'open','',?,?)""",
        (rid,game_id,user["id"],aid,achievement_name,body.report_type.strip(),body.message.strip(),t,t))
    bump_game_live_scope(game_id,"reports")
    return {"ok":True,"id":rid,"achievement_id":aid}


@app.get("/api/games/{game_id}/my/achievement-reports")
def extra_game_my_reports(game_id: str, request: Request):
    game_id=require_extra_game(game_id); user=require_user(request)
    with connect_db() as db:
        rows=db.execute("select * from game_achievement_reports where game_id=? and user_id=? order by created_at desc",(game_id,user["id"])).fetchall()
    return {"ok":True,"reports":[dict(r) for r in rows]}


@app.get("/api/games/{game_id}/admin/achievement-reports")
def extra_game_admin_reports(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        rows=db.execute("""select r.*,u.email from game_achievement_reports r left join users u on u.id=r.user_id
        where r.game_id=? order by case r.status when 'open' then 0 when 'reviewing' then 1 else 2 end,r.created_at desc""",(game_id,)).fetchall()
    return {"ok":True,"reports":[dict(r) for r in rows]}


@app.patch("/api/games/{game_id}/admin/achievement-reports/{report_id}")
@high_risk_operation
def extra_game_admin_update_report(game_id: str, report_id: str, body: AchievementReportUpdate, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request); status=body.status.strip().lower(); note=body.admin_note.strip()
    if status not in {"open","reviewing","resolved","rejected"}:
        raise HTTPException(status_code=400,detail="不支援的回報狀態。")
    with connect_db() as db:
        row=db.execute("select status,admin_note,user_id,achievement_name from game_achievement_reports where game_id=? and id=?",(game_id,report_id)).fetchone()
        if not row: raise HTTPException(status_code=404,detail="找不到回報。")
        if row["status"]==status and (row["admin_note"] or "")==note: return {"ok":True,"changed":False}
        db.execute("update game_achievement_reports set status=?,admin_note=?,updated_at=? where game_id=? and id=?",(status,note,now(),game_id,report_id))
    if row["user_id"]:
        status_name={"open":"待處理","reviewing":"處理中","resolved":"已解決","rejected":"不採納"}.get(status,status)
        body_text=f"你回報的成就「{row['achievement_name']}」狀態已更新為「{status_name}」。"+(f" 管理員說明：{note}" if note else "")
        create_notification("成就資料回報已更新",body_text,"report",f"/_projects/{game_id}/index.html",row["user_id"],admin["id"])
    bump_game_live_scope(game_id,"reports")
    return {"ok":True,"changed":True}


@app.delete("/api/games/{game_id}/admin/achievement-reports/{report_id}")
@high_risk_operation
def extra_game_admin_delete_report(game_id: str, report_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        deleted=db.execute("delete from game_achievement_reports where game_id=? and id=?",(game_id,report_id)).rowcount
    if not deleted: raise HTTPException(status_code=404,detail="找不到回報。")
    bump_game_live_scope(game_id,"reports")
    return {"ok":True}


@app.get("/api/games/{game_id}/achievement-customizations")
def extra_game_customizations(game_id: str):
    game_id=require_extra_game(game_id)
    with connect_db() as db:
        overrides=[serialize_game_override(r) for r in db.execute("select * from game_achievement_overrides where game_id=? order by updated_at",(game_id,)).fetchall()]
        permanently_deleted=[r["achievement_id"] for r in db.execute("select achievement_id from game_deleted_achievements where game_id=? order by deleted_at",(game_id,)).fetchall()]
    return {"ok":True,"overrides":overrides,"permanently_deleted":permanently_deleted}


@app.get("/api/games/{game_id}/admin/achievement-overrides")
def extra_game_admin_overrides(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        rows=db.execute("select * from game_achievement_overrides where game_id=? order by updated_at desc",(game_id,)).fetchall()
    return {"ok":True,"overrides":[serialize_game_override(r) for r in rows]}


@app.put("/api/games/{game_id}/admin/achievements/{achievement_id}")
@high_risk_operation
def extra_game_save_achievement(game_id: str, achievement_id: str, body: AchievementEditPayload, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request); aid=_validate_official_achievement_id(game_id,achievement_id)
    tags=[]
    for tag in body.tags:
        value=str(tag).strip()
        if value and value not in tags: tags.append(value[:80])
    category=canonicalize_wuwa_category(body.category) if game_id=="wuwa" else body.category.strip()
    snapshot=body.model_dump(); snapshot["achievement_id"]=aid; snapshot["tags"]=tags; snapshot["game_id"]=game_id; snapshot["category"]=category
    with connect_db() as db:
        before=_effective_achievement_row(db,game_id,aid)
        db.execute("delete from game_deleted_achievements where game_id=? and achievement_id=?",(game_id,aid))
        existed=db.execute("select achievement_id from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
        db.execute("""insert into game_achievement_overrides(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,is_deleted,source,updated_by,updated_at)
        values(?,?,?,?,?,?,?,?,?,?,?,?,?)
        on conflict(game_id,achievement_id) do update set name=excluded.name,condition=excluded.condition,version=excluded.version,
        category=excluded.category,reward=excluded.reward,hidden=excluded.hidden,tags_json=excluded.tags_json,is_deleted=excluded.is_deleted,
        source=excluded.source,updated_by=excluded.updated_by,updated_at=excluded.updated_at""",
        (game_id,aid,body.name.strip(),body.condition.strip(),body.version.strip(),category,body.reward,1 if body.hidden else 0,json.dumps(tags,ensure_ascii=False),1 if body.is_deleted else 0,body.source.strip() or "admin_override",admin["id"],now()))
        after=_effective_achievement_row(db,game_id,aid)
    record_game_achievement_revision(game_id,aid,"update" if existed or before else "create",snapshot,admin["id"])
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"save_achievement",details=f"game={game_id}; {aid}",category="catalog",game_id=game_id,target_type="achievement",target_id=aid,summary="儲存成就管理資料",before=before or {},after=after or {},actor_ip=client_ip(request),locked=True)
    return {"ok":True,"achievement_id":aid,"created":not bool(before)}


@app.delete("/api/games/{game_id}/admin/achievements/{achievement_id}")
@high_risk_operation
def extra_game_hide_achievement(game_id: str, achievement_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request); aid=_validate_official_achievement_id(game_id,achievement_id)
    with connect_db() as db:
        current=_effective_achievement_row(db,game_id,aid)
        if not current:
            raise HTTPException(status_code=404,detail="找不到可隱藏的成就。")
        row=db.execute("select * from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
        if row:
            db.execute("update game_achievement_overrides set is_deleted=1,updated_by=?,updated_at=? where game_id=? and achievement_id=?",(admin["id"],now(),game_id,aid)); snap=serialize_game_override(row)
        else:
            db.execute("insert into game_achievement_overrides(game_id,achievement_id,name,is_deleted,source,updated_by,updated_at) values(?,?,?,1,'hide-only',?,?)",(game_id,aid,current.get("name") or aid,admin["id"],now())); snap={"achievement_id":aid,"is_deleted":True,"source":"hide-only","game_id":game_id}
    record_game_achievement_revision(game_id,aid,"delete",snap,admin["id"])
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"delete_achievement",details=f"game={game_id}; {aid}",category="catalog",game_id=game_id,target_type="achievement",target_id=aid,summary="從列表隱藏成就",before=current,after={"is_deleted":True},actor_ip=client_ip(request),locked=True)
    return {"ok":True}


def _achievement_delete_snapshot(db: sqlite3.Connection, game_id: str, aid: str) -> dict[str, Any]:
    current=_effective_achievement_row(db,game_id,aid)
    if not current:
        raise HTTPException(status_code=404,detail="找不到可永久刪除的成就。")
    table_queries={
        "catalog":"select * from game_catalog_items where game_id=? and achievement_id=?",
        "override":"select * from game_achievement_overrides where game_id=? and achievement_id=?",
        "progress":"select * from game_progress where game_id=? and achievement_id=? order by user_id",
        "featured":"select * from game_featured_achievements where game_id=? and achievement_id=?",
        "revisions":"select * from game_achievement_revisions where game_id=? and achievement_id=? order by id",
        "deleted":"select * from game_deleted_achievements where game_id=? and achievement_id=?",
        "identity":"select * from achievement_identities where game_id=? and internal_id=?",
        "source_ids":"select * from achievement_source_ids where game_id=? and internal_id=? order by source_name,source_id",
        "source_record":"select * from game_catalog_source_records where game_id=? and achievement_id=?",
        "relations":"select * from game_achievement_choice_groups where game_id=? and achievement_id=? order by relation_type,group_id",
        "reports":"select * from game_achievement_reports where game_id=? and achievement_id=? order by id",
    }
    snapshot={"effective":current,"tables":{}}
    for name,query in table_queries.items():
        snapshot["tables"][name]=[dict(row) for row in db.execute(query,(game_id,aid)).fetchall()]
    return snapshot



def _restore_achievement_database_scope(backup: Path, game_id: str, aid: str) -> dict[str, int]:
    """Restore only one achievement and its dependent rows from a DB backup.

    This intentionally avoids replacing the whole database, so unrelated users,
    messages, audit entries and other games created after the operation began are
    preserved.
    """
    specs = (
        ("game_catalog_items", "game_id=? and achievement_id=?"),
        ("game_achievement_overrides", "game_id=? and achievement_id=?"),
        ("game_progress", "game_id=? and achievement_id=?"),
        ("game_featured_achievements", "game_id=? and achievement_id=?"),
        ("game_achievement_revisions", "game_id=? and achievement_id=?"),
        ("game_deleted_achievements", "game_id=? and achievement_id=?"),
        ("achievement_identities", "game_id=? and internal_id=?"),
        ("achievement_source_ids", "game_id=? and internal_id=?"),
        ("game_catalog_source_records", "game_id=? and achievement_id=?"),
        ("game_achievement_choice_groups", "game_id=? and achievement_id=?"),
        ("game_achievement_reports", "game_id=? and achievement_id=?"),
        ("achievement_delete_backups", "game_id=? and achievement_id=?"),
    )
    source = sqlite3.connect(backup)
    source.row_factory = sqlite3.Row
    restored: dict[str, int] = {}
    try:
        snapshots: dict[str, tuple[list[str], list[sqlite3.Row]]] = {}
        source_tables = {str(row[0]) for row in source.execute("select name from sqlite_master where type='table'").fetchall()}
        for table, where in specs:
            if table not in source_tables:
                snapshots[table] = ([], [])
                continue
            columns = [str(row[1]) for row in source.execute(f"pragma table_info({table})").fetchall()]
            rows = source.execute(f"select * from {table} where {where}", (game_id, aid)).fetchall()
            snapshots[table] = (columns, rows)
        legacy: tuple[list[str], list[sqlite3.Row]] | None = None
        if game_id == "wuwa" and "progress" in source_tables:
            columns = [str(row[1]) for row in source.execute("pragma table_info(progress)").fetchall()]
            legacy = (columns, source.execute("select * from progress where achievement_id=?", (aid,)).fetchall())
        with connect_db() as db:
            db.execute("pragma defer_foreign_keys=on")
            db.execute("begin immediate")
            delete_order = (
                "game_progress", "game_featured_achievements", "game_achievement_revisions",
                "game_achievement_reports", "game_achievement_choice_groups", "achievement_source_ids",
                "game_catalog_source_records", "game_achievement_overrides", "game_deleted_achievements",
                "achievement_delete_backups", "achievement_identities", "game_catalog_items",
            )
            spec_map = dict(specs)
            for table in delete_order:
                db.execute(f"delete from {table} where {spec_map[table]}", (game_id, aid))
            insert_order = (
                "game_catalog_items", "achievement_identities", "achievement_source_ids",
                "game_catalog_source_records", "game_achievement_overrides", "game_deleted_achievements",
                "game_featured_achievements", "game_achievement_revisions", "game_achievement_reports",
                "game_achievement_choice_groups", "game_progress", "achievement_delete_backups",
            )
            for table in insert_order:
                columns, rows = snapshots[table]
                if rows:
                    placeholders = ",".join("?" for _ in columns)
                    db.executemany(
                        f"insert into {table}({','.join(columns)}) values({placeholders})",
                        [tuple(row[column] for column in columns) for row in rows],
                    )
                restored[table] = len(rows)
            if legacy is not None:
                columns, rows = legacy
                db.execute("delete from progress where achievement_id=?", (aid,))
                if rows:
                    placeholders = ",".join("?" for _ in columns)
                    db.executemany(
                        f"insert into progress({','.join(columns)}) values({placeholders})",
                        [tuple(row[column] for column in columns) for row in rows],
                    )
                restored["progress"] = len(rows)
    finally:
        source.close()
    return restored


def _achievement_restore_guard_hash(db: sqlite3.Connection, game_id: str, aid: str) -> str:
    tables={
        "catalog":"select * from game_catalog_items where game_id=? and achievement_id=?",
        "override":"select * from game_achievement_overrides where game_id=? and achievement_id=?",
        "progress":"select * from game_progress where game_id=? and achievement_id=? order by user_id",
        "featured":"select * from game_featured_achievements where game_id=? and achievement_id=?",
        "deleted":"select * from game_deleted_achievements where game_id=? and achievement_id=?",
        "identity":"select * from achievement_identities where game_id=? and internal_id=?",
        "source_ids":"select * from achievement_source_ids where game_id=? and internal_id=? order by source_name,source_id",
        "source_record":"select * from game_catalog_source_records where game_id=? and achievement_id=?",
        "relations":"select * from game_achievement_choice_groups where game_id=? and achievement_id=? order by relation_type,group_id",
        "reports":"select * from game_achievement_reports where game_id=? and achievement_id=? order by id",
    }
    payload={name:[dict(row) for row in db.execute(query,(game_id,aid)).fetchall()] for name,query in tables.items()}
    return governance_hash(payload)


@app.get("/api/games/{game_id}/admin/achievements/{achievement_id}/permanent-preview")
def extra_game_permanent_delete_preview(game_id: str, achievement_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request); aid=_validate_official_achievement_id(game_id,achievement_id)
    with connect_db() as db:
        snapshot=_achievement_delete_snapshot(db,game_id,aid)
    progress_count=len(snapshot["tables"]["progress"])
    relation_count=len(snapshot["tables"]["relations"])
    report_count=len(snapshot["tables"]["reports"])
    confirmation=f"永久刪除 {game_id}:{aid}"
    return {"ok":True,"achievement":snapshot["effective"],"impact":{"progress_count":progress_count,"relation_count":relation_count,"report_count":report_count,"identity_count":len(snapshot["tables"]["identity"]),"source_id_count":len(snapshot["tables"]["source_ids"])},"confirmation_text":confirmation}


@app.delete("/api/games/{game_id}/admin/achievements/{achievement_id}/permanent")
def extra_game_legacy_permanent_delete_disabled(game_id: str, achievement_id: str, request: Request):
    require_extra_game(game_id); require_admin(request)
    raise HTTPException(status_code=409,detail="永久刪除已改為預覽、確認、備份及可回復流程，請重新整理管理頁後操作。")


@app.post("/api/games/{game_id}/admin/achievements/{achievement_id}/permanent")
def extra_game_permanently_delete_achievement(game_id: str, achievement_id: str, body: AchievementPermanentDeletePayload, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request); aid=_validate_official_achievement_id(game_id,achievement_id)
    expected=f"永久刪除 {game_id}:{aid}"
    if body.confirmation_text.strip()!=expected:
        raise HTTPException(status_code=409,detail=f"確認文字不正確，請輸入：{expected}")
    if not GOVERNANCE_OPERATION_GUARD.acquire(blocking=False):
        raise HTTPException(status_code=409,detail="成就管理或治理中心目前正在執行其他高風險操作，請稍後再試。")
    backup=create_database_backup(); delete_id=f"delete-{uuid.uuid4().hex}"; stamp=now()
    try:
        with connect_db() as db:
            db.execute("begin immediate")
            snapshot=_achievement_delete_snapshot(db,game_id,aid)
            name=str(snapshot["effective"].get("name") or aid)
            db.execute("""insert into achievement_delete_backups(id,game_id,achievement_id,achievement_name,admin_user_id,reason,confirmation_text,snapshot_json,backup_name,status,created_at)
                values(?,?,?,?,?,?,?,?,?,'deleted',?)""",(delete_id,game_id,aid,name,admin["id"],body.reason.strip(),expected,json.dumps(snapshot,ensure_ascii=False),backup.name,stamp))
            removed=int(db.execute("delete from game_progress where game_id=? and achievement_id=?",(game_id,aid)).rowcount or 0)
            db.execute("delete from game_featured_achievements where game_id=? and achievement_id=?",(game_id,aid))
            db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid))
            db.execute("""insert into game_deleted_achievements(game_id,achievement_id,achievement_name,deleted_by,deleted_at) values(?,?,?,?,?)
                on conflict(game_id,achievement_id) do update set achievement_name=excluded.achievement_name,deleted_by=excluded.deleted_by,deleted_at=excluded.deleted_at""",(game_id,aid,name,admin["id"],stamp))
            snapshot["restore_guard"]={"post_delete_state_hash":_achievement_restore_guard_hash(db,game_id,aid),"created_at":stamp}
            db.execute("update achievement_delete_backups set snapshot_json=? where id=?",(json.dumps(snapshot,ensure_ascii=False),delete_id))
        bump_game_live_scope(game_id,"catalog"); bump_game_live_scope(game_id,"stats")
        log_admin_action(admin["id"],"permanently_delete_achievement",details=f"game={game_id}; {aid}; progress={removed}; delete_id={delete_id}",category="catalog",game_id=game_id,target_type="achievement",target_id=aid,summary="永久刪除成就（可由專用備份回復）",before=snapshot,after={"deleted":True,"removed_progress":removed},metadata={"delete_id":delete_id,"reason":body.reason},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
        return {"ok":True,"delete_id":delete_id,"removed_progress":removed,"backup":backup.name,"restorable":True}
    except HTTPException:
        raise
    except Exception as exc:
        rollback_error = ""
        try:
            _restore_achievement_database_scope(backup, game_id, aid)
        except Exception as recovery_exc:
            rollback_error = str(recovery_exc)
        if rollback_error:
            raise HTTPException(status_code=500, detail=f"永久刪除失敗，且範圍式安全回復失敗：{exc}；回復錯誤：{rollback_error}") from exc
        raise HTTPException(status_code=409,detail=f"永久刪除失敗，該成就資料已完整回復：{exc}") from exc
    finally:
        GOVERNANCE_OPERATION_GUARD.release()


@app.get("/api/games/{game_id}/admin/achievement-deletions")
def extra_game_list_achievement_deletions(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        rows=db.execute("select id,game_id,achievement_id,achievement_name,reason,backup_name,status,created_at,restored_at,restore_reason from achievement_delete_backups where game_id=? order by created_at desc limit 200",(game_id,)).fetchall()
    return {"ok":True,"deletions":[dict(row) for row in rows]}


@app.post("/api/games/{game_id}/admin/achievement-deletions/{delete_id}/restore")
def extra_game_restore_permanent_deletion(game_id: str, delete_id: str, body: AchievementPermanentRestorePayload, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    if not GOVERNANCE_OPERATION_GUARD.acquire(blocking=False):
        raise HTTPException(status_code=409,detail="成就管理或治理中心目前正在執行其他高風險操作，請稍後再試。")
    safety=create_database_backup()
    try:
        with connect_db() as db:
            db.execute("begin immediate")
            row=db.execute("select * from achievement_delete_backups where id=? and game_id=?",(delete_id,game_id)).fetchone()
            if not row: raise HTTPException(status_code=404,detail="找不到永久刪除備份。")
            if row["status"]!="deleted": raise HTTPException(status_code=409,detail="此永久刪除紀錄目前不可回復。")
            snapshot=_json_object(row["snapshot_json"],{})
            tables=snapshot.get("tables") if isinstance(snapshot,dict) else {}
            aid=str(row["achievement_id"])
            restore_guard=snapshot.get("restore_guard") if isinstance(snapshot,dict) else {}
            expected_hash=str((restore_guard or {}).get("post_delete_state_hash") or "")
            current_hash=_achievement_restore_guard_hash(db,game_id,aid)
            if expected_hash and current_hash!=expected_hash:
                raise HTTPException(status_code=409,detail="永久刪除後，此成就的正式資料、來源、關聯、進度或回報已發生變更；為避免舊快照覆蓋新資料，請先檢查目前資料並以治理中心合併。")
            if not expected_hash:
                current_catalog=db.execute("select 1 from game_catalog_items where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
                current_override=db.execute("select 1 from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
                if current_catalog or current_override:
                    raise HTTPException(status_code=409,detail="此舊版刪除快照沒有狀態指紋，而且目前已有同 ID 資料；不能直接覆蓋，請使用治理中心確認。")
            # Restore dependent rows in FK-safe order. Existing rows are replaced only for this achievement.
            for table,key in (("game_catalog_items","catalog"),("game_achievement_overrides","override"),("game_progress","progress"),("game_featured_achievements","featured"),("game_achievement_revisions","revisions"),("achievement_identities","identity"),("achievement_source_ids","source_ids"),("game_catalog_source_records","source_record"),("game_achievement_choice_groups","relations")):
                rows=list((tables or {}).get(key) or [])
                if table=="achievement_identities": db.execute("delete from achievement_identities where game_id=? and internal_id=?",(game_id,aid))
                elif table=="achievement_source_ids": db.execute("delete from achievement_source_ids where game_id=? and internal_id=?",(game_id,aid))
                else: db.execute(f"delete from {table} where game_id=? and achievement_id=?",(game_id,aid))
                for item in rows:
                    columns=list(item); placeholders=",".join("?" for _ in columns)
                    db.execute(f"insert into {table}({','.join(columns)}) values({placeholders})",[item[c] for c in columns])
            db.execute("delete from game_deleted_achievements where game_id=? and achievement_id=?",(game_id,aid))
            db.execute("update achievement_delete_backups set status='restored',restored_at=?,restored_by=?,restore_reason=? where id=?",(now(),admin["id"],body.reason.strip(),delete_id))
        bump_game_live_scope(game_id,"catalog"); bump_game_live_scope(game_id,"stats")
        log_admin_action(admin["id"],"restore_permanently_deleted_achievement",details=f"game={game_id}; delete_id={delete_id}; achievement={aid}",category="catalog",game_id=game_id,target_type="achievement",target_id=aid,summary="回復永久刪除成就",after={"restored":True,"delete_id":delete_id},backup_name=safety.name,actor_ip=client_ip(request),locked=True)
        return {"ok":True,"achievement_id":aid,"delete_id":delete_id,"safety_backup":safety.name}
    except HTTPException:
        raise
    except Exception as exc:
        rollback_error = ""
        try:
            _restore_achievement_database_scope(safety, game_id, aid if 'aid' in locals() else str(achievement_id if 'achievement_id' in locals() else ""))
        except Exception as recovery_exc:
            rollback_error = str(recovery_exc)
        if rollback_error:
            raise HTTPException(status_code=500, detail=f"永久刪除回復失敗，且範圍式安全回復失敗：{exc}；回復錯誤：{rollback_error}") from exc
        raise HTTPException(status_code=409,detail=f"永久刪除回復失敗，該成就資料已回到操作前狀態：{exc}") from exc
    finally:
        GOVERNANCE_OPERATION_GUARD.release()


@app.post("/api/games/{game_id}/admin/achievements/{achievement_id}/restore")
@high_risk_operation
def extra_game_restore_achievement(game_id: str, achievement_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    with connect_db() as db:
        row=db.execute("select * from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,achievement_id)).fetchone()
        if not row: raise HTTPException(status_code=404,detail="找不到可恢復的成就設定。")
        if not bool(row["is_deleted"]): return {"ok":True,"changed":False}
        tags=json_list(row["tags_json"] or "[]")
        hide_only=(row["source"]=="hide-only" or (row["name"]==achievement_id and not (row["condition"] or "").strip() and (row["version"] or "未標示") in {"","未標示"} and (row["category"] or "未辨識分類") in {"","未辨識分類"} and int(row["reward"] or 0)==0 and not bool(row["hidden"]) and not tags))
        snapshot=serialize_game_override(row); snapshot["is_deleted"]=False
        if hide_only:
            db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,achievement_id)); restored_to_catalog=True
        else:
            db.execute("update game_achievement_overrides set is_deleted=0,updated_by=?,updated_at=? where game_id=? and achievement_id=?",(admin["id"],now(),game_id,achievement_id)); restored_to_catalog=False
    record_game_achievement_revision(game_id,achievement_id,"restore",snapshot,admin["id"])
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"restore_achievement",details=f"game={game_id}; {achievement_id}")
    return {"ok":True,"changed":True,"restored_to_official":restored_to_catalog}


@app.get("/api/games/{game_id}/admin/achievements/{achievement_id}/revisions")
def extra_game_achievement_revisions(game_id: str, achievement_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        rows=db.execute("""select r.*,u.email actor_email from game_achievement_revisions r left join users u on u.id=r.actor_user_id
        where r.game_id=? and r.achievement_id=? order by r.id desc limit 100""",(game_id,achievement_id)).fetchall()
    return {"ok":True,"revisions":[dict(r) for r in rows]}


def _read_relation_document(game_id: str, relation_type: str) -> dict[str,Any]:
    path=game_relation_file(game_id,relation_type)
    if not path.exists():
        return {"schema_version":2,"game_id":game_id,"description":"","groups":[]}
    try:
        value=json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        raise HTTPException(status_code=500,detail=f"關聯資料無法讀取：{exc}")
    if not isinstance(value,dict):
        raise HTTPException(status_code=500,detail="關聯資料格式錯誤。")
    if not isinstance(value.get("groups"),list): value["groups"]=[]
    return value


def _write_relation_document(game_id: str, relation_type: str, document: dict[str,Any]) -> None:
    path=game_relation_file(game_id,relation_type)
    path.parent.mkdir(parents=True,exist_ok=True)
    document={**document,"schema_version":max(2,int(document.get("schema_version") or 1)),"game_id":game_id,"groups":document.get("groups") or []}
    temp=path.with_suffix(path.suffix+".tmp")
    temp.write_text(json.dumps(document,ensure_ascii=False,indent=2)+"\n",encoding="utf-8")
    temp.replace(path)


def _effective_catalog_rows_for_relations(db: sqlite3.Connection, game_id: str) -> list[dict[str,Any]]:
    rows=db.execute("""select c.achievement_id as achievement_id,coalesce(o.name,c.name) as name,coalesce(o.category,c.category) as category,c.source_order as source_order
      from game_catalog_items c left join game_achievement_overrides o on o.game_id=c.game_id and o.achievement_id=c.achievement_id
      left join game_deleted_achievements d on d.game_id=c.game_id and d.achievement_id=c.achievement_id
      where c.game_id=? and d.achievement_id is null and coalesce(o.is_deleted,0)=0
      union all
      select o.achievement_id as achievement_id,o.name as name,o.category as category,999999 as source_order from game_achievement_overrides o
      left join game_catalog_items c on c.game_id=o.game_id and c.achievement_id=o.achievement_id
      left join game_deleted_achievements d on d.game_id=o.game_id and d.achievement_id=o.achievement_id
      where o.game_id=? and c.achievement_id is null and d.achievement_id is null and o.is_deleted=0
      order by 4,1""",(game_id,game_id)).fetchall()
    return [dict(row) for row in rows]


def _validate_relation_documents(game_id: str, documents: dict[str,dict[str,Any]], catalog_ids: set[str]) -> list[dict[str,Any]]:
    _, _, issues = expected_relation_state(documents, catalog_ids)
    return [
        {
            "level": issue.get("severity") or "error",
            "type": issue.get("relation_type") or "",
            "group_id": issue.get("group_id") or "",
            "achievement_id": issue.get("achievement_id") or "",
            "message": issue.get("message") or issue.get("title") or "關聯資料異常",
            "code": issue.get("code") or "",
        }
        for issue in issues
        if issue.get("severity") == "error"
    ]


@app.get("/api/games/{game_id}/admin/relation-groups/manage")
def admin_manage_relation_groups(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    documents={kind:_read_relation_document(game_id,kind) for kind in ("stage","exclusive")}
    with connect_db() as db:
        catalog=_effective_catalog_rows_for_relations(db,game_id)
    catalog_ids={str(row["achievement_id"]) for row in catalog}
    issues=_validate_relation_documents(game_id,documents,catalog_ids)
    groups=[]
    by_id={str(row["achievement_id"]):row for row in catalog}
    for relation_type,document in documents.items():
        for group in document.get("groups") or []:
            members=[]
            for order,achievement_id in enumerate([str(value or "").strip() for value in (group.get("achievement_ids") or []) if str(value or "").strip()],1):
                row=by_id.get(achievement_id,{})
                members.append({"achievement_id":achievement_id,"name":row.get("name") or achievement_id,"category":row.get("category") or "","stage_order":order})
            groups.append({"group_id":str(group.get("id") or ""),"relation_type":relation_type,"name":str(group.get("name") or ""),"basis":str(group.get("basis") or ""),"members":members})
    return {"ok":True,"game_id":game_id,"game_name":game_display_name(game_id),"groups":groups,"catalog":catalog,"issues":issues}


@app.put("/api/games/{game_id}/admin/relation-groups/{relation_type}/{group_id}")
@high_risk_operation
def admin_save_relation_group(game_id: str, relation_type: str, group_id: str, body: RelationGroupPayload, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    relation_type=relation_type.strip().lower()
    if relation_type not in {"stage","exclusive"} or body.relation_type.strip().lower()!=relation_type:
        raise HTTPException(status_code=400,detail="關聯類型無效。")
    clean_group_id=body.group_id.strip()
    if clean_group_id!=group_id:
        raise HTTPException(status_code=400,detail="網址與表單的群組 ID 不一致。")
    if not re.fullmatch(r"[A-Za-z0-9._:-]+",clean_group_id):
        raise HTTPException(status_code=400,detail="群組 ID 只能使用英文、數字、點、底線、冒號與連字號。")
    members=[str(value or "").strip() for value in body.achievement_ids if str(value or "").strip()]
    if len(members)!=len(set(members)):
        raise HTTPException(status_code=400,detail="關聯群組不能包含重複成就。")
    if len(members)<2:
        raise HTTPException(status_code=400,detail="關聯群組至少需要 2 個成就。")
    documents={kind:_read_relation_document(game_id,kind) for kind in ("stage","exclusive")}
    target_doc=documents[relation_type]
    groups=target_doc.get("groups") or []
    existing=next((group for group in groups if str(group.get("id") or "")==clean_group_id),None)
    before_group=copy.deepcopy(existing) if existing is not None else None
    payload=dict(existing or {})
    payload.update({"id":clean_group_id,"type":relation_type,"achievement_ids":members})
    if body.name.strip(): payload["name"]=body.name.strip()
    else: payload.pop("name",None)
    if body.basis.strip(): payload["basis"]=body.basis.strip()
    else: payload.pop("basis",None)
    if existing is None: groups.append(payload)
    else: groups[groups.index(existing)]=payload
    target_doc["groups"]=groups
    with connect_db() as db:
        catalog_ids={str(row["achievement_id"]) for row in _effective_catalog_rows_for_relations(db,game_id)}
    issues=_validate_relation_documents(game_id,documents,catalog_ids)
    if issues:
        raise HTTPException(status_code=400,detail={"message":"關聯資料驗證失敗。","issues":issues[:30]})

    pre_context=_relation_validation_context(game_id)
    backup=create_database_backup()
    batch_id=f"relation-direct-{uuid.uuid4().hex}"
    snapshot_dir=ROOT/"backups"/f"relation-direct-{time.strftime('%Y%m%d-%H%M%S')}-{batch_id[-8:]}"
    snapshot_dir.mkdir(parents=True,exist_ok=True)
    shutil.copy2(game_catalog_file(game_id),snapshot_dir/"catalog.json")
    for kind in ("stage","exclusive"):
        shutil.copy2(game_relation_file(game_id,kind),snapshot_dir/f"{kind}-groups.json")
    try:
        _write_relation_document(game_id,relation_type,target_doc)
        with connect_db() as db:
            db.execute("begin immediate")
            _sync_relation_groups(db,game_id,game_relation_file(game_id,"stage"),"stage")
            _sync_relation_groups(db,game_id,game_relation_file(game_id,"exclusive"),"exclusive")
            _repair_choice_group_progress(db,game_id)
        _,derived_changed=_apply_relation_metadata_to_catalog(game_id,documents)
        post_context=_relation_validation_context(game_id)
        result={
            "operation":"direct_save_relation_group",
            "pre_state_hash":pre_context["state_hash"],
            "post_state_hash":post_context["state_hash"],
            "derived_changed":derived_changed,
            "before":before_group,
            "after":payload,
        }
        plan={"operation":"direct_save_relation_group","relation_type":relation_type,"group_id":clean_group_id,"state_hash":pre_context["state_hash"]}
        with connect_db() as db:
            db.execute("""insert into relation_resolution_batches(id,game_id,validation_id,admin_user_id,status,reason,plan_json,result_json,backup_name,snapshot_dir,created_at,completed_at)
                values(?,?,null,?,'completed',?,?,?,?,?,?,?)""",(batch_id,game_id,admin["id"],"管理員直接編輯關聯群組",json.dumps(plan,ensure_ascii=False),json.dumps(result,ensure_ascii=False),backup.name,str(snapshot_dir),now(),now()))
    except Exception as exc:
        rollback_errors=[]
        for kind in ("stage","exclusive"):
            try:
                shutil.copy2(snapshot_dir/f"{kind}-groups.json",game_relation_file(game_id,kind))
            except Exception as recovery_exc:
                rollback_errors.append(f"{kind}:{recovery_exc}")
        try:
            shutil.copy2(snapshot_dir/"catalog.json",game_catalog_file(game_id))
        except Exception as recovery_exc:
            rollback_errors.append(f"catalog:{recovery_exc}")
        try:
            _restore_relation_database_scope(backup,game_id)
        except Exception as recovery_exc:
            rollback_errors.append(f"database_scope:{recovery_exc}")
        if rollback_errors:
            raise HTTPException(status_code=500,detail=f"關聯群組儲存失敗，且安全回復未完整成功：{exc}；{'；'.join(rollback_errors)}") from exc
        raise HTTPException(status_code=409,detail=f"關聯群組儲存失敗，資料已完整回復：{exc}") from exc
    bump_game_live_scope(game_id,"catalog"); bump_game_live_scope(game_id,"stats")
    log_admin_action(admin["id"],"save_relation_group",details=f"game={game_id}; type={relation_type}; group={clean_group_id}; members={len(members)}; backup={backup.name}; batch={batch_id}",category="relations",game_id=game_id,target_type="relation_group",target_id=clean_group_id,summary="儲存關聯群組（可由批次歷史回復）",before=before_group or {},after=payload,metadata={"batch_id":batch_id,"pre_state_hash":pre_context["state_hash"],"post_state_hash":post_context["state_hash"]},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
    return {"ok":True,"backup":backup.name,"batch_id":batch_id,"restorable":True,"derived_changed":derived_changed}


@app.delete("/api/games/{game_id}/admin/relation-groups/{relation_type}/{group_id}")
@high_risk_operation
def admin_delete_relation_group(game_id: str, relation_type: str, group_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    if relation_type not in {"stage","exclusive"}: raise HTTPException(status_code=400,detail="關聯類型無效。")
    documents={kind:_read_relation_document(game_id,kind) for kind in ("stage","exclusive")}
    document=documents[relation_type]; groups=document.get("groups") or []
    removed_group=next((copy.deepcopy(group) for group in groups if str(group.get("id") or "")==group_id),None)
    next_groups=[group for group in groups if str(group.get("id") or "")!=group_id]
    if len(next_groups)==len(groups): raise HTTPException(status_code=404,detail="找不到關聯群組。")
    document["groups"]=next_groups
    pre_context=_relation_validation_context(game_id)
    backup=create_database_backup()
    batch_id=f"relation-direct-{uuid.uuid4().hex}"
    snapshot_dir=ROOT/"backups"/f"relation-direct-{time.strftime('%Y%m%d-%H%M%S')}-{batch_id[-8:]}"
    snapshot_dir.mkdir(parents=True,exist_ok=True)
    shutil.copy2(game_catalog_file(game_id),snapshot_dir/"catalog.json")
    for kind in ("stage","exclusive"):
        shutil.copy2(game_relation_file(game_id,kind),snapshot_dir/f"{kind}-groups.json")
    try:
        _write_relation_document(game_id,relation_type,document)
        with connect_db() as db:
            db.execute("begin immediate")
            _sync_relation_groups(db,game_id,game_relation_file(game_id,"stage"),"stage")
            _sync_relation_groups(db,game_id,game_relation_file(game_id,"exclusive"),"exclusive")
            _repair_choice_group_progress(db,game_id)
        _,derived_changed=_apply_relation_metadata_to_catalog(game_id,documents)
        post_context=_relation_validation_context(game_id)
        result={"operation":"direct_delete_relation_group","pre_state_hash":pre_context["state_hash"],"post_state_hash":post_context["state_hash"],"derived_changed":derived_changed,"before":removed_group,"after":None}
        plan={"operation":"direct_delete_relation_group","relation_type":relation_type,"group_id":group_id,"state_hash":pre_context["state_hash"]}
        with connect_db() as db:
            db.execute("""insert into relation_resolution_batches(id,game_id,validation_id,admin_user_id,status,reason,plan_json,result_json,backup_name,snapshot_dir,created_at,completed_at)
                values(?,?,null,?,'completed',?,?,?,?,?,?,?)""",(batch_id,game_id,admin["id"],"管理員直接刪除關聯群組",json.dumps(plan,ensure_ascii=False),json.dumps(result,ensure_ascii=False),backup.name,str(snapshot_dir),now(),now()))
    except Exception as exc:
        rollback_errors=[]
        for kind in ("stage","exclusive"):
            try:
                shutil.copy2(snapshot_dir/f"{kind}-groups.json",game_relation_file(game_id,kind))
            except Exception as recovery_exc:
                rollback_errors.append(f"{kind}:{recovery_exc}")
        try:
            shutil.copy2(snapshot_dir/"catalog.json",game_catalog_file(game_id))
        except Exception as recovery_exc:
            rollback_errors.append(f"catalog:{recovery_exc}")
        try:
            _restore_relation_database_scope(backup,game_id)
        except Exception as recovery_exc:
            rollback_errors.append(f"database_scope:{recovery_exc}")
        if rollback_errors:
            raise HTTPException(status_code=500,detail=f"關聯群組刪除失敗，且安全回復未完整成功：{exc}；{'；'.join(rollback_errors)}") from exc
        raise HTTPException(status_code=409,detail=f"關聯群組刪除失敗，資料已完整回復：{exc}") from exc
    bump_game_live_scope(game_id,"catalog"); bump_game_live_scope(game_id,"stats")
    log_admin_action(admin["id"],"delete_relation_group",details=f"game={game_id}; type={relation_type}; group={group_id}; backup={backup.name}; batch={batch_id}",category="relations",game_id=game_id,target_type="relation_group",target_id=group_id,summary="刪除關聯群組（可由批次歷史回復）",before=removed_group or {},after={},metadata={"batch_id":batch_id,"pre_state_hash":pre_context["state_hash"],"post_state_hash":post_context["state_hash"]},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
    return {"ok":True,"backup":backup.name,"batch_id":batch_id,"restorable":True,"derived_changed":derived_changed}


def _relation_achievement_summary(item: dict[str, Any] | None, achievement_id: str, *, stage_order: int = 0, completed: bool = False, completed_at: int = 0) -> dict[str, Any]:
    item = item or {}
    return {
        "achievement_id": achievement_id,
        "name": str(item.get("name") or "找不到成就名稱"),
        "condition": str(item.get("condition") or item.get("description") or ""),
        "version": str(item.get("version") or ""),
        "category": str(item.get("category") or ""),
        "hidden": bool(item.get("hidden")),
        "reward": int(item.get("reward") or 0),
        "stage_order": int(stage_order or 0),
        "completed": bool(completed),
        "completed_at": int(completed_at or 0),
        "exists": bool(item),
    }


def _enrich_relation_validation_result(
    result: dict[str, Any],
    *,
    documents: dict[str, dict[str, Any]],
    catalog_items: list[dict[str, Any]],
    progress_rows: list[dict[str, Any]],
) -> dict[str, Any]:
    """Attach human-readable group/member context to every relation issue.

    The validator intentionally stores compact identifiers. The admin UI needs
    names, conditions, order and completion state so an administrator can make a
    decision without looking up IDs in another page.
    """
    catalog = {
        str(item.get("id") or item.get("achievement_id") or ""): item
        for item in catalog_items
        if str(item.get("id") or item.get("achievement_id") or "")
    }
    progress_lookup = {
        (str(row.get("user_id") or ""), str(row.get("achievement_id") or "")): int(row.get("completed_at") or 0)
        for row in progress_rows
    }
    groups: dict[tuple[str, str], dict[str, Any]] = {}
    for relation_type in ("stage", "exclusive"):
        for group in (documents.get(relation_type) or {}).get("groups") or []:
            if not isinstance(group, dict):
                continue
            group_id = str(group.get("id") or "")
            if not group_id:
                continue
            member_ids = [str(value or "") for value in group.get("achievement_ids") or [] if str(value or "")]
            groups[(relation_type, group_id)] = {
                "relation_type": relation_type,
                "group_id": group_id,
                "name": str(group.get("name") or ""),
                "basis": str(group.get("basis") or ""),
                "members": [
                    _relation_achievement_summary(catalog.get(member_id), member_id, stage_order=index if relation_type == "stage" else 0)
                    for index, member_id in enumerate(member_ids, 1)
                ],
            }

    def add_ids(target: list[str], values: Any) -> None:
        if isinstance(values, str):
            values = [values]
        if not isinstance(values, list):
            return
        for value in values:
            value = str(value or "")
            if value and value not in target:
                target.append(value)

    for issue in result.get("all_issues") or result.get("issues") or []:
        evidence = issue.get("evidence") if isinstance(issue.get("evidence"), dict) else {}
        relation_type = str(issue.get("relation_type") or "")
        group_id = str(issue.get("group_id") or "")
        user_id = str(issue.get("user_id") or "")
        related_ids: list[str] = []
        add_ids(related_ids, issue.get("achievement_id"))
        for key in (
            "completed_achievement_ids", "missing_prior_ids", "member_ids",
            "achievement_ids", "duplicate_members", "related_ids",
        ):
            add_ids(related_ids, evidence.get(key))
        contexts: list[dict[str, Any]] = []
        direct = groups.get((relation_type, group_id))
        if direct:
            contexts.append(direct)
            add_ids(related_ids, [row.get("achievement_id") for row in direct.get("members") or []])
        previous_type = str(evidence.get("previous_type") or "")
        previous_group_id = str(evidence.get("previous_group_id") or "")
        previous = groups.get((previous_type, previous_group_id))
        if previous and previous not in contexts:
            contexts.append(previous)
            add_ids(related_ids, [row.get("achievement_id") for row in previous.get("members") or []])
        related_rows = []
        stage_order_lookup = {
            str(member.get("achievement_id") or ""): int(member.get("stage_order") or 0)
            for context in contexts for member in context.get("members") or []
        }
        for achievement_id in related_ids:
            completed_at = progress_lookup.get((user_id, achievement_id), 0) if user_id else 0
            related_rows.append(_relation_achievement_summary(
                catalog.get(achievement_id), achievement_id,
                stage_order=stage_order_lookup.get(achievement_id, 0),
                completed=bool(completed_at),
                completed_at=completed_at,
            ))
        issue["related_achievements"] = related_rows
        issue["group_contexts"] = contexts
        issue["decision_summary"] = {
            "affected_achievement_count": len(related_rows),
            "affected_group_count": len(contexts),
            "has_user_progress_context": bool(user_id),
        }
    return result


def _relation_validation_context(game_id: str) -> dict[str, Any]:
    documents = {kind: _read_relation_document(game_id, kind) for kind in ("stage", "exclusive")}
    catalog_items = _load_catalog_items_for_health(game_id)
    with connect_db() as db:
        database_rows = [dict(row) for row in db.execute(
            "select group_id,achievement_id,relation_type,stage_order from game_achievement_choice_groups where game_id=? order by relation_type,group_id,stage_order,achievement_id",
            (game_id,),
        ).fetchall()]
        progress_rows = [dict(row) for row in db.execute(
            "select user_id,achievement_id,completed_at from game_progress where game_id=? order by user_id,achievement_id",
            (game_id,),
        ).fetchall()]
        exceptions = {str(row["fingerprint"]) for row in db.execute(
            "select fingerprint from relation_validation_exceptions where game_id=? and active=1",
            (game_id,),
        ).fetchall()}
    result = validate_relation_state(
        game_id=game_id,
        documents=documents,
        catalog_items=catalog_items,
        database_rows=database_rows,
        progress_rows=progress_rows,
        exception_fingerprints=exceptions,
    )
    return _enrich_relation_validation_result(
        result,
        documents=documents,
        catalog_items=catalog_items,
        progress_rows=progress_rows,
    )


def _relation_group_lookup(documents: dict[str, dict[str, Any]], relation_type: str, group_id: str) -> dict[str, Any] | None:
    for group in documents.get(relation_type, {}).get("groups") or []:
        if isinstance(group, dict) and str(group.get("id") or "") == group_id:
            return group
    return None


def _relation_documents_from_database(game_id: str, documents: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    existing = {
        (kind, str(group.get("id") or "")): dict(group)
        for kind, document in documents.items()
        for group in document.get("groups") or []
        if isinstance(group, dict) and str(group.get("id") or "")
    }
    with connect_db() as db:
        rows = [dict(row) for row in db.execute(
            "select group_id,achievement_id,relation_type,stage_order from game_achievement_choice_groups where game_id=? order by relation_type,group_id,stage_order,achievement_id",
            (game_id,),
        ).fetchall()]
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        key = (str(row.get("relation_type") or ""), str(row.get("group_id") or ""))
        grouped.setdefault(key, []).append(row)
    result: dict[str, dict[str, Any]] = {}
    for kind in ("stage", "exclusive"):
        base = dict(documents.get(kind) or {})
        groups = []
        for (relation_type, group_id), members in grouped.items():
            if relation_type != kind:
                continue
            old = existing.get((kind, group_id), {})
            ordered = sorted(members, key=lambda row: (int(row.get("stage_order") or 0), str(row.get("achievement_id") or "")))
            groups.append({
                **old,
                "id": group_id,
                "type": kind,
                "achievement_ids": [str(row.get("achievement_id") or "") for row in ordered],
            })
        base["groups"] = groups
        result[kind] = base
    return result


def _apply_relation_metadata_to_catalog(game_id: str, documents: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    items = _load_catalog_items_for_health(game_id)
    catalog_ids = {str(item.get("id") or item.get("achievement_id") or "") for item in items}
    _, metadata, _ = expected_relation_state(documents, catalog_ids)
    changed = 0
    empty = {
        "choiceGroup": "", "choiceGroupSize": 0, "isChoiceGroup": False,
        "relationGroup": "", "relationGroupSize": 0, "relationType": "", "stageOrder": 0,
    }
    for item in items:
        achievement_id = str(item.get("id") or item.get("achievement_id") or "")
        expected = metadata.get(achievement_id, empty)
        before = {field: item.get(field) for field in DERIVED_RELATION_FIELDS}
        for field, value in expected.items():
            item[field] = value
        if before != {field: item.get(field) for field in DERIVED_RELATION_FIELDS}:
            changed += 1
    _write_catalog_items_payload(game_id, items)
    return items, changed


def _relation_action_label(name: str) -> str:
    return {
        "sync_json_to_database": "以 JSON 同步資料庫",
        "sync_database_to_json": "以資料庫重建 JSON",
        "rebuild_derived_fields": "重建關聯衍生欄位",
        "remove_duplicate_member": "移除群組內重複成員",
        "normalize_group_type": "修正群組類型",
        "move_group_to_type": "移動群組至正確類型",
        "rename_group_id": "更換群組 ID",
        "remove_invalid_member": "移除無效成員",
        "replace_relation_member": "更換錯誤成員",
        "choose_relation_group": "選擇唯一關聯群組",
        "move_relation_member": "移動成員至其他群組",
        "merge_relation_groups": "合併關聯群組",
        "remove_invalid_group": "移除無效群組",
        "set_group_name": "設定群組名稱",
        "set_group_basis": "設定判定依據",
        "keep_exclusive_progress": "保留一項互斥進度",
        "fill_prior_stage_progress": "補齊前置階段進度",
        "remove_later_stage_progress": "移除跳階後續進度",
        "mark_legal_exception": "標記合法例外",
        "repair_document_structure": "修復關聯文件結構",
        "assign_group_id": "指定群組 ID",
        "add_relation_member": "新增群組成員",
        "review": "保留待確認",
    }.get(name, name)


def _build_relation_plan(game_id: str, validation: sqlite3.Row, actions: list[dict[str, Any]], reason: str) -> dict[str, Any]:
    result = json.loads(validation["result_json"] or "{}")
    issues = {str(issue.get("id") or ""): issue for issue in result.get("issues") or []}
    normalized = []
    high_risk = False
    progress_impact = 0
    relation_impact = 0
    for raw in actions:
        issue_id = str(raw.get("issue_id") or "")
        name = str(raw.get("action") or "review")
        params = raw.get("parameters") if isinstance(raw.get("parameters"), dict) else {}
        issue = issues.get(issue_id)
        if issue is None:
            raise HTTPException(status_code=400, detail=f"處置包含不屬於本次驗證的問題：{issue_id}")
        if name not in (issue.get("actions") or []) and name != "review":
            raise HTTPException(status_code=400, detail=f"問題 {issue_id} 不支援處置：{name}")
        if name in {"sync_database_to_json", "repair_document_structure", "assign_group_id", "add_relation_member", "rename_group_id", "replace_relation_member", "choose_relation_group", "move_relation_member", "merge_relation_groups", "remove_invalid_group", "keep_exclusive_progress", "fill_prior_stage_progress", "remove_later_stage_progress"}:
            high_risk = True
        if name in {"keep_exclusive_progress", "fill_prior_stage_progress", "remove_later_stage_progress"}:
            progress_impact += 1
        if name not in {"review", "mark_legal_exception"}:
            relation_impact += 1
        normalized.append({"issue_id": issue_id, "action": name, "action_label": _relation_action_label(name), "parameters": params, "issue": issue})
    return {
        "game_id": game_id,
        "validation_id": validation["id"],
        "state_hash": validation["state_hash"],
        "reason": reason,
        "actions": normalized,
        "impact": {
            "action_count": len(normalized),
            "relation_change_count": relation_impact,
            "progress_change_count": progress_impact,
            "high_risk": high_risk,
            "confirmation_text": "CONFIRM RELATION CHANGE" if high_risk else "",
        },
        "generated_at": now(),
    }


@app.post("/api/games/{game_id}/admin/relation-groups/validate")
def admin_validate_relation_groups(game_id: str, request: Request):
    game_id = require_extra_game(game_id); admin = require_admin(request)
    result = _relation_validation_context(game_id)
    validation_id = f"relation-validation-{uuid.uuid4().hex}"
    stamp = now()
    with connect_db() as db:
        db.execute(
            "insert into relation_validation_runs(id,game_id,admin_user_id,state_hash,result_json,created_at,expires_at) values(?,?,?,?,?,?,?)",
            (validation_id, game_id, admin["id"], result["state_hash"], json.dumps(result, ensure_ascii=False), stamp, stamp + 2 * 60 * 60),
        )
    log_admin_action(admin["id"], "validate_relation_data", details=f"game={game_id}; validation={validation_id}; issues={result['summary']['issue_count']}", category="relations", game_id=game_id, target_type="relation_validation", target_id=validation_id, summary="驗證關聯成就資料", after=result["summary"], actor_ip=client_ip(request), locked=True)
    return {"ok": True, "validation_id": validation_id, "valid": result["summary"]["valid"], **result}


@app.post("/api/games/{game_id}/admin/relation-groups/preview")
def admin_preview_relation_actions(game_id: str, body: RelationValidationPreviewPayload, request: Request):
    game_id = require_extra_game(game_id); admin = require_admin(request)
    with connect_db() as db:
        validation = db.execute("select * from relation_validation_runs where id=? and game_id=? and admin_user_id=?", (body.validation_id, game_id, admin["id"])).fetchone()
    if not validation or int(validation["expires_at"] or 0) <= now():
        raise HTTPException(status_code=404, detail="找不到關聯驗證結果，或結果已過期。")
    current = _relation_validation_context(game_id)
    if current["state_hash"] != validation["state_hash"]:
        raise HTTPException(status_code=409, detail="關聯資料已在驗證後變更，請重新驗證。")
    plan = _build_relation_plan(game_id, validation, body.actions, body.reason)
    batch_id = f"relation-batch-{uuid.uuid4().hex}"
    with connect_db() as db:
        db.execute(
            "insert into relation_resolution_batches(id,game_id,validation_id,admin_user_id,status,reason,plan_json,result_json,backup_name,snapshot_dir,created_at) values(?,?,?,?,?,?,?,?,?,?,?)",
            (batch_id, game_id, body.validation_id, admin["id"], "preview_ready", body.reason, json.dumps(plan, ensure_ascii=False), "{}", "", "", now()),
        )
    log_admin_action(admin["id"], "preview_relation_resolution", details=f"game={game_id}; batch={batch_id}; actions={len(body.actions)}", category="relations", game_id=game_id, target_type="relation_resolution_batch", target_id=batch_id, summary="產生關聯資料處置預覽", after=plan["impact"], actor_ip=client_ip(request), locked=True)
    return {"ok": True, "batch_id": batch_id, "plan": plan}


def _apply_relation_action(game_id: str, action: dict[str, Any], documents: dict[str, dict[str, Any]], db: sqlite3.Connection) -> dict[str, Any]:
    issue = action.get("issue") or {}
    name = str(action.get("action") or "review")
    params = action.get("parameters") or {}
    relation_type = str(issue.get("relation_type") or params.get("relation_type") or "")
    group_id = str(issue.get("group_id") or params.get("group_id") or "")
    achievement_id = str(issue.get("achievement_id") or params.get("achievement_id") or "")
    user_id = str(issue.get("user_id") or params.get("user_id") or "")
    before: Any = {}
    after: Any = {}

    if name == "review":
        return {"status": "review", "before": {}, "after": {}}
    if name == "mark_legal_exception":
        fingerprint = str(issue.get("fingerprint") or "")
        reason = str(params.get("reason") or "管理員確認為合法例外")
        db.execute(
            "insert into relation_validation_exceptions(game_id,fingerprint,reason,active,created_by,created_at,updated_at) values(?,?,?,1,?,?,?) on conflict(game_id,fingerprint) do update set reason=excluded.reason,active=1,created_by=excluded.created_by,updated_at=excluded.updated_at",
            (game_id, fingerprint, reason, params.get("admin_user_id"), now(), now()),
        )
        return {"status": "decision_only", "before": {}, "after": {"fingerprint": fingerprint, "reason": reason}}
    if name == "sync_database_to_json":
        rebuilt = _relation_documents_from_database(game_id, documents)
        documents.clear(); documents.update(rebuilt)
        return {"status": "applied", "before": "json", "after": "database_projection"}
    if name in {"sync_json_to_database", "rebuild_derived_fields"}:
        return {"status": "deferred", "before": {}, "after": {"deferred": name}}

    group = _relation_group_lookup(documents, relation_type, group_id) if relation_type and group_id else None
    if name == "repair_document_structure":
        before = documents.get(relation_type, {}).get("groups")
        if not isinstance(before, list):
            documents.setdefault(relation_type, {})["groups"] = []
        after = documents.get(relation_type, {}).get("groups")
    elif name == "assign_group_id":
        index = int((issue.get("evidence") or {}).get("index") or 0) - 1
        groups = documents.get(relation_type, {}).get("groups") or []
        if index < 0 or index >= len(groups) or not isinstance(groups[index], dict): raise RuntimeError("relation_group_index_not_found")
        new_id = str(params.get("new_group_id") or "").strip()
        if not re.fullmatch(r"[A-Za-z0-9._:-]+", new_id): raise RuntimeError("invalid_new_group_id")
        before = groups[index].get("id"); groups[index]["id"] = new_id; after = new_id
    elif name == "add_relation_member":
        if not group: raise RuntimeError("relation_group_not_found")
        member_id = str(params.get("member_id") or "").strip()
        if not member_id: raise RuntimeError("member_id_required")
        before = list(group.get("achievement_ids") or [])
        group["achievement_ids"] = list(dict.fromkeys([*before, member_id]))
        after = list(group["achievement_ids"])
    elif name == "remove_duplicate_member":
        if not group: raise RuntimeError("relation_group_not_found")
        before = list(group.get("achievement_ids") or [])
        group["achievement_ids"] = list(dict.fromkeys(str(value) for value in before if str(value)))
        after = list(group["achievement_ids"])
    elif name == "normalize_group_type":
        if not group: raise RuntimeError("relation_group_not_found")
        before = group.get("type"); group["type"] = relation_type; after = relation_type
    elif name == "move_group_to_type":
        if not group: raise RuntimeError("relation_group_not_found")
        target_type = str(params.get("target_type") or ("exclusive" if relation_type == "stage" else "stage"))
        if target_type not in {"stage", "exclusive"}: raise RuntimeError("invalid_target_type")
        documents[relation_type]["groups"].remove(group)
        moved = {**group, "type": target_type}
        documents[target_type].setdefault("groups", []).append(moved)
        before = relation_type; after = target_type
    elif name == "rename_group_id":
        if not group: raise RuntimeError("relation_group_not_found")
        new_id = str(params.get("new_group_id") or "").strip()
        if not re.fullmatch(r"[A-Za-z0-9._:-]+", new_id): raise RuntimeError("invalid_new_group_id")
        before = group.get("id"); group["id"] = new_id; after = new_id
    elif name in {"remove_invalid_member", "remove_duplicate_member"}:
        if not group: raise RuntimeError("relation_group_not_found")
        before = list(group.get("achievement_ids") or [])
        group["achievement_ids"] = [str(value) for value in before if str(value) != achievement_id]
        after = list(group["achievement_ids"])
    elif name == "replace_relation_member":
        if not group: raise RuntimeError("relation_group_not_found")
        replacement = str(params.get("replacement_id") or params.get("new_id") or "").strip()
        if not replacement: raise RuntimeError("replacement_id_required")
        before = list(group.get("achievement_ids") or [])
        group["achievement_ids"] = list(dict.fromkeys(replacement if str(value) == achievement_id else str(value) for value in before if str(value)))
        after = list(group["achievement_ids"])
    elif name in {"choose_relation_group", "move_relation_member"}:
        target_type = str(params.get("target_type") or relation_type)
        target_group_id = str(params.get("target_group_id") or group_id)
        if target_type not in {"stage", "exclusive"}: raise RuntimeError("invalid_target_type")
        for kind in ("stage", "exclusive"):
            for candidate in documents[kind].get("groups") or []:
                candidate["achievement_ids"] = [str(value) for value in candidate.get("achievement_ids") or [] if str(value) != achievement_id]
        target = _relation_group_lookup(documents, target_type, target_group_id)
        if not target: raise RuntimeError("target_relation_group_not_found")
        target.setdefault("achievement_ids", []).append(achievement_id)
        target["achievement_ids"] = list(dict.fromkeys(str(value) for value in target["achievement_ids"] if str(value)))
        before = {"achievement_id": achievement_id}; after = {"target_type": target_type, "target_group_id": target_group_id}
    elif name == "merge_relation_groups":
        source_ids = [str(value) for value in params.get("source_group_ids") or [group_id] if str(value)]
        target_type = str(params.get("target_type") or relation_type)
        target_group_id = str(params.get("target_group_id") or "").strip()
        target = _relation_group_lookup(documents, target_type, target_group_id)
        if not target: raise RuntimeError("target_relation_group_not_found")
        merged = list(target.get("achievement_ids") or [])
        for kind in ("stage", "exclusive"):
            retained = []
            for candidate in documents[kind].get("groups") or []:
                cid = str(candidate.get("id") or "")
                if cid in source_ids and candidate is not target:
                    merged.extend(str(value) for value in candidate.get("achievement_ids") or [])
                else:
                    retained.append(candidate)
            documents[kind]["groups"] = retained
        target["achievement_ids"] = list(dict.fromkeys(value for value in merged if value))
        before = source_ids; after = target_group_id
    elif name == "remove_invalid_group":
        if group:
            documents[relation_type]["groups"].remove(group); before = group; after = {}
        else:
            index = int((issue.get("evidence") or {}).get("index") or 0) - 1
            groups = documents.get(relation_type, {}).get("groups") or []
            if index < 0 or index >= len(groups): raise RuntimeError("relation_group_index_not_found")
            before = groups.pop(index); after = {}
    elif name == "set_group_name":
        if not group: raise RuntimeError("relation_group_not_found")
        before = group.get("name"); group["name"] = str(params.get("name") or "").strip(); after = group["name"]
    elif name == "set_group_basis":
        if not group: raise RuntimeError("relation_group_not_found")
        before = group.get("basis"); group["basis"] = str(params.get("basis") or "").strip(); after = group["basis"]
    elif name == "keep_exclusive_progress":
        keep_id = str(params.get("keep_achievement_id") or "").strip()
        members = [str(value) for value in (issue.get("evidence") or {}).get("completed_achievement_ids") or []]
        if not user_id or keep_id not in members: raise RuntimeError("invalid_exclusive_progress_selection")
        removed = [value for value in members if value != keep_id]
        for value in removed:
            db.execute("delete from game_progress where game_id=? and user_id=? and achievement_id=?", (game_id, user_id, value))
            if game_id == "wuwa":
                db.execute("delete from progress where user_id=? and achievement_id=?", (user_id, value))
        before = members; after = [keep_id]
    elif name in {"fill_prior_stage_progress", "remove_later_stage_progress"}:
        evidence = issue.get("evidence") or {}
        member_ids = [str(value) for value in evidence.get("member_ids") or []]
        missing_prior = [str(value) for value in evidence.get("missing_prior_ids") or []]
        if not user_id or not member_ids: raise RuntimeError("invalid_stage_progress_issue")
        if name == "fill_prior_stage_progress":
            timestamp = now()
            for value in missing_prior:
                db.execute("insert into game_progress(game_id,user_id,achievement_id,completed_at) values(?,?,?,?) on conflict(game_id,user_id,achievement_id) do nothing", (game_id, user_id, value, timestamp))
                if game_id == "wuwa":
                    db.execute("insert into progress(user_id,achievement_id,completed_at) values(?,?,?) on conflict(user_id,achievement_id) do nothing", (user_id, value, timestamp))
            before = []; after = missing_prior
        else:
            first_missing_order = min(member_ids.index(value) + 1 for value in missing_prior if value in member_ids)
            removed = member_ids[first_missing_order:]
            for value in removed:
                db.execute("delete from game_progress where game_id=? and user_id=? and achievement_id=?", (game_id, user_id, value))
                if game_id == "wuwa": db.execute("delete from progress where user_id=? and achievement_id=?", (user_id, value))
            before = removed; after = []
    else:
        raise RuntimeError(f"unsupported_relation_action:{name}")
    return {"status": "applied", "before": before, "after": after}



def _restore_relation_catalog_fields_from_snapshot(game_id: str, snapshot_path: Path) -> int:
    before_payload=_json_object(snapshot_path.read_text(encoding="utf-8-sig"),{})
    before_items=before_payload.get("items") if isinstance(before_payload,dict) else None
    if not isinstance(before_items,list):
        raise RuntimeError("relation_snapshot_catalog_invalid")
    before_by_id={str(row.get("id") or row.get("achievement_id") or ""):row for row in before_items if isinstance(row,dict)}
    current_items=_load_catalog_items_for_health(game_id)
    changed=0
    for row in current_items:
        aid=str(row.get("id") or row.get("achievement_id") or "")
        source=before_by_id.get(aid) or {}
        previous={field:(field in row,row.get(field)) for field in DERIVED_RELATION_FIELDS}
        for field in DERIVED_RELATION_FIELDS:
            if field in source:
                row[field]=source.get(field)
            else:
                row.pop(field,None)
        if previous!={field:(field in row,row.get(field)) for field in DERIVED_RELATION_FIELDS}:
            changed+=1
    _write_catalog_items_payload(game_id,current_items)
    return changed


def _restore_relation_database_scope(backup: Path, game_id: str) -> dict[str,int]:
    source=sqlite3.connect(backup)
    source.row_factory=sqlite3.Row
    restored={}
    try:
        with connect_db() as db:
            db.execute("begin immediate")
            for table in ("game_achievement_choice_groups","game_progress","relation_validation_exceptions"):
                columns=[str(row[1]) for row in source.execute(f"pragma table_info({table})").fetchall()]
                rows=source.execute(f"select * from {table} where game_id=?",(game_id,)).fetchall()
                db.execute(f"delete from {table} where game_id=?",(game_id,))
                if rows:
                    placeholders=",".join("?" for _ in columns)
                    db.executemany(f"insert into {table}({','.join(columns)}) values({placeholders})",[tuple(row[col] for col in columns) for row in rows])
                restored[table]=len(rows)
            if game_id=="wuwa":
                columns=[str(row[1]) for row in source.execute("pragma table_info(progress)").fetchall()]
                rows=source.execute("select * from progress").fetchall()
                db.execute("delete from progress")
                if rows:
                    placeholders=",".join("?" for _ in columns)
                    db.executemany(f"insert into progress({','.join(columns)}) values({placeholders})",[tuple(row[col] for col in columns) for row in rows])
                restored["progress"]=len(rows)
    finally:
        source.close()
    return restored


@app.post("/api/games/{game_id}/admin/relation-groups/batches/{batch_id}/execute")
@high_risk_operation
def admin_execute_relation_actions(game_id: str, batch_id: str, body: RelationValidationExecutePayload, request: Request):
    game_id = require_extra_game(game_id); admin = require_admin(request)
    with connect_db() as db:
        batch = db.execute("select * from relation_resolution_batches where id=? and game_id=? and admin_user_id=?", (batch_id, game_id, admin["id"])).fetchone()
    if not batch or batch["status"] != "preview_ready": raise HTTPException(status_code=404, detail="找不到可執行的關聯處置預覽。")
    plan = json.loads(batch["plan_json"] or "{}")
    confirmation = str((plan.get("impact") or {}).get("confirmation_text") or "")
    if confirmation and body.confirmation_text.strip() != confirmation:
        raise HTTPException(status_code=400, detail=f"高風險處置需要輸入確認文字：{confirmation}")
    current = _relation_validation_context(game_id)
    if current["state_hash"] != plan.get("state_hash"):
        raise HTTPException(status_code=409, detail="關聯資料已在預覽後變更，請重新驗證。")

    backup = create_database_backup()
    snapshot_dir = ROOT / "backups" / f"relation-{time.strftime('%Y%m%d-%H%M%S')}-{batch_id[-8:]}"
    snapshot_dir.mkdir(parents=True, exist_ok=True)
    catalog_path = game_catalog_file(game_id)
    relation_paths = {kind: game_relation_file(game_id, kind) for kind in ("stage", "exclusive")}
    shutil.copy2(catalog_path, snapshot_dir / "catalog.json")
    for kind, path in relation_paths.items(): shutil.copy2(path, snapshot_dir / f"{kind}-groups.json")
    documents = {kind: _read_relation_document(game_id, kind) for kind in ("stage", "exclusive")}
    results = []
    try:
        with connect_db() as db:
            db.execute("begin immediate")
            for action in plan.get("actions") or []:
                action.setdefault("parameters", {})["admin_user_id"] = admin["id"]
                result = _apply_relation_action(game_id, action, documents, db)
                results.append({"issue_id": action.get("issue_id"), "action": action.get("action"), **result})
            for kind, document in documents.items():
                _write_relation_document(game_id, kind, document)
            if any(action.get("action") in {"sync_json_to_database", "sync_database_to_json", "repair_document_structure", "assign_group_id", "add_relation_member", "remove_duplicate_member", "normalize_group_type", "move_group_to_type", "rename_group_id", "remove_invalid_member", "replace_relation_member", "choose_relation_group", "move_relation_member", "merge_relation_groups", "remove_invalid_group", "set_group_name", "set_group_basis"} for action in plan.get("actions") or []):
                _sync_relation_groups(db, game_id, game_relation_file(game_id, "stage"), "stage")
                _sync_relation_groups(db, game_id, game_relation_file(game_id, "exclusive"), "exclusive")
        derived_changed = 0
        if any(action.get("action") == "rebuild_derived_fields" for action in plan.get("actions") or []):
            _, derived_changed = _apply_relation_metadata_to_catalog(game_id, documents)
        validation = _relation_validation_context(game_id)
        with connect_db() as db:
            integrity = str(db.execute("pragma integrity_check").fetchone()[0])
            if integrity != "ok": raise RuntimeError(f"database_integrity:{integrity}")
            result_payload = {"actions": results, "derived_changed": derived_changed, "validation": validation["summary"], "backup": backup.name, "pre_state_hash": current["state_hash"], "post_state_hash": validation["state_hash"]}
            db.execute("update relation_resolution_batches set status='completed',result_json=?,backup_name=?,snapshot_dir=?,completed_at=? where id=?", (json.dumps(result_payload, ensure_ascii=False), backup.name, str(snapshot_dir), now(), batch_id))
    except Exception as exc:
        rollback_errors=[]
        try:
            shutil.copy2(snapshot_dir / "catalog.json", catalog_path)
        except Exception as rollback_exc:
            rollback_errors.append(f"catalog:{rollback_exc}")
        for kind, path in relation_paths.items():
            try:
                shutil.copy2(snapshot_dir / f"{kind}-groups.json", path)
            except Exception as rollback_exc:
                rollback_errors.append(f"{kind}:{rollback_exc}")
        try:
            _restore_relation_database_scope(backup, game_id)
        except Exception as rollback_exc:
            rollback_errors.append(f"database_scope:{rollback_exc}")
        rolled_back=not rollback_errors
        try:
            with connect_db() as db:
                db.execute("update relation_resolution_batches set status='failed',result_json=?,backup_name=?,snapshot_dir=?,completed_at=? where id=?", (json.dumps({"error": str(exc), "rolled_back": rolled_back, "rollback_errors": rollback_errors}, ensure_ascii=False), backup.name, str(snapshot_dir), now(), batch_id))
        except Exception as record_exc:
            rollback_errors.append(f"record:{record_exc}")
            rolled_back=False
        log_admin_action(admin["id"], "relation_resolution_failed", details=f"game={game_id}; batch={batch_id}; error={exc}; rollback_errors={rollback_errors}", category="relations", status="failed" if rolled_back else "recovery_failed", game_id=game_id, target_type="relation_resolution_batch", target_id=batch_id, summary="關聯資料處置失敗"+("並已自動回復" if rolled_back else "且自動回復失敗"), error_message=str(exc), metadata={"rollback_errors":rollback_errors}, backup_name=backup.name, actor_ip=client_ip(request), locked=True)
        if rollback_errors:
            raise HTTPException(status_code=500, detail=f"關聯資料處置失敗，且自動回復未完整成功：{exc}；回復錯誤：{'；'.join(rollback_errors)}")
        raise HTTPException(status_code=409, detail=f"關聯資料處置失敗，資料已完整回復：{exc}")
    bump_game_live_scope(game_id, "catalog"); bump_game_live_scope(game_id, "stats")
    log_admin_action(admin["id"], "relation_resolution_completed", details=f"game={game_id}; batch={batch_id}; actions={len(results)}", category="relations", game_id=game_id, target_type="relation_resolution_batch", target_id=batch_id, summary="完成關聯資料處置", after=result_payload, backup_name=backup.name, actor_ip=client_ip(request), locked=True)
    return {"ok": True, "batch_id": batch_id, "result": result_payload}


@app.post("/api/games/{game_id}/admin/relation-groups/batches/{batch_id}/rollback")
@high_risk_operation
def admin_rollback_relation_actions(game_id: str, batch_id: str, body: RelationValidationRollbackPayload, request: Request):
    game_id = require_extra_game(game_id); admin = require_admin(request)
    with connect_db() as db:
        batch = db.execute("select * from relation_resolution_batches where id=? and game_id=?", (batch_id, game_id)).fetchone()
    if not batch or batch["status"] != "completed" or batch["rolled_back_at"]:
        raise HTTPException(status_code=409, detail="此關聯處置批次目前不可回復。")
    backup = ROOT / "backups" / str(batch["backup_name"] or "")
    snapshot_dir = Path(str(batch["snapshot_dir"] or ""))
    if not backup.exists() or not snapshot_dir.exists():
        raise HTTPException(status_code=409, detail="回復所需備份不存在。")
    result=_json_object(batch["result_json"],{})
    expected_post_hash=str(result.get("post_state_hash") or "")
    current=_relation_validation_context(game_id)
    if not expected_post_hash or current["state_hash"]!=expected_post_hash:
        raise HTTPException(status_code=409,detail="關聯處置完成後已有進度、關聯或衍生欄位變更；為避免倒退其他新資料，請建立新的修復批次，不能直接回復舊批次。")
    safety = create_database_backup()
    current_catalog_bytes = game_catalog_file(game_id).read_bytes()
    current_relation_bytes = {kind: game_relation_file(game_id, kind).read_bytes() for kind in ("stage", "exclusive")}
    rollback_errors=[]
    try:
        for kind in ("stage", "exclusive"):
            shutil.copy2(snapshot_dir / f"{kind}-groups.json", game_relation_file(game_id, kind))
        derived_restored=_restore_relation_catalog_fields_from_snapshot(game_id,snapshot_dir/"catalog.json")
        restored_rows=_restore_relation_database_scope(backup,game_id)
        validation=_relation_validation_context(game_id)
        expected_pre_hash=str(result.get("pre_state_hash") or "")
        if expected_pre_hash and validation["state_hash"]!=expected_pre_hash:
            raise RuntimeError("relation_scope_restore_hash_mismatch")
        with connect_db() as db:
            db.execute("update relation_resolution_batches set status='rolled_back',rolled_back_at=?,rollback_reason=?,result_json=? where id=?",(now(),body.reason,json.dumps({**result,"rollback":{"derived_restored":derived_restored,"restored_rows":restored_rows,"state_hash":validation["state_hash"]}},ensure_ascii=False),batch_id))
    except Exception as exc:
        try:
            game_catalog_file(game_id).write_bytes(current_catalog_bytes)
        except Exception as rollback_exc:
            rollback_errors.append(f"catalog:{rollback_exc}")
        for kind, data in current_relation_bytes.items():
            try:
                game_relation_file(game_id, kind).write_bytes(data)
            except Exception as rollback_exc:
                rollback_errors.append(f"{kind}:{rollback_exc}")
        try:
            _restore_relation_database_scope(safety, game_id)
        except Exception as rollback_exc:
            rollback_errors.append(f"database_scope:{rollback_exc}")
        if rollback_errors:
            raise HTTPException(status_code=500, detail=f"回復失敗，且安全回復也失敗：{exc}；{'；'.join(rollback_errors)}") from exc
        raise HTTPException(status_code=409, detail=f"回復失敗，已回到操作前狀態：{exc}") from exc
    bump_game_live_scope(game_id, "catalog"); bump_game_live_scope(game_id, "stats")
    log_admin_action(admin["id"], "relation_resolution_rollback", details=f"game={game_id}; batch={batch_id}; reason={body.reason}", category="relations", game_id=game_id, target_type="relation_resolution_batch", target_id=batch_id, summary="回復關聯資料處置（限批次範圍）", after={"state_hash":validation["state_hash"],"restored_rows":restored_rows}, backup_name=safety.name, actor_ip=client_ip(request), locked=True)
    return {"ok": True, "batch_id": batch_id, "status": "rolled_back", "safety_backup": safety.name, "restored_rows":restored_rows, "derived_restored":derived_restored}


@app.get("/api/games/{game_id}/admin/relation-groups/batches")
def admin_list_relation_batches(game_id: str, request: Request, limit: int = 30):
    game_id = require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        rows = db.execute("select * from relation_resolution_batches where game_id=? order by created_at desc limit ?", (game_id, max(1, min(100, int(limit))))).fetchall()
    return {"ok": True, "batches": [{**dict(row), "plan": json.loads(row["plan_json"] or "{}"), "result": json.loads(row["result_json"] or "{}")} for row in rows]}


@app.post("/api/games/{game_id}/admin/catalog/rebuild")
@high_risk_operation
def extra_game_rebuild_catalog(game_id: str, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    backup=create_database_backup()
    with connect_db() as db:
        db.execute("begin immediate")
        removed_overrides=int(db.execute("select count(*) c from game_achievement_overrides where game_id=?",(game_id,)).fetchone()["c"] or 0)
        removed_featured=int(db.execute("select count(*) c from game_featured_achievements where game_id=?",(game_id,)).fetchone()["c"] or 0)
        removed_deleted=int(db.execute("select count(*) c from game_deleted_achievements where game_id=?",(game_id,)).fetchone()["c"] or 0)
        progress_records=int(db.execute("select count(*) c from game_progress where game_id=?",(game_id,)).fetchone()["c"] or 0)
        db.execute("delete from game_achievement_overrides where game_id=?",(game_id,))
        db.execute("delete from game_featured_achievements where game_id=?",(game_id,))
        db.execute("delete from game_deleted_achievements where game_id=?",(game_id,))
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"rebuild_achievement_catalog",details=f"game={game_id}; overrides={removed_overrides}; featured={removed_featured}; deleted={removed_deleted}; progress={progress_records}; backup={backup.name}",backup_name=backup.name,locked=True)
    return {"ok":True,"removed_overrides":removed_overrides,"removed_featured":removed_featured,"removed_permanent_deletions":removed_deleted,"progress_records":progress_records,"backup":backup.name}


@app.post("/api/games/{game_id}/admin/catalog/validate")
def extra_game_validate_catalog(game_id: str, body: CatalogValidationPayload, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    return {"ok":True,**_scan_catalog_for_admin(game_id,body.items,admin["id"])}


@app.post("/api/games/{game_id}/admin/catalog/repair")
@high_risk_operation
def extra_game_repair_catalog(game_id: str, body: CatalogRepairPayload, request: Request):
    game_id=require_extra_game(game_id); admin=require_admin(request)
    with connect_db() as db: record=db.execute("select * from catalog_scan_previews where id=? and game_id=? and admin_user_id=?",(body.scan_id,game_id,admin["id"])).fetchone()
    if not record: raise HTTPException(status_code=404,detail="找不到檢查結果，請重新檢查。")
    return _apply_catalog_repair(game_id,record,body,admin,request)


def extra_game_completion_stats(game_id: str, include_counts: bool) -> dict[str,Any]:
    """回傳全站完成率；互斥／關聯成就以整組不重複完成帳號數共用完成率。"""
    with connect_db() as db:
        total=int(db.execute(
            "select count(*) c from users where is_active=1 and email_verified=1"
        ).fetchone()["c"] or 0)

        # 一般成就仍使用個別完成帳號數。
        individual_rows=db.execute(
            """select p.achievement_id,count(distinct p.user_id) completed_count
            from game_progress p
            join users u on u.id=p.user_id
            where p.game_id=? and u.is_active=1 and u.email_verified=1
            group by p.achievement_id""",
            (game_id,),
        ).fetchall()
        counts_by_achievement={
            str(row["achievement_id"]):int(row["completed_count"] or 0)
            for row in individual_rows
        }

        # 關聯成就共用同一組完成率；同一帳號即使有舊的異常重複資料也只算一次。
        group_rows=db.execute(
            """select g.group_id,count(distinct p.user_id) completed_count
            from game_achievement_choice_groups g
            join game_progress p
              on p.game_id=g.game_id and p.achievement_id=g.achievement_id
            join users u on u.id=p.user_id
            where g.game_id=? and g.relation_type='exclusive' and u.is_active=1 and u.email_verified=1
            group by g.group_id""",
            (game_id,),
        ).fetchall()
        counts_by_group={
            str(row["group_id"]):int(row["completed_count"] or 0)
            for row in group_rows
        }
        member_rows=db.execute(
            """select group_id,achievement_id
            from game_achievement_choice_groups
            where game_id=? and relation_type='exclusive'
            order by group_id,achievement_id""",
            (game_id,),
        ).fetchall()

    stats_by_id={}
    for achievement_id,count in counts_by_achievement.items():
        stats_by_id[achievement_id]={
            "achievement_id":achievement_id,
            "percentage":round(count*100/total,2) if total else 0,
            **({"completed_count":count} if include_counts else {}),
        }

    for row in member_rows:
        achievement_id=str(row["achievement_id"])
        group_id=str(row["group_id"])
        count=counts_by_group.get(group_id,0)
        stats_by_id[achievement_id]={
            "achievement_id":achievement_id,
            "percentage":round(count*100/total,2) if total else 0,
            "choice_group":group_id,
            **({"completed_count":count} if include_counts else {}),
        }

    stats=sorted(stats_by_id.values(),key=lambda item:str(item["achievement_id"]))
    result={"ok":True,"stats":stats}
    if include_counts:
        result["eligible_users"]=total
    return result



# ----- 成就類別管理 -----
@app.get("/api/games/{game_id}/admin/achievement-categories")
def admin_achievement_categories(game_id: str, request: Request):
    game_id=require_extra_game(game_id);require_admin(request)
    with connect_db() as db:
        rows=_achievement_category_rows(db,game_id)
    return {"ok":True,"game_id":game_id,"categories":rows,"count":len(rows)}


@app.post("/api/games/{game_id}/admin/achievement-categories")
@high_risk_operation
def admin_create_achievement_category(game_id: str, body: AchievementCategoryCreatePayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request);name=_normalize_achievement_category_name(game_id,body.name);stamp=now()
    with connect_db() as db:
        db.execute("begin immediate")
        _sync_achievement_categories(db,game_id)
        duplicate=db.execute("select id from game_achievement_categories where game_id=? and lower(name)=lower(?)",(game_id,name)).fetchone()
        if duplicate:
            raise HTTPException(status_code=409,detail="此遊戲已存在相同名稱的分類。")
        display_order=int(db.execute("select coalesce(max(display_order),-1)+1 from game_achievement_categories where game_id=?",(game_id,)).fetchone()[0])
        category_id=str(uuid.uuid4())
        db.execute(
            """insert into game_achievement_categories(id,game_id,name,display_order,is_custom,created_by,updated_by,created_at,updated_at)
            values(?,?,?,?,1,?,?,?,?)""",
            (category_id,game_id,name,display_order,admin["id"],admin["id"],stamp,stamp),
        )
        rows=_achievement_category_rows(db,game_id)
    log_admin_action(admin["id"],"create_achievement_category",category="catalog",game_id=game_id,target_type="achievement_category",target_id=category_id,summary=f"新增成就分類：{name}",after={"name":name,"display_order":display_order},actor_ip=client_ip(request),locked=True)
    return {"ok":True,"category_id":category_id,"categories":rows}


@app.post("/api/games/{game_id}/admin/achievement-categories/reorder")
@high_risk_operation
def admin_reorder_achievement_categories(game_id: str, body: AchievementCategoryReorderPayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request);requested=[str(value).strip() for value in body.category_ids if str(value).strip()]
    if len(requested)!=len(set(requested)):
        raise HTTPException(status_code=400,detail="分類排序清單包含重複項目。")
    with connect_db() as db:
        db.execute("begin immediate")
        current=_achievement_category_rows(db,game_id)
        current_ids=[str(row["id"]) for row in current]
        if set(requested)!=set(current_ids) or len(requested)!=len(current_ids):
            raise HTTPException(status_code=409,detail="分類清單已變更，請重新整理後再調整順序。")
        stamp=now()
        db.executemany(
            "update game_achievement_categories set display_order=?,updated_by=?,updated_at=? where game_id=? and id=?",
            [(index,admin["id"],stamp,game_id,category_id) for index,category_id in enumerate(requested)],
        )
        rows=_achievement_category_rows(db,game_id)
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"reorder_achievement_categories",category="catalog",game_id=game_id,target_type="achievement_categories",target_id=game_id,summary="調整成就分類順序",before={"category_ids":current_ids},after={"category_ids":requested},actor_ip=client_ip(request),locked=True)
    return {"ok":True,"categories":rows}


@app.patch("/api/games/{game_id}/admin/achievement-categories/{category_id}")
@high_risk_operation
def admin_update_achievement_category(game_id: str, category_id: str, body: AchievementCategoryUpdatePayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request);new_name=_normalize_achievement_category_name(game_id,body.name)
    with connect_db() as db:
        _sync_achievement_categories(db,game_id)
        category=db.execute("select * from game_achievement_categories where game_id=? and id=?",(game_id,category_id)).fetchone()
        if not category:
            raise HTTPException(status_code=404,detail="找不到要修改的分類。")
        old_name=str(category["name"])
        duplicate=db.execute("select id from game_achievement_categories where game_id=? and lower(name)=lower(?) and id<>?",(game_id,new_name,category_id)).fetchone()
        if duplicate:
            raise HTTPException(status_code=409,detail="目標分類已存在，請改用合併分類。")
        affected=sum(1 for row in _effective_category_values(db,game_id) if str(row["category"])==old_name)
    if new_name==old_name:
        return {"ok":True,"changed":False,"affected_achievements":0}
    backup=create_database_backup();original_catalog=None
    try:
        original_catalog,file_changes=_rewrite_catalog_category_names(game_id,{old_name:new_name})
        with connect_db() as db:
            db.execute("begin immediate")
            db.execute("update game_catalog_items set category=?,updated_at=? where game_id=? and category=?",(new_name,now(),game_id,old_name))
            db.execute("update game_achievement_overrides set category=?,updated_by=?,updated_at=? where game_id=? and category=?",(new_name,admin["id"],now(),game_id,old_name))
            db.execute("update game_catalog_source_records set group_name=?,updated_at=? where game_id=? and group_name=?",(new_name,now(),game_id,old_name))
            db.execute("update game_achievement_categories set name=?,is_custom=1,updated_by=?,updated_at=? where game_id=? and id=?",(new_name,admin["id"],now(),game_id,category_id))
            db.execute(
                """insert into game_achievement_category_aliases(game_id,source_name,category_id,created_at) values(?,?,?,?)
                on conflict(game_id,source_name) do update set category_id=excluded.category_id""",
                (game_id,old_name,category_id,now()),
            )
            rows=_achievement_category_rows(db,game_id)
    except Exception:
        if original_catalog is not None:
            _restore_catalog_bytes(game_id,original_catalog)
        raise
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"rename_achievement_category",category="catalog",game_id=game_id,target_type="achievement_category",target_id=category_id,summary=f"成就分類改名：{old_name} → {new_name}",before={"name":old_name,"achievement_count":affected},after={"name":new_name,"achievement_count":affected,"file_changes":file_changes},metadata={"reason":body.reason},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
    return {"ok":True,"changed":True,"affected_achievements":affected,"backup":backup.name,"categories":rows}


@app.post("/api/games/{game_id}/admin/achievement-categories/{category_id}/merge")
@high_risk_operation
def admin_merge_achievement_category(game_id: str, category_id: str, body: AchievementCategoryMergePayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request)
    if category_id==body.target_category_id:
        raise HTTPException(status_code=400,detail="來源分類與目標分類不可相同。")
    with connect_db() as db:
        _sync_achievement_categories(db,game_id)
        source=db.execute("select * from game_achievement_categories where game_id=? and id=?",(game_id,category_id)).fetchone()
        target=db.execute("select * from game_achievement_categories where game_id=? and id=?",(game_id,body.target_category_id)).fetchone()
        if not source or not target:
            raise HTTPException(status_code=404,detail="找不到來源或目標分類。")
        source_name=str(source["name"]);target_name=str(target["name"])
        affected=sum(1 for row in _effective_category_values(db,game_id) if str(row["category"])==source_name)
    backup=create_database_backup();original_catalog=None
    try:
        original_catalog,file_changes=_rewrite_catalog_category_names(game_id,{source_name:target_name})
        with connect_db() as db:
            db.execute("begin immediate")
            db.execute("update game_catalog_items set category=?,updated_at=? where game_id=? and category=?",(target_name,now(),game_id,source_name))
            db.execute("update game_achievement_overrides set category=?,updated_by=?,updated_at=? where game_id=? and category=?",(target_name,admin["id"],now(),game_id,source_name))
            db.execute("update game_catalog_source_records set group_name=?,updated_at=? where game_id=? and group_name=?",(target_name,now(),game_id,source_name))
            db.execute("update game_achievement_category_aliases set category_id=? where game_id=? and category_id=?",(body.target_category_id,game_id,category_id))
            db.execute(
                """insert into game_achievement_category_aliases(game_id,source_name,category_id,created_at) values(?,?,?,?)
                on conflict(game_id,source_name) do update set category_id=excluded.category_id""",
                (game_id,source_name,body.target_category_id,now()),
            )
            db.execute("delete from game_achievement_categories where game_id=? and id=?",(game_id,category_id))
            rows=_achievement_category_rows(db,game_id)
    except Exception:
        if original_catalog is not None:
            _restore_catalog_bytes(game_id,original_catalog)
        raise
    bump_game_live_scope(game_id,"catalog")
    log_admin_action(admin["id"],"merge_achievement_category",category="catalog",game_id=game_id,target_type="achievement_category",target_id=category_id,summary=f"合併成就分類：{source_name} → {target_name}",before={"source":source_name,"achievement_count":affected},after={"target":target_name,"achievement_count":affected,"file_changes":file_changes},metadata={"reason":body.reason},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
    return {"ok":True,"affected_achievements":affected,"backup":backup.name,"categories":rows}


@app.post("/api/games/{game_id}/admin/achievement-categories/{category_id}/delete")
@high_risk_operation
def admin_delete_achievement_category(game_id: str, category_id: str, body: AchievementCategoryDeletePayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request)
    with connect_db() as db:
        rows=_achievement_category_rows(db,game_id)
        category=next((row for row in rows if str(row["id"])==category_id),None)
        if not category:
            raise HTTPException(status_code=404,detail="找不到要刪除的分類。")
        if int(category["achievement_count"] or 0)>0:
            raise HTTPException(status_code=409,detail=f"此分類仍有 {category['achievement_count']} 項成就，請先合併或移動成就。")
    backup=create_database_backup()
    with connect_db() as db:
        db.execute("begin immediate")
        current=next((row for row in _achievement_category_rows(db,game_id) if str(row["id"])==category_id),None)
        if not current:
            raise HTTPException(status_code=404,detail="找不到要刪除的分類。")
        if int(current["achievement_count"] or 0)>0:
            raise HTTPException(status_code=409,detail=f"此分類仍有 {current['achievement_count']} 項成就，請先合併或移動成就。")
        db.execute("delete from game_achievement_categories where game_id=? and id=?",(game_id,category_id))
        updated=_achievement_category_rows(db,game_id)
    log_admin_action(admin["id"],"delete_achievement_category",category="catalog",game_id=game_id,target_type="achievement_category",target_id=category_id,summary=f"刪除空成就分類：{category['name']}",before=category,metadata={"reason":body.reason},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
    return {"ok":True,"backup":backup.name,"categories":updated}


# 20260621-hsr-admin-management-v1
@app.get("/api/games/{game_id}/admin/achievement-management")
def extra_game_admin_achievement_management(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        catalog_rows=db.execute(
            """select c.*,coalesce(s.official_source_id,c.achievement_id) as official_source_id,
            g.group_id as relation_group,g.relation_type,g.stage_order,
            (select count(*) from game_achievement_choice_groups x
             where x.game_id=c.game_id and x.group_id=g.group_id) as relation_group_size
            from game_catalog_items c
            left join game_catalog_source_records s
              on s.game_id=c.game_id and s.achievement_id=c.achievement_id
            left join game_achievement_choice_groups g
              on g.game_id=c.game_id and g.achievement_id=c.achievement_id
            where c.game_id=?
            order by c.source_order,c.achievement_id""",
            (game_id,),
        ).fetchall()
        override_rows=db.execute(
            "select * from game_achievement_overrides where game_id=? order by updated_at",
            (game_id,),
        ).fetchall()
        deleted_rows=db.execute(
            "select achievement_id from game_deleted_achievements where game_id=? order by deleted_at",
            (game_id,),
        ).fetchall()
        category_rows=_achievement_category_rows(db,game_id)
    catalog=[]
    for index,row in enumerate(catalog_rows):
        value=dict(row)
        catalog.append({
            "id":value["achievement_id"],
            "officialId":value.get("official_source_id") or value["achievement_id"],
            "displayId":value.get("official_source_id") or value["achievement_id"],
            "name":value["name"],
            "condition":value["condition"],
            "version":value["version"],
            "category":value["category"],
            "reward":int(value["reward"] or 0),
            "hidden":bool(value["hidden"]),
            "tags":json_list(value.get("tags_json") or "[]"),
            "source":value.get("source") or "catalog",
            "sourceOrder":int(value.get("source_order") or index),
            "relationGroup":value.get("relation_group") or "",
            "relationType":value.get("relation_type") or "",
            "stageOrder":int(value.get("stage_order") or 0),
            "relationGroupSize":int(value.get("relation_group_size") or 0),
            "choiceGroup":value.get("relation_group") if value.get("relation_type")=="exclusive" else "",
            "choiceGroupSize":int(value.get("relation_group_size") or 0) if value.get("relation_type")=="exclusive" else 0,
            "isChoiceGroup":bool(value.get("relation_group") and value.get("relation_type")=="exclusive"),
        })
    with connect_db() as db:
        catalog=_sort_achievement_display_rows(db,game_id,catalog,category_rows)
    return {
        "ok":True,
        "game_id":game_id,
        "catalog":catalog,
        "overrides":[serialize_game_override(row) for row in override_rows],
        "categories":category_rows,
        "permanently_deleted":[row["achievement_id"] for row in deleted_rows],
        "capabilities":{
            "achievement_management":True,
            "achievement_reports":True,
            "manual_achievement_create":True,
        },
    }

@app.get("/api/games/{game_id}/completion-stats")
def extra_game_public_completion_stats(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_user(request)
    return extra_game_completion_stats(game_id,False)


@app.get("/api/games/{game_id}/admin/completion-stats")
def extra_game_admin_completion_stats(game_id: str, request: Request):
    game_id=require_extra_game(game_id); require_admin(request)
    return extra_game_completion_stats(game_id,True)


@app.get("/api/games/{game_id}/progress")
def extra_game_get_progress(game_id: str, request: Request):
    game_id=require_extra_game(game_id); user=require_user(request)
    with connect_db() as db:
        rows=db.execute("select achievement_id from game_progress where game_id=? and user_id=? order by completed_at,achievement_id",(game_id,user["id"])).fetchall()
    return {"ok":True,"completed":[r["achievement_id"] for r in rows]}


@app.post("/api/games/{game_id}/progress/set")
@high_risk_operation
def extra_game_set_progress(game_id: str, body: ProgressSet, request: Request):
    game_id=require_extra_game(game_id); user=require_user(request)
    with connect_db() as db:
        aid=_resolve_effective_achievement_id(db,game_id,body.achievement_id)
        if body.completed:
            if not _stage_prerequisites_completed(db,game_id,user["id"],aid):
                raise HTTPException(status_code=409,detail="前置階段尚未完成。")
            _delete_choice_siblings(db,game_id,user["id"],[aid])
            db.execute("""insert into game_progress(game_id,user_id,achievement_id,completed_at) values(?,?,?,?)
            on conflict(game_id,user_id,achievement_id) do update set completed_at=excluded.completed_at""",(game_id,user["id"],aid,now()))
        else:
            _delete_stage_from(db,game_id,user["id"],aid)
        rows=db.execute("select achievement_id from game_progress where game_id=? and user_id=? order by completed_at,achievement_id",(game_id,user["id"])).fetchall()
    bump_game_live_scope(game_id,"stats")
    return {"ok":True,"completed":[r["achievement_id"] for r in rows]}


@app.post("/api/games/{game_id}/progress/batch")
@high_risk_operation
def extra_game_batch_progress(game_id: str, body: ProgressBatch, request: Request):
    game_id=require_extra_game(game_id); user=require_user(request); requested=validate_ids(body.achievement_ids); t=now()
    with connect_db() as db:
        ids=list(dict.fromkeys(_resolve_effective_achievement_id(db,game_id,aid) for aid in requested))
        if body.completed:
            ids=_normalize_choice_progress_ids(db,game_id,ids)
            existing=[str(row["achievement_id"]) for row in db.execute("select achievement_id from game_progress where game_id=? and user_id=?",(game_id,user["id"])).fetchall()]
            valid_all=_normalize_stage_ids(db,game_id,existing+ids)
            valid_set=set(valid_all); ids=[aid for aid in ids if aid in valid_set]
            _delete_choice_siblings(db,game_id,user["id"],ids)
            db.executemany("""insert into game_progress(game_id,user_id,achievement_id,completed_at) values(?,?,?,?)
            on conflict(game_id,user_id,achievement_id) do update set completed_at=excluded.completed_at""",[(game_id,user["id"],aid,t) for aid in ids])
        else:
            for aid in ids: _delete_stage_from(db,game_id,user["id"],aid)
        rows=db.execute("select achievement_id from game_progress where game_id=? and user_id=? order by completed_at,achievement_id",(game_id,user["id"])).fetchall()
    bump_game_live_scope(game_id,"stats")
    return {"ok":True,"completed":[r["achievement_id"] for r in rows]}


@app.post("/api/games/{game_id}/progress/replace")
@high_risk_operation
def extra_game_replace_progress(game_id: str, body: ProgressReplace, request: Request):
    game_id=require_extra_game(game_id); user=require_user(request); requested=validate_ids(body.achievement_ids); t=now()
    with connect_db() as db:
        ids=list(dict.fromkeys(_resolve_effective_achievement_id(db,game_id,aid) for aid in requested))
        ids=_normalize_choice_progress_ids(db,game_id,ids)
        ids=_normalize_stage_ids(db,game_id,ids)
        db.execute("delete from game_progress where game_id=? and user_id=?",(game_id,user["id"]))
        db.executemany("insert into game_progress(game_id,user_id,achievement_id,completed_at) values(?,?,?,?)",[(game_id,user["id"],aid,t) for aid in ids])
    bump_game_live_scope(game_id,"stats")
    return {"ok":True,"completed":ids}


# Shared admin save endpoint. Uses an existing /api/admin path so IIS forwards
# the request consistently, then dispatches to the selected game internally.
@app.post("/api/admin/game-achievement-save")
async def admin_game_achievement_save(request: Request):
    try:
        payload = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON format error.")
    if not isinstance(payload, dict):
        raise HTTPException(status_code=400, detail="Invalid request body.")
    game_id = str(payload.pop("game_id", "wuwa") or "wuwa").strip()
    achievement_id = str(payload.pop("achievement_id", "") or "").strip()
    try:
        body = AchievementEditPayload(**payload)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))
    return extra_game_save_achievement(require_extra_game(game_id), achievement_id, body, request)

@app.post("/api/convert-traditional")
async def convert_traditional(request: Request):
    raw=await request.body()
    if len(raw)>10_000_000: raise HTTPException(status_code=413,detail="內容過大。")
    try: value=json.loads(raw)
    except Exception: raise HTTPException(status_code=400,detail="JSON 格式錯誤。")
    return {"ok":True,"payload":convert_to_traditional(value)}

# ----- 成就資料治理中心與共用訊息中心 -----
GOVERNANCE_RULES_VERSION = "2026.06.26-final-governance-v1"
GOVERNANCE_ACTIVE_STATES = {"new", "waiting_review", "assigned", "ready", "processing", "reopened", "failed"}
GOVERNANCE_TERMINAL_STATES = {"resolved", "accepted_current", "ignored", "legal_exception", "rolled_back"}
GOVERNANCE_OPERATION_GUARD = HIGH_RISK_OPERATION_GUARD



def _json_object(value: str | None, fallback: Any) -> Any:
    try:
        parsed = json.loads(value or "")
        return parsed
    except Exception:
        return fallback


def _governance_scan_profile(rules_version: str, options: dict[str, Any] | None) -> dict[str, Any]:
    normalized_options = copy.deepcopy(options if isinstance(options, dict) else {})
    profile = {
        "rules_version": str(rules_version or ""),
        "options": normalized_options,
    }
    profile["profile_hash"] = governance_hash(profile)
    return profile


def _governance_snapshot_hash(entity_ids: list[str] | tuple[str, ...], evidence: Any) -> str:
    return governance_hash({
        "entities": sorted(str(value) for value in entity_ids if str(value)),
        "evidence": evidence if isinstance(evidence, (dict, list)) else {},
    })


GOVERNANCE_VOLATILE_SNAPSHOT_FIELDS = {"updated_at", "override_updated_at"}
GOVERNANCE_DECISION_HASH_SCHEMA = "semantic-v2"


def _governance_semantic_entity_snapshots(entity_snapshots: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for entity_id in sorted(entity_snapshots):
        snapshot = entity_snapshots[entity_id]
        if isinstance(snapshot, dict):
            snapshot = {
                str(key): value
                for key, value in snapshot.items()
                if str(key) not in GOVERNANCE_VOLATILE_SNAPSHOT_FIELDS
            }
        normalized[str(entity_id)] = snapshot
    return normalized


def _governance_legacy_decision_snapshot_hash(
    *, kind: str, entity_ids: list[str] | tuple[str, ...], evidence: Any,
    entity_snapshots: dict[str, Any], progress_count: int, relation_count: int
) -> str:
    return governance_hash({
        "kind": str(kind or ""),
        "entities": sorted(str(value) for value in entity_ids if str(value)),
        "evidence": evidence if isinstance(evidence, (dict, list)) else {},
        "entity_snapshots": {str(key): entity_snapshots[key] for key in sorted(entity_snapshots)},
        "progress_count": int(progress_count or 0),
        "relation_count": int(relation_count or 0),
    })


def _governance_decision_snapshot_hash(
    *, kind: str, entity_ids: list[str] | tuple[str, ...], evidence: Any,
    entity_snapshots: dict[str, Any], progress_count: int, relation_count: int
) -> str:
    return governance_hash({
        "kind": str(kind or ""),
        "entities": sorted(str(value) for value in entity_ids if str(value)),
        "evidence": evidence if isinstance(evidence, (dict, list)) else {},
        "entity_snapshots": _governance_semantic_entity_snapshots(entity_snapshots),
        "progress_count": int(progress_count or 0),
        "relation_count": int(relation_count or 0),
    })


def _governance_issue_snapshot_from_db(db: sqlite3.Connection, issue: sqlite3.Row | dict[str, Any]) -> tuple[str, list[str], dict[str, Any], dict[str, Any]]:
    issue_id = str(issue["id"] if isinstance(issue, sqlite3.Row) else issue.get("id") or "")
    entity_rows = db.execute(
        "select entity_id,snapshot_json from achievement_issue_entities where issue_id=? order by entity_id",
        (issue_id,),
    ).fetchall()
    entities = [str(row["entity_id"]) for row in entity_rows]
    entity_snapshots = {str(row["entity_id"]): _json_object(row["snapshot_json"], {}) for row in entity_rows}
    evidence_raw = issue["evidence_json"] if isinstance(issue, sqlite3.Row) else issue.get("evidence_json")
    if evidence_raw is None and not isinstance(issue, sqlite3.Row):
        evidence = issue.get("evidence") if isinstance(issue.get("evidence"), dict) else {}
    else:
        evidence = _json_object(evidence_raw, {})
    kind = str(issue["kind"] if isinstance(issue, sqlite3.Row) else issue.get("kind") or "")
    progress_count = int(issue["progress_count"] if isinstance(issue, sqlite3.Row) else issue.get("progress_count") or 0)
    relation_count = int(issue["relation_count"] if isinstance(issue, sqlite3.Row) else issue.get("relation_count") or 0)
    snapshot_hash = _governance_decision_snapshot_hash(
        kind=kind, entity_ids=entities, evidence=evidence, entity_snapshots=entity_snapshots,
        progress_count=progress_count, relation_count=relation_count,
    )
    return snapshot_hash, entities, evidence, entity_snapshots


def _deactivate_governance_decision(
    db: sqlite3.Connection, game_id: str, fingerprint: str, *, reason: str
) -> None:
    stamp = now()
    db.execute(
        """update achievement_governance_decisions
           set active=0,invalidated_at=?,invalidation_reason=?,updated_at=?
           where game_id=? and fingerprint=? and active=1""",
        (stamp, reason, stamp, game_id, fingerprint),
    )


def _upsert_accepted_current_decision(
    db: sqlite3.Connection, *, game_id: str, issue: sqlite3.Row | dict[str, Any],
    admin_id: str | None, reason: str, source_basis: str, batch_id: str = ""
) -> dict[str, Any]:
    fingerprint = str(issue["fingerprint"] if isinstance(issue, sqlite3.Row) else issue.get("fingerprint") or "")
    snapshot_hash, entities, evidence, entity_snapshots = _governance_issue_snapshot_from_db(db, issue)
    stamp = now()
    decision_id = f"decision-{uuid.uuid4().hex}"
    issue_kind = str(issue["kind"] if isinstance(issue, sqlite3.Row) else issue.get("kind") or "")
    issue_title = str(issue["title"] if isinstance(issue, sqlite3.Row) else issue.get("title") or "")
    payload = {
        "decision_type": "accepted_current",
        "reason": reason,
        "source_basis": source_basis,
        "batch_id": batch_id,
        "snapshot_hash": snapshot_hash,
        "issue_kind": issue_kind,
        "issue_title": issue_title,
        "entity_ids": entities,
        "evidence": evidence,
        "entity_snapshots": entity_snapshots,
        "hash_schema": GOVERNANCE_DECISION_HASH_SCHEMA,
        "accepted_at": stamp,
    }
    db.execute(
        """insert into achievement_governance_decisions(
               id,game_id,fingerprint,decision_type,reason,snapshot_hash,evidence_json,active,created_by,created_at,updated_at,invalidated_at,invalidation_reason
           ) values(?,?,?,?,?,?,?,1,?,?,?,null,'')
           on conflict(game_id,fingerprint,decision_type) do update set
               reason=excluded.reason,snapshot_hash=excluded.snapshot_hash,evidence_json=excluded.evidence_json,active=1,
               created_by=excluded.created_by,updated_at=excluded.updated_at,invalidated_at=null,invalidation_reason=''""",
        (
            decision_id, game_id, fingerprint, "accepted_current", reason, snapshot_hash,
            json.dumps(payload, ensure_ascii=False), admin_id or None, stamp, stamp,
        ),
    )
    stored = db.execute(
        "select id from achievement_governance_decisions where game_id=? and fingerprint=? and decision_type='accepted_current'",
        (game_id, fingerprint),
    ).fetchone()
    payload["decision_id"] = str(stored["id"] if stored else decision_id)
    return payload


def _migrate_governance_semantic_decision_hashes(db: sqlite3.Connection) -> dict[str, int]:
    migration = "2026-07-12-governance-semantic-decision-hash-v1"
    if db.execute("select 1 from schema_migrations where name=?", (migration,)).fetchone():
        return {"converted": 0, "reactivated": 0}
    converted = 0
    reactivated = 0
    decisions = db.execute(
        """select * from achievement_governance_decisions
           where decision_type='accepted_current'
             and (active=1 or invalidation_reason='issue_evidence_changed')"""
    ).fetchall()
    for decision in decisions:
        issue = db.execute(
            "select * from achievement_issues where game_id=? and fingerprint=?",
            (decision["game_id"], decision["fingerprint"]),
        ).fetchone()
        if not issue:
            continue
        payload = _json_object(decision["evidence_json"], {})
        old_snapshots = payload.get("entity_snapshots") if isinstance(payload.get("entity_snapshots"), dict) else {}
        old_entities = sorted(str(value) for value in payload.get("entity_ids") or [] if str(value))
        old_evidence = payload.get("evidence") if isinstance(payload.get("evidence"), (dict, list)) else {}
        current_hash, current_entities, current_evidence, current_snapshots = _governance_issue_snapshot_from_db(db, issue)
        if str(payload.get("issue_kind") or "") != str(issue["kind"] or ""):
            continue
        if old_entities != sorted(current_entities) or old_evidence != current_evidence:
            continue
        legacy_hash = _governance_legacy_decision_snapshot_hash(
            kind=str(issue["kind"] or ""), entity_ids=current_entities, evidence=current_evidence,
            entity_snapshots=old_snapshots, progress_count=int(issue["progress_count"] or 0),
            relation_count=int(issue["relation_count"] or 0),
        )
        if legacy_hash != str(decision["snapshot_hash"] or ""):
            continue
        if governance_hash(_governance_semantic_entity_snapshots(old_snapshots)) != governance_hash(
            _governance_semantic_entity_snapshots(current_snapshots)
        ):
            continue
        payload.update({
            "snapshot_hash": current_hash,
            "entity_snapshots": current_snapshots,
            "hash_schema": GOVERNANCE_DECISION_HASH_SCHEMA,
            "migrated_by": migration,
        })
        was_inactive = not bool(decision["active"])
        stamp = now()
        db.execute(
            """update achievement_governance_decisions
               set snapshot_hash=?,evidence_json=?,active=1,updated_at=?,invalidated_at=null,invalidation_reason=''
               where id=?""",
            (current_hash, json.dumps(payload, ensure_ascii=False), stamp, decision["id"]),
        )
        resolution = _json_object(issue["resolution_json"], {})
        resolution.update({
            "action": "keep",
            "completion_type": "accepted_current",
            "verification": "same_semantic_evidence_suppressed_until_change",
            "decision_id": decision["id"],
            "snapshot_hash": current_hash,
            "hash_schema": GOVERNANCE_DECISION_HASH_SCHEMA,
            "migration": migration,
        })
        db.execute(
            """update achievement_issues
               set state='accepted_current',resolution_json=?,resolved_by=coalesce(resolved_by,?),resolved_at=coalesce(resolved_at,?)
               where id=?""",
            (json.dumps(resolution, ensure_ascii=False), decision["created_by"], stamp, issue["id"]),
        )
        converted += 1
        reactivated += 1 if was_inactive else 0
    details = {"converted": converted, "reactivated": reactivated, "hash_schema": GOVERNANCE_DECISION_HASH_SCHEMA}
    db.execute(
        "insert into schema_migrations(name,applied_at,details_json) values(?,?,?)",
        (migration, now(), json.dumps(details, ensure_ascii=False)),
    )
    return {"converted": converted, "reactivated": reactivated}


def _repair_final_governance_lifecycle(db: sqlite3.Connection) -> dict[str, int]:
    """Repair legacy keep decisions and invalidate pre-final governance drafts.

    A legacy keep is only converted when the batch post-state hash still equals
    the issue's latest scan state.  If any protected data changed afterwards,
    the issue remains reopened for administrator review.
    """
    migration = "2026-06-26-final-governance-v1"
    if db.execute("select 1 from schema_migrations where name=?", (migration,)).fetchone():
        return {"accepted_current_repaired": 0, "drafts_expired": 0}
    rows = db.execute(
        """select i.*,a.id action_id,a.status action_status,b.id batch_id,b.admin_user_id,b.result_json,
                  b.completed_at,d.reason draft_reason,s.summary_json scan_summary
           from achievement_issues i
           join achievement_resolution_actions a on a.issue_id=i.id and a.action='keep'
           join achievement_resolution_batches b on b.id=a.batch_id and b.status='completed'
           left join achievement_resolution_drafts d on d.id=b.draft_id
           left join achievement_scan_runs s on s.id=i.last_scan_id
           where not exists (
             select 1 from achievement_resolution_actions newer
             join achievement_resolution_batches nb on nb.id=newer.batch_id and nb.status='completed'
             where newer.issue_id=i.id and (coalesce(nb.completed_at,0)>coalesce(b.completed_at,0)
               or (coalesce(nb.completed_at,0)=coalesce(b.completed_at,0) and newer.id>a.id))
           )"""
    ).fetchall()
    repaired = 0
    for issue in rows:
        batch_result = _json_object(issue["result_json"], {})
        scan_summary = _json_object(issue["scan_summary"], {})
        batch_hash = str(batch_result.get("post_state_hash") or "")
        scan_hash = str(scan_summary.get("state_hash") or "")
        if not batch_hash or not scan_hash or batch_hash != scan_hash:
            continue
        decision = _upsert_accepted_current_decision(
            db, game_id=str(issue["game_id"]), issue=issue,
            admin_id=str(issue["admin_user_id"] or issue["resolved_by"] or "") or None,
            reason=str(issue["draft_reason"] or "沿用管理員已確認的保留目前資料決策"),
            source_basis="final_governance_migration", batch_id=str(issue["batch_id"] or ""),
        )
        resolution = {
            "action": "keep",
            "completion_type": "accepted_current",
            "verification": "same_evidence_suppressed_until_change",
            "decision_id": decision.get("decision_id"),
            "snapshot_hash": decision.get("snapshot_hash"),
            "batch_id": str(issue["batch_id"] or ""),
            "migration": migration,
        }
        db.execute(
            "update achievement_issues set state='accepted_current',resolution_json=?,resolved_by=?,resolved_at=? where id=?",
            (json.dumps(resolution, ensure_ascii=False), issue["admin_user_id"] or issue["resolved_by"], now(), issue["id"]),
        )
        db.execute(
            "update achievement_resolution_actions set status='accepted_current',after_json=? where id=?",
            (json.dumps({"decision_state": "accepted_current", "data_changed": False, "migration": migration}, ensure_ascii=False), issue["action_id"]),
        )
        repaired += 1
    expired = 0
    for draft in db.execute("select id,plan_json from achievement_resolution_drafts where status='preview_ready'").fetchall():
        plan = _json_object(draft["plan_json"], {})
        if str(plan.get("rules_version") or "") != GOVERNANCE_RULES_VERSION:
            db.execute("update achievement_resolution_drafts set status='expired',updated_at=? where id=?", (now(), draft["id"]))
            expired += 1
    details = {"accepted_current_repaired": repaired, "drafts_expired": expired, "rules_version": GOVERNANCE_RULES_VERSION}
    db.execute(
        "insert into schema_migrations(name,applied_at,details_json) values(?,?,?)",
        (migration, now(), json.dumps(details, ensure_ascii=False)),
    )
    return details


def _governance_refresh_scan_summary(db: sqlite3.Connection, game_id: str, scan_id: str) -> dict[str, Any]:
    row = db.execute(
        "select summary_json from achievement_scan_runs where id=? and game_id=?",
        (scan_id, game_id),
    ).fetchone()
    if not row:
        return {}
    summary = _json_object(row["summary_json"], {})
    states = {
        str(state): int(count)
        for state, count in db.execute(
            "select state,count(*) from achievement_issues where game_id=? and last_scan_id=? group by state",
            (game_id, scan_id),
        ).fetchall()
    }
    active_count = sum(states.get(state, 0) for state in GOVERNANCE_ACTIVE_STATES)
    summary.update({
        "active_issue_count": active_count,
        "accepted_current_count": states.get("accepted_current", 0),
        "ignored_count": states.get("ignored", 0),
        "legal_exception_count": states.get("legal_exception", 0),
        "resolved_count": states.get("resolved", 0),
        "state_counts": states,
    })
    status = "no_issues" if active_count == 0 else "completed"
    db.execute(
        "update achievement_scan_runs set status=?,summary_json=? where id=? and game_id=?",
        (status, json.dumps(summary, ensure_ascii=False), scan_id, game_id),
    )
    return summary


def _migrate_completed_keep_decisions(db: sqlite3.Connection) -> int:
    migration = "2026-06-26-governance-accepted-current-v1"
    if db.execute("select 1 from schema_migrations where name=?", (migration,)).fetchone():
        return 0
    rows = db.execute(
        """select i.*,b.id batch_id,b.admin_user_id,d.reason draft_reason,b.completed_at,a.id action_id
           from achievement_issues i
           join achievement_resolution_actions a on a.issue_id=i.id and a.action='keep' and a.status in ('no_change','accepted_current','decision_only')
           join achievement_resolution_batches b on b.id=a.batch_id and b.status='completed'
           left join achievement_resolution_drafts d on d.id=b.draft_id
           where not exists (
               select 1 from achievement_resolution_actions newer
               join achievement_resolution_batches newer_batch on newer_batch.id=newer.batch_id and newer_batch.status='completed'
               where newer.issue_id=i.id and (coalesce(newer_batch.completed_at,0)>coalesce(b.completed_at,0)
                   or (coalesce(newer_batch.completed_at,0)=coalesce(b.completed_at,0) and newer.id>a.id))
           )
           order by coalesce(b.completed_at,0),a.id"""
    ).fetchall()
    migrated = 0
    for issue in rows:
        if str(issue["state"] or "") == "legal_exception":
            continue
        active_exception = db.execute(
            "select 1 from achievement_exception_rules where game_id=? and fingerprint=? and active=1",
            (issue["game_id"], issue["fingerprint"]),
        ).fetchone()
        if active_exception:
            continue
        payload = _upsert_accepted_current_decision(
            db, game_id=str(issue["game_id"]), issue=issue,
            admin_id=str(issue["admin_user_id"] or issue["resolved_by"] or "") or None,
            reason=str(issue["draft_reason"] or "沿用已完成的保留目前資料決策"),
            source_basis="upgrade_migration", batch_id=str(issue["batch_id"] or ""),
        )
        db.execute(
            "update achievement_issues set state='accepted_current',resolution_json=?,resolved_by=?,resolved_at=? where id=?",
            (json.dumps(payload, ensure_ascii=False), issue["admin_user_id"], int(issue["completed_at"] or now()), issue["id"]),
        )
        migrated += 1
    db.execute(
        "insert into schema_migrations(name,applied_at,details_json) values(?,?,?)",
        (migration, now(), json.dumps({"migrated_keep_decisions": migrated}, ensure_ascii=False)),
    )
    return migrated


def _governance_context(game_id: str, options: dict[str, Any] | None = None) -> dict[str, Any]:
    options = options or {}
    raw_catalog_items = _load_catalog_items_for_health(game_id)
    with connect_db() as db:
        raw_database_items = [dict(row) for row in db.execute(
            "select achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order from game_catalog_items where game_id=? order by source_order,achievement_id",
            (game_id,),
        ).fetchall()]
        effective_catalog_items = _effective_catalog_items(db, game_id)
        progress_rows = [dict(row) for row in db.execute(
            "select user_id,achievement_id,completed_at from game_progress where game_id=? order by user_id,achievement_id",
            (game_id,),
        ).fetchall()]
        relation_rows = [dict(row) for row in db.execute(
            "select group_id,achievement_id,relation_type,stage_order from game_achievement_choice_groups where game_id=? order by relation_type,group_id,stage_order,achievement_id",
            (game_id,),
        ).fetchall()]
        aliases = [dict(row) for row in db.execute(
            "select alias_id,canonical_id,reason,created_at from achievement_id_aliases where game_id=? order by alias_id",
            (game_id,),
        ).fetchall()]
        overrides = [dict(row) for row in db.execute(
            "select * from game_achievement_overrides where game_id=? order by achievement_id",
            (game_id,),
        ).fetchall()]
        deleted_rows = [dict(row) for row in db.execute(
            "select * from game_deleted_achievements where game_id=? order by achievement_id",
            (game_id,),
        ).fetchall()]
        identities = [dict(row) for row in db.execute(
            "select * from achievement_identities where game_id=? order by internal_id",
            (game_id,),
        ).fetchall()]
        source_ids = [dict(row) for row in db.execute(
            "select * from achievement_source_ids where game_id=? order by source_name,source_id",
            (game_id,),
        ).fetchall()]
        registered_fields = {
            str(row["field_name"])
            for row in db.execute(
                "select field_name from achievement_field_registry where game_id=? and active=1",
                (game_id,),
            ).fetchall()
        }

    # Governance uses the same effective rows that the administrator and public
    # pages show.  This makes an administrator override or a legitimate manual
    # row visible to both management and governance instead of creating two
    # contradictory data layers.
    catalog_items: list[dict[str, Any]] = []
    database_items: list[dict[str, Any]] = []
    for row in effective_catalog_items:
        value = dict(row)
        value["id"] = str(value.get("achievement_id") or value.get("id") or "")
        value["tags"] = value.get("tags") if isinstance(value.get("tags"), list) else _json_object(value.get("tags_json"), [])
        value["hidden"] = bool(value.get("hidden"))
        catalog_items.append(value)
        database_items.append(dict(value))

    config = get_game_config(game_id) or {}
    result = scan_governance(
        game_id=game_id,
        catalog_items=catalog_items,
        database_items=database_items,
        progress_rows=progress_rows,
        relation_rows=relation_rows,
        aliases=aliases,
        # Effective management rows are already part of catalog_items. Passing
        # overrides again would incorrectly label legitimate manual rows as
        # orphan overrides.
        overrides=[],
        deleted_rows=deleted_rows,
        minimum_catalog_count=max(1, int(config.get("minimumCatalogCount") or 1)),
        similarity_threshold=float(options.get("similarity_threshold") or 0.94),
        registered_fields=registered_fields,
        identity_rows=identities,
        source_id_rows=source_ids,
    )

    def canonical_storage_row(row: dict[str, Any]) -> dict[str, Any]:
        tags = row.get("tags") if isinstance(row.get("tags"), list) else _json_object(row.get("tags_json"), [])
        return {
            "achievement_id": str(row.get("id") or row.get("achievement_id") or ""),
            "name": str(row.get("name") or ""),
            "condition": str(row.get("condition") or ""),
            "version": str(row.get("version") or "未標示"),
            "category": str(row.get("category") or "未辨識分類"),
            "reward": int(row.get("reward") or 0),
            "hidden": bool(row.get("hidden")),
            "tags": tags,
            "source": str(row.get("source") or "catalog"),
            "source_order": int(row.get("sourceOrder") if row.get("sourceOrder") is not None else row.get("source_order") or 0),
        }

    # Keep raw JSON/SQLite consistency checks separate from effective overrides.
    # A management override should not make governance claim that the official
    # JSON and SQLite are inconsistent.
    raw_json_by_id = {str(row.get("id") or row.get("achievement_id") or ""): canonical_storage_row(row) for row in raw_catalog_items if str(row.get("id") or row.get("achievement_id") or "")}
    raw_db_by_id = {str(row.get("achievement_id") or ""): canonical_storage_row(row) for row in raw_database_items if str(row.get("achievement_id") or "")}
    storage_issues: list[dict[str, Any]] = []
    for aid in sorted(set(raw_json_by_id) - set(raw_db_by_id)):
        storage_issues.append(make_governance_issue(game_id,"json_only","error","blocked","JSON 有資料但資料庫沒有",f"成就 {aid} 尚未寫入 SQLite。",[aid],{"json":raw_json_by_id[aid]},["sync_json_to_database"],auto_fixable=True))
    for aid in sorted(set(raw_db_by_id) - set(raw_json_by_id)):
        storage_issues.append(make_governance_issue(game_id,"database_only","error","blocked","資料庫有資料但 JSON 沒有",f"成就 {aid} 只存在 SQLite。",[aid],{"database":raw_db_by_id[aid]},["sync_database_to_json","archive_database_row"]))
    for aid in sorted(set(raw_json_by_id) & set(raw_db_by_id)):
        differences={key:{"json":raw_json_by_id[aid][key],"database":raw_db_by_id[aid][key]} for key in raw_json_by_id[aid] if raw_json_by_id[aid][key]!=raw_db_by_id[aid].get(key)}
        if differences:
            storage_issues.append(make_governance_issue(game_id,"json_database_mismatch","warning","needs_review","JSON 與資料庫內容不一致",f"成就 {aid} 有 {len(differences)} 個欄位不一致。",[aid],{"differences":differences},["sync_json_to_database","sync_database_to_json","manual_edit"]))
    # Remove duplicate storage issues emitted by the effective-row scanner and
    # append the authoritative raw-storage result.
    storage_kinds={"json_only","database_only","json_database_mismatch"}
    result["issues"]=[issue for issue in result.get("issues") or [] if issue.get("kind") not in storage_kinds]+storage_issues

    category_by_kind = {
        "progress": {"orphan_progress", "deleted_with_progress"},
        "relations": {"orphan_relation", "relation_too_small", "duplicate_relation_member", "invalid_stage_order", "multiple_relation_groups"},
        "aliases": {"alias_self_reference", "alias_dangling", "alias_cycle", "alias_chain", "alias_source_still_exists"},
        "sources": {"duplicate_source_id", "orphan_override", "orphan_identity", "orphan_source_mapping"},
    }
    disabled_kinds: set[str] = set()
    if not bool(options.get("include_progress", True)): disabled_kinds |= category_by_kind["progress"]
    if not bool(options.get("include_relations", True)): disabled_kinds |= category_by_kind["relations"]
    if not bool(options.get("include_aliases", True)): disabled_kinds |= category_by_kind["aliases"]
    if not bool(options.get("include_sources", True)): disabled_kinds |= category_by_kind["sources"]
    if not bool(options.get("include_catalog", True)):
        all_special=set().union(*category_by_kind.values())
        disabled_kinds |= {str(issue.get("kind") or "") for issue in result.get("issues") or [] if str(issue.get("kind") or "") not in all_special}
    if disabled_kinds:
        result["issues"]=[issue for issue in result.get("issues") or [] if str(issue.get("kind") or "") not in disabled_kinds]

    by_severity={name:sum(1 for issue in result["issues"] if issue.get("severity")==name) for name in ("error","warning","info")}
    by_risk={name:sum(1 for issue in result["issues"] if issue.get("risk")==name) for name in ("blocked","needs_review","safe")}
    by_kind: dict[str,int]={}
    for issue in result["issues"]:
        kind=str(issue.get("kind") or "unknown"); by_kind[kind]=by_kind.get(kind,0)+1
    result.update({"by_severity":by_severity,"by_risk":by_risk,"by_kind":by_kind})

    state_components={
        "catalog":catalog_items,
        "raw_database":raw_database_items,
        "progress":progress_rows,
        "relations":relation_rows,
        "aliases":aliases,
        "overrides":overrides,
        "deleted":deleted_rows,
        "identities":identities,
        "source_ids":source_ids,
    }
    component_hashes={key:governance_hash(value) for key,value in state_components.items()}
    result.update({
        "catalog_items": catalog_items,
        "database_items": database_items,
        "raw_catalog_items": raw_catalog_items,
        "raw_database_items": raw_database_items,
        "progress_rows": progress_rows,
        "relation_rows": relation_rows,
        "aliases": aliases,
        "overrides": overrides,
        "deleted_rows": deleted_rows,
        "identities": identities,
        "source_ids": source_ids,
        "registered_fields": sorted(registered_fields),
        "catalog_hash": component_hashes["catalog"],
        "database_hash": component_hashes["raw_database"],
        "state_hash": governance_hash(component_hashes),
        "state_component_hashes": component_hashes,
    })
    return result


def _governance_issue_payload(row: sqlite3.Row) -> dict[str, Any]:
    return {
        "id": row["id"],
        "game_id": row["game_id"],
        "fingerprint": row["fingerprint"],
        "kind": row["kind"],
        "severity": row["severity"],
        "risk": row["risk"],
        "title": row["title"],
        "message": row["message"],
        "state": row["state"],
        "first_seen_at": int(row["first_seen_at"] or 0),
        "last_seen_at": int(row["last_seen_at"] or 0),
        "occurrence_count": int(row["occurrence_count"] or 0),
        "last_scan_id": row["last_scan_id"] or "",
        "progress_count": int(row["progress_count"] or 0),
        "relation_count": int(row["relation_count"] or 0),
        "auto_fixable": bool(row["auto_fixable"]),
        "evidence": _json_object(row["evidence_json"], {}),
        "suggested_actions": _json_object(row["actions_json"], []),
        "resolution": _json_object(row["resolution_json"], {}),
        "resolved_by": row["resolved_by"] or "",
        "resolved_at": row["resolved_at"],
    }


GOVERNANCE_EXECUTOR = ThreadPoolExecutor(max_workers=2, thread_name_prefix="governance-scan")

def _governance_scan_payload(row: sqlite3.Row) -> dict[str, Any]:
    return {
        **dict(row),
        "options": _json_object(row["options_json"], {}),
        "summary": _json_object(row["summary_json"], {}),
    }


def _update_governance_scan_progress(scan_id: str, *, status: str | None = None, phase: str = "", percent: int | None = None, message: str = "", extra: dict[str, Any] | None = None) -> None:
    with connect_db() as db:
        row = db.execute("select summary_json from achievement_scan_runs where id=?", (scan_id,)).fetchone()
        summary = _json_object(row["summary_json"], {}) if row else {}
        progress = dict(summary.get("progress") or {})
        if phase:
            progress["phase"] = phase
        if percent is not None:
            progress["percent"] = max(0, min(100, int(percent)))
        if message:
            progress["message"] = message
        progress["updated_at"] = now()
        summary["progress"] = progress
        if extra:
            summary.update(extra)
        if status:
            db.execute("update achievement_scan_runs set status=?,summary_json=? where id=?", (status, json.dumps(summary, ensure_ascii=False), scan_id))
        else:
            db.execute("update achievement_scan_runs set summary_json=? where id=?", (json.dumps(summary, ensure_ascii=False), scan_id))


def _execute_governance_scan(scan_id: str, game_id: str, admin_id: str, options: dict[str, Any], started: int) -> None:
    if not GOVERNANCE_OPERATION_GUARD.acquire(blocking=False):
        failure={"error":"governance_operation_busy","progress":{"phase":"failed","percent":100,"message":"另一個成就管理或治理操作正在執行。","updated_at":now()}}
        with connect_db() as db:
            db.execute("update achievement_scan_runs set status='failed',summary_json=?,completed_at=? where id=?",(json.dumps(failure,ensure_ascii=False),now(),scan_id))
        return
    try:
        _update_governance_scan_progress(scan_id, status="scanning", phase="load_data", percent=10, message="正在載入成就、進度與關聯資料……")
        context = _governance_context(game_id, options)
        _update_governance_scan_progress(scan_id, phase="analyze", percent=55, message="正在分析重複、異常與資料一致性……")
        current_fingerprints = {str(issue["fingerprint"]) for issue in context["issues"]}
        catalog_by_id = {
            str(row.get("id") or row.get("achievement_id") or ""): row
            for row in context["catalog_items"]
            if str(row.get("id") or row.get("achievement_id") or "")
        }
        with connect_db() as db:
            exceptions = {
                str(row["fingerprint"]): dict(row)
                for row in db.execute(
                    "select * from achievement_exception_rules where game_id=? and active=1",
                    (game_id,),
                ).fetchall()
            }
            accepted_decisions = {
                str(row["fingerprint"]): dict(row)
                for row in db.execute(
                    """select * from achievement_governance_decisions
                       where game_id=? and decision_type='accepted_current' and active=1""",
                    (game_id,),
                ).fetchall()
            }
            accepted_by_signature: dict[tuple[str, tuple[str, ...]], dict[str, Any]] = {}
            for decision in accepted_decisions.values():
                decision_payload = _json_object(decision.get("evidence_json"), {})
                signature = (
                    str(decision_payload.get("issue_kind") or ""),
                    tuple(sorted(str(value) for value in decision_payload.get("entity_ids") or [] if str(value))),
                )
                if signature[0] and signature[1]:
                    accepted_by_signature[signature] = decision
            matched_accepted_decision_ids: set[str] = set()
            persisted_count = 0
            active_count = 0
            legal_exception_count = 0
            accepted_current_count = 0
            ignored_count = 0
            resolved_count = 0
            severity_counts = {"error": 0, "warning": 0, "info": 0}
            safe_fix_count = 0
            for issue in context["issues"]:
                fingerprint = str(issue["fingerprint"])
                issue_entity_ids = [str(value) for value in issue.get("entity_ids") or [] if str(value)]
                exception_snapshot_hash = _governance_snapshot_hash(issue_entity_ids, issue.get("evidence") or {})
                decision_snapshot_hash = _governance_decision_snapshot_hash(
                    kind=str(issue.get("kind") or ""),
                    entity_ids=issue_entity_ids,
                    evidence=issue.get("evidence") or {},
                    entity_snapshots={entity_id: catalog_by_id.get(entity_id, {}) for entity_id in issue_entity_ids},
                    progress_count=int(issue.get("progress_count") or 0),
                    relation_count=int(issue.get("relation_count") or 0),
                )
                exception = exceptions.get(fingerprint)
                signature = (
                    str(issue.get("kind") or ""),
                    tuple(sorted(str(value) for value in issue.get("entity_ids") or [] if str(value))),
                )
                accepted_decision = accepted_decisions.get(fingerprint) or accepted_by_signature.get(signature)
                state = "new"
                if exception:
                    same_snapshot = str(exception.get("snapshot_hash") or "") == exception_snapshot_hash
                    if bool(exception.get("permanent")) or same_snapshot or not bool(exception.get("recheck_on_change")):
                        state = "legal_exception"
                    else:
                        state = "reopened"
                        db.execute("update achievement_exception_rules set active=0,updated_at=? where id=?", (started, exception["id"]))
                elif accepted_decision:
                    matched_accepted_decision_ids.add(str(accepted_decision["id"]))
                    same_fingerprint = str(accepted_decision.get("fingerprint") or "") == fingerprint
                    same_snapshot = str(accepted_decision.get("snapshot_hash") or "") == decision_snapshot_hash
                    if same_fingerprint and same_snapshot:
                        state = "accepted_current"
                    else:
                        state = "reopened"
                        db.execute(
                            """update achievement_governance_decisions set active=0,invalidated_at=?,
                               invalidation_reason='issue_evidence_changed',updated_at=? where id=?""",
                            (started, started, accepted_decision["id"]),
                        )
                        if str(accepted_decision.get("fingerprint") or "") != fingerprint:
                            db.execute(
                                """update achievement_issues set state='resolved',resolved_at=?,resolution_json=?
                                   where game_id=? and fingerprint=? and state='accepted_current'""",
                                (
                                    started,
                                    json.dumps({
                                        "reason": "accepted_issue_evidence_changed",
                                        "reopened_as_fingerprint": fingerprint,
                                        "scan_id": scan_id,
                                    }, ensure_ascii=False),
                                    game_id,
                                    accepted_decision["fingerprint"],
                                ),
                            )
                existing = db.execute(
                    "select * from achievement_issues where game_id=? and fingerprint=?",
                    (game_id, fingerprint),
                ).fetchone()
                issue_id = existing["id"] if existing else f"issue-{uuid.uuid4().hex}"
                if existing:
                    previous_state = str(existing["state"] or "new")
                    if state == "new":
                        state = "reopened" if previous_state in GOVERNANCE_TERMINAL_STATES else previous_state
                        if state not in GOVERNANCE_ACTIVE_STATES:
                            state = "waiting_review"
                    db.execute(
                        """update achievement_issues set kind=?,severity=?,risk=?,title=?,message=?,state=?,last_seen_at=?,
                           occurrence_count=occurrence_count+1,last_scan_id=?,progress_count=?,relation_count=?,auto_fixable=?,
                           evidence_json=?,actions_json=? where id=?""",
                        (
                            issue["kind"], issue["severity"], issue["risk"], issue["title"], issue["message"], state,
                            started, scan_id, int(issue.get("progress_count") or 0), int(issue.get("relation_count") or 0),
                            1 if issue.get("auto_fixable") else 0, json.dumps(issue.get("evidence") or {}, ensure_ascii=False),
                            json.dumps(issue.get("suggested_actions") or [], ensure_ascii=False), issue_id,
                        ),
                    )
                else:
                    db.execute(
                        """insert into achievement_issues(id,game_id,fingerprint,kind,severity,risk,title,message,state,first_seen_at,last_seen_at,
                           occurrence_count,last_scan_id,progress_count,relation_count,auto_fixable,evidence_json,actions_json,resolution_json)
                           values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            issue_id, game_id, fingerprint, issue["kind"], issue["severity"], issue["risk"], issue["title"],
                            issue["message"], state, started, started, 1, scan_id, int(issue.get("progress_count") or 0),
                            int(issue.get("relation_count") or 0), 1 if issue.get("auto_fixable") else 0,
                            json.dumps(issue.get("evidence") or {}, ensure_ascii=False),
                            json.dumps(issue.get("suggested_actions") or [], ensure_ascii=False), "{}",
                        ),
                    )
                db.execute("delete from achievement_issue_entities where issue_id=?", (issue_id,))
                for entity_id in issue.get("entity_ids") or []:
                    snapshot = catalog_by_id.get(str(entity_id), {})
                    db.execute(
                        "insert into achievement_issue_entities(issue_id,entity_type,entity_id,snapshot_json) values(?,?,?,?)",
                        (issue_id, "achievement", str(entity_id), json.dumps(snapshot, ensure_ascii=False)),
                    )
                persisted_count += 1
                if state == "legal_exception":
                    legal_exception_count += 1
                elif state == "accepted_current":
                    accepted_current_count += 1
                elif state == "ignored":
                    ignored_count += 1
                elif state == "resolved":
                    resolved_count += 1
                elif state in GOVERNANCE_ACTIVE_STATES:
                    active_count += 1
                    severity_counts[issue["severity"]] = severity_counts.get(issue["severity"], 0) + 1
                    if issue.get("auto_fixable"):
                        safe_fix_count += 1
            for decision in accepted_decisions.values():
                decision_id = str(decision["id"])
                if decision_id in matched_accepted_decision_ids:
                    continue
                db.execute(
                    """update achievement_governance_decisions set active=0,invalidated_at=?,
                       invalidation_reason='issue_not_reproduced',updated_at=? where id=?""",
                    (started, started, decision_id),
                )
                db.execute(
                    """update achievement_issues set state='resolved',resolved_at=?,resolution_json=?
                       where game_id=? and fingerprint=? and state='accepted_current'""",
                    (
                        started,
                        json.dumps({"reason": "accepted_issue_not_reproduced", "scan_id": scan_id}, ensure_ascii=False),
                        game_id,
                        decision["fingerprint"],
                    ),
                )
            if current_fingerprints:
                placeholders = ",".join("?" for _ in current_fingerprints)
                db.execute(
                    f"""update achievement_issues set state='resolved',resolved_at=?,resolution_json=?
                        where game_id=? and fingerprint not in ({placeholders}) and state in ('new','waiting_review','assigned','ready','processing','reopened','failed')""",
                    (started, json.dumps({"reason": "not_reproduced_in_latest_scan", "scan_id": scan_id}, ensure_ascii=False), game_id, *sorted(current_fingerprints)),
                )
            else:
                db.execute(
                    """update achievement_issues set state='resolved',resolved_at=?,resolution_json=?
                       where game_id=? and state in ('new','waiting_review','assigned','ready','processing','reopened','failed')""",
                    (started, json.dumps({"reason": "not_reproduced_in_latest_scan", "scan_id": scan_id}, ensure_ascii=False), game_id),
                )
            summary = {
                "scan_profile": _governance_scan_profile(GOVERNANCE_RULES_VERSION, options),
                "issue_count": persisted_count,
                "active_issue_count": active_count,
                "legal_exception_count": legal_exception_count,
                "accepted_current_count": accepted_current_count,
                "ignored_count": ignored_count,
                "resolved_count": resolved_count,
                "error_count": severity_counts.get("error", 0),
                "warning_count": severity_counts.get("warning", 0),
                "info_count": severity_counts.get("info", 0),
                "safe_fix_count": safe_fix_count,
                "catalog_count": context["catalog_count"],
                "database_count": context["database_count"],
                "progress_count": context["progress_count"],
                "relation_count": context["relation_count"],
                "rule_count": context["rule_count"],
                "state_hash": context["state_hash"],
                "state_component_hashes": context["state_component_hashes"],
                "duration_ms": max(0, int((time.time() - started) * 1000)),
                "progress": {"phase": "completed", "percent": 100, "message": "掃描完成。", "updated_at": now()},
            }
            status = "no_issues" if not active_count else "completed"
            db.execute(
                "update achievement_scan_runs set status=?,catalog_hash=?,database_hash=?,summary_json=?,completed_at=? where id=?",
                (status, context["catalog_hash"], context["database_hash"], json.dumps(summary, ensure_ascii=False), now(), scan_id),
            )
        log_admin_action(
            admin_id, "governance_scan_completed", details=f"game={game_id}; scan={scan_id}; active={active_count}",
            category="catalog", game_id=game_id, target_type="governance_scan", target_id=scan_id,
            summary="完成成就資料治理掃描" if active_count else "完成掃描，未發現重複或異常",
            after=summary, metadata={"rules_version": GOVERNANCE_RULES_VERSION, "options": options}, locked=True,
        )
    except Exception as exc:
        failure = {
            "error": f"{type(exc).__name__}: {exc}",
            "progress": {"phase": "failed", "percent": 100, "message": "掃描失敗。", "updated_at": now()},
        }
        with connect_db() as db:
            db.execute(
                "update achievement_scan_runs set status='failed',summary_json=?,completed_at=? where id=?",
                (json.dumps(failure, ensure_ascii=False), now(), scan_id),
            )
        log_admin_action(
            admin_id, "governance_scan_failed", details=f"game={game_id}; scan={scan_id}; error={exc}",
            category="catalog", status="failed", game_id=game_id, target_type="governance_scan", target_id=scan_id,
            summary="成就資料治理掃描失敗", error_message=str(exc), locked=True,
        )
    finally:
        GOVERNANCE_OPERATION_GUARD.release()


def _create_governance_scan(game_id: str, admin_id: str, options: dict[str, Any], *, background: bool = True) -> dict[str, Any]:
    started = now()
    scan_id = f"scan-{uuid.uuid4().hex}"
    scan_profile = _governance_scan_profile(GOVERNANCE_RULES_VERSION, options)
    initial_summary = {
        "scan_profile": scan_profile,
        "progress": {"phase": "queued", "percent": 0, "message": "掃描已排入處理佇列。", "updated_at": started}
    }
    with connect_db() as db:
        db.execute(
            """insert into achievement_scan_runs(id,game_id,status,admin_user_id,rules_version,catalog_hash,database_hash,
               options_json,summary_json,started_at,completed_at,expires_at) values(?,?,?,?,?,?,?,?,?,?,?,?)""",
            (scan_id, game_id, "queued", admin_id, GOVERNANCE_RULES_VERSION, "", "",
             json.dumps(options, ensure_ascii=False), json.dumps(initial_summary, ensure_ascii=False), started, None, started + 7 * 86400),
        )
    if background:
        GOVERNANCE_EXECUTOR.submit(_execute_governance_scan, scan_id, game_id, admin_id, options, started)
    else:
        _execute_governance_scan(scan_id, game_id, admin_id, options, started)
    with connect_db() as db:
        row = db.execute("select * from achievement_scan_runs where id=?", (scan_id,)).fetchone()
    return _governance_scan_payload(row)


@app.post("/api/games/{game_id}/admin/governance/scans")
def governance_create_scan(game_id: str, body: GovernanceScanPayload, request: Request):
    game_id = require_extra_game(game_id)
    admin = require_admin(request)
    options = body.model_dump(exclude={"background"})
    scan = _create_governance_scan(game_id, admin["id"], options, background=bool(body.background))
    return {"ok": True, "scan": scan, "scan_id": scan["id"], "status": scan["status"], "summary": scan["summary"]}


@app.get("/api/games/{game_id}/admin/governance/scans")
def governance_list_scans(game_id: str, request: Request, limit: int = 20):
    game_id = require_extra_game(game_id); require_admin(request)
    limit = max(1, min(100, int(limit)))
    with connect_db() as db:
        rows = db.execute(
            "select * from achievement_scan_runs where game_id=? order by started_at desc limit ?",
            (game_id, limit),
        ).fetchall()
    return {"ok": True, "scans": [_governance_scan_payload(row) for row in rows]}


@app.get("/api/games/{game_id}/admin/governance/scans/{scan_id}")
def governance_get_scan(game_id: str, scan_id: str, request: Request):
    game_id = require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        scan = db.execute("select * from achievement_scan_runs where id=? and game_id=?", (scan_id, game_id)).fetchone()
    if not scan:
        raise HTTPException(status_code=404, detail="找不到掃描批次。")
    return {"ok": True, "scan": _governance_scan_payload(scan)}


def _governance_issue_summary_payload(row: sqlite3.Row, entity_ids: list[str]) -> dict[str, Any]:
    return {
        "id": row["id"],
        "game_id": row["game_id"],
        "kind": row["kind"],
        "severity": row["severity"],
        "risk": row["risk"],
        "title": row["title"],
        "message": row["message"],
        "state": row["state"],
        "first_seen_at": int(row["first_seen_at"] or 0),
        "last_seen_at": int(row["last_seen_at"] or 0),
        "occurrence_count": int(row["occurrence_count"] or 0),
        "last_scan_id": row["last_scan_id"] or "",
        "progress_count": int(row["progress_count"] or 0),
        "relation_count": int(row["relation_count"] or 0),
        "auto_fixable": bool(row["auto_fixable"]),
        "suggested_actions": _json_object(row["actions_json"], []),
        "entity_ids": entity_ids,
    }


@app.get("/api/games/{game_id}/admin/governance/issues")
def governance_list_issues(
    game_id: str,
    request: Request,
    scan_id: str = "",
    state: str = "",
    severity: str = "",
    kind: str = "",
    query: str = "",
    offset: int = 0,
    limit: int = 25,
):
    game_id = require_extra_game(game_id); require_admin(request)
    offset = max(0, int(offset)); limit = max(1, min(100, int(limit)))
    # By default show only issues from the latest completed scan and only active states.
    # Historical/resolved rows remain available when a scan or state is explicitly selected.
    requested_scan_id = scan_id
    with connect_db() as lookup_db:
        if not scan_id:
            latest = lookup_db.execute(
                "select id from achievement_scan_runs where game_id=? and status in ('completed','no_issues') order by started_at desc limit 1",
                (game_id,),
            ).fetchone()
            if latest:
                scan_id = str(latest["id"])
    clauses = ["game_id=?"]; params: list[Any] = [game_id]
    if scan_id: clauses.append("last_scan_id=?"); params.append(scan_id)
    if state:
        clauses.append("state=?"); params.append(state)
    else:
        clauses.append("state in ('new','waiting_review','assigned','ready','processing','reopened','failed')")
    if severity: clauses.append("severity=?"); params.append(severity)
    if kind: clauses.append("kind=?"); params.append(kind)
    if query:
        clauses.append("(title like ? or message like ? or id in (select issue_id from achievement_issue_entities where entity_id like ?))")
        term = f"%{query}%"; params.extend([term, term, term])
    where = " and ".join(clauses)
    with connect_db() as db:
        all_rows = db.execute(
            f"select * from achievement_issues where {where}", params,
        ).fetchall()
        total = len(all_rows)
        all_issue_ids = [str(row["id"]) for row in all_rows]
        entities: dict[str, list[str]] = {}
        if all_issue_ids:
            placeholders = ",".join("?" for _ in all_issue_ids)
            for entity in db.execute(
                f"select issue_id,entity_id from achievement_issue_entities where issue_id in ({placeholders}) order by issue_id,entity_id",
                all_issue_ids,
            ).fetchall():
                entities.setdefault(str(entity["issue_id"]), []).append(str(entity["entity_id"]))
        catalog_by_id = {
            str(row["achievement_id"]): dict(row)
            for row in db.execute(
                "select achievement_id,category,tags_json from game_catalog_items where game_id=?",
                (game_id,),
            ).fetchall()
        }
        severity_rank = {"error": 0, "warning": 1, "info": 2}
        def issue_order(row: sqlite3.Row) -> tuple[Any, ...]:
            entity_ids = entities.get(str(row["id"]), [])
            if game_id == "zzz":
                entity_keys = [
                    catalog_sort_key(game_id, {**catalog_by_id.get(entity_id, {}), "achievement_id": entity_id})
                    for entity_id in entity_ids
                ]
                entity_key = min(entity_keys) if entity_keys else (2, 1, 2**63 - 1, "")
                return (*entity_key, severity_rank.get(str(row["severity"]), 3), -int(row["last_seen_at"] or 0), str(row["id"]))
            return (severity_rank.get(str(row["severity"]), 3), -int(row["last_seen_at"] or 0), str(row["id"]))
        ordered_rows = sorted(all_rows, key=issue_order)
        rows = ordered_rows[offset:offset + limit]
        facet_rows = db.execute(
            f"select severity,state,kind,count(*) c from achievement_issues where {where} group by severity,state,kind",
            params,
        ).fetchall()
    facets = {"severity": {}, "state": {}, "kind": {}}
    for row in facet_rows:
        facets["severity"][row["severity"]] = facets["severity"].get(row["severity"], 0) + int(row["c"])
        facets["state"][row["state"]] = facets["state"].get(row["state"], 0) + int(row["c"])
        facets["kind"][row["kind"]] = facets["kind"].get(row["kind"], 0) + int(row["c"])
    return {
        "ok": True,
        "issues": [_governance_issue_summary_payload(row, entities.get(row["id"], [])) for row in rows],
        "total": total,
        "offset": offset,
        "limit": limit,
        "has_more": offset + len(rows) < total,
        "facets": facets,
        "scan_id": scan_id,
        "active_only": not bool(state),
    }


@app.get("/api/games/{game_id}/admin/governance/actions")
def governance_action_contract(game_id: str, request: Request):
    require_extra_game(game_id);require_admin(request)
    return {"ok":True,"actions":governance_action_public_payload(),"rules_version":GOVERNANCE_RULES_VERSION}


@app.get("/api/games/{game_id}/admin/governance/issues/{issue_id}")
def governance_issue_detail(game_id: str, issue_id: str, request: Request):
    game_id = require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        issue = db.execute("select * from achievement_issues where id=? and game_id=?", (issue_id, game_id)).fetchone()
        if not issue:
            raise HTTPException(status_code=404, detail="找不到問題。")
        entity_rows = db.execute(
            "select entity_id,snapshot_json from achievement_issue_entities where issue_id=? order by entity_id",
            (issue_id,),
        ).fetchall()
        entities: list[dict[str, Any]] = []
        for entity_row in entity_rows:
            entity_id = str(entity_row["entity_id"] or "")
            snapshot = _json_object(entity_row["snapshot_json"], {})
            current_db = db.execute(
                "select achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order from game_catalog_items where game_id=? and achievement_id=?",
                (game_id, entity_id),
            ).fetchone()
            progress_count = int(db.execute(
                "select count(*) from game_progress where game_id=? and achievement_id=?",
                (game_id, entity_id),
            ).fetchone()[0])
            relation_rows = [dict(row) for row in db.execute(
                "select group_id,relation_type,stage_order from game_achievement_choice_groups where game_id=? and achievement_id=? order by relation_type,group_id,stage_order",
                (game_id, entity_id),
            ).fetchall()]
            aliases = [dict(row) for row in db.execute(
                "select alias_id,canonical_id,reason,created_at from achievement_id_aliases where game_id=? and (alias_id=? or canonical_id=?) order by alias_id",
                (game_id, entity_id, entity_id),
            ).fetchall()]
            override = db.execute(
                "select * from game_achievement_overrides where game_id=? and achievement_id=?",
                (game_id, entity_id),
            ).fetchone()
            current = dict(current_db) if current_db else {}
            if current:
                current["id"] = current.get("achievement_id")
                current["tags"] = _json_object(current.pop("tags_json", "[]"), [])
                current["hidden"] = bool(current.get("hidden"))
            entities.append({
                "achievement_id": entity_id,
                "snapshot": snapshot,
                "current": current,
                "progress_count": progress_count,
                "relations": relation_rows,
                "aliases": aliases,
                "override": dict(override) if override else None,
            })
    fields = [
        ("achievement_id", "成就 ID"), ("name", "名稱"), ("condition", "達成條件"),
        ("version", "版本"), ("category", "分類"), ("hidden", "隱藏狀態"),
        ("reward", "獎勵"), ("source_order", "排序"), ("source", "資料來源"),
    ]
    comparisons: list[dict[str, Any]] = []
    if entities:
        for field,label in fields:
            values=[]
            for entity in entities:
                source=entity.get("current") or entity.get("snapshot") or {}
                value=entity["achievement_id"] if field=="achievement_id" else source.get(field,source.get("sourceOrder") if field=="source_order" else None)
                values.append({"achievement_id":entity["achievement_id"],"value":value})
            distinct={json.dumps(row["value"],ensure_ascii=False,sort_keys=True,default=str) for row in values}
            comparisons.append({"field":field,"label":label,"values":values,"a":values[0]["value"] if values else None,"b":values[1]["value"] if len(values)>1 else None,"same":len(distinct)<=1})
    payload = _governance_issue_payload(issue)
    payload["entity_ids"] = [row["achievement_id"] for row in entities]
    return {"ok": True, "issue": payload, "entities": entities, "comparisons": comparisons}


@app.post("/api/games/{game_id}/admin/governance/issues/{issue_id}/state")
def governance_set_issue_state(game_id: str, issue_id: str, body: GovernanceIssueStatePayload, request: Request):
    game_id = require_extra_game(game_id); admin = require_admin(request)
    allowed = {"waiting_review", "assigned", "ready", "accepted_current", "ignored", "legal_exception", "reopened"}
    if body.state not in allowed:
        if body.state == "resolved":
            raise HTTPException(status_code=409, detail="已解決只能由實際資料修復並通過處置後驗證產生；請使用治理處置流程。")
        raise HTTPException(status_code=400, detail="不支援的問題狀態。")
    with connect_db() as db:
        issue = db.execute("select * from achievement_issues where id=? and game_id=?", (issue_id, game_id)).fetchone()
        if not issue:
            raise HTTPException(status_code=404, detail="找不到問題。")
        stamp = now()
        resolution: dict[str, Any] = {"reason": body.reason, "state": body.state, "updated_at": stamp}
        if body.state == "accepted_current":
            resolution.update(_upsert_accepted_current_decision(
                db, game_id=game_id, issue=issue, admin_id=admin["id"],
                reason=body.reason or "管理員接受目前資料", source_basis="administrator_state_change",
            ))
            db.execute(
                "update achievement_exception_rules set active=0,updated_at=? where game_id=? and fingerprint=? and active=1",
                (stamp, game_id, issue["fingerprint"]),
            )
        elif body.state == "legal_exception":
            _deactivate_governance_decision(db, game_id, str(issue["fingerprint"]), reason="replaced_by_legal_exception")
            _decision_hash, entities, evidence, _entity_snapshots = _governance_issue_snapshot_from_db(db, issue)
            snapshot_hash = _governance_snapshot_hash(entities, evidence)
            rule_id = f"exception-{uuid.uuid4().hex}"
            db.execute(
                """insert into achievement_exception_rules(id,game_id,fingerprint,reason,source_basis,snapshot_hash,permanent,recheck_on_change,active,created_by,created_at,updated_at)
                   values(?,?,?,?,?,?,?,?,1,?,?,?) on conflict(game_id,fingerprint) do update set reason=excluded.reason,source_basis=excluded.source_basis,
                   snapshot_hash=excluded.snapshot_hash,permanent=excluded.permanent,recheck_on_change=excluded.recheck_on_change,active=1,created_by=excluded.created_by,updated_at=excluded.updated_at""",
                (rule_id, game_id, issue["fingerprint"], body.reason, "administrator_review", snapshot_hash, 1 if body.permanent else 0, 1 if body.recheck_on_change else 0, admin["id"], stamp, stamp),
            )
            resolution.update({"snapshot_hash": snapshot_hash, "entity_ids": entities, "evidence": evidence})
        elif body.state == "reopened":
            _deactivate_governance_decision(db, game_id, str(issue["fingerprint"]), reason="administrator_reopened")
            db.execute(
                "update achievement_exception_rules set active=0,updated_at=? where game_id=? and fingerprint=? and active=1",
                (stamp, game_id, issue["fingerprint"]),
            )
        else:
            _deactivate_governance_decision(db, game_id, str(issue["fingerprint"]), reason=f"state_changed_to_{body.state}")
        resolved_at = stamp if body.state in GOVERNANCE_TERMINAL_STATES else None
        db.execute(
            "update achievement_issues set state=?,resolution_json=?,resolved_by=?,resolved_at=? where id=?",
            (body.state, json.dumps(resolution, ensure_ascii=False), admin["id"], resolved_at, issue_id),
        )
        _governance_refresh_scan_summary(db, game_id, str(issue["last_scan_id"] or ""))
    log_admin_action(admin["id"], "governance_issue_state", details=f"game={game_id}; issue={issue_id}; state={body.state}; reason={body.reason}", category="catalog", game_id=game_id, target_type="governance_issue", target_id=issue_id, summary="更新成就治理問題狀態", metadata=resolution, actor_ip=client_ip(request), locked=True)
    return {"ok": True, "issue_id": issue_id, "state": body.state, "resolution": resolution}


def _governance_validate_simulated_state(game_id: str, catalog_items: list[dict[str, Any]], relation_documents: dict[str, dict[str, Any]], db: sqlite3.Connection) -> dict[str, Any]:
    ids=[str(row.get("id") or row.get("achievement_id") or "").strip() for row in catalog_items]
    if any(not OFFICIAL_ACHIEVEMENT_ID_PATTERN.fullmatch(value or "") for value in ids):
        invalid=next(value for value in ids if not OFFICIAL_ACHIEVEMENT_ID_PATTERN.fullmatch(value or ""))
        raise RuntimeError(f"non_numeric_official_id:{invalid}")
    if len(ids)!=len(set(ids)):
        raise RuntimeError("duplicate_catalog_id_after_action")
    catalog_ids=set(ids)
    invalid_relations=[]
    relation_members=0
    for relation_type,document in relation_documents.items():
        groups=document.get("groups") if isinstance(document,dict) else None
        if not isinstance(groups,list):
            raise RuntimeError(f"invalid_relation_document:{relation_type}")
        seen_group_ids=set()
        for group in groups:
            if not isinstance(group,dict):
                raise RuntimeError(f"invalid_relation_group:{relation_type}")
            group_id=str(group.get("id") or "").strip()
            if not group_id or group_id in seen_group_ids:
                raise RuntimeError(f"invalid_or_duplicate_relation_group_id:{relation_type}:{group_id}")
            seen_group_ids.add(group_id)
            members=[str(value).strip() for value in group.get("achievement_ids") or [] if str(value).strip()]
            if len(members)!=len(set(members)):
                raise RuntimeError(f"duplicate_relation_member:{relation_type}:{group_id}")
            missing=[value for value in members if value not in catalog_ids]
            if missing:
                invalid_relations.append({"relation_type":relation_type,"group_id":group_id,"missing":missing})
            relation_members+=len(members)
    if invalid_relations:
        first=invalid_relations[0]
        raise RuntimeError(f"orphan_relation_after_action:{first['relation_type']}:{first['group_id']}:{','.join(first['missing'][:5])}")
    progress_orphans=int(db.execute("select count(*) from game_progress p left join game_catalog_items c on c.game_id=p.game_id and c.achievement_id=p.achievement_id where p.game_id=? and c.achievement_id is null",(game_id,)).fetchone()[0])
    return {"catalog_count":len(catalog_ids),"relation_member_count":relation_members,"progress_orphans_before_persist":progress_orphans}


def _governance_build_plan(game_id: str, scan_id: str, actions: list[dict[str, Any]], admin_id: str) -> dict[str, Any]:
    requested_ids=[str(action.get("issue_id") or "") for action in actions if str(action.get("issue_id") or "")]
    if not requested_ids:
        raise HTTPException(status_code=400,detail="尚未選擇任何處置。")
    if len(requested_ids)!=len(set(requested_ids)):
        raise HTTPException(status_code=400,detail="同一問題不可在同一草稿中重複處置。")
    with connect_db() as db:
        scan=db.execute("select * from achievement_scan_runs where id=? and game_id=?",(scan_id,game_id)).fetchone()
        if not scan:
            raise HTTPException(status_code=404,detail="找不到掃描批次。")
        if str(scan["status"] or "") not in {"completed","no_issues"}:
            raise HTTPException(status_code=409,detail="掃描尚未完成，不能建立處置預覽。")
        if str(scan["rules_version"] or "") != GOVERNANCE_RULES_VERSION:
            raise HTTPException(status_code=409,detail="此掃描使用舊版治理規則，不能建立處置；請重新執行完整掃描。")
        placeholders=",".join("?" for _ in requested_ids)
        issue_rows=db.execute(f"select * from achievement_issues where game_id=? and last_scan_id=? and id in ({placeholders})",(game_id,scan_id,*requested_ids)).fetchall()
        entities:dict[str,list[str]]={}
        for row in db.execute(f"select issue_id,entity_id from achievement_issue_entities where issue_id in ({placeholders}) order by issue_id,entity_id",requested_ids).fetchall():
            entities.setdefault(str(row["issue_id"]),[]).append(str(row["entity_id"]))
    scan_options=_json_object(scan["options_json"],{})
    scan_profile=_governance_scan_profile(str(scan["rules_version"] or ""),scan_options)
    current=_governance_context(game_id, scan_options)
    scan_summary=_json_object(scan["summary_json"],{})
    expected_state_hash=str(scan_summary.get("state_hash") or "")
    if not expected_state_hash or expected_state_hash!=current["state_hash"]:
        raise HTTPException(status_code=409,detail="成就、進度、關聯、別名或身分資料已在掃描後變更，請重新掃描。")
    issues={str(row["id"]):{**_governance_issue_payload(row),"entity_ids":entities.get(str(row["id"]),[])} for row in issue_rows}
    if len(issues)!=len(requested_ids):
        missing=sorted(set(requested_ids)-set(issues))
        raise HTTPException(status_code=400,detail=f"處置包含不屬於此掃描的問題：{', '.join(missing)}")
    normalized_actions=[]
    for raw in actions:
        issue_id=str(raw.get("issue_id") or "")
        raw_name=str(raw.get("action") or "").strip()
        if raw_name not in GOVERNANCE_SUPPORTED_ACTIONS:
            raise HTTPException(status_code=422,detail=f"問題 {issue_id} 使用未支援的處置：{raw_name or '(空白)'}")
        issue=issues[issue_id]
        allowed_actions=governance_allowed_actions_for_issue(issue.get("suggested_actions") or [])
        if raw_name not in allowed_actions:
            allowed_labels="、".join(str((GOVERNANCE_ACTION_SPECS.get(name) or {}).get("label") or name) for name in sorted(allowed_actions))
            raise HTTPException(status_code=422,detail=f"問題「{issue.get('title') or issue_id}」不允許使用此處置：{raw_name}。可用處置：{allowed_labels}")
        parameters=copy.deepcopy(raw.get("parameters") if isinstance(raw.get("parameters"),dict) else {})
        normalized_actions.append({
            "issue_id":issue_id,"action":raw_name,"parameters":parameters,
            "entity_ids":list(issue.get("entity_ids") or []),"evidence":copy.deepcopy(issue.get("evidence") or {}),
            "issue_kind":str(issue.get("kind") or ""),
        })
    impact=summarize_plan(normalized_actions,issues)
    previews=[]
    simulated_validation: dict[str, Any] = {}
    temp_root=Path(tempfile.mkdtemp(prefix="governance-dry-run-"))
    simulated_db=temp_root/"app.db"
    sim: sqlite3.Connection | None = None
    primary_error: Exception | None = None
    primary_http_error: HTTPException | None = None
    cleanup_warning=""
    try:
        # sqlite3.Connection.__exit__ only commits or rolls back; it does not close
        # the connection. Explicit close is required before deleting the temporary
        # database on Windows, otherwise app.db remains locked by this process.
        _copy_sqlite_database(DB_FILE,simulated_db)
        sim_catalog=[copy.deepcopy(dict(row)) for row in current["raw_catalog_items"]]
        sim_relations={kind:copy.deepcopy(_read_relation_document(game_id,kind)) for kind in ("stage","exclusive")}
        sim=sqlite3.connect(simulated_db)
        sim.row_factory=sqlite3.Row
        sim.execute("pragma foreign_keys=on")
        sim.execute("begin immediate")
        for action in normalized_actions:
            issue=issues[action["issue_id"]]
            action_result=_governance_apply_action(sim,game_id,sim_catalog,sim_relations,action,admin_id)
            previews.append({
                "issue":issue,"action":action,"before":action_result.get("before") or {},
                "after":action_result.get("after") or {},"status":action_result.get("status") or "",
                "risk":str((GOVERNANCE_ACTION_SPECS.get(action["action"]) or {}).get("risk") or "normal"),
            })
        simulated_validation=_governance_validate_simulated_state(game_id,sim_catalog,sim_relations,sim)
    except HTTPException as exc:
        primary_http_error=exc
    except Exception as exc:
        primary_error=exc
    finally:
        if sim is not None:
            try:
                sim.rollback()
            except Exception:
                pass
            try:
                sim.close()
            except Exception:
                pass
        sim=None
        gc.collect()
        cleanup_warning=_remove_temporary_tree(temp_root)
        if cleanup_warning:
            print(f"[governance] Dry Run temporary directory cleanup deferred: {temp_root} ({cleanup_warning})",flush=True)
    if primary_http_error is not None:
        raise primary_http_error
    if primary_error is not None:
        # The original simulation error is authoritative. A later Windows cleanup
        # failure is intentionally not allowed to replace it.
        raise HTTPException(status_code=422,detail=f"Dry Run 已阻擋此處置：{primary_error}") from primary_error
    return {
        "scan_id":scan_id,"game_id":game_id,"scan_options":scan_options,"scan_profile":scan_profile,
        "rules_version":scan_profile["rules_version"],"scan_profile_hash":scan_profile["profile_hash"],"state_hash":current["state_hash"],
        "state_component_hashes":current["state_component_hashes"],
        "catalog_hash":current["catalog_hash"],"database_hash":current["database_hash"],
        "impact":impact,"actions":normalized_actions,"previews":previews,
        "simulated_validation":simulated_validation,"temporary_cleanup":"ok" if not cleanup_warning else "deferred",
        "generated_at":now(),
    }


@app.post("/api/games/{game_id}/admin/governance/drafts")
def governance_create_draft(game_id: str, body: GovernanceDraftPayload, request: Request):
    game_id = require_extra_game(game_id); admin = require_admin(request)
    plan = _governance_build_plan(game_id, body.scan_id, body.actions, admin["id"])
    draft_id = f"draft-{uuid.uuid4().hex}"; stamp = now()
    with connect_db() as db:
        db.execute(
            "insert into achievement_resolution_drafts(id,game_id,scan_id,admin_user_id,name,reason,actions_json,plan_json,status,created_at,updated_at) values(?,?,?,?,?,?,?,?,?,?,?)",
            (draft_id, game_id, body.scan_id, admin["id"], body.name, body.reason, json.dumps(body.actions, ensure_ascii=False), json.dumps(plan, ensure_ascii=False), "preview_ready", stamp, stamp),
        )
    log_admin_action(admin["id"], "governance_dry_run", details=f"game={game_id}; draft={draft_id}; actions={len(body.actions)}", category="catalog", game_id=game_id, target_type="governance_draft", target_id=draft_id, summary="產生成就治理處置預覽", after=plan["impact"], metadata={"scan_id": body.scan_id}, actor_ip=client_ip(request), locked=True)
    return {"ok": True, "draft_id": draft_id, "plan": plan}


@app.get("/api/games/{game_id}/admin/governance/drafts/{draft_id}")
def governance_get_draft(game_id: str, draft_id: str, request: Request):
    game_id = require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        row = db.execute("select * from achievement_resolution_drafts where id=? and game_id=?", (draft_id, game_id)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="找不到處置草稿。")
    return {"ok": True, "draft": {**dict(row), "actions": _json_object(row["actions_json"], []), "plan": _json_object(row["plan_json"], {})}}


def _replace_relation_ids_in_document(document: dict[str, Any], mapping: dict[str, str]) -> bool:
    changed = False
    for group in document.get("groups") or []:
        if not isinstance(group, dict):
            continue
        original = [str(value) for value in group.get("achievement_ids") or []]
        replaced: list[str] = []
        for value in original:
            candidate = mapping.get(value, value)
            if candidate and candidate not in replaced:
                replaced.append(candidate)
        if replaced != original:
            group["achievement_ids"] = replaced
            changed = True
    return changed


def _governance_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text_value=str(value or "").strip().casefold()
    if text_value in {"1","true","yes","on","是","隱藏"}: return True
    if text_value in {"0","false","no","off","否","顯示",""}: return False
    raise RuntimeError(f"invalid_boolean:{value}")


def _governance_typed_value(target_field: str, value: Any) -> Any:
    if target_field in {"reward","sourceOrder","source_order"}:
        try: return int(value)
        except Exception as exc: raise RuntimeError(f"invalid_integer:{target_field}:{value}") from exc
    if target_field == "hidden": return _governance_bool(value)
    if target_field == "tags":
        if isinstance(value,list): return list(dict.fromkeys(str(v).strip() for v in value if str(v).strip()))
        return list(dict.fromkeys(v.strip() for v in re.split(r"[,，]",str(value or "")) if v.strip()))
    if target_field in {"name","condition","version","category","source","source_id","official_id"}:
        return str(value or "").strip()
    return value


def _governance_action_entity_ids(action: dict[str, Any]) -> list[str]:
    params=action.get("parameters") if isinstance(action.get("parameters"),dict) else {}
    ids=[str(v).strip() for v in action.get("entity_ids") or [] if str(v).strip()]
    for key in ("achievement_id","keep_id","source_id","target_id","alias_id","canonical_id","old_id","new_id","member_id"):
        value=str(params.get(key) or "").strip()
        if value: ids.append(value)
    ids.extend(str(v).strip() for v in params.get("remove_ids") or [] if str(v).strip())
    ids.extend(str(v).strip() for v in params.get("achievement_ids") or [] if str(v).strip())
    return list(dict.fromkeys(ids))


def _materialize_governance_overrides(db: sqlite3.Connection, game_id: str, catalog_items: list[dict[str, Any]], action: dict[str, Any]) -> None:
    by_id={str(row.get("id") or row.get("achievement_id") or ""):row for row in catalog_items}
    for aid in _governance_action_entity_ids(action):
        override=db.execute("select * from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
        if not override or bool(override["is_deleted"]):
            continue
        ov=dict(override); row=by_id.get(aid)
        if row is None:
            row={"id":aid,"achievement_id":aid,"name":aid,"condition":"","version":"未標示","category":"未辨識分類","reward":0,"hidden":False,"tags":[],"source":"admin_manual","sourceOrder":official_id_number(aid)}
            catalog_items.append(row); by_id[aid]=row
        for key in ("name","condition","version","category","reward","hidden","source"):
            if ov.get(key) is not None: row[key]=bool(ov[key]) if key=="hidden" else ov[key]
        row["tags"]=_json_object(ov.get("tags_json"),[])
        row["sourceOrder"]=official_id_number(aid)


def _governance_relation_group(relation_documents: dict[str, dict[str, Any]], relation_type: str, group_id: str) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, Any] | None]:
    if relation_type not in {"stage","exclusive"}: raise RuntimeError(f"invalid_relation_type:{relation_type}")
    document=relation_documents[relation_type]
    groups=document.setdefault("groups",[])
    group=next((g for g in groups if isinstance(g,dict) and str(g.get("id") or "")==group_id),None)
    return document,groups,group


def _governance_apply_action(db: sqlite3.Connection, game_id: str, catalog_items: list[dict[str, Any]], relation_documents: dict[str, dict[str, Any]], action: dict[str, Any], admin_id: str) -> dict[str, Any]:
    name=canonical_governance_action(action.get("action"))
    if name not in GOVERNANCE_SUPPORTED_ACTIONS: raise RuntimeError(f"unsupported_action:{name}")
    params=action.get("parameters") if isinstance(action.get("parameters"),dict) else {}
    before: dict[str,Any]={}; after: dict[str,Any]={}
    if name in GOVERNANCE_DECISION_ONLY_ACTIONS:
        decision_status = {
            "keep": "accepted_current",
            "review": "waiting_review",
            "keep_pending": "waiting_review",
            "resync_source": "waiting_review",
            "ignore_once": "ignored_once",
            "mark_legal_exception": "legal_exception",
        }.get(name, "decision_only")
        return {
            "status": decision_status,
            "before": {},
            "after": {
                **dict(params),
                "decision_state": decision_status,
                "data_changed": False,
            },
        }
    if name in {"manual_edit","merge_fields","keep_selected","recalculate_order","normalize_hidden","normalize_tags","map_unknown_field","remove_unknown_fields","source_fill"}:
        _materialize_governance_overrides(db,game_id,catalog_items,action)
    by_id={str(row.get("id") or row.get("achievement_id") or ""):row for row in catalog_items}
    if name in {"register_field","keep_unknown_field"}:
        field=str(params.get("field") or "").strip()
        if not field: raise RuntimeError("field_name_required")
        classification=str(params.get("classification") or ("retained_metadata" if name=="keep_unknown_field" else "registered")).strip()
        old=db.execute("select * from achievement_field_registry where game_id=? and field_name=?",(game_id,field)).fetchone(); before=dict(old) if old else {}
        db.execute("""insert into achievement_field_registry(game_id,field_name,classification,mapped_field,active,created_by,created_at,updated_at)
        values(?,?,?,?,1,?,?,?) on conflict(game_id,field_name) do update set classification=excluded.classification,mapped_field=excluded.mapped_field,active=1,created_by=excluded.created_by,updated_at=excluded.updated_at""",(game_id,field,classification,"",admin_id,now(),now()))
        after={"field":field,"classification":classification,"active":True}
    elif name=="map_unknown_field":
        field=str(params.get("field") or "").strip(); target=str(params.get("target_field") or "").strip()
        allowed={"name","condition","version","category","reward","hidden","tags","source","sourceOrder","source_order","source_id","official_id"}
        if not field or target not in allowed: raise RuntimeError("invalid_field_mapping")
        affected=0;samples=[]
        for row in catalog_items:
            if field not in row: continue
            value=row.get(field)
            if len(samples)<10:samples.append({"achievement_id":str(row.get("id") or row.get("achievement_id") or ""),"value":value})
            if row.get(target) in (None,"",[]): row[target]=_governance_typed_value(target,value)
            row.pop(field,None);affected+=1
        before={"field":field,"target_field":target,"samples":samples}
        db.execute("""insert into achievement_field_registry(game_id,field_name,classification,mapped_field,active,created_by,created_at,updated_at)
        values(?,?,'mapped',?,0,?,?,?) on conflict(game_id,field_name) do update set classification='mapped',mapped_field=excluded.mapped_field,active=0,created_by=excluded.created_by,updated_at=excluded.updated_at""",(game_id,field,target,admin_id,now(),now()))
        after={"field":field,"target_field":target,"affected":affected}
    elif name=="manual_edit":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip(); row=by_id.get(aid)
        if not row: raise RuntimeError(f"achievement_not_found:{aid}")
        changes=params.get("changes") if isinstance(params.get("changes"),dict) else {}
        if not changes: raise RuntimeError("manual_edit_changes_required")
        before=dict(row)
        for key,value in changes.items():
            if key not in {"name","condition","version","category","reward","hidden","sourceOrder","source_order","tags","source"}: raise RuntimeError(f"manual_edit_field_not_allowed:{key}")
            typed=_governance_typed_value(key,value); row["sourceOrder" if key=="source_order" else key]=typed
        db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid));after=dict(row)
    elif name in {"merge_fields","keep_selected"}:
        keep_id=str(params.get("keep_id") or (action.get("entity_ids") or [""])[0]).strip()
        remove_ids=[str(v).strip() for v in (params.get("remove_ids") or (action.get("entity_ids") or [])[1:]) if str(v).strip() and str(v).strip()!=keep_id]
        keep=by_id.get(keep_id)
        if not keep or not remove_ids: raise RuntimeError("merge_ids_required")
        missing=[rid for rid in remove_ids if rid not in by_id]
        if missing: raise RuntimeError(f"merge_remove_not_found:{','.join(missing)}")
        # Distinct confirmed official IDs are separate upstream identities. They
        # must not be silently collapsed by a duplicate-name heuristic.
        confirmed=[]
        for rid in [keep_id,*remove_ids]:
            count=int(db.execute("select count(*) from achievement_source_ids where game_id=? and internal_id=? and is_primary=1 and match_status='confirmed'",(game_id,rid)).fetchone()[0])
            if count: confirmed.append(rid)
        if len(confirmed)>1 and not bool(params.get("confirm_identity_merge")):
            raise RuntimeError("confirmed_official_id_merge_requires_explicit_confirmation")
        before={"keep":dict(keep),"remove":{rid:dict(by_id[rid]) for rid in remove_ids}}
        final_values=params.get("final_values") if isinstance(params.get("final_values"),dict) else {}
        for key,value in final_values.items():
            if key in {"name","condition","version","category","reward","hidden","tags","source","sourceOrder","source_order"}: keep["sourceOrder" if key=="source_order" else key]=_governance_typed_value(key,value)
        remove_set=set(remove_ids);catalog_items[:]=[row for row in catalog_items if str(row.get("id") or row.get("achievement_id") or "") not in remove_set]
        mapping={rid:keep_id for rid in remove_ids}
        for document in relation_documents.values():_replace_relation_ids_in_document(document,mapping)
        for rid in remove_ids:
            db.execute("insert or ignore into game_progress(game_id,user_id,achievement_id,completed_at) select game_id,user_id,?,completed_at from game_progress where game_id=? and achievement_id=?",(keep_id,game_id,rid));db.execute("delete from game_progress where game_id=? and achievement_id=?",(game_id,rid))
            db.execute("update game_achievement_reports set achievement_id=? where game_id=? and achievement_id=?",(keep_id,game_id,rid))
            db.execute("delete from game_achievement_overrides where game_id=? and achievement_id in (?,?)",(game_id,keep_id,rid))
            db.execute("update game_achievement_revisions set achievement_id=? where game_id=? and achievement_id=?",(keep_id,game_id,rid))
            db.execute("delete from game_deleted_achievements where game_id=? and achievement_id=?",(game_id,rid))
            db.execute("update or ignore game_featured_achievements set achievement_id=? where game_id=? and achievement_id=?",(keep_id,game_id,rid));db.execute("delete from game_featured_achievements where game_id=? and achievement_id=?",(game_id,rid))
            db.execute("insert into achievement_id_aliases(game_id,alias_id,canonical_id,reason,created_by,created_at) values(?,?,?,?,?,?) on conflict(game_id,alias_id) do update set canonical_id=excluded.canonical_id,reason=excluded.reason,created_by=excluded.created_by,created_at=excluded.created_at",(game_id,rid,keep_id,str(params.get("reason") or "governance_merge"),admin_id,now()))
            # Remove stale identity/source records for the removed formal row.
            db.execute("delete from achievement_source_ids where game_id=? and internal_id=?",(game_id,rid));db.execute("delete from achievement_identities where game_id=? and internal_id=?",(game_id,rid))
        after={"canonical":dict(keep),"aliases":mapping,"removed_identity_ids":remove_ids}
    elif name=="repair_historical_id":
        source_id=str(params.get("source_id") or "").strip();target_id=str(params.get("target_id") or "").strip()
        if not source_id or target_id not in by_id or source_id==target_id:raise RuntimeError("invalid_historical_id_mapping")
        rows=db.execute("select user_id,completed_at from game_progress where game_id=? and achievement_id=?",(game_id,source_id)).fetchall();before={"source_id":source_id,"target_id":target_id,"progress_count":len(rows)}
        for pr in rows:
            db.execute("insert into game_progress(game_id,user_id,achievement_id,completed_at) values(?,?,?,?) on conflict(game_id,user_id,achievement_id) do update set completed_at=min(game_progress.completed_at,excluded.completed_at)",(game_id,pr["user_id"],target_id,pr["completed_at"]))
        db.execute("delete from game_progress where game_id=? and achievement_id=?",(game_id,source_id));db.execute("update game_achievement_reports set achievement_id=? where game_id=? and achievement_id=?",(target_id,game_id,source_id));db.execute("update game_achievement_revisions set achievement_id=? where game_id=? and achievement_id=?",(target_id,game_id,source_id))
        for doc in relation_documents.values():_replace_relation_ids_in_document(doc,{source_id:target_id})
        if str(params.get("create_alias","true")).casefold() not in {"false","0","no"}:db.execute("insert into achievement_id_aliases(game_id,alias_id,canonical_id,reason,created_by,created_at) values(?,?,?,?,?,?) on conflict(game_id,alias_id) do update set canonical_id=excluded.canonical_id,reason=excluded.reason,created_by=excluded.created_by,created_at=excluded.created_at",(game_id,source_id,target_id,str(params.get("reason") or "governance_historical_id_repair"),admin_id,now()))
        after={"source_id":source_id,"target_id":target_id,"progress_transferred":len(rows)}
    elif name in {"create_alias","flatten_alias_chain","repair_alias"}:
        alias_id=str(params.get("alias_id") or "").strip();canonical_id=str(params.get("canonical_id") or "").strip()
        if not alias_id or canonical_id not in by_id or alias_id==canonical_id:raise RuntimeError("invalid_alias_mapping")
        old=db.execute("select * from achievement_id_aliases where game_id=? and alias_id=?",(game_id,alias_id)).fetchone();before=dict(old) if old else {}
        db.execute("insert into achievement_id_aliases(game_id,alias_id,canonical_id,reason,created_by,created_at) values(?,?,?,?,?,?) on conflict(game_id,alias_id) do update set canonical_id=excluded.canonical_id,reason=excluded.reason,created_by=excluded.created_by,created_at=excluded.created_at",(game_id,alias_id,canonical_id,str(params.get("reason") or "governance_alias"),admin_id,now()))
        scope=str(params.get("transfer_scope") or "none")
        transferred_progress=0
        if scope in {"progress","all"}:
            progress_rows=db.execute("select user_id,completed_at from game_progress where game_id=? and achievement_id=?",(game_id,alias_id)).fetchall();transferred_progress=len(progress_rows)
            for progress_row in progress_rows:
                db.execute("insert into game_progress(game_id,user_id,achievement_id,completed_at) values(?,?,?,?) on conflict(game_id,user_id,achievement_id) do update set completed_at=min(game_progress.completed_at,excluded.completed_at)",(game_id,progress_row["user_id"],canonical_id,progress_row["completed_at"]))
            db.execute("delete from game_progress where game_id=? and achievement_id=?",(game_id,alias_id))
        if scope=="all":
            db.execute("update game_achievement_reports set achievement_id=? where game_id=? and achievement_id=?",(canonical_id,game_id,alias_id))
            db.execute("update game_achievement_revisions set achievement_id=? where game_id=? and achievement_id=?",(canonical_id,game_id,alias_id))
            for document in relation_documents.values():_replace_relation_ids_in_document(document,{alias_id:canonical_id})
        after={"alias_id":alias_id,"canonical_id":canonical_id,"transfer_scope":scope,"progress_transferred":transferred_progress}
    elif name in {"delete_alias","remove_alias","break_alias_cycle"}:
        alias_id=str(params.get("alias_id") or (action.get("entity_ids") or [""])[0]).strip();old=db.execute("select * from achievement_id_aliases where game_id=? and alias_id=?",(game_id,alias_id)).fetchone()
        if not old:raise RuntimeError(f"alias_not_found:{alias_id}")
        before=dict(old);db.execute("delete from achievement_id_aliases where game_id=? and alias_id=?",(game_id,alias_id));after={}
    elif name=="delete_orphan_progress":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();count=int(db.execute("select count(*) from game_progress where game_id=? and achievement_id=?",(game_id,aid)).fetchone()[0]);before={"achievement_id":aid,"progress_count":count};db.execute("delete from game_progress where game_id=? and achievement_id=?",(game_id,aid));after={"achievement_id":aid,"progress_count":0}
    elif name in {"delete_orphan_identity","delete_orphan_source_mapping"}:
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip()
        if not aid:raise RuntimeError("achievement_id_required")
        if aid in by_id:raise RuntimeError(f"catalog_item_still_exists:{aid}")
        progress_count=int(db.execute("select count(*) from game_progress where game_id=? and achievement_id=?",(game_id,aid)).fetchone()[0])
        relation_count=int(db.execute("select count(*) from game_achievement_choice_groups where game_id=? and achievement_id=?",(game_id,aid)).fetchone()[0])
        if progress_count or relation_count:raise RuntimeError(f"orphan_identity_still_referenced:progress={progress_count};relations={relation_count}")
        identity=db.execute("select * from achievement_identities where game_id=? and internal_id=?",(game_id,aid)).fetchone()
        mappings=db.execute("select * from achievement_source_ids where game_id=? and internal_id=? order by source_name,source_id",(game_id,aid)).fetchall()
        before={"achievement_id":aid,"identity":dict(identity) if identity else None,"source_ids":[dict(row) for row in mappings]}
        if name=="delete_orphan_source_mapping" and identity:raise RuntimeError(f"identity_still_exists:{aid}")
        if not identity and not mappings:raise RuntimeError(f"orphan_identity_not_found:{aid}")
        db.execute("delete from achievement_source_ids where game_id=? and internal_id=?",(game_id,aid))
        if name=="delete_orphan_identity":db.execute("delete from achievement_identities where game_id=? and internal_id=?",(game_id,aid))
        after={"achievement_id":aid,"identity":None,"source_id_count":0}
    elif name=="transfer_progress":
        source_id=str(params.get("source_id") or (action.get("entity_ids") or [""])[0]).strip();target_id=str(params.get("target_id") or "").strip()
        if target_id not in by_id:raise RuntimeError(f"target_not_found:{target_id}")
        count=int(db.execute("select count(*) from game_progress where game_id=? and achievement_id=?",(game_id,source_id)).fetchone()[0]);before={"source_id":source_id,"source_count":count}
        db.execute("insert or ignore into game_progress(game_id,user_id,achievement_id,completed_at) select game_id,user_id,?,completed_at from game_progress where game_id=? and achievement_id=?",(target_id,game_id,source_id));db.execute("delete from game_progress where game_id=? and achievement_id=?",(game_id,source_id));after={"source_id":source_id,"source_count":0,"target_id":target_id}
    elif name=="recalculate_order":
        before={"count":len(catalog_items),"rule":"official_id"}
        for row in catalog_items:aid=str(row.get("id") or row.get("achievement_id") or "").strip();row["sourceOrder"]=official_id_number(aid);row.pop("source_order",None)
        catalog_items.sort(key=lambda row:(official_id_number(row.get("id") or row.get("achievement_id")),str(row.get("id") or row.get("achievement_id"))));after={"count":len(catalog_items),"rule":"official_id"}
    elif name in {"normalize_hidden","normalize_tags"}:
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();row=by_id.get(aid)
        if not row:raise RuntimeError(f"achievement_not_found:{aid}")
        if name=="normalize_hidden":before={"hidden":row.get("hidden")};row["hidden"]=_governance_bool(row.get("hidden"));after={"hidden":row["hidden"]}
        else:before={"tags":row.get("tags")};row["tags"]=_governance_typed_value("tags",row.get("tags"));after={"tags":row["tags"]}
        db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid))
    elif name=="remove_unknown_fields":
        field=str(params.get("field") or "").strip();aid=str(params.get("achievement_id") or "").strip()
        if field:
            affected=0;samples=[]
            for row in catalog_items:
                if field in row:
                    if len(samples)<10:samples.append({"achievement_id":str(row.get("id") or row.get("achievement_id") or ""),"value":row.get(field)})
                    row.pop(field,None);affected+=1
            before={"field":field,"samples":samples,"occurrence_count":affected};after={"field":field,"removed_from":affected};db.execute("update achievement_field_registry set active=0,classification='removed',updated_at=? where game_id=? and field_name=?",(now(),game_id,field))
        else:
            row=by_id.get(aid)
            if not row:raise RuntimeError(f"achievement_not_found:{aid}")
            before=dict(row);allowed={"id","achievement_id","name","condition","version","category","reward","hidden","tags","source","sourceOrder","source_order","source_id","official_id","arcade","is_deleted","content_id","raw_id","source_url","source_page","source_name","source_mode","source_purpose","source_version","source_category","source_meta","updated_at","official_name","official_condition","match_method","match_confidence","choiceGroup","choiceGroupSize","isChoiceGroup","relationGroup","relationGroupSize","relationType","stageOrder"}
            for key in list(row):
                if key not in allowed:row.pop(key,None)
            after=dict(row)
    elif name in {"normalize_stage_order","deduplicate_relation_group","delete_relation_group","remove_relation_member","replace_relation_member","add_relation_member","create_stage_group","create_exclusive_group","choose_relation_group"}:
        relation_type=str(params.get("relation_type") or ("stage" if name=="create_stage_group" else "exclusive" if name=="create_exclusive_group" else (action.get("evidence") or {}).get("relation_type") or "stage"));group_id=str(params.get("group_id") or (action.get("evidence") or {}).get("group_id") or "").strip()
        if name in {"create_stage_group","create_exclusive_group"}:
            relation_type="stage" if name=="create_stage_group" else "exclusive";group_id=group_id or f"{game_id}-{relation_type}-{uuid.uuid4().hex[:10]}";members=[str(v).strip() for v in (params.get("achievement_ids") or action.get("entity_ids") or []) if str(v).strip()]
            if len(set(members))<2 or any(v not in by_id for v in members):raise RuntimeError("relation_group_requires_two_valid_members")
            document,groups,group=_governance_relation_group(relation_documents,relation_type,group_id)
            if group:raise RuntimeError(f"relation_group_already_exists:{group_id}")
            before={};group={"id":group_id,"name":str(params.get("name") or "").strip(),"basis":str(params.get("basis") or "governance_review").strip(),"achievement_ids":list(dict.fromkeys(members))};groups.append(group);after=dict(group)
        elif name=="choose_relation_group":
            aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();target_type=str(params.get("target_type") or relation_type);target_group=str(params.get("target_group_id") or group_id).strip()
            document,groups,target=_governance_relation_group(relation_documents,target_type,target_group)
            if not target:raise RuntimeError(f"relation_group_not_found:{target_group}")
            before={"achievement_id":aid,"memberships":[]}
            for typ,doc in relation_documents.items():
                for grp in doc.get("groups") or []:
                    ids=[str(v) for v in grp.get("achievement_ids") or []]
                    if aid in ids:before["memberships"].append({"relation_type":typ,"group_id":grp.get("id")});grp["achievement_ids"]=[v for v in ids if v!=aid]
            target["achievement_ids"]=list(dict.fromkeys([*(str(v) for v in target.get("achievement_ids") or []),aid]));after={"achievement_id":aid,"relation_type":target_type,"group_id":target_group}
        else:
            document,groups,group=_governance_relation_group(relation_documents,relation_type,group_id)
            if not group:raise RuntimeError(f"relation_group_not_found:{group_id}")
            before=copy.deepcopy(group)
            if name=="delete_relation_group":document["groups"]=[g for g in groups if str(g.get("id") or "")!=group_id];after={}
            elif name in {"deduplicate_relation_group","normalize_stage_order"}:group["achievement_ids"]=list(dict.fromkeys(str(v) for v in group.get("achievement_ids") or [] if str(v)));after=copy.deepcopy(group)
            elif name=="remove_relation_member":aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();group["achievement_ids"]=[str(v) for v in group.get("achievement_ids") or [] if str(v)!=aid];after=copy.deepcopy(group)
            elif name=="replace_relation_member":old_id=str(params.get("old_id") or (action.get("entity_ids") or [""])[0]).strip();new_id=str(params.get("new_id") or "").strip();
            if name=="replace_relation_member":
                if new_id not in by_id:raise RuntimeError(f"achievement_not_found:{new_id}")
                group["achievement_ids"]=list(dict.fromkeys(new_id if str(v)==old_id else str(v) for v in group.get("achievement_ids") or []));after=copy.deepcopy(group)
            elif name=="add_relation_member":
                member=str(params.get("member_id") or params.get("achievement_id") or "").strip()
                if member not in by_id:raise RuntimeError(f"achievement_not_found:{member}")
                group["achievement_ids"]=list(dict.fromkeys([*(str(v) for v in group.get("achievement_ids") or []),member]));after=copy.deepcopy(group)
    elif name=="delete_override":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();old=db.execute("select * from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
        if not old:raise RuntimeError(f"override_not_found:{aid}")
        before=dict(old);db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid));after={}
    elif name=="transfer_override":
        source_id=str(params.get("source_id") or (action.get("entity_ids") or [""])[0]).strip();target_id=str(params.get("target_id") or "").strip()
        if target_id not in by_id:raise RuntimeError(f"achievement_not_found:{target_id}")
        old=db.execute("select * from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,source_id)).fetchone()
        if not old:raise RuntimeError(f"override_not_found:{source_id}")
        before=dict(old);columns=[key for key in dict(old) if key not in {"game_id","achievement_id"}];values=[old[key] for key in columns]
        db.execute(f"insert or replace into game_achievement_overrides(game_id,achievement_id,{','.join(columns)}) values(?,?,{','.join('?' for _ in columns)})",[game_id,target_id,*values]);db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,source_id));after={"source_id":source_id,"target_id":target_id}
    elif name=="sync_json_to_database":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();row=by_id.get(aid)
        if not row:raise RuntimeError(f"json_row_not_found:{aid}")
        old=db.execute("select * from game_catalog_items where game_id=? and achievement_id=?",(game_id,aid)).fetchone();before=dict(old) if old else {};after=dict(row)
    elif name=="sync_database_to_json":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();dbrow=db.execute("select * from game_catalog_items where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
        if not dbrow:raise RuntimeError(f"database_row_not_found:{aid}")
        before=dict(by_id.get(aid) or {});value=dict(dbrow);value.update({"id":aid,"tags":_json_object(value.pop("tags_json","[]"),[]),"hidden":bool(value.get("hidden")),"sourceOrder":int(value.pop("source_order",official_id_number(aid)) or official_id_number(aid))});value.pop("game_id",None);value.pop("achievement_id",None);value.pop("updated_at",None)
        if aid in by_id:by_id[aid].clear();by_id[aid].update(value)
        else:catalog_items.append(value);by_id[aid]=value
        after=dict(value)
    elif name=="archive_database_row":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();progress=int(db.execute("select count(*) from game_progress where game_id=? and achievement_id=?",(game_id,aid)).fetchone()[0]);relations=int(db.execute("select count(*) from game_achievement_choice_groups where game_id=? and achievement_id=?",(game_id,aid)).fetchone()[0])
        if progress or relations:raise RuntimeError(f"database_row_has_references:progress={progress};relations={relations}")
        old=db.execute("select * from game_catalog_items where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
        if not old:raise RuntimeError(f"database_row_not_found:{aid}")
        before=dict(old);db.execute("delete from game_catalog_items where game_id=? and achievement_id=?",(game_id,aid));db.execute("delete from game_catalog_source_records where game_id=? and achievement_id=?",(game_id,aid));after={"archived":True}
    elif name=="restore_catalog_item":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();before=dict(by_id.get(aid) or {})
        db.execute("delete from game_deleted_achievements where game_id=? and achievement_id=?",(game_id,aid))
        if aid not in by_id:
            current=_effective_achievement_row(db,game_id,aid)
            dbrow=db.execute("select * from game_catalog_items where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
            source=db.execute("select raw_json from game_catalog_source_records where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
            candidate=current or (dict(dbrow) if dbrow else None)
            if not candidate and source:
                raw=_json_object(source["raw_json"],{});candidate=raw if isinstance(raw,dict) else None
            if not candidate:raise RuntimeError(f"restore_source_not_found:{aid}")
            value={"id":aid,"name":str(candidate.get("name") or aid),"condition":str(candidate.get("condition") or ""),"version":str(candidate.get("version") or "未標示"),"category":str(candidate.get("category") or "未辨識分類"),"reward":int(candidate.get("reward") or 0),"hidden":bool(candidate.get("hidden")),"tags":candidate.get("tags") if isinstance(candidate.get("tags"),list) else _json_object(candidate.get("tags_json"),[]),"source":str(candidate.get("source") or "restored"),"sourceOrder":official_id_number(aid)};catalog_items.append(value);by_id[aid]=value
        after=dict(by_id[aid])
    elif name=="source_fill":
        aid=str(params.get("achievement_id") or (action.get("entity_ids") or [""])[0]).strip();row=by_id.get(aid)
        if not row:raise RuntimeError(f"achievement_not_found:{aid}")
        source=db.execute("select raw_json,provenance_json from game_catalog_source_records where game_id=? and achievement_id=?",(game_id,aid)).fetchone()
        if not source:raise RuntimeError(f"source_record_not_found:{aid}")
        raw=_json_object(source["raw_json"],{});before=dict(row);filled=[]
        if isinstance(raw,dict):
            for key in ("name","condition","version","category","reward","hidden","tags"):
                if row.get(key) in (None,"",[],"未標示","未辨識分類") and raw.get(key) not in (None,"",[]):row[key]=_governance_typed_value(key,raw[key]);filled.append(key)
        if not filled:raise RuntimeError("source_record_has_no_safe_fill_fields")
        db.execute("delete from game_achievement_overrides where game_id=? and achievement_id=?",(game_id,aid));after={"achievement":dict(row),"filled_fields":filled}
    else:
        raise RuntimeError(f"unsupported_action:{name}")
    return {"status":"applied","before":before,"after":after}


GOVERNANCE_CATALOG_ACTIONS=frozenset({
    "manual_edit","merge_fields","keep_selected","recalculate_order","normalize_hidden","normalize_tags",
    "map_unknown_field","remove_unknown_fields","sync_json_to_database","sync_database_to_json",
    "archive_database_row","restore_catalog_item","source_fill",
})
GOVERNANCE_RELATION_ACTIONS=frozenset({
    "normalize_stage_order","deduplicate_relation_group","delete_relation_group","remove_relation_member",
    "replace_relation_member","add_relation_member","create_stage_group","create_exclusive_group","choose_relation_group",
    "merge_fields","keep_selected","repair_historical_id","create_alias",
})


def _governance_persist_catalog_rows(db: sqlite3.Connection,game_id: str,rows: list[dict[str,Any]]) -> None:
    normalized=[];desired=set();stamp=now()
    for index,row in enumerate(rows):
        aid=str(row.get("id") or row.get("achievement_id") or "").strip()
        if not OFFICIAL_ACHIEVEMENT_ID_PATTERN.fullmatch(aid or ""):raise RuntimeError(f"non_numeric_official_id:{aid}")
        if aid in desired:raise RuntimeError(f"duplicate_catalog_id:{aid}")
        desired.add(aid)
        normalized.append((game_id,aid,str(row.get("name") or "").strip(),str(row.get("condition") or "").strip(),str(row.get("version") or "未標示").strip(),str(row.get("category") or "未辨識分類").strip(),int(row.get("reward") or 0),1 if row.get("hidden") else 0,json.dumps(row.get("tags") if isinstance(row.get("tags"),list) else [],ensure_ascii=False),str(row.get("source") or "catalog"),int(row.get("sourceOrder") if row.get("sourceOrder") is not None else row.get("source_order") or official_id_number(aid)),stamp))
    existing={str(row[0]) for row in db.execute("select achievement_id from game_catalog_items where game_id=?",(game_id,)).fetchall()}
    removed=sorted(existing-desired)
    for aid in removed:
        # Removing a formal row is explicit; its dependent identity rows must not
        # survive as an apparently valid mapping to a missing achievement.
        db.execute("delete from achievement_source_ids where game_id=? and internal_id=?",(game_id,aid))
        db.execute("delete from achievement_identities where game_id=? and internal_id=?",(game_id,aid))
        db.execute("delete from game_catalog_source_records where game_id=? and achievement_id=?",(game_id,aid))
        db.execute("delete from game_catalog_items where game_id=? and achievement_id=?",(game_id,aid))
    db.executemany("""insert into game_catalog_items(game_id,achievement_id,name,condition,version,category,reward,hidden,tags_json,source,source_order,updated_at)
        values(?,?,?,?,?,?,?,?,?,?,?,?)
        on conflict(game_id,achievement_id) do update set name=excluded.name,condition=excluded.condition,version=excluded.version,
        category=excluded.category,reward=excluded.reward,hidden=excluded.hidden,tags_json=excluded.tags_json,source=excluded.source,
        source_order=excluded.source_order,updated_at=excluded.updated_at""",normalized)


@app.post("/api/games/{game_id}/admin/governance/drafts/{draft_id}/execute")
def governance_execute_draft(game_id: str, draft_id: str, body: GovernanceExecutePayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request)
    if not GOVERNANCE_OPERATION_GUARD.acquire(blocking=False):
        raise HTTPException(status_code=409,detail="另一項成就治理或掃描作業正在執行，請完成後再試。")
    batch_id="";backup:Path|None=None;snapshot_dir:Path|None=None
    catalog_path=game_catalog_file(game_id)
    relation_paths={kind:game_relation_file(game_id,kind) for kind in ("stage","exclusive")}
    try:
        with connect_db() as db:
            draft=db.execute("select * from achievement_resolution_drafts where id=? and game_id=?",(draft_id,game_id)).fetchone()
        if not draft:raise HTTPException(status_code=404,detail="找不到處置草稿。")
        if str(draft["status"] or "")!="preview_ready":raise HTTPException(status_code=409,detail="此處置草稿已執行或已失效。")
        plan=_json_object(draft["plan_json"],{})
        scan_options=plan.get("scan_options") if isinstance(plan.get("scan_options"),dict) else {}
        plan_profile=_governance_scan_profile(str(plan.get("rules_version") or ""),scan_options)
        if str(plan.get("rules_version") or "")!=GOVERNANCE_RULES_VERSION or str(plan.get("scan_profile_hash") or "")!=plan_profile["profile_hash"]:
            raise HTTPException(status_code=409,detail="此治理草稿的掃描規則或選項已過期，請重新掃描並建立預覽。")
        current=_governance_context(game_id, scan_options)
        if str(plan.get("state_hash") or "")!=current["state_hash"]:
            raise HTTPException(status_code=409,detail="成就、進度、關聯、別名或身分資料已在 Dry Run 後變更，請重新建立預覽。")
        confirmation=str((plan.get("impact") or {}).get("confirmation_text") or "")
        if confirmation and body.confirmation_text.strip()!=confirmation:
            raise HTTPException(status_code=400,detail=f"高風險處置需要輸入確認文字：{confirmation}")
        batch_id=f"batch-{uuid.uuid4().hex}";started=now()
        snapshot_dir=ROOT/"backups"/f"governance-{time.strftime('%Y%m%d-%H%M%S')}-{batch_id[-8:]}"
        snapshot_dir.mkdir(parents=True,exist_ok=True)
        with connect_db() as db:
            db.execute("insert into achievement_resolution_batches(id,game_id,draft_id,admin_user_id,status,plan_json,result_json,started_at,snapshot_dir) values(?,?,?,?,?,?,?,?,?)",(batch_id,game_id,draft_id,admin["id"],"running",json.dumps(plan,ensure_ascii=False),"{}",started,str(snapshot_dir)))
        backup=create_database_backup()
        shutil.copy2(catalog_path,snapshot_dir/"catalog.json")
        for kind,path in relation_paths.items():shutil.copy2(path,snapshot_dir/f"{kind}-groups.json")
        with connect_db() as db:
            db.execute("update achievement_resolution_batches set backup_name=? where id=?",(backup.name,batch_id))
            for snapshot_type,path in [("database",backup),("catalog",snapshot_dir/"catalog.json"),("stage_relations",snapshot_dir/"stage-groups.json"),("exclusive_relations",snapshot_dir/"exclusive-groups.json")]:
                checksum=hashlib.sha256(path.read_bytes()).hexdigest() if path.exists() else ""
                db.execute("insert into achievement_resolution_snapshots(id,batch_id,snapshot_type,file_path,payload_json,checksum,created_at) values(?,?,?,?,?,?,?)",(f"snapshot-{uuid.uuid4().hex}",batch_id,snapshot_type,str(path),"",checksum,now()))
        catalog_items=[copy.deepcopy(dict(row)) for row in current["raw_catalog_items"]]
        relation_documents={kind:copy.deepcopy(_read_relation_document(game_id,kind)) for kind in ("stage","exclusive")}
        results=[]
        with connect_db() as db:
            db.execute("begin immediate")
            for action in plan.get("actions") or []:
                raw_name=str(action.get("action") or "")
                if raw_name not in GOVERNANCE_SUPPORTED_ACTIONS:raise RuntimeError(f"unsupported_action:{raw_name}")
                action_result=_governance_apply_action(db,game_id,catalog_items,relation_documents,action,admin["id"])
                result_row={"issue_id":action.get("issue_id"),"action":raw_name,**action_result};results.append(result_row)
                db.execute("insert into achievement_resolution_actions(batch_id,issue_id,action,status,before_json,after_json,error_message,created_at) values(?,?,?,?,?,?,?,?)",(batch_id,action.get("issue_id"),raw_name,action_result["status"],json.dumps(action_result.get("before") or {},ensure_ascii=False),json.dumps(action_result.get("after") or {},ensure_ascii=False),"",now()))
                if raw_name=="mark_legal_exception":
                    issue=db.execute("select * from achievement_issues where id=? and game_id=?",(action.get("issue_id"),game_id)).fetchone()
                    if issue:
                        params=action.get("parameters") or {};reason=str(params.get("reason") or draft["reason"] or "administrator_review")
                        entities=[r["entity_id"] for r in db.execute("select entity_id from achievement_issue_entities where issue_id=? order by entity_id",(issue["id"],)).fetchall()]
                        snapshot_hash=governance_hash({"entities":entities,"evidence":_json_object(issue["evidence_json"],{})})
                        permanent=1 if _governance_bool(params.get("permanent",False)) else 0
                        recheck=1 if _governance_bool(params.get("recheck_on_change",True)) else 0
                        db.execute("""insert into achievement_exception_rules(id,game_id,fingerprint,reason,source_basis,snapshot_hash,permanent,recheck_on_change,active,created_by,created_at,updated_at)
                        values(?,?,?,?,?,?,?,?,1,?,?,?) on conflict(game_id,fingerprint) do update set reason=excluded.reason,source_basis=excluded.source_basis,snapshot_hash=excluded.snapshot_hash,permanent=excluded.permanent,recheck_on_change=excluded.recheck_on_change,active=1,created_by=excluded.created_by,updated_at=excluded.updated_at""",(f"exception-{uuid.uuid4().hex}",game_id,issue["fingerprint"],reason,"resolution_batch",snapshot_hash,permanent,recheck,admin["id"],now(),now()))
            _governance_validate_simulated_state(game_id,catalog_items,relation_documents,db)
            action_names={str(action.get("action") or "") for action in plan.get("actions") or []}
            catalog_changed=bool(action_names & GOVERNANCE_CATALOG_ACTIONS)
            relations_changed=bool(action_names & GOVERNANCE_RELATION_ACTIONS)
            if catalog_changed:
                _governance_persist_catalog_rows(db,game_id,catalog_items)
            if relations_changed:
                for kind,document in relation_documents.items():
                    _write_relation_document(game_id,kind,document);_sync_relation_groups(db,game_id,game_relation_file(game_id,kind),kind)
                _repair_choice_group_progress(db,game_id)
        if catalog_changed:
            _write_catalog_items_payload(game_id,catalog_items)
        post_context=_governance_context(game_id, scan_options)
        post_issues_by_fingerprint={str(issue.get("fingerprint") or ""):issue for issue in post_context.get("issues") or []}
        pre_fingerprints={str(issue.get("fingerprint") or "") for issue in current.get("issues") or []}
        new_post_issues=[issue for issue in post_context.get("issues") or [] if str(issue.get("fingerprint") or "") not in pre_fingerprints]
        severity_rank={"info":1,"warning":2,"error":3}
        completion_results=[]
        with connect_db() as db:
            integrity=str(db.execute("pragma integrity_check").fetchone()[0])
            if integrity!="ok":raise RuntimeError(f"database_integrity:{integrity}")
            orphan_progress=int(db.execute("select count(*) from game_progress p left join game_catalog_items c on c.game_id=p.game_id and c.achievement_id=p.achievement_id where p.game_id=? and c.achievement_id is null",(game_id,)).fetchone()[0])
            orphan_relations=int(db.execute("select count(*) from game_achievement_choice_groups g left join game_catalog_items c on c.game_id=g.game_id and c.achievement_id=g.achievement_id where g.game_id=? and c.achievement_id is null",(game_id,)).fetchone()[0])
            orphan_identities=int(db.execute("select count(*) from achievement_identities i left join game_catalog_items c on c.game_id=i.game_id and c.achievement_id=i.internal_id where i.game_id=? and c.achievement_id is null",(game_id,)).fetchone()[0])
            orphan_source_ids=int(db.execute("select count(*) from achievement_source_ids s left join game_catalog_items c on c.game_id=s.game_id and c.achievement_id=s.internal_id where s.game_id=? and c.achievement_id is null",(game_id,)).fetchone()[0])
            if orphan_progress or orphan_relations or orphan_identities or orphan_source_ids:raise RuntimeError(f"referential_validation_failed:progress={orphan_progress};relations={orphan_relations};identities={orphan_identities};source_ids={orphan_source_ids}")
            for action in plan.get("actions") or []:
                issue_id=str(action.get("issue_id") or "")
                action_name=str(action.get("action") or "")
                issue=db.execute("select * from achievement_issues where id=? and game_id=?",(issue_id,game_id)).fetchone()
                if not issue:raise RuntimeError(f"governance_issue_missing_after_action:{issue_id}")
                fingerprint=str(issue["fingerprint"] or "")
                entity_ids=[str(row["entity_id"]) for row in db.execute("select entity_id from achievement_issue_entities where issue_id=? order by entity_id",(issue_id,)).fetchall()]
                post_issue=post_issues_by_fingerprint.get(fingerprint)
                resolution={"batch_id":batch_id,"action":action_name,"verified_at":now()}
                resolved_at=now()
                if action_name=="keep":
                    # Decision-only keep never pretends that data was repaired.
                    # It records the exact issue/evidence snapshot and remains
                    # accepted until a later scan observes changed evidence.
                    next_state="accepted_current"
                    decision=_upsert_accepted_current_decision(
                        db,game_id=game_id,issue=issue,admin_id=admin["id"],
                        reason=str(draft["reason"] or "管理員接受目前資料"),
                        source_basis="resolution_batch",batch_id=batch_id,
                    )
                    resolution.update(decision)
                    resolution.update({"completion_type":"accepted_current","verification":"same_evidence_suppressed_until_change","scan_options":scan_options})
                    db.execute("update achievement_exception_rules set active=0,updated_at=? where game_id=? and fingerprint=? and active=1",(now(),game_id,fingerprint))
                elif action_name in {"review","keep_pending","resync_source"}:
                    next_state="waiting_review";resolved_at=None
                    resolution.update({"completion_type":"pending_review","verification":"administrator_follow_up_required"})
                    _deactivate_governance_decision(db,game_id,fingerprint,reason=f"changed_to_{action_name}")
                elif action_name=="ignore_once":
                    next_state="ignored"
                    resolution.update({"completion_type":"ignored_once","verification":"hidden_until_next_explicit_scan"})
                    _deactivate_governance_decision(db,game_id,fingerprint,reason="ignored_once")
                elif action_name=="mark_legal_exception":
                    next_state="legal_exception"
                    resolution.update({"completion_type":"legal_exception","verification":"exception_rule_recorded"})
                    _deactivate_governance_decision(db,game_id,fingerprint,reason="replaced_by_legal_exception")
                else:
                    if post_issue is not None:
                        raise RuntimeError(f"resolution_verification_failed:issue_still_present:{issue_id}:{issue['title']}")
                    related_new=[]
                    entity_set=set(entity_ids)
                    original_rank=severity_rank.get(str(issue["severity"] or "info"),1)
                    for candidate in new_post_issues:
                        candidate_entities={str(value) for value in candidate.get("entity_ids") or []}
                        if entity_set & candidate_entities and severity_rank.get(str(candidate.get("severity") or "info"),1)>=original_rank:
                            related_new.append(str(candidate.get("title") or candidate.get("kind") or candidate.get("fingerprint") or "新問題"))
                    if related_new:
                        raise RuntimeError(f"resolution_verification_failed:new_related_issue:{issue_id}:{'、'.join(related_new[:3])}")
                    next_state="resolved"
                    resolution.update({"completion_type":"verified_repair","verification":"original_fingerprint_absent","new_related_issue_count":0})
                    _deactivate_governance_decision(db,game_id,fingerprint,reason="verified_repair_completed")
                    db.execute("update achievement_exception_rules set active=0,updated_at=? where game_id=? and fingerprint=? and active=1",(now(),game_id,fingerprint))
                db.execute("update achievement_issues set state=?,resolution_json=?,resolved_by=?,resolved_at=? where id=? and game_id=?",(next_state,json.dumps(resolution,ensure_ascii=False),admin["id"],resolved_at,issue_id,game_id))
                completion_results.append({"issue_id":issue_id,"action":action_name,"state":next_state,"completion_type":resolution.get("completion_type"),"verification":resolution.get("verification")})
            scan_summary=_governance_refresh_scan_summary(db,game_id,str(plan.get("scan_id") or draft["scan_id"] or ""))
            remaining_issue_count=int(db.execute("select count(*) from achievement_issues where game_id=? and state in ('new','waiting_review','assigned','ready','processing','reopened','failed')",(game_id,)).fetchone()[0])
            result={"actions":results,"completions":completion_results,"validation":{"integrity":integrity,"orphan_progress":orphan_progress,"orphan_relations":orphan_relations,"orphan_identities":orphan_identities,"orphan_source_ids":orphan_source_ids,"remaining_issue_count":remaining_issue_count,"post_action_scan_summary":scan_summary},"backup":backup.name,"pre_state_hash":current["state_hash"],"post_state_hash":post_context["state_hash"],"post_state_component_hashes":post_context["state_component_hashes"]}
            db.execute("update achievement_resolution_batches set status='completed',result_json=?,completed_at=? where id=?",(json.dumps(result,ensure_ascii=False),now(),batch_id))
            db.execute("update achievement_resolution_drafts set status='executed',updated_at=? where id=?",(now(),draft_id))
        bump_game_live_scope(game_id,"catalog");bump_game_live_scope(game_id,"stats")
        log_admin_action(admin["id"],"governance_execute_completed",details=f"game={game_id}; batch={batch_id}; actions={len(results)}",category="catalog",game_id=game_id,target_type="governance_batch",target_id=batch_id,summary="完成成就資料治理處置",before={"state_hash":current["state_hash"]},after=result,metadata=plan.get("impact") or {},backup_name=backup.name,actor_ip=client_ip(request),locked=True)
        return {"ok":True,"batch_id":batch_id,"result":result}
    except HTTPException:
        raise
    except Exception as exc:
        rollback_errors=[]
        rolled_back=False
        if backup is not None and snapshot_dir is not None:
            try:
                shutil.copy2(snapshot_dir/"catalog.json",catalog_path)
            except Exception as rollback_exc:
                rollback_errors.append(f"catalog:{rollback_exc}")
            for kind,path in relation_paths.items():
                try:
                    shutil.copy2(snapshot_dir/f"{kind}-groups.json",path)
                except Exception as rollback_exc:
                    rollback_errors.append(f"{kind}:{rollback_exc}")
            try:
                _restore_governance_database_scope(backup,game_id)
            except Exception as rollback_exc:
                rollback_errors.append(f"database_scope:{rollback_exc}")
            rolled_back=not rollback_errors
            try:
                with connect_db() as db:
                    db.execute("update achievement_resolution_batches set status='failed',result_json=?,completed_at=? where id=?",(json.dumps({"error":str(exc),"rolled_back":rolled_back,"rollback_errors":rollback_errors},ensure_ascii=False),now(),batch_id))
            except Exception as record_exc:
                rollback_errors.append(f"record:{record_exc}")
                rolled_back=False
        if batch_id:
            log_admin_action(admin["id"],"governance_execute_failed",details=f"game={game_id}; batch={batch_id}; error={exc}; rollback_errors={rollback_errors}",category="catalog",status="failed" if rolled_back else "recovery_failed",game_id=game_id,target_type="governance_batch",target_id=batch_id,summary="成就治理處置失敗"+("並已自動回復" if rolled_back else "且自動回復失敗"),error_message=str(exc),metadata={"rollback_errors":rollback_errors},backup_name=backup.name if backup else "",actor_ip=client_ip(request),locked=True)
        if rollback_errors:
            raise HTTPException(status_code=500,detail=f"處置未套用，且自動回復未完整成功：{exc}；回復錯誤：{'；'.join(rollback_errors)}") from exc
        raise HTTPException(status_code=409,detail=f"處置未套用，資料已完整回復：{exc}") from exc
    finally:
        GOVERNANCE_OPERATION_GUARD.release()


@app.get("/api/games/{game_id}/admin/governance/batches")
def governance_list_batches(game_id: str, request: Request, limit: int = 50):
    game_id = require_extra_game(game_id); require_admin(request)
    with connect_db() as db:
        rows = db.execute("select * from achievement_resolution_batches where game_id=? order by started_at desc limit ?", (game_id, max(1, min(200, int(limit))))).fetchall()
    return {"ok": True, "batches": [{**dict(row), "plan": _json_object(row["plan_json"], {}), "result": _json_object(row["result_json"], {})} for row in rows]}



def _restore_governance_database_scope(backup: Path, game_id: str) -> dict[str,int]:
    tables=(
        "game_catalog_items","achievement_identities","achievement_source_ids","game_catalog_source_records",
        "game_achievement_overrides","game_deleted_achievements","game_featured_achievements",
        "game_achievement_reports","game_achievement_revisions","achievement_id_aliases",
        "achievement_field_registry","achievement_exception_rules","achievement_governance_decisions",
        "game_achievement_choice_groups","game_progress",
    )
    source=sqlite3.connect(backup)
    source.row_factory=sqlite3.Row
    restored={}
    try:
        snapshots={}
        for table in tables:
            columns=[str(row[1]) for row in source.execute(f"pragma table_info({table})").fetchall()]
            rows=source.execute(f"select * from {table} where game_id=?",(game_id,)).fetchall()
            snapshots[table]=(columns,rows)
        legacy_progress=None
        if game_id=="wuwa":
            columns=[str(row[1]) for row in source.execute("pragma table_info(progress)").fetchall()]
            legacy_progress=(columns,source.execute("select * from progress").fetchall())
        with connect_db() as db:
            db.execute("pragma defer_foreign_keys=on")
            db.execute("begin immediate")
            # Delete dependants before parents.
            delete_order=(
                "game_progress","game_achievement_choice_groups","achievement_source_ids","game_catalog_source_records",
                "game_achievement_overrides","game_deleted_achievements","game_featured_achievements",
                "game_achievement_reports","game_achievement_revisions","achievement_id_aliases",
                "achievement_identities","achievement_field_registry","achievement_exception_rules",
                "achievement_governance_decisions","game_catalog_items",
            )
            for table in delete_order:
                db.execute(f"delete from {table} where game_id=?",(game_id,))
            insert_order=(
                "game_catalog_items","achievement_identities","achievement_source_ids","game_catalog_source_records",
                "game_achievement_overrides","game_deleted_achievements","game_featured_achievements",
                "game_achievement_reports","game_achievement_revisions","achievement_id_aliases",
                "achievement_field_registry","achievement_exception_rules","achievement_governance_decisions",
                "game_achievement_choice_groups","game_progress",
            )
            for table in insert_order:
                columns,rows=snapshots[table]
                if rows:
                    placeholders=",".join("?" for _ in columns)
                    db.executemany(f"insert into {table}({','.join(columns)}) values({placeholders})",[tuple(row[col] for col in columns) for row in rows])
                restored[table]=len(rows)
            if legacy_progress is not None:
                columns,rows=legacy_progress
                db.execute("delete from progress")
                if rows:
                    placeholders=",".join("?" for _ in columns)
                    db.executemany(f"insert into progress({','.join(columns)}) values({placeholders})",[tuple(row[col] for col in columns) for row in rows])
                restored["progress"]=len(rows)
    finally:
        source.close()
    return restored


@app.post("/api/games/{game_id}/admin/governance/batches/{batch_id}/rollback")
def governance_rollback_batch(game_id: str, batch_id: str, body: GovernanceRollbackPayload, request: Request):
    game_id=require_extra_game(game_id);admin=require_admin(request)
    if not GOVERNANCE_OPERATION_GUARD.acquire(blocking=False):raise HTTPException(status_code=409,detail="另一項成就治理或掃描作業正在執行，請完成後再試。")
    try:
        with connect_db() as db:batch=db.execute("select * from achievement_resolution_batches where id=? and game_id=?",(batch_id,game_id)).fetchone()
        if not batch:raise HTTPException(status_code=404,detail="找不到處置批次。")
        if batch["status"]!="completed" or batch["rolled_back_at"]:raise HTTPException(status_code=409,detail="此批次目前不可回復。")
        plan=_json_object(batch["plan_json"],{})
        scan_options=plan.get("scan_options") if isinstance(plan.get("scan_options"),dict) else {}
        plan_profile=_governance_scan_profile(str(plan.get("rules_version") or ""),scan_options)
        if str(plan.get("scan_profile_hash") or "") and str(plan.get("scan_profile_hash"))!=plan_profile["profile_hash"]:
            raise HTTPException(status_code=409,detail="治理批次的掃描規則指紋不一致，為避免錯誤回復已停止操作。")
        result=_json_object(batch["result_json"],{})
        current=_governance_context(game_id,scan_options)
        expected_post=str(result.get("post_state_hash") or "")
        if not expected_post or expected_post!=current["state_hash"]:
            raise HTTPException(status_code=409,detail="批次完成後已有進度、關聯、目錄或身分資料變更；為避免倒退新資料，已阻擋整批回復。請建立新的治理處置。")
        backup=ROOT/"backups"/str(batch["backup_name"] or "");snapshot_dir=Path(str(batch["snapshot_dir"] or ""))
        if not backup.exists() or not snapshot_dir.exists():raise HTTPException(status_code=409,detail="回復所需備份不存在。")
        current_backup=create_database_backup()
        current_catalog_bytes=game_catalog_file(game_id).read_bytes()
        current_relation_bytes={kind:game_relation_file(game_id,kind).read_bytes() for kind in ("stage","exclusive")}
        with connect_db() as db:
            original_issue_rows=[dict(row) for row in db.execute("select id,state,resolution_json,resolved_by,resolved_at from achievement_issues where resolution_json like ? and game_id=?",(f'%\"batch_id\": \"{batch_id}\"%',game_id)).fetchall()]
        original_batch={key:batch[key] for key in batch.keys()}
        rollback_errors=[]
        try:
            shutil.copy2(snapshot_dir/"catalog.json",game_catalog_file(game_id))
            for kind in ("stage","exclusive"):
                shutil.copy2(snapshot_dir/f"{kind}-groups.json",game_relation_file(game_id,kind))
            restored_rows=_restore_governance_database_scope(backup,game_id)
            restored_context=_governance_context(game_id,scan_options)
            expected_pre=str(result.get("pre_state_hash") or plan.get("state_hash") or "")
            if expected_pre and restored_context["state_hash"]!=expected_pre:
                raise RuntimeError("governance_scope_restore_hash_mismatch")
            with connect_db() as db:
                db.execute("update achievement_resolution_batches set status='rolled_back',rolled_back_at=?,rollback_reason=?,result_json=? where id=?",(now(),body.reason,json.dumps({**result,"rollback":{"restored_rows":restored_rows,"state_hash":restored_context["state_hash"]}},ensure_ascii=False),batch_id))
                db.execute("update achievement_issues set state='reopened',resolved_at=null,resolved_by=null where resolution_json like ? and game_id=?",(f'%\"batch_id\": \"{batch_id}\"%',game_id))
        except Exception as exc:
            try:
                game_catalog_file(game_id).write_bytes(current_catalog_bytes)
            except Exception as rollback_exc:
                rollback_errors.append(f"catalog:{rollback_exc}")
            for kind,data in current_relation_bytes.items():
                try:
                    game_relation_file(game_id,kind).write_bytes(data)
                except Exception as rollback_exc:
                    rollback_errors.append(f"{kind}:{rollback_exc}")
            try:
                _restore_governance_database_scope(current_backup,game_id)
            except Exception as rollback_exc:
                rollback_errors.append(f"database_scope:{rollback_exc}")
            try:
                with connect_db() as db:
                    db.execute("update achievement_resolution_batches set status=?,result_json=?,rolled_back_at=?,rollback_reason=? where id=?",(original_batch.get("status"),original_batch.get("result_json"),original_batch.get("rolled_back_at"),original_batch.get("rollback_reason"),batch_id))
                    for issue_row in original_issue_rows:
                        db.execute("update achievement_issues set state=?,resolution_json=?,resolved_by=?,resolved_at=? where id=?",(issue_row.get("state"),issue_row.get("resolution_json"),issue_row.get("resolved_by"),issue_row.get("resolved_at"),issue_row.get("id")))
            except Exception as rollback_exc:
                rollback_errors.append(f"governance_metadata:{rollback_exc}")
            if rollback_errors:
                raise HTTPException(status_code=500,detail=f"回復失敗，且安全回復也失敗：{exc}；{'；'.join(rollback_errors)}") from exc
            raise HTTPException(status_code=409,detail=f"回復失敗，已回到操作前狀態：{exc}") from exc
        bump_game_live_scope(game_id,"catalog");bump_game_live_scope(game_id,"stats")
        log_admin_action(admin["id"],"governance_batch_rollback",details=f"game={game_id}; batch={batch_id}; reason={body.reason}",category="catalog",game_id=game_id,target_type="governance_batch",target_id=batch_id,summary="回復成就治理處置批次（限遊戲資料範圍）",after={"restored_rows":restored_rows,"state_hash":restored_context["state_hash"]},backup_name=current_backup.name,actor_ip=client_ip(request),locked=True)
        return {"ok":True,"batch_id":batch_id,"status":"rolled_back","safety_backup":current_backup.name,"restored_rows":restored_rows}
    finally:
        GOVERNANCE_OPERATION_GUARD.release()


@app.get("/api/message-center")
def message_center_list(request: Request, item_type: str = "all"):
    user = require_user(request); stamp = now()
    clauses, params = message_center_visible_clauses(user["id"], stamp, item_type, exclude_deleted=True)
    with connect_db() as db:
        rows = db.execute(f"""select m.*,case when r.item_id is not null and r.read_at>=coalesce(m.updated_at,m.created_at,0) then 1 else 0 end is_read
            from message_center_items m
            left join message_center_reads r on r.item_id=m.id and r.user_id=?
            left join message_center_deletions d on d.item_id=m.id and d.user_id=?
            where {' and '.join(clauses)} order by m.pinned desc,m.created_at desc""", [user["id"], user["id"], *params]).fetchall()
    items = [dict(row) for row in rows]
    return {"ok": True, "items": items, "unread": sum(1 for row in items if not row["is_read"]), "unread_announcements": sum(1 for row in items if row["item_type"] == "announcement" and not row["is_read"]), "unread_notifications": sum(1 for row in items if row["item_type"] == "notification" and not row["is_read"])}


def message_center_visible_clauses(user_id: str, stamp: int, item_type: str = "all", *, exclude_deleted: bool = False) -> tuple[list[str], list[Any]]:
    clauses = ["m.is_active=1", "(m.target_user_id is null or m.target_user_id=?)", "(m.starts_at is null or m.starts_at<=?)", "(m.ends_at is null or m.ends_at>?)"]
    params: list[Any] = [user_id, stamp, stamp]
    if exclude_deleted:
        clauses.append("d.item_id is null")
    if item_type in {"announcement", "notification"}:
        clauses.append("m.item_type=?"); params.append(item_type)
    return clauses, params


@app.post("/api/message-center/{item_id}/read")
def message_center_read(item_id: str, request: Request):
    user = require_user(request); stamp = now()
    clauses, params = message_center_visible_clauses(user["id"], stamp, exclude_deleted=True)
    clauses.insert(0, "m.id=?")
    with connect_db() as db:
        item = db.execute(f"""select m.id,m.target_user_id,m.created_at,m.updated_at from message_center_items m
            left join message_center_deletions d on d.item_id=m.id and d.user_id=?
            where {' and '.join(clauses)}""", [user["id"], item_id, *params]).fetchone()
        if not item or (item["target_user_id"] and item["target_user_id"] != user["id"]): raise HTTPException(status_code=404, detail="找不到訊息。")
        db.execute("insert into message_center_reads(user_id,item_id,read_at) values(?,?,?) on conflict(user_id,item_id) do update set read_at=excluded.read_at", (user["id"], item_id, max(stamp,int(item["updated_at"] or item["created_at"] or 0))))
    return {"ok": True}


@app.post("/api/message-center/read-all")
def message_center_read_all(request: Request, item_type: str = "all"):
    user = require_user(request); stamp = now()
    clauses, params = message_center_visible_clauses(user["id"], stamp, item_type, exclude_deleted=True)
    with connect_db() as db:
        rows = db.execute(f"""select m.id,m.created_at,m.updated_at from message_center_items m
            left join message_center_deletions d on d.item_id=m.id and d.user_id=?
            where {' and '.join(clauses)}""", [user["id"], *params]).fetchall()
        for row in rows: db.execute("insert into message_center_reads(user_id,item_id,read_at) values(?,?,?) on conflict(user_id,item_id) do update set read_at=excluded.read_at", (user["id"], row["id"], max(stamp,int(row["updated_at"] or row["created_at"] or 0))))
    return {"ok": True, "count": len(rows)}


@app.get("/api/admin/dashboard/overview")
def admin_dashboard_overview(request: Request):
    require_admin(request); stamp = now()
    with connect_db() as db:
        game_cards = []
        for project in enabled_projects():
            game_id = str(project.get("id") or "")
            total = int(db.execute("select count(*) from game_catalog_items where game_id=?", (game_id,)).fetchone()[0])
            hidden = int(db.execute("select count(*) from game_catalog_items where game_id=? and hidden=1", (game_id,)).fetchone()[0])
            categories = int(db.execute("select count(distinct category) from game_catalog_items where game_id=?", (game_id,)).fetchone()[0])
            relations = int(db.execute("select count(*) from game_achievement_choice_groups where game_id=?", (game_id,)).fetchone()[0])
            progress = int(db.execute("select count(*) from game_progress where game_id=?", (game_id,)).fetchone()[0])
            pending = int(db.execute("select count(*) from achievement_issues where game_id=? and state in ('new','waiting_review','assigned','ready','reopened','failed')", (game_id,)).fetchone()[0])
            scan = db.execute("select * from achievement_scan_runs where game_id=? order by started_at desc limit 1", (game_id,)).fetchone()
            preview = int(db.execute("select count(*) from game_sync_previews where game_id=? and expires_at>?", (game_id, stamp)).fetchone()[0])
            latest_sync = db.execute("select id,status,source_id,source_mode,summary_json,created_at,rolled_back_at from source_sync_history where game_id=? order by created_at desc limit 1", (game_id,)).fetchone()
            latest_source_test = db.execute("select status,created_at,error_message from admin_audit_logs where game_id=? and action='test_official_source' order by created_at desc limit 1", (game_id,)).fetchone()
            preview_row = db.execute("select diff_json,created_at,expires_at from game_sync_previews where game_id=? and expires_at>? order by created_at desc limit 1", (game_id, stamp)).fetchone()
            severity_rows = db.execute("select severity,count(*) count from achievement_issues where game_id=? and state in ('new','waiting_review','assigned','ready','reopened','failed') group by severity", (game_id,)).fetchall()
            severity = {str(row["severity"] or "unknown"): int(row["count"] or 0) for row in severity_rows}
            last_activity = max(int(latest_sync["created_at"] or 0) if latest_sync else 0, int(latest_source_test["created_at"] or 0) if latest_source_test else 0)
            freshness = "never" if not last_activity else "stale" if stamp-last_activity>14*86400 else "warning" if stamp-last_activity>7*86400 else "current"
            game_cards.append({"id": game_id, "name": project.get("name") or game_id, "achievement_count": total, "hidden_count": hidden, "category_count": categories, "relation_count": relations, "progress_count": progress, "pending_issue_count": pending, "pending_issue_severity": severity, "pending_sync_preview_count": preview, "active_preview_summary": _json_object(preview_row["diff_json"], {}).get("summary", {}) if preview_row else {}, "last_scan": {**dict(scan), "summary": _json_object(scan["summary_json"], {})} if scan else None, "latest_sync": {**dict(latest_sync), "summary": _json_object(latest_sync["summary_json"], {})} if latest_sync else None, "latest_source_test": dict(latest_source_test) if latest_source_test else None, "source_freshness": freshness, "source_last_activity_at": last_activity or None, "source_policy": get_source_policy(game_id)})
        pending_reports = int(db.execute("select count(*) from game_achievement_reports where status in ('open','reviewing')").fetchone()[0])
        pending_tickets = int(db.execute("select count(*) from support_tickets where status in ('open','pending','reviewing')").fetchone()[0])
        unread_announcements = int(db.execute("select count(*) from message_center_items where item_type='announcement' and is_active=1 and (starts_at is null or starts_at<=?) and (ends_at is null or ends_at>?)", (stamp, stamp)).fetchone()[0])
        redeem_row = db.execute("""select count(*) total,
            sum(case when enabled=1 and (end_at is null or end_at>=?) then 1 else 0 end) open,
            sum(case when enabled=1 and end_at is not null and end_at<? then 1 else 0 end) expired,
            sum(case when enabled=1 and end_at>=? and end_at<=? then 1 else 0 end) expiring,
            sum(case when trim(coalesce(source,''))='' and trim(coalesce(description,''))='' then 1 else 0 end) missing_source,
            sum(case when trim(coalesce(reward,''))='' then 1 else 0 end) missing_reward,
            sum(case when trim(coalesce(redeem_url,''))='' then 1 else 0 end) missing_url,
            sum(case when end_at is null then 1 else 0 end) missing_end
            from redeem_codes""", (stamp, stamp, stamp, stamp+3*86400)).fetchone()
        redeem_summary = {key: int(redeem_row[key] or 0) for key in redeem_row.keys()}
        redeem_summary["pending_import_batches"] = int(db.execute("select count(*) from redeem_import_batches where status='preview_ready'").fetchone()[0])
        recent_logs = []
        for row in db.execute("select event_id,actor_email_snapshot,action,category,status,game_id,target_type,target_id,summary,created_at from admin_audit_logs order by created_at desc limit 8").fetchall():
            item = dict(row)
            item["summary"] = sanitize_legacy_id_display(str(item.get("summary") or ""))
            item["target_id"] = sanitize_legacy_id_display(str(item.get("target_id") or ""))
            target_game = str(item.get("game_id") or "")
            target_id = str(item.get("target_id") or "")
            if target_game in {"wuwa", "hsr", "genshin", "zzz"} and target_id.isdigit():
                achievement = db.execute(
                    "select name from game_catalog_items where game_id=? and achievement_id=?",
                    (target_game, target_id),
                ).fetchone()
                if achievement and not item["summary"]:
                    item["summary"] = f"{achievement['name']}（{target_id}）"
            recent_logs.append(item)
        latest_validation = db.execute("select * from achievement_scan_runs order by started_at desc limit 1").fetchone()
    health_status = "ok"
    health_messages: list[str] = []
    try:
        with connect_db() as db:
            integrity = str(db.execute("pragma quick_check").fetchone()[0])
            if integrity != "ok": health_status = "error"; health_messages.append(f"資料庫：{integrity}")
    except Exception as exc:
        health_status = "error"; health_messages.append(str(exc))
    if any(card["pending_issue_count"] for card in game_cards) and health_status == "ok":
        health_status = "warning"; health_messages.append("存在待處理的成就資料問題")
    return {"ok": True, "health_summary": {"status": health_status, "message": "；".join(health_messages) or "核心檢查正常", "checked_at": stamp}, "games": game_cards, "redeem": redeem_summary, "work": {"pending_reports": pending_reports, "pending_tickets": pending_tickets, "active_announcements": unread_announcements, "pending_issues": sum(card["pending_issue_count"] for card in game_cards), "pending_sync_previews": sum(card["pending_sync_preview_count"] for card in game_cards), "pending_redeem_imports": redeem_summary["pending_import_batches"]}, "recent_logs": recent_logs, "latest_scan": dict(latest_validation) if latest_validation else None}


@app.get("/api/admin/system-health/detail")
def admin_system_health_detail(request: Request):
    require_admin(request); stamp = now(); checks: list[dict[str, Any]] = []
    def add(name: str, status: str, message: str, details: Any = None):
        checks.append({"name": name, "status": status, "message": message, "details": details})
    journal=""; checkpoint: dict[str,Any]={}
    try:
        with connect_db() as db:
            integrity = str(db.execute("pragma integrity_check").fetchone()[0]); add("SQLite 完整性", "ok" if integrity == "ok" else "error", integrity)
            journal = str(db.execute("pragma journal_mode").fetchone()[0]); add("資料庫日誌模式", "ok" if journal.casefold() == "wal" else "warning", journal)
            migrations = int(db.execute("select count(*) from schema_migrations").fetchone()[0]); add("資料庫遷移", "ok", f"已套用 {migrations} 筆遷移標記")
            try:
                cp=db.execute("pragma wal_checkpoint(passive)").fetchone()
                if cp:
                    checkpoint={"busy":int(cp[0] or 0),"log_frames":int(cp[1] or 0),"checkpointed_frames":int(cp[2] or 0)}
            except Exception as exc:
                checkpoint={"error":str(exc)}
    except Exception as exc:
        add("資料庫連線", "error", str(exc))
    add("後端 API", "ok", "目前請求已由後端正常回應", {"backend_running":True,"checked_at":stamp})
    add("資料庫檔案", "ok" if DB_FILE.exists() else "error", str(DB_FILE), {"size_bytes": DB_FILE.stat().st_size if DB_FILE.exists() else 0})
    for suffix in ("-wal", "-shm"):
        path = Path(str(DB_FILE) + suffix)
        exists=path.exists(); size=path.stat().st_size if exists else 0
        modified=int(path.stat().st_mtime) if exists else None
        abnormal=False; reasons=[]
        if suffix=="-wal" and size>64*1024*1024:
            abnormal=True; reasons.append("WAL 檔案超過 64 MB")
        if checkpoint.get("busy",0)>0:
            abnormal=True; reasons.append("checkpoint 正在被資料庫鎖定")
        status="warning" if abnormal else "ok"
        message="；".join(reasons) if reasons else ("後端執行期間存在屬正常現象，無須處理" if exists else "目前不存在，屬正常狀態")
        add(f"SQLite {suffix[1:].upper()}",status,message,{"path":str(path),"exists":exists,"size_bytes":size,"modified_at":modified,"backend_running":True,"journal_mode":journal,"checkpoint":checkpoint,"action_required":abnormal})
    required = [HUB_INDEX, ADMIN_INDEX, game_catalog_file("wuwa"), game_catalog_file("hsr"), game_catalog_file("genshin"), game_catalog_file("zzz")]
    missing = [str(path) for path in required if not path.exists()]; add("必要檔案", "ok" if not missing else "error", "必要檔案完整" if not missing else f"缺少 {len(missing)} 個檔案", missing)
    backup_files = sorted((ROOT / "backups").glob("*.db"), key=lambda p: p.stat().st_mtime, reverse=True)
    add("資料庫備份", "ok" if backup_files else "warning", f"共 {len(backup_files)} 份", {"latest": str(backup_files[0]) if backup_files else ""})
    log_size = sum(path.stat().st_size for path in LOG_DIR.rglob("*") if path.is_file()) if LOG_DIR.exists() else 0
    add("日誌空間", "ok", f"{log_size} bytes", {"size_bytes": log_size})
    disk = shutil.disk_usage(ROOT); add("磁碟空間", "ok" if disk.free > 512 * 1024 * 1024 else "warning", f"可用 {disk.free} bytes", {"total": disk.total, "used": disk.used, "free": disk.free})
    overall = "error" if any(row["status"] == "error" for row in checks) else "warning" if any(row["status"] == "warning" for row in checks) else "ok"
    return {"ok": True, "status": overall, "checked_at": stamp, "checks": checks}



@app.get("/api/game-projects")
def game_projects():
    return load_registry()


@app.get("/api/home-game-icons")
def home_game_icons():
    games_dir = SITE_DIR / "assets" / "games"
    icons = []
    if games_dir.exists():
        for folder in sorted(games_dir.iterdir(), key=lambda path: path.name.casefold()):
            if not folder.is_dir() or not re.fullmatch(r"[A-Za-z0-9_-]+", folder.name):
                continue
            icon = folder / "icon.png"
            if icon.is_file():
                icons.append({
                    "game_id": folder.name,
                    "url": f"/assets/games/{folder.name}/icon.png",
                })
    return {"ok": True, "icons": icons, "count": len(icons)}

@app.get("/game-manifest.json")
def public_game_manifest():
    manifest = SITE_DIR / "game-manifest.json"
    if not manifest.is_file():
        raise HTTPException(status_code=404, detail="game manifest not found")
    return FileResponse(manifest, media_type="application/json", headers={"Cache-Control": "no-store"})


@app.get("/assets/{asset_path:path}")
def public_site_asset(asset_path: str):
    assets_root = (SITE_DIR / "assets").resolve()
    candidate = (assets_root / asset_path).resolve()
    if assets_root not in candidate.parents or not candidate.is_file():
        raise HTTPException(status_code=404, detail="asset not found")
    return FileResponse(candidate)


@app.get("/robots.txt")
def public_robots():
    path = SITE_DIR / "robots.txt"
    if not path.is_file():
        raise HTTPException(status_code=404, detail="robots.txt not found")
    return FileResponse(path, media_type="text/plain")


@app.get("/sitemap.xml")
def public_sitemap():
    path = SITE_DIR / "sitemap.xml"
    if not path.is_file():
        raise HTTPException(status_code=404, detail="sitemap.xml not found")
    return FileResponse(path, media_type="application/xml")


def _html_no_store_response(index_file: Path):
    return FileResponse(index_file,headers={
        "Cache-Control":"no-store, no-cache, must-revalidate, max-age=0",
        "Pragma":"no-cache",
        "Expires":"0",
    })

def _game_project_response(game_id: str):
    index_file=resolve_game_index(game_id)
    if not index_file:
        raise HTTPException(status_code=404,detail="找不到或尚未啟用此遊戲專案。")
    return _html_no_store_response(index_file)

@app.get("/_projects/account")
def account_project():
    if not ACCOUNT_INDEX.exists():
        raise HTTPException(status_code=404, detail="找不到帳號中心專案。")
    return _html_no_store_response(ACCOUNT_INDEX)

@app.get("/_projects/account/")
def account_project_slash():
    return account_project()

@app.get("/_projects/account/index.html")
def account_project_index():
    return account_project()

@app.get("/_projects/admin")
def admin_project():
    if not ADMIN_INDEX.exists():
        raise HTTPException(status_code=404, detail="找不到後台管理專案。")
    return _html_no_store_response(ADMIN_INDEX)

@app.get("/_projects/admin/")
def admin_project_slash():
    return admin_project()

@app.get("/_projects/admin/index.html")
def admin_project_index():
    return admin_project()

def redeem_project():
    index_file = PROJECTS_DIR / "redeem" / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="找不到兌換碼中心專案。")
    return _html_no_store_response(index_file)

@app.get("/_projects/redeem")
def redeem_project_page():
    return redeem_project()

@app.get("/_projects/redeem/")
def redeem_project_slash():
    return redeem_project()

@app.get("/_projects/redeem/index.html")
def redeem_project_index():
    return redeem_project()


def home_project():
    index_file = PROJECTS_DIR / "home" / "index.html"
    if not index_file.exists():
        raise HTTPException(status_code=404, detail="找不到主頁。")
    return _html_no_store_response(index_file)


@app.get("/_projects/home")
def home_project_page():
    return home_project()


@app.get("/_projects/home/")
def home_project_slash():
    return home_project()


@app.get("/_projects/home/index.html")
def home_project_index():
    return home_project()

@app.get("/_projects/{game_id}")
def internal_game_project(game_id: str):
    return _game_project_response(game_id)

@app.get("/_projects/{game_id}/")
def internal_game_project_slash(game_id: str):
    return _game_project_response(game_id)

@app.get("/_projects/{game_id}/index.html")
def internal_game_project_index(game_id: str):
    return _game_project_response(game_id)

# 舊版內部網址相容：既有書籤與快取仍可載入遊戲專案。
@app.get("/account/index.html")
def legacy_account_project_index():
    return account_project()

@app.get("/games/{game_id}")
def legacy_game_project(game_id: str):
    return _game_project_response(game_id)

@app.get("/games/{game_id}/")
def legacy_game_project_slash(game_id: str):
    return _game_project_response(game_id)

@app.get("/games/{game_id}/index.html")
def legacy_game_project_index(game_id: str):
    return _game_project_response(game_id)


def _hub_response():
    if not HUB_INDEX.exists():
        raise HTTPException(status_code=404, detail="找不到遊戲成就紀錄器首頁。")
    return _html_no_store_response(HUB_INDEX)

@app.get("/")
def index():
    return _hub_response()

@app.get("/wuwa")
@app.get("/wuwa/")
def wuwa_page():
    return _hub_response()

@app.get("/hsr")
@app.get("/hsr/")
def hsr_page():
    return _hub_response()

@app.get("/genshin")
@app.get("/genshin/")
def genshin_page():
    return _hub_response()

@app.get("/zzz")
@app.get("/zzz/")
def zzz_page():
    return _hub_response()

@app.get("/hna")
@app.get("/hna/")
def hna_page():
    return _hub_response()

@app.get("/redeem")
@app.get("/redeem/")
def redeem_page():
    return _hub_response()

@app.get("/account")
@app.get("/account/")
def account_page():
    return _hub_response()

@app.get("/admin")
@app.get("/admin/")
def admin_page():
    return _hub_response()
